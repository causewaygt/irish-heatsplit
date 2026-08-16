#!/usr/bin/env python3
"""
Convert EirGrid's half-hourly wind dispatch-down files into the compact
monthly series the site ships.

    python3 tools/dd_convert.py DD-HH-*.xlsx --prices prices.csv

Download the files from the "DD Half-Hourly Data" section of
https://www.eirgrid.ie/grid/system-and-renewable-data-reports - they sit
behind a JavaScript accordion, so they cannot be fetched by URL pattern:
the version suffix changes without notice (V7, v10) and a guessed URL
would rot silently rather than fail loudly. Closed years never change;
re-run this only when the current year's file is refreshed.

Wind only. Solar coverage does not start until 2023 and solar is about a
tenth of the volume, so including it would give a series whose
denominator changes shape midway.

With --prices, an hourly or half-hourly SEM day-ahead series (UTC
stamps, "Timestamp_UTC,Price"), it also emits the VOLUME-WEIGHTED price
in the half-hours each reason was actually spilling, and the plain
average across all periods in the month. Those two are the chart: the
spilled volume is worth what it would have earned in ITS OWN hours, not
what power averaged that month.

THE JOIN IS BY UTC, DERIVED PER ROW. EirGrid stamps are Irish local
clock with an explicit GMT_OFFSET column that changes at the DST
boundary, so a fixed offset misaligns half the year. Verified by
alignment test rather than by trusting either file's labels: one
published price series was found to be a full day out, and the error
was invisible because it pulled the weighted figure toward the mean,
which looks like a plausible answer rather than a broken one.
"""
import collections
import csv
import datetime as dt
import json
import sys

import openpyxl

COLS = {"avail": 4, "output": 5, "hifrq": 6, "rocof": 7, "snsp": 8,
        "trans": 9, "test": 10, "dd": 11, "curt": 12, "cons": 13,
        "other": 14}
OUT = "docs/dispatch_down_monthly.json"


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_prices(path):
    """UTC-stamped price series -> {datetime: EUR/MWh}."""
    out = {}
    with open(path) as fh:
        for row in csv.DictReader(fh):
            stamp = (row.get("Timestamp_UTC") or "").strip()
            try:
                out[dt.datetime.strptime(stamp[:16], "%Y-%m-%d %H:%M")] = \
                    float(row["Price"])
            except (ValueError, KeyError, TypeError):
                continue
    if not out:
        raise SystemExit("no usable rows in " + path)
    return out


def main(paths, prices=None):
    agg = collections.defaultdict(lambda: collections.defaultdict(float))
    pnum = collections.defaultdict(float)
    pden = collections.defaultdict(float)
    seen = collections.defaultdict(dict)
    gaps = collections.Counter()
    for path in sorted(paths):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        # The schema has held across every year and version to date. If
        # it ever moves, fail here rather than write silent nonsense.
        assert header[0] == "UT_TYPE" and header[2] == "HH_TIMESTAMP" \
            and header[9] == "Sum of TRANS_CONSTR_MWH", (path, header)
        n = 0
        for r in rows:
            if not r[0] or r[0] != "Wind":
                continue
            month = str(r[2])[:7]
            key = (month, r[1])
            for name, i in COLS.items():
                agg[key][name] += num(r[i])
            n += 1
            if prices is None:
                continue
            # UTC derived PER ROW from the file's own offset column
            local = dt.datetime.strptime(str(r[2])[:19], "%Y-%m-%d %H:%M:%S")
            utc = local - dt.timedelta(hours=int(num(r[3])))
            p = prices.get(utc)
            if p is None:
                p = prices.get(utc.replace(minute=0))
            if p is None:
                gaps[month] += 1
                continue
            seen[month][utc] = p
            for name in ("dd", "cons", "curt", "trans", "hifrq", "snsp"):
                v = num(r[COLS[name]])
                if v > 0:
                    pnum[(month, r[1], name)] += v * p
                    pden[(month, r[1], name)] += v
        wb.close()
        print(f"{path}: {n} wind rows")
    months = sorted({k[0] for k in agg})
    out = {"months": months, "unit": "GWh", "technology": "Wind",
           "jurisdictions": {}}
    for j in ("IE", "NI"):
        block = {name: [round(agg[(m, j)][name] / 1000.0, 2) for m in months]
                 for name in COLS}
        if prices is not None:
            for name in ("dd", "cons", "curt", "trans", "hifrq", "snsp"):
                block["price_" + name] = [
                    round(pnum[(m, j, name)] / pden[(m, j, name)], 2)
                    if pden[(m, j, name)] > 0 else None for m in months]
        out["jurisdictions"][j] = block
    if prices is not None:
        # The comparator: every period in the month, unweighted. What
        # the spilled energy would have been worth if it had been
        # spilled at random rather than when the wind blows.
        out["price_month_mean"] = [
            round(sum(seen[m].values()) / len(seen[m]), 2) if seen[m] else None
            for m in months]
        if gaps:
            print("price gaps by month: "
                  + ", ".join(f"{m} {n}" for m, n in sorted(gaps.items())[:6]))
    with open(OUT, "w") as fh:
        json.dump(out, fh, separators=(",", ":"))
    print(f"wrote {OUT}: {len(months)} months, {months[0]}..{months[-1]}")


if __name__ == "__main__":
    args = sys.argv[1:]
    prices = None
    if "--prices" in args:
        i = args.index("--prices")
        prices = load_prices(args[i + 1])
        del args[i:i + 2]
    if not args:
        sys.exit(__doc__)
    main(args, prices)
