#!/usr/bin/env python3
"""
Irish Heat Split - daily build pipeline. PIPELINE_VERSION 1.0.0.

Changelog:
  1.0.0 - launch. Geothermal register + WGC2026 ROI anchors + per-capita
          what-if derivation with European reference points; per-
          jurisdiction hero; climate-modelled ASHP SPF; cool-side panel;
          events, flags and soft-feed registers.
  0.x   - see below.
  0.5.0 - ONS GB heating-oil feed (kj5u - same-tax control for the NI-GB
          market-structure gap); EVENTS register rendered as chart
          annotations; FEED_FLAGS register for value-level caveats distinct
          from fetch status; tariff anchors re-based on the July 2026
          sourced pass (Power NI/UR review, ROI standard rates);
          derive_heat_gap() - cost of useful heat by route per jurisdiction
          with break-even SPF vs the incumbent oil boiler.
  0.4.0 - oil bulletin fetches with- AND without-taxes files; ex-tax series.
  0.3.0 - gni_live implemented against the probed gasconsumption JSON API.
  0.2.0 - ANCHORS + derive_hero() weekly four-stat and geothermal what-if.
  0.1.x - scaffold, feed fixes from first-run logs.

House rules: self-resolve IDs/links at runtime; dump available names on
failure; every feed try/except with previous values retained and status
"stale"; fetch health vs data recency tracked separately; unit
autodetection; clip future-dated rows; en dashes in user-facing strings.
"""

import datetime as dt
import html as html_mod
import io
import json
import math
import random
import re
import statistics
import sys
import time
import traceback
import urllib.parse
from pathlib import Path
from xml.etree import ElementTree

import requests

# ---------------------------------------------------------------- constants

# VERSIONING, and a temporary rule. Normally x.y.z means x a new
# source or panel, y a source update, z wording or format. While the
# site is under construction the x is FROZEN at 5 and only y and z
# move - the panels are changing weekly and an x that tracked every
# new one would say nothing. The "Under Construction" label on the
# masthead and this freeze come off together.
PIPELINE_VERSION = "6.0.0"
# 5.25.0: THE DEMAND SERIES DEFINITION, WRITTEN DOWN AND ENFORCED.
#   EirGrid's demandactual is "the electricity production required to
#   meet national electricity consumption" - so grid-connected solar
#   is ALREADY INSIDE it - and solaractual is "large scale solar farms
#   ... small scale embedded solar is not included". Adding one to the
#   other is a double count, NOT a reconstruction.
#   The UK sibling's NESO series is the opposite case: its demand
#   excludes embedded generation and must have it added back once.
#   Porting that reasoning across manufactured a temperature signal in
#   exactly the daylight hours where cooling would appear. It cost the
#   UK three published claims and cost this site a withdrawn cooling
#   analysis - a diurnal comfort/process separation that was entirely
#   artefact.
#   THE IRISH PROBLEM IS THE REVERSE AND HAS NO FIX IN THIS DATA:
#   ~310 MW of small-scale embedded solar IS invisible to demandactual
#   and no published series exists to add back, so Irish underlying
#   demand is understated on bright days by an unknown and growing
#   amount. Daylight cooling estimates are biased DOWN. Night-time
#   estimators are unaffected, solar being zero - which is why the
#   weekday/weekend night placebo is the route that survives.
# 5.23.0: WORKED EXAMPLES of absorbing the constrained wind. A
#   different kind of claim from every other panel - it sizes the SINK
#   ("what scale of load would it take") rather than claiming a
#   benefit, because sizing survives the coincidence objection and a
#   benefit claim does not. Half the spill falls outside the heating
#   season and only 44% between midnight and six.
#   THE RESULT IS THE POINT AND IT IS NOT THE ONE EXPECTED: NI's
#   constrained wind in 2025 was 563 GWh, or 2,392 GWh of heat at
#   network SPF - 22% of all the building heat NI uses. That needs 344
#   hospitals of 7.0 GWh against an acute estate of about ten. So
#   institutional anchor loads CANNOT absorb this volume, which is why
#   the published answer is 250,000 households. The hospital is the
#   demonstrator; the aggregation is the scale.
#   Domestic figures are Agbonaye, Keatley, Huang, Odiase & Hewitt
#   (2022) Renewable Energy 190:487-500, quoted not re-derived - Prof
#   Hewitt is a named peer reviewer for this work.
# 5.22.0: WHAT THE SPILLED ENERGY WAS WORTH. dd_convert.py now takes
#   --prices, an hourly SEM day-ahead series, and emits the VOLUME-
#   WEIGHTED price in the half-hours each reason was actually spilling
#   alongside the month's plain average. The join derives UTC PER ROW
#   from the file's own GMT_OFFSET column.
#   THE RESULT IS THE POINT: spilled wind clears at roughly half the
#   average price, because it is spilled when the wind blows and power
#   is cheap. Island total since 2021 is EUR 721m at the prices of its
#   own hours against EUR 1,218m at monthly averages - the naive figure
#   overstates by 40%. Constraint hours clear ~40% above curtailment
#   hours, so the volume a local heat load can absorb is also the more
#   valuable volume.
#   NOT A PAYMENT. Constrained wind with firm access is already
#   compensated, so absorbing it saves the system operator and
#   consumers; curtailed wind is not, so absorbing it is revenue the
#   generator keeps. The panel is titled "what the energy was worth"
#   for that reason and carries the split in its note.
# 5.21.0: WIND DISPATCH-DOWN, monthly by jurisdiction and REASON, from
#   EirGrid's own half-hourly DD files, 2021 to date. Shipped as a
#   static docs/dispatch_down_monthly.json rather than fetched: the
#   downloads sit behind a JavaScript accordion and carry version
#   suffixes that change without notice (V7, v10), so a guessed URL
#   would rot silently. tools/dd_convert.py regenerates it and asserts
#   the schema rather than trusting it. Closed years never change.
#   Stacked by reason, not by the constraint/curtailment fold, because
#   the fold hides the finding: NI spills 22-30% of its wind and ~85%
#   of that is transmission constraint - the local kind a local heat
#   load can absorb - while the Republic runs 10-13% at about half.
#   COLUMN RELATIONSHIP, corrected: DD = CURTAILMENTS + CONSTRAINTS
#   exactly, and OTHER sits OUTSIDE dispatch-down (DSO/DNO, developer
#   outage and testing). I first read 1,314 non-reconciling rows as a
#   data quirk; the formula was mine and wrong.
# 5.20.1: THE GRID LAYER NEVER SHIPPED. data.json is serialised before
#   the hourly block runs, so derived["tightest_hour"] and
#   derived["grid_views"] were assigned to a dict already written to
#   disk. Panel 3 drew its decline messages for two bundles while the
#   log showed B.2.1 and the falcon computing perfectly - the
#   renderers were right and the payload was empty. The tell was there
#   and missed: data.json stayed at 1,436 kB when the grid views
#   should have added ~34 kB. main() now writes AGAIN once the grid
#   keys exist, and says in the log when they are absent. The first
#   write stays put so a payload still lands if the hourly step
#   throws.
# 5.20.0: THE FALCON, and a correction. I said it needed two winters
#   and deferred it; it does not. The UK sibling builds it as a
#   CALENDAR YEAR with each month filled by the LATEST COMPLETE
#   instance of that month - Jan to Jul from this year, Aug to Dec
#   from last - so twelve complete months is enough and a 13-month
#   store already carries a full one. derive_grid_views publishes
#   `falcon`, twelve rows ordered by calendar month with the month
#   each came from, plus falcon_complete so a short store draws a
#   partial curve rather than a wrong one.
# 5.19.0: derive_grid_views publishes the three series Panel 3 plots -
#   168 hourly rows, 90 daily, up to 24 monthly, about 280 against the
#   store's 9,000-plus hours, so the payload carries what is DRAWN
#   rather than what was computed (~34 kB). Each row is the island's
#   delivered heat and the electricity the site's own 20% what-if
#   would draw by route, netted of the resistive heating already in
#   observed demand, on the SAME hourly air COP the binding-hour panel
#   uses.
#   TWO WHAT-IFS, KEPT APART: the binding-hour panel SOLVES for the
#   share that fits; these views plot the FIXED 20% the rest of the
#   site uses. A test asserts the solve is not pinned to the constant,
#   because reading them as one number would be the easy mistake.
# 5.18.0: "What heat emits" - gCO2e per USEFUL kWh by route, the
#   Irish answer to the UK sibling's sub-panel, under panel 2 with its
#   own method fold. ALL-ISLAND and deliberately without a
#   jurisdiction toggle: the combustion factors do not change at the
#   border, the efficiencies are shared and the grid is a single
#   all-island market, so a split would draw the same bars three
#   times. That is the finding rather than a gap - the PRICE answer
#   differs sharply across the border and the CARBON answer does not
#   at all. The network route's jurisdictional SPFs (5.0 NI, 4.0 ROI)
#   combine as a heat-weighted HARMONIC mean, which is how
#   efficiencies average; the arithmetic mean would flatter it.
# 5.17.0: the calibration is PUBLISHED, not just logged. CALIBRATION
#   carries each route's solved Carnot fraction beside the SPF anchor
#   it was solved to reproduce and the source temperature it saw, per
#   jurisdiction, with the spread and the gate. It rides in
#   derived["calibration"] and the site draws it in panel 1's method
#   fold. A fraction without its anchor means nothing, so the two
#   travel together. This is the exhibit that answers "how do you know
#   the COP model is right" - it belongs on the page rather than in a
#   run log nobody reads.
# 5.16.0: two places where the code did not say what it does.
#   - build_history offered a LITERAL 60 weeks while HISTORY_MAX was
#     120, and the cap was applied afterwards to a list that could
#     never exceed 60. The record sat at 60, the log line "60 weeks
#     built, none skipped" reported the loop bound rather than any
#     data limit, and panel 1's sparklines were short because they
#     slice whatever the record holds. Now driven by HISTORY_MAX, with
#     a test that fails if a literal comes back.
#   - the carbon-reach diagnostic sat INSIDE the backfill branch, so
#     on any run where the backfill did not fire it said nothing - and
#     nothing is indistinguishable from "the block never ran", which
#     is the one case the line exists for. It now reports every run.
# 5.15.0: the masthead ticker now comes from the SAME engine as panel
#   2. derive_heat_gap is the original calculation and never got the
#   panel's changes - one geothermal SPF of 4.0 for both jurisdictions
#   against the panel's 5.0/4.0, a single oil-boiler efficiency of
#   0.82 against 0.71 on hot water, and no hot-water blending at all -
#   so the headline disagreed with the chart underneath it by 13-20%
#   on identical routes. It is kept as heat_gap_diagnostic for its
#   breakeven-SPF figures and its tests; the ticker reads
#   heat_gap_from_cost_series. That also carries the gap's MEDIAN over
#   the window, because a spot gap reads as a standing fact and on
#   this record oil climbs steeply from Nov 2025.
# 5.14.0: THE CALIBRATION WAS DAY-WEIGHTED, NOT HEAT-WEIGHTED.
#   calibrate_eta's docstring has said heat-weighted since it was
#   written, and the caller passed the day's space/hot-water SHARES -
#   two numbers summing to 1.0 on every day - so a mild August day
#   weighed exactly as much as a January one. That flatters air
#   source, whose advantage is concentrated in mild weather, and it
#   was most of the calibration spread the 15% gate has been firing
#   on. The caller now passes day_delivered_heat's actual GWh.
#   Comparison with the UK sibling is what surfaced it: the UK's three
#   fractions agree to 1.35% on the same constants and the same
#   anchors, so the difference had to be in the weighting rather than
#   in the physics.
# 5.12.0: THE BACK-LOOK REACHES APRIL 2024. Both sides of the tariff
#   table move together, because they had to - the sterling side is a
#   dated table but the euro side was derived at call time from a
#   single S2 2025 anchor, so extending one alone would have priced
#   every earlier ROI day at the 2025 level. Silently.
#   - three sterling rows from UREGNI's tariff-review releases, which
#     publish the annual bill at this site's own consumption basis;
#   - IE_DOMESTIC_SEMESTER, the ROI domestic band series by semester
#     (band DC electricity, band D2 gas), replacing the derived euro
#     side. ie_domestic_eur refuses below the first semester rather
#     than clamping, the same rule as tariffs_for and nondom_for.
#   - electricity is CREDIT-FREE: Ireland credited domestic
#     electricity accounts €1,500 in lump sums per meter, which SEAI
#     books into the effective unit price. A lump sum never changed
#     the cost of the next kWh, so it is added back. Gas accounts
#     were never eligible and need no correction.
#   The divisor for that add-back is a JUDGEMENT (dagger) - three
#   controls bracket it 1,500-2,000 kWh a semester and the run logs
#   the sensitivity every time rather than burying it.
# 5.11.0: NI oil is STEP-HELD across its week, capped at
#   NI_OIL_HOLD_DAYS. 5.10.0 gave the series depth back to 2023 but
#   priced only the survey days, so behind the daily checker's start
#   the NI line was one day in seven - 144 of 375 on the first live
#   run against ROI's 375. That reads as missing data rather than as a
#   weekly survey. Held forward it is the same treatment, and the same
#   claim, as the ROI bulletin week has always had. The cap stops one
#   reading being smeared across the archive's real gaps.
# 5.10.0: the CCNI WEEKLY ARCHIVE joins the daily checker as a second
#   source for NI oil. The archive page embeds a chart array of the
#   same shape as the daily page, so it parses through the existing
#   extract_chart_data_arrays / parse_ccni_series with no new parser -
#   277 points to April 2021 against the daily chart's few months.
#   This is what lets NI oil reach a 24-month window. The archive
#   fetch is SOFT: a failure loses reach, not the run. Series merge
#   into one dict (every consumer unchanged) with the daily reading
#   winning any overlap, provenance recorded in daily_page_days, and
#   the overlap disagreement logged every run rather than averaged.
#   ccni_ratio_gate names rows whose litre ratios cannot be right -
#   three exist in CCNI's own record and only one is visible in the
#   900 L series the site prices on.
# 5.8.0: heat_cost_series rows carry vol_roi and vol_ni - GWh of
#   DELIVERED heat that day, split space / hot water - so the cost
#   panel can draw the quantity its per-MWh axis is charged on. Shaped
#   identically to day_dhw_share (hot water flat, space heat by the
#   day's share of the trailing year's degree days) and scaled by the
#   sector anchors converted from fuel input to delivered heat at each
#   jurisdiction's own fuel mix, the same conversion hourly_heat_mw()
#   uses. Shape is island-wide, scale is jurisdictional. Additive: a
#   front end that has not been updated ignores the fields.
# 5.9.0: three guards and an encoding, all ahead of the back-look
#   extension rather than after it.
#   - nondom_for REFUSES below the first published REMM semester
#     instead of clamping to it. Clamping was safe only while nothing
#     reached back that far; extending the window is what removes that
#     safety. Callers decline the day or the week, by name.
#   - retention_span_gate asserts the retained record covers the
#     widest window plus its trailing year. Margin is 55 days.
#   - heat_cost_series is written columnar, same wire format and same
#     encoder as the history block. Measured 32% of flat.
#   - WINDOW_MAX_DAYS records the widest window the site offers; the
#     60-month window was withdrawn at site 5.6.0 because HISTORY_MAX
#     is 120 weeks and it could never fill.
ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "docs" / "data.json"
# The hourly store lives in its OWN file: a malformed hourly write can
# never corrupt the weekly tracker's data, the grid panel can be absent
# while everything else renders, and the store versions independently
# of HISTORY_SCHEMA / ANCHOR_EPOCH. Fetched by the page only when the
# grid panel needs it.
HOURLY_PATH = ROOT / "docs" / "hourly.json"
# 2 (7 Aug 2026): temp_ai added - population-weighted island air
# temperature per hour, from Open-Meteo, on the SAME local clock as
# the EirGrid series. Without it the grid layer cannot be computed at
# all: hourly heat needs an hourly temperature, and the only hourly
# temperature this pipeline previously touched was a trailing 60 days
# reduced to daily ODH26 before anything was persisted. The store
# holds temperature rather than degree hours so that hourly HDD (base
# 15.5), ODH26 and the Carnot source temperature all derive from one
# retained series instead of three.
# 3 (7 Aug 2026): the store is written as flat arrays against a base
# hour instead of one key per value. At schema 2 the file reached
# 1,025 kB rewritten daily, and the repeated 13-character keys were
# most of it. Readers must tolerate BOTH shapes - expand_hourly()
# accepts a schema 1/2 document so the first run after this change
# inherits its own history rather than refilling from empty.
# 4 (8 Aug 2026): price_ai added - SEMOpx day-ahead, EUR/MWh, for
# B.2.3 (the coincidence premium). Fills FORWARD only; see
# feed_semopx and semopx_history_probe for why.
HOURLY_SCHEMA = 4
HOURLY_MONTHS = 13          # rolling window the store keeps
HOURLY_CHUNKS_PER_RUN = 16  # walk-back budget; 14 fills 13 months
# Open-Meteo serves long hourly spans in one request, so the
# temperature series walks in far bigger chunks than EirGrid's
# ~28-day window; 4 x 120 days covers 13 months with room over.
HOURLY_TEMP_CHUNK_DAYS = 120
HOURLY_TEMP_CHUNKS_PER_RUN = 5
# Raised 400 -> 1150 (27 Jul 2026) so the back-look can reach the
# tariff-confidence floor with a full trailing HDD year beneath its
# earliest week.
SERIES_KEEP_DAYS = 1150
# Earliest week the back-look will price: all four tariff anchors are
# verified from this date (Power NI +4% and Firmus -7.86% both
# effective 1 Oct 2025; Electric Ireland electricity frozen through
# the winter, gas -4% from Sep 2025).
# 8 Aug 2026: extended BACK eight weeks to complete a year on record.
# Why this date is reachable and 1 Oct 2025 was the previous floor:
#  - ROI needs no new anchor at all. Electric Ireland held prices from
#    Oct 2022 to 1 Jul 2026, and SEAI/Eurostat semester S2 2025 runs
#    1 Jul - 31 Dec 2025, so Aug-Sep 2025 is inside the same period
#    the Oct 2025 row already describes.
#  - NI needs one prior row, and both regulated suppliers publish it:
#    SSE's maximum average price was set on 1 Apr 2024 and unchanged
#    until 1 Oct 2025; Power NI's tariff ran from 1 Dec 2024.
#  - Grid carbon for these weeks is NOT in the daily EirGrid feed
#    (50-day retention). It is taken from the hourly store, whose
#    floor advances daily - see daily_ci_from_hourly().
HISTORY_START = "2025-02-02"
# Moved back from 2025-08-06 at 5.13.0, once the tariff table reached
# April 2024. It did NOT move all the way to match: the binding
# constraint on the weekly record is no longer tariffs but each week's
# own grid carbon, and the EirGrid probe's demonstrated reach is 18
# months. Weeks are refused rather than priced at the carbon anchor,
# so the record extends itself as the backfill reaches further and
# never contains a week whose emissions are today's grid intensity
# wearing last year's date.
# The date from which the record has been observed as it happened.
# Weeks before it are reconstructed from published tariffs and the
# hourly store, and are counted and labelled separately - the 52-live
# -weeks milestone is not reached by backfilling.
LIVE_FROM = "2025-10-01"
# Weeks retained. Was 60, which a 52-week record plus forward growth
# would have hit inside two months - silently dropping the weeks
# backfilled here. 120 leaves room for the 24-month view.
HISTORY_MAX = 120
# History schema (splits handover, 1 Aug 2026). Bump triggers a
# one-time restatement: stored weeks recompute through derive_hero
# with their STORED per-entry inputs (ef_electricity injected, oil
# and fx from the retained series) - existing fields must re-round
# identically, only new fields appear. Unlike the UK, the Irish
# recompute window is the price feeds at 1,150-day retention, so all
# weeks are restatable and stranding risk is ~nil - measured, not
# assumed, by the drift check below.
# 4 (7 Aug 2026): forces a re-run of the schema-3 migration, which
# used setdefault() for the ni/roi blocks and therefore left frozen
# weeks with jurisdiction blocks predating the per-fuel addition -
# so the windowed energy bars worked all-island but fell back to the
# live week for NI and ROI.
# 5 (7 Aug 2026): the per-jurisdiction sub-blocks gain the heat/cold
# splits the island entry has carried since schema 2. Without them
# every windowed view under the NI or ROI toggle failed the front
# end's split guard, so the 4w/12w/12m cards and what-if silently
# dropped their "(heat ... cooling ... saves ...)" breakdowns while
# all-island showed them - a jurisdiction-only gap that looked like a
# copy bug and was a missing field.
HISTORY_SCHEMA = 5
# Anchor epoch. Bump whenever a change to ANCHORS alters what a past
# week WOULD have computed - as distinct from HISTORY_SCHEMA, which
# tracks new FIELDS. A schema bump adds fields and leaves stored
# values alone; an epoch bump rewrites every recomputable week onto
# the new basis, using that week's own stored prices, tariffs, fx and
# emission factor. Without it a basis change leaves the series half on
# one footing and half on another, and the heat/cold splits stop
# reconciling with the totals they were derived from.
#   1 - launch basis
#   2 - 6 Aug 2026: data-centre line repriced to its cooling share
#       (SEAI NHS R1) and hot water re-anchored 18.3% -> 22.4%
#   3 - 7 Aug 2026: cooling service factors re-anchored to the SEAI
#       National Heat Study supporting data (useful-to-final ratios)
#   4 - 7 Aug 2026: heat-pump stock split out of the electricity line
#       so the bars can show heat-pump electricity and the ambient heat
#       it harvests, on census anchors
# 5 (8 Aug 2026): NI tariff basis rebuilt (see TARIFF_HISTORY) and
# the back-look floor moved to 2025-08-06.
# 6 (8 Aug 2026): ROI domestic moved onto the same all-in basis as NI
# - Eurostat band prices, standing charges included.
# 7 (8 Aug 2026): Irish anchors moved to semester 2 2025 (credit-free,
# see IE_PUBLISHED_P_PER_KWH) and the euro conversion moved from a
# hard-coded rate to the ECB semester mean computed from the feed.
# 8 (8 Aug 2026): non-domestic rates step by semester instead of being
# held at one, so the services share of each week is priced at the
# semester that week falls in. Every stored week reprices.
ANCHOR_EPOCH = 8
# The heat/cold split fields carried by both the island entry and,
# from history schema 5, each jurisdiction sub-block.
JUR_SPLIT_KEYS = ("heat_gwh", "cold_gwh",
                  "bill_heat_eur_m", "bill_cold_eur_m",
                  "bill_heat_gbp_m", "bill_cold_gbp_m",
                  "emissions_heat_kt", "emissions_cold_kt")

# ---------------------------------------------------------------- Irish
# anchors: published in sterling, converted at a FETCHED rate.
#
# WHY SEMESTER 2 2025 AND NOT AN EARLIER ONE. The Irish domestic
# electricity series carries government credits as negative taxes -
# UREGNI says so explicitly, EUR1,500 of them since 2022 with the last
# EUR125 in Jan/Feb 2025. That is why Ireland reads 31.3 p/kWh in S2
# 2024, 27.5 in S1 2025 and 35.2 in S2 2025: a 28% jump in one
# semester caused by a credit ending, not by a price moving. This site
# prices the REAL cost of heat, so the credit-bearing semesters are
# unusable and S2 2025 - the first clean one - is the anchor. It also
# happens to be the semester the back-look starts in.
#
# WHY THE RATE IS FETCHED. UREGNI publishes Ireland in sterling,
# having converted Eurostat's euro at the semester average. Recovering
# the euro figure therefore needs that same average, and it scales
# EVERY Irish anchor on the site. It was a hard-coded 0.84 until now,
# which was my estimate for the wrong semester - S2 2025 ran nearer
# 0.87, so the estimate was overstating every ROI anchor by about 4%.
# It is now the ECB semester mean, computed from the daily reference
# rates this pipeline already retains and logged on every run.
IE_SEMESTER = "2025-S2"

# ------------------------------------------------- non-domestic by semester
# Published REMM band prices, pence per kWh, incl CCL excl VAT.
# Electricity is consumption-weighted across every band BELOW Large +
# Very Large, using each semester's own published consumption shares -
# services buildings are not the industrial tail, and the shares move
# enough between semesters (very small 6.1 -> 6.9 -> 5.1% of I&C GWh)
# to be worth taking per semester rather than once. Gas is band I1,
# where services buildings sit; the REMM price bands do not map onto
# the network bands the gas consumption split is published in.
#
# WHY THIS STEPS AND DOMESTIC DOES NOT. There are no regulated
# non-domestic announcements to give finer timing, so the semester IS
# the resolution. Domestic keeps dated steps because the regulator and
# the incumbents publish them.
#
# CONVERSION. Each Irish figure converts at the ECB mean for ITS OWN
# semester - UREGNI sterlings Eurostat's euro at that semester's
# average, so that is the rate that recovers the original. Not the
# week's semester: the figure belongs to the semester it was published
# for, whichever week is being priced.
#
# TAIL. REMM lags about nine months, so weeks past the last published
# semester hold at it rather than extrapolating - flagged, not guessed.
#
# Sources: 2024 AREMM (S2 2024), Q3 2025 QREMM (S1 2025), Q1 2026
# QREMM (S2 2025). Domestic credits do not contaminate these: the
# Irish credits were an electricity-only, household-only measure.
NONDOM_SEMESTERS = {
    #             NI elec  NI gas   IE elec  IE gas
    "2024-S2": (23.457,   8.665,   23.727,  8.579),
    "2025-S1": (22.706,   8.576,   23.221,  8.280),
    "2025-S2": (23.808,   7.995,   22.692,  9.345),
}


def _nondom(a, week_ctx):
    """The week's own non-domestic rates if it carries them, else the
    live anchors. Historic weeks carry them so the services share is
    priced at the semester the week falls in, not at today's."""
    nd = (week_ctx or {}).get("nondom")
    return nd if nd else {"eur": a["nondom_eur_per_kwh"],
                          "gbp": a["nondom_gbp_per_kwh"]}


def semester_of(date_iso):
    """Which semester a week belongs to, by WEEK ENDING. A week
    straddling 30 June or 31 December lands wholly in the semester it
    ends in - two weeks a year, disclosed rather than split."""
    return f"{date_iso[:4]}-S{1 if int(date_iso[5:7]) <= 6 else 2}"


def nondom_for(date_iso, fx_by_semester=None):
    """
    Non-domestic rates in force for a week: NI in sterling, ROI in
    euro, converted at each published figure's OWN semester rate.

    Returns (rates, semester_used), or (None, None) for a week before
    the first published semester.

    REFUSE, DO NOT CLAMP. This clamped until 5.9.0, which was safe
    only because nothing reached past 2024-S2. Extending the back-look
    is exactly what makes it unsafe: a clamped week is priced at a
    semester it does not belong to and looks identical to one that
    does. That is the fault already fixed for carbon and for tariffs,
    and it is cheaper to close here while it still cannot fire than
    after something reaches through it. Weeks after the last published
    semester still HOLD at it - that is the REMM lag, a known ~9-month
    publication delay, not a guess about a period nobody has measured.
    """
    keys = sorted(NONDOM_SEMESTERS)
    want = semester_of(date_iso)
    if want < keys[0]:
        return None, None
    use = keys[-1] if want > keys[-1] else want
    if use not in NONDOM_SEMESTERS:
        use = max(k for k in keys if k <= want)
    ni_e, ni_g, ie_e, ie_g = NONDOM_SEMESTERS[use]
    rate = (fx_by_semester or {}).get(use) or IE_FX["rate"]
    return {
        "gbp": {"electricity": round(ni_e / 100, 4),
                "gas": round(ni_g / 100, 4)},
        "eur": {"electricity": round(ie_e / 100 / rate, 4),
                "gas": round(ie_g / 100 / rate, 4)},
    }, use
# Used ONLY if the semester mean cannot be computed. Logged as
# unverified when it fires, because a silent fallback here would move
# the whole ROI side without saying so.
IE_FX_FALLBACK = 0.87
IE_FX = {"rate": IE_FX_FALLBACK, "source": "fallback (UNVERIFIED)",
         "semester": IE_SEMESTER}

# Published Irish figures, pence per kWh, semester 2 2025, as UREGNI
# prints them. Domestic incl all taxes; non-domestic incl CCL excl VAT.
# Electricity non-domestic is consumption-weighted across the bands
# below Large + Very Large using UREGNI's own published consumption
# shares (5.1 / 32.9 / 19.1 / 26.8 / 16.1%); gas non-domestic is band
# I1, where services buildings sit.
# ROI DOMESTIC BY SEMESTER, euro cent/kWh, all taxes included.
# Electricity is Eurostat/SEAI band DC (2,500-5,000 kWh a year), gas
# is band D2 (20-200 GJ) - the bands the existing anchors were shown
# to sit on: band DC S2 2025 is 40.4 c, and the anchor 35.2p at the
# S2 2025 ECB mean of 0.87073 is 40.43 c. Gas band D2 S2 2025 is
# 13.0 c against the anchor's 11.3p = 12.98 c.
#
# ELECTRICITY IS CREDIT-FREE, GAS IS AS PUBLISHED. Ireland credited
# domestic ELECTRICITY accounts €1,500 between 2022 and 2025, and
# SEAI states the credits "are accounted for in the residential
# electricity prices for the relevant semester" - they reduce the
# effective unit price. Gas accounts were never eligible, so the gas
# figures need no correction at all.
#
# The credit was a LUMP SUM PER METER, not a change in unit rate: a
# household's cost of the next kWh never moved when it landed. A
# per-MWh-of-delivered-heat axis needs the unit rate, or the electric
# routes look cheap exactly while a subsidy runs and jump when it
# stops. So the credit is added back. Proof of the mechanism is in
# the band gradient itself - the 12-month change at S2 2025 runs
# DA +93%, DB +79%, DC +33%, DD +8%, DE -4%, monotone in band size,
# which is what a fixed sum per meter does.
IE_DOMESTIC_SEMESTER = {
    # semester: (electricity as published, gas as published, credits
    #            paid to a domestic account in that semester, EUR)
    "2024-S1": (25.9, 12.7, 300.0),   # two €150 credits, Jan and Mar
    "2024-S2": (30.4, 13.4, 125.0),   # one €125 credit, Nov
    "2025-S1": (32.6, 12.2, 125.0),   # one €125 credit, Jan
    "2025-S2": (40.4, 13.0, 0.0),     # first clean semester
}
# Semester consumption the lump sum is divided by. DAGGER. Three
# controls disagree: band DE flatness implies ~1,500, the S2 2024
# reconciliation ~1,810, and the S1 2024 double-credit check ~2,000.
# The site's own 3,200 kWh/yr basis (1,600 here) was the first
# recommendation and is NOT used, because at 1,600 the corrected
# S1 2025 lands exactly on S2 2025 - implying no market movement
# across an autumn in which Energia, Pinergy, SSE Airtricity and
# Flogas all raised prices. The midpoint is carried and the
# sensitivity is logged every run rather than buried.
IE_CREDIT_KWH_PER_SEMESTER = 1750.0
IE_CREDIT_SENSITIVITY = (1500.0, 2000.0)

IE_PUBLISHED_P_PER_KWH = {
    "domestic_electricity": 35.2,
    "domestic_gas": 11.3,
    "nondom_electricity": 22.68,
    "nondom_gas": 9.3,
}
# Level from the semester, timing from the announcements. Electric
# Ireland held from Oct 2022 to 1 Jul 2026, then +8% / +7.7%.
IE_STEPS = [("2026-07-01", {"domestic_electricity": 1.08,
                            "domestic_gas": 1.077})]


def ie_domestic_eur(fuel, date_iso):
    """
    ROI domestic EUR/kWh on the credit-free band series.

    Returns None below the first published semester - REFUSE, do not
    clamp, the same rule tariffs_for and nondom_for follow. Above the
    last published semester the level holds and IE_STEPS carries the
    announcements forward, which is how the 1 Jul 2026 change is
    applied.
    """
    keys = sorted(IE_DOMESTIC_SEMESTER)
    want = semester_of(date_iso)
    if want < keys[0]:
        return None
    use = want if want in IE_DOMESTIC_SEMESTER else keys[-1]
    elec, gas, credit = IE_DOMESTIC_SEMESTER[use]
    if fuel == "domestic_gas":
        v = gas / 100.0
    else:
        # add the credit back: lump sum per meter over a semester's
        # consumption at the stated basis. The credit is in EUROS and
        # the band figures in CENTS, hence the 100.
        v = (elec + 100.0 * credit / IE_CREDIT_KWH_PER_SEMESTER) / 100.0
    for frm, steps in IE_STEPS:
        if date_iso >= frm and fuel in steps:
            v *= steps[fuel]
    return round(v, 4)


def ie_eur(key, date_iso=None):
    """Published Irish sterling figure -> EUR/kWh at the fetched
    semester rate, stepped by any announcement on or before the date.
    Domestic fuels now come from the dated band series instead, so
    this carries the non-domestic anchors only."""
    if date_iso and key in ("domestic_electricity", "domestic_gas"):
        v = ie_domestic_eur(key, date_iso)
        if v is not None:
            return v
    v = IE_PUBLISHED_P_PER_KWH[key] / 100.0 / IE_FX["rate"]
    for frm, steps in IE_STEPS:
        if date_iso and date_iso >= frm and key in steps:
            v *= steps[key]
    return round(v, 4)


def semester_means(daily, min_days=110):
    """{date: rate} -> {'YYYY-S1'|'YYYY-S2': mean}. Semesters with
    fewer than min_days observations are DROPPED, not averaged thin -
    a part-semester mean is indistinguishable from a whole one and
    would silently mis-scale every Irish anchor."""
    buckets = {}
    for d, v in (daily or {}).items():
        try:
            y, m = int(d[:4]), int(d[5:7])
        except (ValueError, IndexError):
            continue
        buckets.setdefault(f"{y}-S{1 if m <= 6 else 2}", []).append(v)
    return {k: round(sum(v) / len(v), 5)
            for k, v in buckets.items() if len(v) >= min_days}
UA = {"User-Agent": "ioi-heatsplit/0.5 (contact@causewaygt.com)"}
TIMEOUT = 90
RETRIES = 3

# Best-effort context feeds - failure marks stale but never pages or
# fails the run; regressions stay visible via status badges and flags.
SOFT_FEEDS = {"gb_oil", "entsog_probe", "sem_mix",
              "eirgrid_probe"}

# Feeds known broken for reasons outside this pipeline - marked stale,
# logged, but neither paged nor allowed to fail the run.
EXPECTED_DOWN = {}
# eirgrid was expected-down 09-18 Jul 2026 (dashboard redesign); restored
# via the /api/chart/ endpoint captured by browser probe.

# Population weights for degree days - Causeway judgement figures (dagger).
# Challenge and input welcome at contact@causewaygt.com.
STATIONS = {
    #  name        lat      lon     weight  jurisdiction
    "Dublin":    (53.35,  -6.26,   0.40, "ROI"),
    "Belfast":   (54.60,  -5.93,   0.20, "NI"),
    "Cork":      (51.90,  -8.47,   0.12, "ROI"),
    "Galway":    (53.27,  -9.05,   0.08, "ROI"),
    "Limerick":  (52.66,  -8.63,   0.08, "ROI"),
    "Derry":     (54.99,  -7.31,   0.06, "NI"),
    "Waterford": (52.26,  -7.11,   0.06, "ROI"),
}
HDD_BASE_C = 15.5

# Fill after browser XHR probe (the one remaining probe)
EIRGRID_ENDPOINT = "https://www.smartgriddashboard.com/api/chart/"
# Captured by browser XHR probe, 18 Jul 2026. Response schema is the
# pre-redesign DashboardService shape unchanged: {"Rows":[{"EffectiveTime":
# "18-Jul-2026 00:15:00","FieldName":"SYSTEM_DEMAND","Region":"NI",
# "Value":573}, ...]} - 15-min actuals with nulls for future intervals,
# half-hourly DEMAND_FORECAST_VALUE rows appended.

NTFY_TOPIC = None

# Set by main() before the feed loop - lets feeds merge history across runs
PREVIOUS_FEEDS: dict = {}
PREVIOUS_DERIVED: dict = {}

# ------------------------------------------------- annual anchors
# Sourced figures cite their publication; every judgement figure is marked
# with a dagger and is a current Causeway Energies estimate - challenge and
# input welcome at contact@causewaygt.com.

ANCHORS = {
    "year": 2024,
    "roi": {
        "residential_heat_twh": 22.3,      # SEAI Energy in Ireland 2025
        "services_heat_twh": 8.5,          # dagger - from SEAI sector shares
        "fuel_shares": {"oil": 0.565, "gas": 0.251, "peat": 0.067,
                        "electricity": 0.08, "other": 0.037},
        #             oil/gas/peat SEAI 2024; electricity/other dagger
        "gas_indigenous": 0.18,            # SEAI H1-2025, Corrib, falling
        "elec_indigenous": 0.413,          # SEAI RES-E 2024
    },
    "ni": {
        "residential_heat_twh": 10.0,      # dagger - NISRA stock x intensity
        "services_heat_twh": 3.0,          # dagger
        "fuel_shares": {"oil": 0.62, "gas": 0.26, "peat": 0.0,
                        "electricity": 0.08, "other": 0.04},
        #             oil NISRA CHS 2024/25; gas/electricity/other dagger
        "gas_indigenous": 0.0,             # all NI gas arrives via Moffat
        "elec_indigenous": 0.46,           # dagger - DfE yr-to-Mar-2026 ~48%
    },
    "efficiency": {"oil": 0.82, "gas": 0.85, "peat": 0.60,
                   "electricity": 1.0, "other": 0.70},          # dagger
    "geothermal_spf": 4.0,                                       # dagger
    # Air-source heat pump model - all dagger. Carnot-fraction COP against
    # the HDD-derived, demand-weighted outdoor temperature; defrost derate
    # for humid Irish winters; DHW share at higher flow. Calibrated to the
    # GB Electrification of Heat field-trial median (~2.8-2.9) rather than
    # laboratory SCOP figures.
    # PROVENANCE (audit, 2 Aug 2026 - these become the v7 hourly COP
    # engine's foundation, so every parameter is declared here):
    #   flow_c 45.0 - mid-range radiator flow for a retrofit onto
    #     existing emitters (weather compensation 30-50 C in v7);
    #     dagger, convention not measurement.
    #   carnot_fraction 0.38 - typical field-observed fraction of
    #     Carnot for modern inverter ASHP; consistent with the SPF
    #     anchors below and with UK field-trial ranges 0.35-0.45.
    #     Dagger; the SPF anchor pins the annual, so this shapes the
    #     seasonal SHAPE, not the level.
    #   defrost_derate 0.90 - performance penalty in the 0-7 C humid
    #     band that dominates the island's winter; dagger.
    #   dhw_share 0.20 - hot water as a share of heat-pump duty in
    #     the modelled route; sits alongside the 22.4% DHW share of
    #     national heat input (SEAI National Heat Study re-anchoring,
    #     Aug 2026); dagger.
    "ashp": {"flow_c": 45.0, "carnot_fraction": 0.38,
             "defrost_derate": 0.90, "dhw_share": 0.20,
             "dhw_flow_c": 55.0, "dhw_source_c": 10.0},
    # Fuel emission factors, g CO2e per kWh of fuel INPUT (not
    # delivered): oil/gas/peat from the SEAI emission-factor series
    # for kerosene, natural gas and milled peat; "other" is a
    # biomass-weighted residual (dagger); electricity is the anchor
    # replaced at runtime by live all-island grid intensity.
    "ef_g_per_kwh": {"oil": 257, "gas": 205, "peat": 340,
                     "other": 100, "electricity": 280},          # dagger -
    # electricity factor replaced by live grid intensity once eirgrid returns
    "indigenous": {"oil": 0.0, "peat": 1.0, "other": 0.9},      # dagger
    # CROSS-CALIBRATED 18 Jul 2026 against the UK sibling (input
    # basis, buildings only): UK 430.6 TWh / 68m = 6.3 MWh/person;
    # island 43.8 / 7.1 = 6.2 - per-capita parity, ratio 0.98.
    # RE-ANCHORED 6 Aug 2026 to an Irish source. SEAI National Heat
    # Study Report 1 puts residential hot water at 25% of residential
    # heat demand (space heating 75%). Applied to this file's own
    # sector split - island residential 32.3 TWh, services 11.5 TWh -
    # with services hot water at 15% (dagger; SEAI does not publish
    # the services split), the island hot-water share is 22.4% and
    # the weather-driven space share 77.6%. This REPLACES the
    # UK-aligned 18.3% adopted in July, which was explicitly marked
    # as pending an SEAI-specific figure. Caption symmetry with the
    # UK sibling is now broken by design: both sites shape space heat
    # by degree days and carry hot water flat, but each uses its own
    # national hot-water share. Weekly = A x [0.224/52 + 0.776 x Hw/Hy].
    "space_heat_fraction": 0.776,   # = 1 - DHW share; see above
    "kerosene_kwh_per_litre": 10.35,   # industry standard figure
    # Tariff anchors, July 2026 pass. Sourced bands, dagger on the point:
    #  ROI electricity: standard 24h ~35c (Electric Ireland, May 2026);
    #    Eurostat H2-2025 all-in ~40c; anchor 36c.
    #  ROI gas: standard unit ~11-12c incl 9% VAT; anchor 11.5c.
    #  NI electricity: Power NI 1 Jul 2026 review - GBP1,093/yr at 3,200 kWh
    #    -> ~32.5p unit ex-standing; anchor 32.5p.
    #  NI gas: Ten Towns GBP972/yr at 12,000 kWh (1 Jul 2026) -> ~7.5p unit.
    # Standard-tariff basis - time-of-use/night rates materially lower for
    # heat-pump households; see heat_gap basis note.
    # Sector blend (UK-pattern, Irish placeholders - dagger until read
    # from the SEAI national energy balance / BER end-use work and
    # Eurostat band prices nrg_pc_203 (gas) / nrg_pc_205 (electricity);
    # Eurostat publishes on ~half-year lag, stated in methodology).
    # dom_share: domestic fraction of each fuel's buildings-heat input.
    # Oil is priced identically across sectors (tanker market), so only
    # gas and electricity blend. Cooling is wholly non-domestic.
    "dom_share": {
        "roi": {"gas": 0.60, "electricity": 0.55},
        "ni": {"gas": 0.55, "electricity": 0.55},
    },
    # NON-DOMESTIC RATES - Utility Regulator Retail Energy Market
    # Monitoring, semester 2 2025 (Q1 2026 QREMM, published 16 Jun 2026).
    #
    # WHAT WAS WRONG. Both jurisdictions carried LARGE-USER prices.
    # The NI pair (24.0p / 5.5p) was the GB QEP manufacturing average
    # lifted verbatim from the UK sibling - wrong jurisdiction AND
    # wrong sector. This site's non-domestic scope is SERVICES
    # BUILDINGS - offices, retail, hospitality, public estate - which
    # sit at the small end of the distribution where prices run near
    # domestic, not at the industrial end where they are little more
    # than half that. We had been pricing offices at smelter rates.
    #
    # ELECTRICITY - consumption-weighted across the published bands,
    # excluding only Large + Very Large. That top band is seventeen NI
    # connections and 683 GWh of unambiguously heavy industry and data
    # centres; every band below it is where services buildings live.
    # Weighting is UREGNI's published I&C consumption share per band,
    # applied to both jurisdictions because Ireland's own split is not
    # published alongside - a dagger, but a much smaller one than the
    # band choice, and the two ladders track each other closely.
    # Semester 2 2025 (Q1 2026 QREMM), with UREGNI's own published
    # consumption shares - no proxy weighting needed any more:
    #   band (MWh/yr)      NI p/kWh   IE p/kWh   % of I&C GWh
    #   very small <20        28.6       26.2       5.1
    #   small 20-500          26.3       25.4      32.9
    #   small/med 500-2k      23.3       22.2      19.1
    #   medium 2k-20k         20.2       19.0      26.8
    #   large+VL >20k         17.8       16.1      16.1   <- excluded
    #   weighted (services)   23.8       22.7      83.9
    # Ireland has crossed BELOW NI since 2024 on both fuels.
    # Ireland converted back to euro at 0.84, the semester average -
    # UREGNI converts Eurostat's euro figures to sterling for the
    # comparison, so this recovers the original.
    #
    # GAS - band I1 (<278,000 kWh/yr), which is where services
    # buildings overwhelmingly sit. NOT consumption-weighted, because
    # the REMM price bands (I1/I2/I3&I4) do not map onto the network
    # EUC bands the consumption split is published in, and NI I&C gas
    # consumption is ~68% daily-metered heavy industry which would
    # drag a weighted figure to an industrial number. I1 for
    # reference: NI 8.67, Ireland 8.58; I2 7.55 / 7.76; I3&I4 5.82 /
    # 3.23. Dagger on the band choice.
    #
    # THE BASIS. UREGNI derives these the way Eurostat does - volume
    # sold and revenue submitted per size band, revenue over volume -
    # so standing charges are INCLUDED by construction. That makes
    # them like-for-like with the SEAI/Eurostat semester prices on the
    # ROI side, which is what closes the cross-jurisdiction basis gap
    # for the services share. I&C figures include CCL and EXCLUDE VAT,
    # which is the convention this site already applies to services.
    #
    # VINTAGE - and a live consequence. Held at S2 2025, the latest
    # published semester (Q1 2026 QREMM, 16 Jun 2026), and they do NOT move week to week. REMM lags
    # about nine months. Since the domestic rates ARE stepped through
    # to 2026 and NI gas fell about 15% over that span, NI
    # non-domestic gas (8.67p, 2024) now prints ABOVE NI domestic gas
    # (7.70p, Jul 2026). That is a vintage artefact, not a claim that
    # small business gas is cheaper than household gas - at a common
    # vintage the ordering is right (domestic 9.88p incl VAT, 9.41p
    # excl, against 8.67p non-domestic). It is disclosed rather than
    # escalated away, because applying regulated domestic steps to
    # unregulated business contracts would be a dagger on a dagger.
    # It closes when REMM publishes newer semesters.
    # Level from the semester series and timing from tariff
    # announcements is the intended design; until that is built the
    # services share of every week is priced at a 2024 semester.
    # set per run by apply_ie_fx() from the fetched semester rate
    "nondom_eur_per_kwh": {"electricity": 0.284, "gas": 0.102},
    "nondom_gbp_per_kwh": {"electricity": 0.2381, "gas": 0.080},
    "retail_eur_per_kwh": {"gas": 0.115, "electricity": 0.36},
    "retail_gbp_per_kwh": {"gas": 0.075, "electricity": 0.325},
    # The cold economy - island cooling loads, electricity basis.
    #  dc: CSO metered consumption ~22% of ROI electricity (sourced);
    #    29% projected by 2028. Other loads are Causeway anchors, dagger:
    #  refrigeration: dairy (farm milk cooling + processing), meat, cold
    #    stores, retail - the food-export cold chain;
    #  process: pharma/semiconductor chilled-water and process cooling;
    #  comfort: commercial space cooling and ventilation (summer-peaked,
    #    shaped by live overheating degree-hours when available);
    #  ni_cooling: food processing, retail refrigeration, small DC.
    #  rejection: refrigeration/comfort reject compressor work PLUS the
    #    heat they pump - factors dagger; DC rejects ~all electricity.
    # CORRECTION 6 Aug 2026 (SEAI National Heat Study, Report 1):
    # the census previously carried TOTAL data-centre electricity as
    # cold-economy cooling load. It is not: SEAI finds cooling is
    # ~14% of data-centre electricity (SEAI; see the share note in the
    # cold-economy anchors), the balance being IT load that
    # BECOMES heat rather than removing it. The DC line is now priced
    # at its cooling share; the heat it rejects is unchanged, so the
    # rejection and service factors rise to match (see below) and the
    # delivered cooling service is preserved. SEAI's absolute (0.4
    # TWh) is a 2019-base figure and is stale against today's fleet -
    # the transferable parameter is the SHARE, applied to the current
    # metered total.
    # Heat-pump stock, from the two censuses - the anchor that lets the
    # electricity line split into resistive and heat-pump, and the free
    # ambient heat appear on the out-bar.
    #   ROI: Census 2022 - 71,000 households with heat pumps, of which
    #        57,000 air source and 14,000 ground source. A FLOOR: NZEB
    #        has made heat pumps near-universal in new dwellings since
    #        2019 and four further years of SEAI grants have followed,
    #        so 2026 stock is materially higher (dagger uplift 1.5).
    #   NI:  Census 2021 Table 27 - air source 0.09% and geothermal
    #        0.08% of 768,808 households = ~692 and ~615. The 615
    #        confirms the 500-700 dagger this file has carried for the
    #        NI domestic ground-source count.
    #   Non-domestic tail: WGC2026 gives ROI 20,128 ground-source
    #        systems against the census's 14,000 households; the ~6,000
    #        difference is the non-domestic and communal tail. At the
    #        fleet's ~11 kW average the stock is overwhelmingly
    #        domestic-scale. SSRH, the ROI non-domestic register, had
    #        supported only ~EUR0.5m of heat-pump projects by 2023.
    #   Communal schemes serving apartment blocks are UNDER-COUNTED by
    #        both censuses (a shared plant reads as many households);
    #        district heat is ~1% of island heat, so the error is small
    #        but real - dagger.
    "heat_pumps": {
        "roi": {"households": 71000, "census_uplift": 1.5,
                "nondom_equivalent": 6000},
        "ni": {"households": 1307, "census_uplift": 1.5,
               "nondom_equivalent": 200},
        # delivered heat per heat-pumped dwelling, MWh/yr - below the
        # stock average because heat-pumped homes are newer and better
        # insulated; dagger.
        "delivered_mwh_per_dwelling": 9.0,
    },
    "cool": {"dc_share_of_roi_elec": 0.22, "dc_share_2028": 0.29,
             # HELD AT 0.14, and the reasoning is recorded because it
             # was nearly changed on a misreading. SEAI's 2025
             # Comprehensive Assessment states data-centre cooling of
             # 0.9 TWh; read against its 2025 ELECTRICITY figure of 9+
             # TWh that looks like ~10%, but the 0.9 is a 2023 quantity
             # and 2023 data-centre electricity was ~6.4 TWh - which
             # implies 14%, matching the 2022 archetypes. The two SEAI
             # publications agree; the 10% was a 2023 numerator over a
             # 2025 denominator. Resolve against the archetype table
             # before moving this.
             # reads COOL_DC_COOLING_SHARE so the two panels cannot
             # diverge; the reasoning above is why it is 0.14
             # 0.14 here and in COOL_DC_COOLING_SHARE. Asserted, not
             # assigned: ANCHORS is read at import time by code that
             # runs before any later assignment, and setting this to
             # None to fill in afterwards broke the hero.
             "dc_cooling_share": 0.14,
             "roi_elec_twh": 31.0,
             "loads_twh": {"dc": None,          # computed: share x elec
                           # Cold-economy load census, TWh/yr of electricity. DC from CSO
    # data-centre metered consumption (22% of ROI electricity);
    # the remainder are Causeway estimates pending component-level
    # sourcing in the autumn pass - each a dagger:
    #   refrigeration 2.3 - food/dairy/pharma cold chain, scaled from
    #     the sector's share of industrial electricity
    #   process 0.8 - industrial process cooling excl. cold chain
    #   comfort 1.0 - comfort cooling AND ventilation in services
    #     buildings (the UK sibling's whole cooling scope, roughly)
    #   ni_all 1.2 - the NI cold economy, all categories
    "refrigeration": 2.3,
                           "process": 0.8,
                           "comfort": 1.0,
                           "ni_all": 1.2},
             # CENSUS VALIDATED 7 Aug 2026 against the SEAI
             # National Heat Study supporting data (final-energy
             # table). SEAI ROI cooling electricity: commercial
             # 2,868 + public 221 + agriculture 93 + industry 804
             # + data centres 1,055 (their 2026 projection) =
             # 5,041 GWh. Our ROI lines (dc 955 + refrigeration
             # 2,300 + process 800 + comfort 1,000) = 5,055 GWh -
             # 14 GWh apart on 5 TWh, with industry/process
             # matching to 4 GWh. The earlier suspicion that
             # refrigeration was light came from comparing SEAI's
             # USEFUL demand against our ELECTRICITY; the
             # like-for-like comparison holds. The open question is
             # now the internal split between refrigeration and
             # comfort within the commercial total, not the total.
             # AUDIT 18 Jul 2026: census scope is the full cold economy
             # (DC + cold chain + process + comfort) - deliberately
             # wider than comfort-only national cooling lines (e.g. the
             # UK tracker's ECUK-anchored figure); the two are not
             # ratio-comparable and the hero basis says so. Component
             # source hardening scheduled for the autumn cycle.
             # dc 7.1: with the line repriced to cooling electricity
             # only, the heat rejected is unchanged (the whole DC draw
             # ends as heat), so heat rejected per unit COOLING
             # electricity = 1/0.14 ~ 7.1. Physical output preserved
             # across the 6 Aug 2026 correction; only the purchased
             # side moved.
             # dc is COUPLED to dc_cooling_share and must move with it:
             # essentially ALL data-centre electricity leaves as heat,
             # so heat rejected = cooling electricity / cooling share,
             # i.e. dc = 1 / dc_cooling_share. 7.1 is 1/0.14. Changing
             # one without the other silently loses or invents rejected
             # heat, and the suite asserts the identity - it caught
             # exactly that on 18 Aug 2026.
             "rejection_factor": {"dc": 7.1, "refrigeration": 2.5,
                                  # Heat-rejection factors: kWh of heat rejected per kWh of
    # electricity drawn - approximately (COP + 1) for vapour
    # compression, near 1.0 for DC where most draw is IT load
    # rejected as heat. All dagger, all Causeway judgement.
    "process": 2.0, "comfort": 3.0,
                                  "ni_all": 1.8},
             # District heat as a share of national heat - order 1% on the
    # island against ~2% GB and 50%+ in Denmark; dagger, and the
    # number the geothermal panel's district-heat argument rests on.
    "dh_share_of_national_heat": 0.01,
             # dagger: electricity saving on cooling load moved to
             # ground-coupled systems (free cooling / high-EER ATES)
             "ground_cooling_saving": 0.70,
             # dagger: cooling service delivered per unit electricity
             # (seasonal system EER). DC ~1.0: its service is heat
             # removal, roughly its electricity; vapour-compression
             # loads deliver a multiple of theirs.
             # Service factors: kWh of COOLING SERVICE delivered per kWh of
    # electricity - the delivered-basis multiplier that lets the
    # out-bar legitimately exceed the in-bar. Dagger throughout.
    # dc 6.1: cooling service (heat removed from IT equipment) per
    # unit cooling electricity - an effective seasonal EER, high
    # because Ireland's climate permits extensive free cooling.
    # Raised from 1.0 with the 6 Aug 2026 repricing.
    # RE-ANCHORED 7 Aug 2026 to the SEAI National Heat Study
    # supporting data, which publishes useful cooling demand AND
    # final cooling consumption per sector - their ratio IS the
    # service factor. SEAI's ROI implied ratios: commercial 2.07,
    # public 2.57, industry 1.00, agriculture 1.00, aggregate 1.86
    # (excluding data centres). We adopt 2.07 for the commercial-type
    # loads - refrigeration and comfort - and hold process at 2.2
    # dagger rather than copying SEAI's 1.00, which is a modelling
    # pass-through (a COP of exactly 1.0 for process chilling is not
    # a physical claim). NI carries the commercial ratio. The data
    # centre factor stays 6.1 on its own provenance: Irish free
    # cooling genuinely delivers an effective EER above 6, and SEAI
    # models data centres in a separate sheet at 14% cooling share,
    # which our census already follows.
    # COOLING SERVICE FACTORS - the EER at which each load converts
    # electricity into delivered cooling. Panel 1 and Panel 4 read the
    # SAME figures; a module-level assert below fails the build if they
    # drift apart. Each is sourced or reasoned, and one is weak:
    #
    #   dc 6.1  - the arithmetic, not a judgement: 14% of data centre
    #             electricity on cooling means 5.5 TWh of heat removed
    #             for 0.9 TWh, and the IT load at a measured Irish PUE
    #             of 1.15-1.25 gives the same. In this climate that is
    #             free cooling expressed as a number.
    #   refrigeration / comfort 2.07 - SEAI's own commercial ratio,
    #             7.5 TWh of service on 3.6 TWh of electricity.
    #   process 3.0 - OURS, dagger. Was 2.2, which put industrial
    #             process chillers barely above part-load office air
    #             conditioning; process plant runs steady, high-
    #             utilisation duty, which is where chillers perform
    #             best. Bracketed by SEAI's own anchors (2.08
    #             commercial, 2.50 public) with Barth et al. (2025)
    #             putting Manhattan at 3.5. Note it cuts both ways: a
    #             higher figure enlarges the service bar but SHRINKS
    #             the geothermal what-if, because less electricity is
    #             displaced per unit of service moved.
    #   public 2.50 - SEAI's own, 0.5 TWh of service on 0.2 of
    #             electricity. Used by the cooling panel only: this
    #             panel slices FUNCTIONALLY (refrigeration, process,
    #             comfort) where that one slices by SEAI SECTOR, so
    #             public comfort cooling is inside "comfort" here and
    #             has nothing of its own to attach to. That is a
    #             difference of cut, not a drift.
    #   ni_all 2.07 - THE WEAKEST FIGURE HERE, dagger. The commercial
    #             ratio applied to an entire jurisdiction with no
    #             split, because DfE publishes no cooling line and NI
    #             is not separable inside UK statistics.
    "cooling_service_factor": {"dc": 6.1, "refrigeration": 2.07,
                                        "process": 3.0, "comfort": 2.07,
                                        "public": 2.50,
                                        "ni_all": 2.07}},
}

# Policy events rendered as chart annotations - date, jurisdiction, label.
EVENTS = [
    {"date": "2025-10-08", "jur": "ROI",
     "label": "Carbon tax to \u20ac71/t \u2013 motor fuels"},
    {"date": "2026-03-16", "jur": "UK",
     "label": "UK \u00a350m oil support; CMA review"},
    {"date": "2026-04-01", "jur": "ROI",
     "label": "NORA levy paused (two months)"},
    {"date": "2026-05-01", "jur": "ROI",
     "label": "Heating-fuel carbon increase postponed"},
    {"date": "2026-07-13", "jur": "ROI",
     "label": "Heat Bill advanced \u2013 district heating framework"},
    {"date": "2026-10-14", "jur": "ROI",
     "label": "Carbon tax \u20ac63.50\u2192\u20ac71/t \u2013 heating fuels (due)"},
]

# Value-level caveats, distinct from fetch status - machine-carried nuance.
FEED_FLAGS = {
    "ccni_oil": ["Verification pending that parsed values are the NI "
                 "average series, not a council-area series"],
    "oil_bulletin": ["Bulletin heading is gas oil; treated as ROI "
                     "heating-oil (kerosene) price level - Causeway "
                     "judgement"],
    "hdd": ["Population weights are Causeway estimates"],
    "gb_oil": ["BoilerJuice site average of lowest quotes, not a survey "
               "average - basis differs from CCNI; ONS kj5u discontinued "
               "Jan 2025"],
}




# ------------------------------------------------- geothermal register
# NI >60 kW register: Causeway research pass, June 2026, updated July 2026
# Martinstown (325 kW, 2015) is ROI - confirmed by S. Todd, counted in
# the WGC ROI totals, not carried here.
# (Ryan Daly pers. comm. - Randalstown 44 kW and Strabane 18 kW confirmed
# sub-threshold, logged as exclusions). ROI anchors: WGC2026 Country Update
# (Ireland, Blake, Pasquali, Dunphy & Hunter Williams) - sourced.
# Corrections welcome at contact@causewaygt.com.
GEO = {
    # Northern Ireland register, Causeway Energies, circulating for
    # comment among NI practitioners as at 20 Aug 2026. Ten documented
    # systems at or above 45 kW.
    #
    # EFFECTIVE, NOT NAMEPLATE. The register carries both, and the gap
    # matters: 532 kW of heating was built and 460 kW is delivered, so
    # 14% of installed capacity never reached the building. Our earlier
    # anchors took nameplate and missed three failures entirely, which
    # made the record look better than it is - five of seven confirmed,
    # against two clean operational systems of ten.
    "ni_register": [
        {"id": "R1", "site": "R1", "year": 2009, "kw": None,
         "kw_nameplate": None, "duty": "cooling",
         "status": "candidate - unconfirmed", "confirmed": False},
        {"id": "R2", "site": "R2", "year": 2011, "kw": 120,
         "kw_nameplate": None, "duty": "cooling",
         "status": "operational - minor issues", "confirmed": True},
        {"id": "R3", "site": "R3", "year": 2012, "kw": 72,
         "kw_nameplate": 72, "duty": "heating",
         "status": "operational", "confirmed": True},
        {"id": "R4", "site": "R4", "year": 2016, "kw": 108,
         "kw_nameplate": 180, "duty": "heating",
         "status": "operational - impaired", "confirmed": True},
        {"id": "R5", "site": "R5", "year": 2018, "kw": 0,
         "kw_nameplate": None, "duty": "cooling",
         "status": "never commissioned", "confirmed": True},
        {"id": "R6", "site": "R6", "year": 2023, "kw": 280,
         "kw_nameplate": 280, "duty": "heating + DHW",
         "status": "operational", "confirmed": True},
        {"id": "R7", "site": "R7", "year": None, "kw": 0,
         "kw_nameplate": None, "duty": None,
         "status": "dead - completion failure", "confirmed": True},
        {"id": "R8", "site": "R8", "year": None, "kw": 0,
         "kw_nameplate": None, "duty": None,
         "status": "inoperable", "confirmed": True},
        {"id": "R9", "site": "R9", "year": None, "kw": 0,
         "kw_nameplate": None, "duty": None,
         "status": "never completed", "confirmed": True},
        {"id": "R10", "site": "R10", "year": None, "kw": None,
         "kw_nameplate": None, "duty": "heating",
         "status": "candidate - unconfirmed", "confirmed": False},
    ],
    # Site names are deliberately NOT held here. The register is out
    # for comment and the site must not publish its entries; keeping
    # the names out of the payload makes that structural rather than a
    # matter of remembering.
    "ni_register_totals": {
        "documented": 10, "operational_clean": 2, "operational_any": 4,
        "failed": 4, "unconfirmed": 2,
        "delivered_heating_kw": 460, "delivered_cooling_kw": 120,
        "nameplate_heating_kw": 532, "heating_shortfall_kw": 72,
    },
    # 386 MCS-registered units at 10-12 kW, per the register's small
    # tier. Previously carried as 500-700, a triangulation with no
    # stated basis that put NI's total more than a megawatt too high.
    # MCS-registered domestic only: it excludes non-MCS installs, and
    # NI has no domestic RHI, so this is a floor rather than a count.
    "ni_domestic": {"units": 386, "kw_each": (10, 12),
                    "mw_low": 3.5, "mw_high": 4.5,
                    "note": ("386 MCS-registered small ground and water "
                             "source units, March 2023, at 10-12 kW "
                             "assumed mean output - a floor, since it "
                             "excludes non-MCS installs and NI has no "
                             "domestic RHI")},
    # 224.4 and 291.9, matching EGC 2025 Table 4 exactly. Previously
    # carried as 225 and 293 - rounded somewhere between the source and
    # this file, and ours should be the source's numbers.
    "roi": {"capacity_mwth": 224.4, "heat_gwh": 291.9,
            "cooling_gwh": 11.9, "units": 20128, "new_2024_mwth": 7.4, "proj_2028_mwth": 261,
            "deep_plants": 0,
            # sector shares per WGC2026 text (approximate - sum > 100 in
            # the source; presented as reported)
            "sector_share_pct": {"residential": 85, "commercial": 14,
                                 "industrial": 4},
            "gshp_share_of_hp_market_pct": 4,
            "source": ("WGC2026 Country Update: Ireland - Ireland, Blake, "
                       "Pasquali, Dunphy & Hunter Williams, June 2026")},
    # DERIVED below from capacity and population, not typed. Carried
    # as a literal until 20 Aug 2026, when re-anchoring NI from 6.6 to
    # 4.5 MWth left the 3 W figure stale - the same trap as the
    # "eleven terawatt-hours" sub-heading on the cooling panel.
    "per_capita_w": {"note": "installed Wth per person, derived"},
    "population_m": {"roi": 5.3, "ni": 1.92},   # dagger, mid-2026
    "eflh_h": 2000,   # equivalent full-load heating hours - dagger
    # European reference points, installed GSHP Wth per person -
    # shallow basis, EGC 2025 capacities over mid-2020s populations,
    # dagger: Sweden 8.12 GWth/10.5m; NL 2.49 GWth/17.8m (ATES-heavy);
    # France 2.29 GWth/68m.
    "reference_w_pp": {"Sweden": 773, "Netherlands": 140, "France": 34},
    # REPORTED ANNUAL OUTPUT, not derived from a load-hour convention.
    # EGC 2025 Country Update Summary (Sanner, Antics, Baresi,
    # Urchueguia & Dumas), Table 4: Ground Source Heat Pump Use in
    # Europe in 2024 - units, capacity, production and the full-load
    # hours the source itself calculates from them.
    #
    # THE 2,000-HOUR CONVENTION WAS WRONG FOR EVERY COUNTRY, and wrong
    # in opposite directions: it overstated Ireland by 54% and
    # understated Sweden by 43%, so the gap between them was drawn
    # roughly 2.7 times narrower than it is. Both errors flattered
    # Ireland. Using reported output removes the assumption entirely.
    "reference_output": {
        "Ireland": {"units": 20128, "mwth": 224.4, "gwh": 291.9,
                    "flh": 1301},
        "Sweden": {"units": 690000, "mwth": 8120.0, "gwh": 28400.0,
                   "flh": 3498},
        "Netherlands": {"units": 163169, "mwth": 2486.0, "gwh": 2722.0,
                        "flh": 1095},
        "France": {"units": 209021, "mwth": 2293.0, "gwh": 4750.0,
                   "flh": 2072},
        "United Kingdom": {"units": 55210, "mwth": 861.0, "gwh": 1430.0,
                           "flh": 1661},
    },
    "reference_output_source": (
        "EGC 2025 Country Update Summary, Sanner et al., Table 4: "
        "Ground Source Heat Pump Use in Europe in 2024"),
    # Comparator installed capacity, MWth - EGC 2025 country updates
    # (data year 2024), replacing the older WGC2023-lineage values on
    # 27 Jul 2026 audit: Sweden had understated ~18% (8,120 shallow,
    # 690k units, plus Lund's 47 MW heat-pump-coupled deep), the
    # Netherlands ~16% (doublet fleet 367 MWth deep, ~28 doublets,
    # mostly horticulture), France was ~8% high (older inflated
    # shallow estimate; true split is less shallow, more Paris Basin
    # Dogger district heating). ROI's own 225 stays on its WGC2026
    # citation - EGC 2025 carries 224, same lineage, rounding-level.
    "reference_mwth": {
        "Sweden": {"shallow": 8120, "deep": 47},
        "Netherlands": {"shallow": 2486, "deep": 367},
        "France": {"shallow": 2293, "deep": 724}},
    # EGEC Geothermal Market Report 2025, Key Findings - sourced.
    # Note: EGEC counts units SOLD in 2025; the WGC2026 paper reports
    # capacity COMMISSIONED in 2024 (+7.4 MWth). Different measures and
    # years - both carried, not reconciled.
    "egec_2025": {
        "eu_ghp_units_m": 2.55, "eu_ghp_gwth": 39.2,
        "eu_ghp_heat_twh": 88, "eu_people_served_m": 10.6,
        "eu_sales_2025": 123000, "eu_sales_growth_pct": 10,
        "ghp_sales_2025": {"Ireland": 1409, "Sweden": 26785,
                           "Netherlands": 22148, "Germany": 19125,
                           "United Kingdom": 4070},
        "dhc_systems": 434, "dhc_gwth": 6.0,
        "dhc_rank": "second-largest renewable source for district heat",
        "source": "EGEC Geothermal Market Report 2025, Key Findings"},
    # 4.5 MWth: 0.46 MW of DELIVERED large-tier heating plus the
    # midpoint of the 3.5-4.5 MW domestic tier. Previously 6.6, which
    # took nameplate rather than effective capacity and a 500-700 unit
    # domestic estimate - about 40% too high, and flattering in the
    # same direction as every other error found this week. Cooling
    # capacity (120 kW) is excluded: this figure is compared against
    # heat demand.
    "ni_capacity_mwth_est": 4.5,
    "island_today_twh": 0.30,
    "pipeline": [
        "GEMINI (EUR 20m, PEACEPLUS): 3 shallow demos Sligo + Belfast "
        "(NIHE, NI Water); deep 2 km at Grangegorman, drilling late 2027",
        "GeoEnergy NI: Stormont shallow boreholes drilled 2024; CAFRE "
        "Greenmount deep doublet consented (LA03/2025/0443/F)",
        "GSI deep scientific boreholes: BHT 22.6-38 C at 1 km, five "
        "completed 2022-2026",
    ],
}


# ------------------------------------------------- why heat? (whole economy)
# The zoom-out panel: annual, all-island, whole-economy anchors for the
# three energy services. Sourced where a publication exists; allocations
# are Causeway derivations, kept deliberately round and dagger-marked.
#  services: ROI TFC ~143 TWh (SEAI Energy in Ireland 2025) + NI ~48
#    dagger. Heat includes industrial process heat (dagger allocation);
#    power = non-heat electricity.
#  spend: retail-basis, EUR bn - dagger throughout.
#  imports: allocation of ~179 TWh imported primary energy (81.2% of
#    220.7 TWh PES - Causeway island Sankey, 2024).
#  emissions: ROI ~32 Mt energy CO2 (SEAI) + NI ~11 dagger.
WHY_HEAT = {
    "tfc_twh": 191,
    "services_twh": {"heat": 64, "transport": 83, "power": 35},
    "spend_eur_bn": {"heat": 6.0, "transport": 15.0, "power": 10.5},
    "imports_twh": {"heat": 52, "transport": 79, "power": 28},
    "emissions_mt": {"heat": 16, "transport": 17, "power": 9},
    "basis": ("Annual, all-island, whole-economy - dagger throughout. "
              "Services: SEAI Energy in Ireland 2025 (ROI TFC ~143 TWh) "
              "+ NI dagger; heat includes industrial heat. Imports: "
              "allocation of 81.2%-imported primary energy, 220.7 TWh "
              "PES (Causeway island Sankey 2024). Emissions: SEAI + NI "
              "dagger. Challenge and input welcome at "
              "contact@causewaygt.com"),
}


# ------------------------------------------------ per-week tariff history
# No regulated cap exists in either jurisdiction's oil-heated world and
# none in ROI electricity - representative standard domestic unit rates
# by period, dagger throughout (the representativeness caveat the UK
# series does not carry). Backfilled weeks freeze at first computation:
# a wrong early row is permanent, so rows are added only once verified
# against SEAI/CSO/CRU archives. Single verified period at porting.
# NI TARIFF BASIS - rebuilt 8 Aug 2026. Read this before changing a
# number.
#
# WHAT CHANGED. The NI rows were previously a percentage chain off a
# single supplier - Firmus Energy (Ten Towns) - with the standing
# charge stripped out to leave a bare unit rate. Two problems. Firmus
# regulates about 75,756 customers while SSE Airtricity Gas Supply
# covers roughly 195,000 in Greater Belfast plus 3,200 in the West, so
# the smaller market was setting the NI gas bill. And the two
# suppliers are not structurally comparable: SSE's domestic tariff is
# banded with NO standing charge (its two published unit rates alone
# reproduce the typical bill exactly), while Firmus charges a unit
# rate plus a standing charge. Stripping "the standing charge" from
# one and not the other imports a tariff structure as a price.
#
# THE BASIS NOW. Effective all-in pence per kWh at the Utility
# Regulator's own stated consumption - 12,000 kWh for gas, 3,200 kWh
# for electricity - taken from the published annual bills, INCLUDING
# 5% VAT, and weighted across suppliers by regulated customer count
# (SSE 0.7235 / Firmus 0.2765). All-in rather than unit-only is also
# the right aggregate: the national bill is households x (standing +
# unit x consumption), and average NI domestic consumption sits near
# the Regulator's basis.
#
# EFFECT. NI gas rises about 12% for Oct 2025 and 14% for Apr 2026
# against the old rows, roughly half from restoring the standing
# charge and half from weighting in the larger and dearer supplier.
# NI electricity rises about 5%. These are corrections, not new
# estimates - but they move the NI headline, so revert here first if
# an NI figure looks wrong after this release.
#
# SOURCES. UREGNI tariff review briefing papers. SSE maximum average
# price 246.78 p/therm (eff 1 Apr 2024, unchanged to 30 Sep 2025) ->
# 225.87 (1 Oct 2025, -8.47%) -> 207.59 (1 Apr 2026, -8.10%), not
# reviewed in Jul 2026 - Annex 1 of each paper carries the series back
# to Apr 2015 if the record is ever extended further. Firmus bills
# GBP1,014 -> 934 (-7.86%, 1 Oct 2025) -> 840 (-10.1%, 1 Apr 2026) ->
# 972 (+15.7%, 1 Jul 2026); the three announced GBP steps reproduce
# those percentages to 0.04pp, which is the check that the chain is
# sound. Power NI GBP989 -> 1,029 (+4%, 1 Oct 2025) -> 1,093 (+6.2%,
# 1 Jul 2026), no change in Apr 2026.
#
# ROI DOMESTIC - rebuilt 8 Aug 2026 onto the same KIND of basis as
# NI. Previously a standard unit rate incl 9% VAT with no standing
# charge, which meant the two jurisdictions' domestic bills were not
# comparable at the component level: NI carried the standing charge
# and ROI did not.
#
# The like-for-like artefact is the Eurostat band price, which is
# computed as total revenue over volume for a consumption band and
# therefore includes standing charges by construction - the same
# property that made the REMM non-domestic bands work. Band D2
# (5,557-55,557 kWh gas) and medium domestic (2,500-4,999 kWh
# electricity), incl all taxes, semester 2 2024, from the 2024 AREMM
# data publication: Ireland 31.33 p/kWh electricity and 11.30 p/kWh
# gas, converted back to euro at 0.84, the semester average UREGNI
# used to sterling them in the first place.
#
# LEVEL from the semester, TIMING from the announcements - the same
# split the NI oil bridge uses. Electric Ireland held prices from Oct
# 2022 to 1 Jul 2026, so the S2 2024 level carries unchanged through
# the backfill row, Oct 2025 and Apr 2026, then steps +8% electricity
# and +7.7% gas on 1 Jul 2026.
#
# TWO DAGGERS, both worth knowing. (1) The band price is a market-wide
# average across every supplier and tariff, including discounts, while
# the ROI steps come from the incumbent's standard tariff - so
# non-incumbent movements between Dec 2024 and Jul 2026 are not
# captured, and cannot be until REMM publishes newer semesters. (2)
# The residual asymmetry with NI is now scope, not basis: NI is
# incumbent-weighted regulated bills, ROI is a market-wide average.
# Measured against the same tables, NI's regulated electricity runs
# about 7% ABOVE its own market band and its regulated gas about 18%
# BELOW - so the two do not bias in one direction and neither is a
# simple correction on the other.
#
# VAT CONVENTION (applies to both jurisdictions and to the UK sibling)
# Domestic rates INCLUDE VAT - 5% NI, 9% ROI. Non-domestic rates
# EXCLUDE it, because businesses recover input VAT and it is not a
# cost to them. That is Eurostat level 3 for households and level 2
# for non-households.
#
# NO CLAMP. tariffs_for() REFUSES any date before the first row rather
# than resolving it to that row. The clamp was live and wrong: once
# the price panel's floor came off it reached April 2024, and 68 weeks
# were quietly priced with August-2025 tariffs. Extending this table
# backwards is what restores them - UREGNI briefing papers carry dated
# annexes (SSE Airtricity to 2008, firmus to 2015) and SEAI publishes
# ROI cent/kWh to 2008 Q1.
TARIFF_HISTORY = [
    # (from_date, {eur: {electricity, gas}, gbp: {electricity, gas}})
    # BOTH are now all-in effective rates at a stated consumption,
    # including VAT and standing charges. gbp: UREGNI regulated bills,
    # gas weighted SSE/Firmus by customers. eur: Eurostat band prices
    # (S2 2024) stepped by the Electric Ireland announcements.
    # Rows below added at 5.12.0, from UREGNI's tariff-review news
    # releases, which publish the annual bill at this site's exact
    # basis (12,000 kWh gas, 3,200 kWh electricity, standard tariff,
    # VAT in) split by supplier pairing. Same construction as the
    # rows that follow: bill divided by consumption, gas weighted
    # SSE 0.7235 / Firmus 0.2765 by regulated customer count. The
    # 2025-04-01 row reproduces the old 2025-08-06 floor row exactly,
    # which is the check that the derivation is the right one.
    ("2024-04-01", {"eur": None,
                    "gbp": {"electricity": 0.2972, "gas": 0.0916}}),
    ("2024-12-01", {"eur": None,
                    "gbp": {"electricity": 0.3091, "gas": 0.0916}}),
    ("2025-04-01", {"eur": None,
                    "gbp": {"electricity": 0.3091, "gas": 0.0884}}),
    ("2025-08-06", {"eur": None,
                    "gbp": {"electricity": 0.3091, "gas": 0.0884}}),
    ("2025-10-01", {"eur": None,
                    "gbp": {"electricity": 0.3216, "gas": 0.0809}}),
    ("2026-04-01", {"eur": None,
                    "gbp": {"electricity": 0.3216, "gas": 0.0739}}),
    ("2026-07-01", {"eur": None,
                    "gbp": {"electricity": 0.3416, "gas": 0.0770}}),
]


def tariffs_for(date_iso):
    """
    Resolve the tariff period in force on a date. Dagger.

    The sterling side is a table of regulated all-in rates. The euro
    side is DERIVED at call time from the published Irish figures and
    the fetched semester rate, because a hard-coded euro row would go
    stale silently the moment the exchange rate moved - and it did,
    by about 4%, between the semester I first guessed and the one the
    anchors actually belong to.
    """
    if date_iso < TARIFF_HISTORY[0][0]:
        # REFUSE, do not clamp. The price panel reached back to April
        # 2024 the moment its floor came off, and 68 of those weeks
        # were being priced with August-2025 gas and electricity
        # tariffs - in both jurisdictions, silently, exactly the fault
        # just fixed for carbon. A missing figure is visible; a wrong
        # one is not. Extending TARIFF_HISTORY backwards is what
        # restores those weeks: UREGNI's briefing papers carry dated
        # annexes (SSE Airtricity to 2008, firmus to 2015) and SEAI
        # publishes ROI cent/kWh to 2008 Q1.
        return None
    row = TARIFF_HISTORY[0][1]
    for frm, rates in TARIFF_HISTORY:
        if date_iso >= frm:
            row = rates
    return {"gbp": row["gbp"],
            "eur": {"electricity": ie_eur("domestic_electricity", date_iso),
                    "gas": ie_eur("domestic_gas", date_iso)}}


def apply_ie_fx(feeds):
    """Set the euro conversion from the ECB semester mean and refresh
    the euro anchors that depend on it. Called once per run, before
    anything prices a week."""
    sem = ((feeds.get("ecb_fx") or {}).get("eur_gbp_semester") or {})
    rate = sem.get(IE_SEMESTER)
    if rate:
        IE_FX.update(rate=rate,
                     source=f"ECB semester mean {IE_SEMESTER}")
    else:
        IE_FX.update(rate=IE_FX_FALLBACK,
                     source="fallback (UNVERIFIED)")
        log(f"fx: WARNING {IE_SEMESTER} semester mean unavailable - "
            f"every Irish anchor is on the {IE_FX_FALLBACK} fallback")
    nd, used = nondom_for(today_utc().isoformat(), sem)
    ANCHORS["nondom_eur_per_kwh"] = nd["eur"]
    ANCHORS["nondom_gbp_per_kwh"] = nd["gbp"]
    log(f"fx: EUR/GBP {IE_FX['rate']} ({IE_FX['source']}); "
        f"ROI domestic {tariffs_for(HISTORY_START)['eur']}")
    lo, hi = IE_CREDIT_SENSITIVITY
    for sem in sorted(IE_DOMESTIC_SEMESTER):
        e, g, credit = IE_DOMESTIC_SEMESTER[sem]
        if not credit:
            log(f"roi domestic: {sem} elec {e:.1f}c (clean), gas {g:.1f}c")
            continue
        log(f"roi domestic: {sem} elec {e:.1f}c published -> "
            f"{e + 100 * credit / IE_CREDIT_KWH_PER_SEMESTER:.1f}c "
            f"credit-free (\u20ac{credit:.0f} over "
            f"{IE_CREDIT_KWH_PER_SEMESTER:.0f} kWh) \u2020; at "
            f"{lo:.0f}/{hi:.0f} kWh it would be "
            f"{e + 100 * credit / lo:.1f}/{e + 100 * credit / hi:.1f}c; "
            f"gas {g:.1f}c (no credit ever applied to gas)")
    log(f"nondom: live week on semester {used} "
        f"(latest published; REMM lags ~9 months) - "
        f"NI {nd['gbp']}, ROI {nd['eur']}")
    return IE_FX["rate"]


# ---------------------------------------------------------------- utilities

def log(*a):
    print(f"[{dt.datetime.now(dt.timezone.utc):%H:%M:%S}]", *a, flush=True)


def http_get(url, *, params=None, timeout=TIMEOUT, retries=RETRIES, headers=None):
    """GET with retries + exponential backoff + jitter. Raises on final failure."""
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, params=params, timeout=timeout,
                             headers=headers or UA)
            if r.status_code in (429, 500, 502, 503, 504):
                raise requests.HTTPError(f"{r.status_code} {r.reason}", response=r)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            last = e
            if attempt == retries:
                break
            wait = 2 ** attempt + random.uniform(0, 1.5)
            log(f"retry {attempt}/{retries - 1} after "
                f"{e.__class__.__name__} - sleeping {wait:.1f}s: {url[:80]}")
            time.sleep(wait)
    raise last


def today_utc():
    return dt.datetime.now(dt.timezone.utc).date()


def clip_days(day_values: dict) -> dict:
    """Drop future-dated keys (NESO lesson)."""
    cut = today_utc().isoformat()
    return {k: v for k, v in day_values.items() if k <= cut}


def trim_series(day_values: dict) -> dict:
    keep = (today_utc() - dt.timedelta(days=SERIES_KEEP_DAYS)).isoformat()
    return dict(sorted((k, v) for k, v in day_values.items() if k >= keep))


def autodetect_scale_to_gwh(values):
    med = statistics.median(abs(v) for v in values if v is not None) if values else 0
    if med > 1e6:
        return 1e-6, "kWh->GWh"
    if med > 1e3:
        return 1e-3, "MWh->GWh"
    return 1.0, "GWh"


def recency_status(latest_day: str | None, fresh_within_days: int) -> str:
    if not latest_day:
        return "stale"
    age = (today_utc() - dt.date.fromisoformat(latest_day)).days
    return "ok" if age <= fresh_within_days else "lagging"


def ddmmyyyy_to_iso(s: str) -> str | None:
    m = re.match(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s.strip())
    if not m:
        return None
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"


def find_date_in_text(s: str) -> str | None:
    """dd/mm/yyyy anywhere inside a longer string (bulletin title cells)."""
    m = re.search(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s)
    if not m:
        return None
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"


def prev_series(feed: str, *keys) -> dict:
    """Previously stored series for cross-run history accumulation."""
    node = PREVIOUS_FEEDS.get(feed, {})
    for k in keys:
        node = node.get(k, {}) if isinstance(node, dict) else {}
    return dict(node) if isinstance(node, dict) else {}


def _num(s) -> float | None:
    """Parse a number that may use a decimal comma (SEMOpx CSV convention)."""
    try:
        return float(str(s).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


# ------------------------------------------------- pure parsers (unit tested)

def extract_chart_data_arrays(page_html: str) -> list:
    """Bracket-matched '"data":[[...]]' chart payloads, entity-unescaped."""
    text = html_mod.unescape(page_html)
    out = []
    for m in re.finditer(r'"data"\s*:\s*(\[\[)', text):
        i = m.start(1)
        depth = 0
        for j in range(i, min(len(text), i + 2_000_000)):
            if text[j] == "[":
                depth += 1
            elif text[j] == "]":
                depth -= 1
                if depth == 0:
                    try:
                        out.append(json.loads(text[i:j + 1]))
                    except json.JSONDecodeError:
                        pass
                    break
    return out


def parse_ccni_series(page_html: str) -> dict:
    """{"300l": {iso: gbp}, ...} from litre-labelled embedded charts."""
    series = {"300l": {}, "500l": {}, "900l": {}}
    for arr in extract_chart_data_arrays(page_html):
        if not arr or not isinstance(arr[0], list):
            continue
        header = [str(c).lower() for c in arr[0]]
        if not any("litre" in h for h in header):
            continue
        cols = {}
        for idx, h in enumerate(header):
            for litres in ("300", "500", "900"):
                if litres in h:
                    cols[idx] = f"{litres}l"
        for row in arr[1:]:
            if not row:
                continue
            d = ddmmyyyy_to_iso(str(row[0]))
            if not d:
                continue
            for idx, key in cols.items():
                try:
                    v = float(row[idx])
                except (TypeError, ValueError, IndexError):
                    continue
                series[key][d] = round(v, 2)
    return series


def resolve_oil_bulletin_url(page_html: str, with_tax: bool = True) -> str | None:
    """Weekly prices xlsx, with or without taxes - unquote before matching."""
    for m in re.finditer(r'href="([^"]+)"', page_html):
        u = m.group(1)
        decoded = urllib.parse.unquote(u).lower()
        if ".xlsx" not in decoded:
            continue
        has_without = "without" in decoded
        has_with = re.search(r"with[ _]tax", decoded) is not None
        hit = (has_with and not has_without) if with_tax \
            else (has_without and "tax" in decoded)
        if hit:
            return u if u.startswith("http") else "https://energy.ec.europa.eu" + u
    return None


def parse_bulletin_rows(rows) -> tuple:
    """
    One-week-snapshot layout confirmed live (14 Jul 2026): a title cell
    carries the bulletin date; a header row names products ('...Heating...'
    at some column); country rows carry values with no per-row dates.
    Returns (iso_date | None, ireland_heating_value | None).
    Tolerates per-row dates too, should the layout ever grow them.
    """
    bulletin_date, idx_heat, value = None, None, None
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        joined = " ".join(strs).lower()

        if bulletin_date is None:
            for c in cells:
                if isinstance(c, dt.datetime):
                    bulletin_date = c.date().isoformat()
                    break
                if isinstance(c, dt.date):
                    bulletin_date = c.isoformat()
                    break
                iso = find_date_in_text(str(c)) if c is not None else None
                if iso:
                    bulletin_date = iso
                    break

        if idx_heat is None and ("heating" in joined or "chauffage" in joined):
            for i, c in enumerate(strs):
                if "heating" in c.lower() or "chauffage" in c.lower():
                    idx_heat = i
                    break
            continue

        if idx_heat is not None and value is None and (
                "ireland" in joined
                or (strs and strs[0].strip().upper() in ("IE", "EI"))):
            row_date = None
            for c in cells:
                if isinstance(c, (dt.datetime, dt.date)):
                    row_date = (c.date() if isinstance(c, dt.datetime)
                                else c).isoformat()
                    break
            try:
                v = float(cells[idx_heat])
            except (TypeError, ValueError, IndexError):
                continue
            value = round(v, 2)
            if row_date:
                bulletin_date = row_date
    return bulletin_date, value


def parse_bulletin_history_rows(rows, colpat=r"^IE_.*heating"):
    """
    EC oil bulletin price-history workbook - WIDE format (verified from a
    live run dump, 16 Jul 2026): one row per date in column 0, countries
    spread across columns headed e.g. 'IE_price_with_tax_heating_oil'.
    Falls back to long-table / country-block scanning if no wide header
    is found. Returns {iso_date: value}.
    """
    pat = re.compile(colpat, re.I)
    buffered = []
    idx = None
    for row in rows:
        cells = list(row)
        buffered.append(cells)
        for i, c in enumerate(cells):
            if c is not None and pat.match(str(c).strip()):
                idx = i
                break
        if idx is not None:
            break
    series = {}
    if idx is not None:
        def take(cells):
            if not cells:
                return
            d = cells[0]
            if isinstance(d, dt.datetime):
                d = d.date()
            if not isinstance(d, dt.date):
                return
            try:
                v = float(cells[idx])
            except (TypeError, ValueError, IndexError):
                return
            if 300 <= v <= 3000:
                series[d.isoformat()] = round(v, 2)
        for cells in rows:          # continue the iterator
            take(list(cells))
        return series
    # fallback: long-table / country-block layouts
    return _parse_history_longtable(buffered)


def _parse_history_longtable(rows) -> dict:
    EU = {"belgium", "bulgaria", "czechia", "czech republic", "denmark",
          "germany", "estonia", "greece", "spain", "france", "croatia",
          "italy", "cyprus", "latvia", "lithuania", "luxembourg", "hungary",
          "malta", "netherlands", "austria", "poland", "portugal",
          "romania", "slovenia", "slovakia", "finland", "sweden",
          "united kingdom", "uk"}
    idx_heat, series, in_ireland = None, {}, False
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        joined = " ".join(strs).lower()
        if idx_heat is None and ("heating" in joined or "chauffage" in joined):
            for i, c in enumerate(strs):
                if "heating" in c.lower() or "chauffage" in c.lower():
                    idx_heat = i
                    break
            continue
        if idx_heat is None:
            continue
        is_ireland_row = ("ireland" in joined
                          or (strs and strs[0].strip().upper() in ("IE", "EI")))
        if is_ireland_row:
            in_ireland = True
        elif any(c in joined for c in EU):
            in_ireland = False
        if not (is_ireland_row or in_ireland):
            continue
        row_date = None
        for c in cells:
            if isinstance(c, dt.datetime):
                row_date = c.date().isoformat()
                break
            if isinstance(c, dt.date):
                row_date = c.isoformat()
                break
            iso = find_date_in_text(str(c)) if c is not None else None
            if iso:
                row_date = iso
                break
        if row_date is None:
            continue
        try:
            v = float(cells[idx_heat])
        except (TypeError, ValueError, IndexError):
            continue
        if 300 <= v <= 3000:
            series[row_date] = round(v, 2)
    return series


def semopx_hour_keys(stamps, trade_day):
    """
    Delivery stamps -> Irish local-clock hour keys, ANCHORED on the
    trade day rather than converted by an assumed timezone.

    Two guesses have now been wrong. The stamps carry a Z suffix, so
    the first pass treated them as UTC and added an hour in summer;
    the span still came out one hour below local midnight, which
    pointed at CET - and a third guess would have been a third
    coin-toss. The document does not need decoding: a trade day runs
    00:00 to 23:00 local by definition, and the resource name states
    which day it is. So the offset is measured from the document -
    first stamp against local midnight - and applied to the rest.

    Returns [] if the trade day is unknown or nothing parses, so a
    document that cannot be anchored contributes nothing rather than
    contributing something misaligned.
    """
    parsed = []
    for t in stamps:
        m = re.match(r"(20\d\d)-(\d\d)-(\d\d)T(\d\d)", str(t))
        parsed.append(dt.datetime(*(int(g) for g in m.groups()))
                      if m else None)
    real = [x for x in parsed if x]
    if not real or not trade_day:
        return []
    try:
        anchor = dt.datetime.strptime(trade_day, "%Y-%m-%d")
    except ValueError:
        return []
    offset = anchor - min(real)
    return [(x + offset).strftime("%Y-%m-%dT%H") if x else None
            for x in parsed]


def parse_semopx_csv(text: str, trade_day: str = None) -> dict:
    """
    SEMOpx MarketResult CSV, format confirmed live (14 Jul 2026):
    semicolon-delimited, decimal commas, sections -
        Auction;SEM-DA
        FX rates
        EUR;GBP;0,85506627
        Market;NI-DA
        Index prices;30;EUR
        <row of ISO delivery timestamps>
        <row of prices>
    Returns {"fx_eur_gbp", "day", "auction", "markets"}.
    """
    fx, day, auction = None, None, None
    markets: dict = {}
    series: dict = {}
    stamps: dict = {}
    market, currency, expect_series = None, None, False
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        key = parts[0].lower()

        if key == "auction" and len(parts) > 1:
            auction = parts[1]
            continue
        if key == "market" and len(parts) > 1:
            market = parts[1]
            expect_series = False
            continue
        if key.startswith("index prices"):
            currency = parts[-1].upper() if len(parts) >= 2 else "EUR"
            expect_series = True
            continue
        if key == "eur" and len(parts) >= 3 and parts[1].upper() == "GBP":
            fx = _num(parts[2])
            continue
        if expect_series:
            if re.match(r"20\d\d-\d\d-\d\dT", parts[0]):
                if day is None:
                    day = parts[0][:10]
                # The delivery stamps were being read and discarded.
                # They are the only thing that makes a price hourly,
                # and B.2.3 asks what price applied in ONE hour.
                stamps[(market, currency)] = list(parts)
                continue
            nums = [n for n in (_num(p) for p in parts) if n is not None]
            if nums and market and currency:
                markets.setdefault(market, {}).setdefault(
                    currency, []).extend(nums)
                ts = stamps.get((market, currency)) or []
                for k, v in zip(semopx_hour_keys(ts, trade_day or day),
                                nums):
                    if k:
                        series.setdefault(market, {}).setdefault(
                            currency, {})[k] = v
                expect_series = False
    return {"fx_eur_gbp": fx, "day": day, "auction": auction,
            "markets": markets, "series": series}


def parse_gni_series(series_list) -> dict:
    """
    Pure parser for the GNI gasconsumption JSON API. Input: list of series
    objects {name, location, group, ..., data: [[unix_ms, value], ...]}.
    Output: {location: {iso_date: value}}. Keys off `location`, tolerates
    missing fields, skips unparseable points.
    """
    out = {}
    for s in series_list or []:
        loc = (s or {}).get("location")
        if not loc:
            continue
        pts = {}
        for pair in s.get("data") or []:
            try:
                ms, val = pair[0], pair[1]
                d = dt.datetime.fromtimestamp(
                    ms / 1000.0, tz=dt.timezone.utc).date().isoformat()
                pts[d] = float(val)
            except (TypeError, ValueError, IndexError, OSError):
                continue
        if pts:
            out.setdefault(loc, {}).update(pts)
    return out




def parse_gb_oil_page(text: str) -> tuple:
    """
    BoilerJuice prices page - server-rendered sentence:
      'Our average heating oil price for today, Saturday 25th March 2017
       is 40.32 pence per litre (inc. VAT)'
    Returns (iso_date | None, pence_per_litre | None). Date falls back to
    None if unparseable - caller may substitute the run date.
    """
    m = re.search(
        r"average (?:kerosene|heating oil) price for today[^0-9]*?"
        r"(?:(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4}))?"
        r"[^0-9]*?is\s*(\d{1,3}\.\d{1,2})\s*pence per litre",
        text, re.I | re.S)
    if not m:
        # modern template (observed 18 Jul 2026): dated sentence gone,
        # but the current price is server-rendered beside the chart -
        # digits, optional markup, then 'pence per litre'. Undated;
        # the caller stamps the run date. Fossils cannot reach here:
        # they carry the dated 2021 sentence, matched above and then
        # rejected by the caller's freshness gate.
        m2 = re.search(
            r"(\d{2,3}\.\d{1,2})\s*(?:</?[a-z][^>]*>\s*|&nbsp;)*"
            r"pence per litre", text, re.I)
        if m2:
            v = float(m2.group(1))
            if 40.0 <= v <= 150.0:
                return None, v
        return None, None
    day, month_name, year, ppl = m.group(1), m.group(2), m.group(3), m.group(4)
    iso = None
    if day and month_name and year:
        months = {"january": 1, "february": 2, "march": 3, "april": 4,
                  "may": 5, "june": 6, "july": 7, "august": 8,
                  "september": 9, "october": 10, "november": 11,
                  "december": 12}
        mo = months.get(month_name.lower())
        if mo:
            iso = f"{int(year):04d}-{mo:02d}-{int(day):02d}"
    return iso, float(ppl)


# ---------------------------------------------------------------- feeds

def parse_eirgrid_rows(payload, field="SYSTEM_DEMAND", daily="gwh",
                       min_intervals=48):
    """
    /api/chart/ response -> {iso_date: value}. Rows for the requested
    field are aggregated per day: daily="gwh" averages MW and converts
    (mean x 24 / 1000); daily="mean" returns the plain daily mean (used
    for CO2 intensity, gCO2/kWh). Nulls and other fields are ignored;
    days with fewer than min_intervals observations are dropped as
    incomplete. Pure, unit tested.
    """
    acc = {}
    for row in (payload or {}).get("Rows", []):
        if row.get("FieldName") != field or row.get("Value") is None:
            continue
        et = str(row.get("EffectiveTime", ""))
        try:
            iso = dt.datetime.strptime(et.split()[0], "%d-%b-%Y")\
                    .date().isoformat()
        except (ValueError, IndexError):
            iso = find_date_in_text(et)
        if not iso:
            continue
        s, n = acc.get(iso, (0.0, 0))
        acc[iso] = (s + float(row["Value"]), n + 1)
    if daily == "mean":
        return {d: round(s / n, 1)
                for d, (s, n) in acc.items() if n >= min_intervals}
    return {d: round(s / n * 24.0 / 1000.0, 2)
            for d, (s, n) in acc.items() if n >= min_intervals}


def feed_eirgrid():
    """
    Smart Grid Dashboard via the /api/chart/ endpoint (probed 18 Jul
    2026). Daily electricity demand in GWh for ALL / NI / ROI, merged
    across runs. Today is always incomplete and excluded by the parser's
    48-interval rule; each run therefore backfills the last full days.
    """
    def dmy(d):
        return d.strftime("%d-%b-%Y")

    out = {"source": "EirGrid Smart Grid Dashboard (/api/chart/)",
           "demand_gwh_daily": {}}
    end = today_utc()
    start = end - dt.timedelta(days=8)
    for region, key in (("ALL", "island"), ("NI", "ni"), ("ROI", "roi")):
        try:
            payload = http_get(EIRGRID_ENDPOINT, params={
                "region": region, "chartType": "demand",
                "dateRange": "month", "dateFrom": dmy(start),
                "dateTo": dmy(end), "areas": "demandactual",
            }, timeout=90).json()
        except Exception as e:
            log(f"eirgrid: {region} span fetch failed "
                f"({e.__class__.__name__}: {e})")
            payload = {}
        got = parse_eirgrid_rows(payload)
        if not got:
            rows = (payload or {}).get("Rows", [])
            log(f"eirgrid: {region} span returned {len(rows)} rows, "
                f"0 complete days - falling back to per-day calls")
            got = {}
            for i in range(1, 8):
                d = end - dt.timedelta(days=i)
                try:
                    pl = http_get(EIRGRID_ENDPOINT, params={
                        "region": region, "chartType": "demand",
                        "dateRange": "day", "dateFrom": dmy(d),
                        "dateTo": dmy(d), "areas": "demandactual",
                    }, timeout=60).json()
                    got.update(parse_eirgrid_rows(pl))
                except Exception as e:
                    log(f"eirgrid: {region} {d} failed ({e})")
        series = prev_series("eirgrid", "demand_gwh_daily", key)
        series.update(got)
        out["demand_gwh_daily"][key] = trim_series(series)
        log(f"eirgrid: {region} {len(got)} days this run, "
            f"{len(out['demand_gwh_daily'][key])} retained")

    # CO2 intensity - schema confirmed in the 18 Jul 2026 run log
    # (CO2_INTENSITY rows, same shape). Daily mean gCO2/kWh, island,
    # merged across runs; soft.
    try:
        pl = http_get(EIRGRID_ENDPOINT, params={
            "region": "ALL", "chartType": "co2",
            "dateRange": "month", "dateFrom": dmy(end - dt.timedelta(days=150)),
            "dateTo": dmy(end), "areas": "co2intensity"},
            timeout=90).json()
        got = parse_eirgrid_rows(pl, field="CO2_INTENSITY",
                                 daily="mean", min_intervals=40)
        if not got:
            for i in range(1, 8):
                d = end - dt.timedelta(days=i)
                pl = http_get(EIRGRID_ENDPOINT, params={
                    "region": "ALL", "chartType": "co2",
                    "dateRange": "day", "dateFrom": dmy(d),
                    "dateTo": dmy(d), "areas": "co2intensity"},
                    timeout=60).json()
                got.update(parse_eirgrid_rows(
                    pl, field="CO2_INTENSITY", daily="mean",
                    min_intervals=40))
        ser = prev_series("eirgrid", "co2_intensity_g_per_kwh")
        ser.update(got)
        # DEEP BACKFILL, once. Extending the weekly record needs each
        # week's OWN carbon, and the daily feed retains 50 days while
        # the hourly store starts 13 months back - so weeks older than
        # the store have no carbon at all unless it is fetched. The
        # probe has shown co2intensity walking back 18 months. This
        # walks monthly chunks until the series reaches HISTORY_START
        # or a chunk comes back empty, and then never runs again,
        # because the retained series already covers the floor.
        if ser and min(ser) > HISTORY_START:
            edge = dt.date.fromisoformat(min(ser))
            walked = 0
            for _ in range(24):
                if edge.isoformat() <= HISTORY_START:
                    break
                a = edge - dt.timedelta(days=28)
                chunk = parse_eirgrid_rows(http_get(
                    EIRGRID_ENDPOINT, params={
                        "region": "ALL", "chartType": "co2",
                        "dateRange": "month", "dateFrom": dmy(a),
                        "dateTo": dmy(edge - dt.timedelta(days=1)),
                        "areas": "co2intensity"}, timeout=90).json(),
                    field="CO2_INTENSITY", daily="mean", min_intervals=40)
                if not chunk:
                    log(f"eirgrid: carbon backfill stopped at "
                        f"{edge.isoformat()} - chunk from {a.isoformat()} "
                        f"came back empty (0-row results have been "
                        f"flakiness before, so this may clear itself)")
                    break
                ser.update(chunk)
                walked += 1
                edge = a
            log(f"eirgrid: carbon backfill walked {walked} chunk(s)")
        out["co2_intensity_g_per_kwh"] = trim_series(ser)
        # UNCONDITIONAL. This line used to sit inside the branch above,
        # so on any run where the backfill did not fire it said nothing
        # at all - and "nothing" is indistinguishable from "the block
        # never ran". A diagnostic that goes quiet in the case you
        # cannot otherwise tell apart is the one case it is for. It now
        # reports where the carbon record reaches every run, whether
        # the backfill walked or not.
        cser = out["co2_intensity_g_per_kwh"]
        if cser:
            reach = min(cser)
            log(f"eirgrid: carbon reaches {reach} against HISTORY_START "
                f"{HISTORY_START} - "
                + ("covered" if reach <= HISTORY_START else
                   "SHORT, weeks before it will be refused for want of "
                   "their own carbon"))
        log(f"eirgrid: co2 intensity {len(got)} days this run, "
            f"{len(out['co2_intensity_g_per_kwh'])} retained")
    except Exception as e:
        log(f"eirgrid: co2 parse failed ({e.__class__.__name__}: {e})")

    isl = out["demand_gwh_daily"].get("island") or {}
    out["latest_day"] = max(isl) if isl else None
    return out, recency_status(out["latest_day"], 4)


def odh26_from_hourly(payload, names, weights, base_c=26.0):
    """
    Population-weighted overheating degree-hours per day from Open-Meteo
    hourly payloads: sum over hours of max(0, T - 26), weighted across
    stations, keyed by UTC date. Pure, unit tested.
    """
    locs = payload if isinstance(payload, list) else [payload]
    per_day = {}
    for name, loc in zip(names, locs):
        w = weights.get(name, 0.0)
        hh = loc.get("hourly", {})
        for ts, t in zip(hh.get("time", []), hh.get("temperature_2m", [])):
            if t is None:
                continue
            d = ts[:10]
            per_day[d] = per_day.get(d, 0.0) + w * max(0.0, t - base_c)
    return {d: round(v, 2) for d, v in per_day.items()}


# MIDLANDS PROBE - RAN AND RETIRED, 14 Aug 2026. All seven weighted
# stations sit on or near the coast, so the island series is entirely
# maritime, and Ireland's interior runs colder on winter nights. The
# worry was that the weighted HDD is biased low, and with it the
# space-heat share of every week.
#
# Measured rather than argued: Athlone was fetched alongside the seven
# at the midland counties' share of island population (Longford,
# Westmeath, Offaly, Laois ~ 4%) over 1,151 days. Island HDD moved
# 4.65 -> 4.66, +0.28%; winter alone 8.90 -> 8.92, +0.27%. Immaterial,
# so the all-coastal set stands and the probe is retired rather than
# carried for ever.
#
# It also left a station in the fetch that STATIONS did not know
# about, which broke the ODH26 loop with a KeyError - a reason to
# retire a probe once it has answered rather than leave it running.
PROBE_STATIONS: dict = {}


def feed_hdd():
    """Open-Meteo, batched; forecast tail optional (degrades to lagging)."""
    names = list(STATIONS) + list(PROBE_STATIONS)
    _all = {**STATIONS, **PROBE_STATIONS}
    lats = ",".join(str(_all[n][0]) for n in names)
    lons = ",".join(str(_all[n][1]) for n in names)

    def unpack(payload):
        locs = payload if isinstance(payload, list) else [payload]
        per_station = {}
        for name, loc in zip(names, locs):
            d = loc.get("daily", {})
            per_station[name] = {
                day: t for day, t in zip(d.get("time", []),
                                         d.get("temperature_2m_mean", []))
                if t is not None
            }
        return per_station

    arch = unpack(http_get(
        "https://archive-api.open-meteo.com/v1/archive", params={
            "latitude": lats, "longitude": lons,
            "start_date": (today_utc()
                           - dt.timedelta(days=SERIES_KEEP_DAYS)).isoformat(),
            "end_date": today_utc().isoformat(),
            "daily": "temperature_2m_mean", "timezone": "UTC",
        }, timeout=120).json())

    tail_ok = True
    try:
        tail = unpack(http_get(
            "https://api.open-meteo.com/v1/forecast", params={
                "latitude": lats, "longitude": lons, "past_days": 10,
                "forecast_days": 1, "daily": "temperature_2m_mean",
                "timezone": "UTC",
            }).json())
        for n in names:
            for d, t in tail.get(n, {}).items():
                arch[n].setdefault(d, t)
    except Exception as e:
        tail_ok = False
        log(f"hdd: forecast tail unavailable ({e.__class__.__name__}) - "
            "continuing archive-only, expect 'lagging'")

    daily_by_station = {n: clip_days(v) for n, v in arch.items()}

    def weighted_temp(subset):
        """Population-weighted daily mean AIR TEMPERATURE.

        HDD is max(0, base - t), so it throws the temperature away
        above the base - and the COP engine needs the temperature in
        summer as much as in winter, because that is when hot water is
        the whole load. Same weights, same days, one line of extra
        state.
        """
        wsum = sum(STATIONS[n][2] for n in subset)
        days = set.intersection(*(set(daily_by_station[n]) for n in subset))
        return {d: round(sum(daily_by_station[n][d] * STATIONS[n][2]
                             for n in subset) / wsum, 2)
                for d in sorted(days)}

    def weighted(subset):
        wsum = sum(STATIONS[n][2] for n in subset)
        days = set.intersection(*(set(daily_by_station[n]) for n in subset))
        return {
            d: round(max(0.0, HDD_BASE_C - sum(
                daily_by_station[n][d] * STATIONS[n][2] for n in subset) / wsum), 2)
            for d in sorted(days)
        }

    weighted_names = [n for n in names if n in STATIONS]
    roi = [n for n in weighted_names if STATIONS[n][3] == "ROI"]
    ni = [n for n in weighted_names if STATIONS[n][3] == "NI"]
    out = {
        "hdd_island": trim_series(weighted(weighted_names)),
        "temp_island": trim_series(weighted_temp(weighted_names)),
        "temp_roi": trim_series(weighted_temp(roi)),
        "temp_ni": trim_series(weighted_temp(ni)),
        "hdd_roi": trim_series(weighted(roi)),
        "hdd_ni": trim_series(weighted(ni)),
        "base_c": HDD_BASE_C,
        "forecast_tail": tail_ok,
        "weights_note": ("Population weights are current Causeway Energies "
                         "estimates - challenge and input welcome at "
                         "contact@causewaygt.com"),
        "source": "ERA5 via Open-Meteo, population-weighted HDD",
    }

    # --- midlands probe (log-only, changes nothing; empty once answered)
    try:
        if "Athlone" in daily_by_station and daily_by_station["Athlone"]:
            share = PROBE_STATIONS["Athlone"][2]
            days = set(daily_by_station["Athlone"])
            for n in weighted_names:
                days &= set(daily_by_station[n])
            days = sorted(days)
            if len(days) > 300:
                wsum = sum(STATIONS[n][2] for n in weighted_names)
                base, withm = [], []
                for d in days:
                    t7 = sum(daily_by_station[n][d] * STATIONS[n][2]
                             for n in weighted_names) / wsum
                    t8 = ((1 - share) * t7
                          + share * daily_by_station["Athlone"][d])
                    base.append(max(0.0, HDD_BASE_C - t7))
                    withm.append(max(0.0, HDD_BASE_C - t8))
                a, b = sum(base) / len(base), sum(withm) / len(withm)
                pct = 100 * (b - a) / a if a else 0.0
                # Winter is what matters: a summer difference cannot
                # move a space-heat share that is already near zero.
                wi = [i for i, d in enumerate(days)
                      if d[5:7] in ("12", "01", "02")]
                wa = sum(base[i] for i in wi) / max(len(wi), 1)
                wb = sum(withm[i] for i in wi) / max(len(wi), 1)
                wpct = 100 * (wb - wa) / wa if wa else 0.0
                log(f"hdd: MIDLANDS PROBE over {len(days)} days - island "
                    f"HDD {a:.2f} without Athlone, {b:.2f} with it at "
                    f"{share:.0%} ({pct:+.2f}%); winter only {wa:.2f} -> "
                    f"{wb:.2f} ({wpct:+.2f}%)")
                log("hdd:   log-only. Under ~1% the all-coastal station "
                    "set is fine as it stands; more than that and the "
                    "weighted series needs a midlands member, which "
                    "would move every HDD-shaped figure on the site")
    except Exception as exc:
        log(f"hdd: midlands probe failed ({exc.__class__.__name__}) - "
            "log-only, the feed is unaffected")

    # ODH26 groundwork: hourly overheating-degree-hours (base 26 C),
    # population-weighted, trailing 60 days per run, merged across runs.
    # Collected for the future comfort/cooling metric; not yet displayed.
    # Soft: any failure leaves the hdd feed intact.
    try:
        hp = http_get(
            "https://archive-api.open-meteo.com/v1/archive", params={
                "latitude": lats, "longitude": lons,
                "start_date": (today_utc()
                               - dt.timedelta(days=60)).isoformat(),
                "end_date": today_utc().isoformat(),
                "hourly": "temperature_2m", "timezone": "UTC",
            }, timeout=180).json()
        odh_new = odh26_from_hourly(
            # weighted_names, not names: a probe station is fetched
            # but is not in STATIONS, and passing it here raised a
            # KeyError that silently stopped ODH collection on
            # 14 Aug 2026. Anything downstream of the fetch reads the
            # WEIGHTED set.
            hp, weighted_names,
            {n: STATIONS[n][2] for n in weighted_names})
        odh = prev_series("hdd", "odh26_island")
        odh.update(odh_new)
        out["odh26_island"] = trim_series(odh)
        log(f"hdd: odh26 {len(odh_new)} days this run, "
            f"{len(out['odh26_island'])} retained, "
            f"nonzero {sum(1 for v in out['odh26_island'].values() if v>0)}")
    except Exception as e:
        log(f"hdd: odh26 groundwork skipped ({e.__class__.__name__}: {e})")
    latest = max(out["hdd_island"] or {"": None})
    out["latest_day"] = latest or None
    return out, recency_status(out["latest_day"], 3 if tail_ok else 7)


def feed_ecb_fx():
    r = http_get("https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml")
    root = ElementTree.fromstring(r.content)
    ns = {"e": "http://www.ecb.int/vocabulary/2002-08-01/eurofxref"}
    cube_day = root.find(".//e:Cube[@time]", ns)
    rate = None
    for c in cube_day.findall("e:Cube", ns):
        if c.get("currency") == "GBP":
            rate = float(c.get("rate"))
    if rate is None:
        avail = [c.get("currency") for c in cube_day.findall("e:Cube", ns)]
        log("ecb_fx: GBP missing - available currencies:", avail)
        raise ValueError("GBP not in ECB daily cube")
    out = {"eur_gbp": rate, "gbp_eur": round(1 / rate, 5),
           "rate_date": cube_day.get("time"), "latest_day": cube_day.get("time"),
           "source": "ECB euro foreign exchange reference rates (daily)"}
    # 90-day history for per-week pricing of the back-look. Soft.
    try:
        r90 = http_get("https://www.ecb.europa.eu/stats/eurofxref/"
                       "eurofxref-hist-90d.xml")
        root90 = ElementTree.fromstring(r90.content)
        ser = prev_series("ecb_fx", "eur_gbp_daily")
        for cube in root90.findall(".//e:Cube[@time]", ns):
            for c in cube.findall("e:Cube", ns):
                if c.get("currency") == "GBP":
                    ser[cube.get("time")] = float(c.get("rate"))
        sem = dict(prev_series("ecb_fx", "eur_gbp_semester"))
        sem.update(semester_means(ser))
        if not ser or min(ser) > HISTORY_START or IE_SEMESTER not in sem:
            # Deep backfill to the back-look floor. Also fires when the
            # Irish anchor semester is not yet covered - the daily
            # series starts at HISTORY_START, which can sit inside the
            # semester and leave the mean short of its day count.
            import zipfile
            zr = http_get("https://www.ecb.europa.eu/stats/eurofxref/"
                          "eurofxref-hist.zip", timeout=120)
            with zipfile.ZipFile(io.BytesIO(zr.content)) as z:
                txt = z.read(z.namelist()[0]).decode("utf8")
            lines = [l for l in txt.splitlines() if l.strip()]
            cols = lines[0].split(",")
            gi = cols.index("GBP")
            added = 0
            # The zip is the FULL series back to 1999, so semester
            # means come off it whole rather than off the trimmed
            # daily window - which is the point of fetching it here.
            whole = {}
            for line in lines[1:]:
                parts = line.split(",")
                d = parts[0].strip()
                try:
                    whole[d] = float(parts[gi])
                except (ValueError, IndexError):
                    pass
                if d < HISTORY_START or d in ser:
                    continue
                try:
                    ser[d] = float(parts[gi])
                    added += 1
                except (ValueError, IndexError):
                    continue
            sem.update(semester_means(whole))
            log(f"ecb_fx: deep history backfill +{added} days, "
                f"{len(whole)} days seen, {len(sem)} semester means")
        out["eur_gbp_daily"] = trim_series(ser)
        out["eur_gbp_semester"] = sem
        log(f"ecb_fx: history {len(out['eur_gbp_daily'])} days retained, "
            f"{len(sem)} semester means; {IE_SEMESTER}="
            f"{sem.get(IE_SEMESTER, 'MISSING')}")
    except Exception as e:
        out["eur_gbp_daily"] = prev_series("ecb_fx", "eur_gbp_daily")
        out["eur_gbp_semester"] = prev_series("ecb_fx", "eur_gbp_semester")
        log(f"ecb_fx: history skipped ({e.__class__.__name__})")
    return out, recency_status(out["latest_day"], 5)


def feed_gni_ckan():
    """data.gov.ie CKAN - GNI daily demand by sector, CC BY 4.0, quarterly."""
    pkg = http_get("https://data.gov.ie/api/3/action/package_search",
                   params={"q": "daily gas demand", "rows": 10}).json()
    results = pkg.get("result", {}).get("results", [])
    resource_url = None
    for ds in results:
        if "gas networks ireland" not in json.dumps(
                ds.get("organization", {})).lower() \
           and "gas" not in ds.get("title", "").lower():
            continue
        for res in ds.get("resources", []):
            if res.get("format", "").upper() == "CSV" \
               and "demand" in (res.get("name", "") + ds.get("title", "")).lower():
                resource_url = res.get("url")
                break
        if resource_url:
            break
    if not resource_url:
        log("gni_ckan: no CSV resource matched - datasets found:",
            [d.get("title") for d in results])
        raise ValueError("CKAN resolution failed")
    log("gni_ckan: resolved", resource_url)

    csv_text = http_get(resource_url).text
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    header = [h.strip().strip('"') for h in lines[0].split(",")]
    log("gni_ckan: header:", header)

    def col(*needles):
        for i, h in enumerate(header):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    cols = {
        "ndm": col("ndm") if col("ndm") is not None else col("non", "daily"),
        "dm": None,
        "ldm": col("ldm"),
        "power": col("power"),
        "total_roi": col("total"),
    }
    for i, h in enumerate(header):
        if h.lower().startswith("daily metered"):
            cols["dm"] = i
            break
    if cols["ndm"] is None:
        raise ValueError(f"gni_ckan: NDM column not found in {header}")

    series = {k: {} for k in cols if cols[k] is not None}
    i_date = col("date") if col("date") is not None else 0
    for ln in lines[1:]:
        parts = [p.strip().strip('"') for p in ln.split(",")]
        raw_d = parts[i_date][:10]
        d = ddmmyyyy_to_iso(raw_d) or raw_d.replace("/", "-")
        try:
            d = dt.date.fromisoformat(d).isoformat()
        except ValueError:
            continue
        for k, i in cols.items():
            if i is None:
                continue
            try:
                series[k][d] = float(parts[i])
            except (ValueError, IndexError):
                continue

    scale, unit_note = autodetect_scale_to_gwh(list(series["ndm"].values()))
    out = {"unit_detection": unit_note,
           "latest_day": max(series["ndm"]) if series["ndm"] else None,
           "source": ("Gas Networks Ireland via data.gov.ie, CC BY 4.0 - "
                      "quarterly refresh, calibration series")}
    for k, vals in series.items():
        out[f"{k}_gwh"] = clip_days(
            {d: round(v * scale, 2) for d, v in vals.items()})
    return out, recency_status(out["latest_day"], 100)


def feed_gni_live():
    """
    GNI Data Transparency JSON API - probed 14 Jul 2026, daily/hourly/
    monthly all 200. The CSV export route 503s and is not used. Window per
    call unknown: anchor four dates a month apart, log the observed span,
    merge across runs. Values arrive in kWh by default - unit autodetected.
    JURISDICTION FLAG: DM/NDM carry no ROI prefix while LDM and Power Gen
    do - whether unprefixed series include NI exits is unconfirmed; the
    space-heat regression stays on the confirmed-ROI gni_ckan series until
    resolved.
    """
    base = "https://www.gasnetworks.ie/api/v1/gasconsumption"
    raw = {}
    for back in range(0, 92, 7):   # API window ~8 trailing days
        time.sleep(0.3)
        anchor = (today_utc() - dt.timedelta(days=back)).isoformat()
        try:
            body = http_get(base, params={
                "date": anchor, "frequency": "daily", "unit": ""}).json()
        except Exception as e:
            log(f"gni_live: anchor {anchor} failed - {e.__class__.__name__}: {e}")
            continue
        if not isinstance(body, list):
            log("gni_live: unexpected shape for", anchor, "-", str(body)[:300])
            continue
        parsed = parse_gni_series(body)
        if back == 0:
            log("gni_live: locations seen:", sorted(parsed))
        ndm = parsed.get("NDM", {})
        log(f"gni_live: anchor {anchor} -> NDM {len(ndm)} pts "
            f"{min(ndm) if ndm else '-'}..{max(ndm) if ndm else '-'}")
        for loc, pts in parsed.items():
            raw.setdefault(loc, {}).update(pts)
    if not raw:
        raise ValueError("gni_live: no series parsed from any anchor")

    spans = {loc: (min(p), max(p), len(p)) for loc, p in raw.items()}
    log("gni_live: spans:", spans)

    # Jurisdiction signal (standing reminder, 24 Jul 2026): if Total LDM
    # minus ROI LDM is ~zero, "Total" is ROI-total naming and the
    # unprefixed DM/NDM series are very likely ROI-scoped; a material,
    # stable difference means the feed carries NI exits. Logged every
    # run so a convention change announces itself.
    tot, roi = raw.get("Total LDM", {}), raw.get("ROI LDM", {})
    common = sorted(set(tot) & set(roi))[-28:]
    if len(common) >= 7:
        diffs = [tot[d] - roi[d] for d in common]
        mean_d = sum(diffs) / len(diffs)
        mean_t = sum(tot[d] for d in common) / len(common)
        pct = 100 * mean_d / mean_t if mean_t else 0.0
        log(f"gni_live: jurisdiction check - Total LDM minus ROI LDM "
            f"mean {mean_d:.3f} ({pct:.2f}% of Total) over "
            f"{len(common)}d "
            + ("-> ~zero: unprefixed series read as ROI-scoped"
               if abs(pct) < 1.0 else
               "-> MATERIAL: feed appears to include NI exits - "
               "review the jurisdiction note"))

    ndm_vals = list(raw.get("NDM", {}).values())
    scale, unit_note = autodetect_scale_to_gwh(
        ndm_vals or [v for p in raw.values() for v in p.values()])

    def keyname(loc):
        return loc.lower().replace(" ", "_")

    out = {"unit_detection": unit_note,
           "jurisdiction_note": ("DM/NDM carry no ROI prefix while LDM and "
                                 "Power Gen do - whether unprefixed series "
                                 "include NI exits is unconfirmed; regression "
                                 "remains on the confirmed-ROI calibration "
                                 "series until resolved"),
           "source": ("Gas Networks Ireland Data Transparency - "
                      "gasconsumption API, daily by market sector")}
    latest = None
    for loc, pts in raw.items():
        merged = prev_series("gni_live", f"{keyname(loc)}_gwh")
        merged.update({d: round(v * scale, 3) for d, v in pts.items()})
        merged = trim_series(clip_days(merged))
        out[f"{keyname(loc)}_gwh"] = merged
        if merged:
            latest = max(latest or "", max(merged))
    out["latest_day"] = latest
    return out, recency_status(latest, 3)


def semopx_trade_day(item):
    """Trade day from a SEMOpx resource name, e.g.
    MarketResult_SEM-DA_PWR-MRC-D+1_20260806100000_... -> 2026-08-06.

    The first probe sliced this positionally and cut mid-timestamp,
    which turned every day tag into nonsense and made the verdict line
    say the opposite of what the data showed. Match the field, do not
    count characters into it."""
    m = re.search(r"_(20\d{6})\d{6}_", str((item or {}).get("ResourceName")
                                            or item or ""))
    return f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}" \
        if m else None


def semo_dispatch_probe():
    """
    Where does per-UNIT downward dispatch live, and how far back? Log
    only. Nothing is stored, nothing is parsed into the payload.

    WHY THIS IS URGENT AND THE FEED IS NOT. Fintan Devenney (Montel)
    confirmed on 9 Aug 2026 that per-unit dispatch-down volumes are
    published on SEMO and that building against SEMO's reports would
    be fine for live data - but that SEMO does NOT retain the full
    history, which is why Montel serve their own store back to 2018.
    So the series exists only from the day capture starts. Every day
    without a feed is a day that cannot be recovered later, which
    makes finding the report the most time-critical thing in B.2.2
    even though the panel itself does not depend on it.

    WHAT THE ANSWER HAS TO CONTAIN before a parser is worth writing:
      - which report ID carries per-unit volumes (SEMO publishes
        dozens; the naming is not self-describing)
      - the retention window, since that sets how often the feed must
        run to lose nothing
      - the resolution (half-hourly imbalance periods, most likely)
      - what identifies a unit. This is the one that shapes the
        register: SEMO will name resource codes, not wind farms, so
        the map needs code -> farm -> coordinates -> capacity and
        only the first hop comes from here.

    Two families are tried because SEMO runs two: the BM/balancing
    reporting on the main site, and the semopx static-report API this
    pipeline already uses for prices. Candidates are listed, not
    chosen - a guessed report ID is how the first semopx probe wasted
    a day on 400s.
    """
    hits = []

    # --- family 1: the static-report listing this pipeline knows
    base = "https://reports.semopx.com/api/v1/documents/static-reports"
    for label, params in (
            ("semopx listing, unfiltered", {"page_size": 50}),
            ("semopx listing, dispatch text", {"page_size": 50,
                                               "search_text": "dispatch"}),
    ):
        try:
            items = http_get(base, params=params).json().get("items", [])
            ids = sorted({str(it.get("DPuG_ID") or "?") for it in items})
            names = [str(it.get("ResourceName") or "")[:60]
                     for it in items[:3]]
            log(f"semo_probe: {label} -> {len(items)} items, "
                f"report IDs {ids[:12]}")
            for n in names:
                log(f"semo_probe:     e.g. {n}")
            hits.append((label, len(items), ids))
        except Exception as e:
            log(f"semo_probe: {label} -> {e.__class__.__name__}")

    # --- family 2: the SEMO market-data API. Report IDs are listed
    # rather than assumed; whichever returns is the evidence.
    semo = "https://reports.sem-o.com/api/v1/documents/static-reports"
    for label, params in (
            ("sem-o listing, unfiltered", {"page_size": 50}),
            ("sem-o listing, page 2", {"page_size": 50, "page": 2}),
    ):
        try:
            items = http_get(semo, params=params).json().get("items", [])
            ids = sorted({str(it.get("DPuG_ID") or "?") for it in items})
            log(f"semo_probe: {label} -> {len(items)} items, "
                f"report IDs {ids[:12]}")
            for it in items[:3]:
                log(f"semo_probe:     e.g. "
                    f"{str(it.get('ResourceName') or '')[:60]}")
            hits.append((label, len(items), ids))
        except Exception as e:
            log(f"semo_probe: {label} -> {e.__class__.__name__}")

    # A response is not an answer. The first live run (13 Aug 2026)
    # returned 50 items from every trial - all one report ID, all
    # IDC_Statistic documents from 2019 - and the probe congratulated
    # itself on finding "the catalogue". This endpoint lists
    # DOCUMENTS, ascending by date, not report types, and search_text
    # is ignored. So the test is whether more than one report ID came
    # back, not whether rows did.
    ids = sorted({i for _, _, got in hits for i in got})
    if len(ids) > 1:
        log(f"semo_probe: {len(ids)} report IDs seen {ids[:12]} - "
            "look for balancing-market dispatch or unit-level "
            "availability, and write the feed against whichever names "
            "resource codes and half-hourly periods")
    else:
        log(f"semo_probe: INCONCLUSIVE - every trial returned the same "
            f"{'single report ' + ids[0] if ids else 'nothing'}, so "
            "this endpoint lists documents rather than report types "
            "and cannot enumerate the catalogue. STOP PROBING: read "
            "the report catalogue by hand at sem-o.com/market-data, "
            "then write the feed against the ID found. Two rounds of "
            "guessing have produced two wrong answers")
    log("semo_probe: reminder - SEMO does not retain full history, so "
        "the captured series starts the day the feed ships. The B.2.2 "
        "panel does not wait on it: the annual regional report and "
        "EirGrid's 30-minute jurisdiction series are already published")


def semopx_history_probe():
    """
    Can a historic SEMOpx trade day be resolved? Log-only.

    price_ai fills forward, which leaves B.2.3 unable to price the
    binding hour already found. Before writing a backfill this asks
    the listing API whether it will hand over an older day at all, and
    which parameter does it - by trying candidates and reporting what
    each returns, rather than picking one and hoping. Parsers here are
    written against evidence from live logs, never guessed; the same
    rule applies to query parameters.
    """
    base = "https://reports.semopx.com/api/v1/documents/static-reports"
    want = (today_utc() - dt.timedelta(days=120)).isoformat()
    trials = [
        ("page_size only, deep page", {"DPuG_ID": "EA-001",
                                       "page_size": 100, "page": 5,
                                       "sort_by": "Date",
                                       "order_by": "DESC"}),
        ("Date filter", {"DPuG_ID": "EA-001", "page_size": 20,
                         "Date": want}),
        ("date range", {"DPuG_ID": "EA-001", "page_size": 20,
                        "Date_from": want, "Date_to": want}),
        ("publish range", {"DPuG_ID": "EA-001", "page_size": 20,
                           "PublishDateFrom": want,
                           "PublishDateTo": want}),
    ]
    for label, params in trials:
        try:
            items = http_get(base, params=params).json().get("items", [])
            da = [it for it in items
                  if re.search(r"_SEM-DA_", str(it.get("ResourceName") or ""))]
            days = sorted(x for x in (semopx_trade_day(it) for it in da) if x)
            log(f"semopx_probe: {label} -> {len(items)} items, "
                f"{len(da)} DA, days "
                f"{days[0] if days else '-'}..{days[-1] if days else '-'} "
                f"[{'REACHES ' + days[0] if days and days[0] <= want else 'recent window only'}]")
        except Exception as e:
            log(f"semopx_probe: {label} -> {e.__class__.__name__}")
    log(f"semopx_probe: target was {want} (120 days back); if nothing "
        "reaches it, price_ai fills forward and B.2.3 waits or moves "
        "to an hour the store has priced")


SEMOPX_BACKFILL_PAGES = 6      # listing pages walked per run
SEMOPX_BACKFILL_DAYS = 12      # documents fetched per run


def semopx_backfill(prev_hourly, base, want_from):
    """
    Walk the SEMOpx listing backwards and fetch trade days the store
    has not priced. Bounded per run - the same converging-walk pattern
    the hourly chunks and the temperature archive already use, rather
    than 400 requests in one build.

    Paging is the route, not a date filter: the probe showed `Date`
    returns nothing while a deep DESC page reaches months back and an
    unsorted listing returns the oldest documents in the archive. So
    the archive holds the days; they just have to be walked to.
    """
    have_days = {k[:10] for k in prev_hourly}
    added, seen, fetched = {}, set(), 0
    for page in range(1, SEMOPX_BACKFILL_PAGES + 1):
        if fetched >= SEMOPX_BACKFILL_DAYS:
            break
        try:
            items = http_get(base, params={
                "DPuG_ID": "EA-001", "page_size": 100, "page": page,
                "sort_by": "Date", "order_by": "DESC"}).json().get("items", [])
        except Exception as e:
            log(f"semopx: backfill page {page} {e.__class__.__name__}")
            break
        if not items:
            break
        for it in items:
            name = str(it.get("ResourceName") or "")
            if "_SEM-DA_" not in name:
                continue
            day = semopx_trade_day(it)
            if not day or day in seen or day in have_days or day < want_from:
                continue
            seen.add(day)
            if fetched >= SEMOPX_BACKFILL_DAYS:
                break
            try:
                doc = parse_semopx_csv(
                    http_get(f"https://reports.semopx.com/documents/{name}").text,
                    trade_day=day)
            except Exception as e:
                log(f"semopx: backfill {day} {e.__class__.__name__}")
                continue
            fresh = {}
            for cur in (doc.get("series") or {}).values():
                for k, v in (cur.get("EUR") or {}).items():
                    fresh.setdefault(k, []).append(v)
            for k, vals in fresh.items():
                added[k] = round(statistics.mean(vals), 2)
            fetched += 1
            time.sleep(0.3)
    log(f"semopx: backfill walked {min(page, SEMOPX_BACKFILL_PAGES)} page(s), "
        f"fetched {fetched} trade day(s), +{len(added)} priced hours")
    return added


def feed_semopx():
    """
    SEMOpx DAM results. DPuG_ID=EA-001 listing confirmed live but mixes DA
    with IDA1/2/3 - select _SEM-DA_ resources explicitly, widening the page
    if the first window holds none. Document is semicolon-CSV with decimal
    commas - see parse_semopx_csv().
    """
    base = "https://reports.semopx.com/api/v1/documents/static-reports"

    def da_items(items):
        return [it for it in items
                if re.search(r"_SEM-DA_", str(it.get("ResourceName") or ""))]

    chosen = None
    for page_size in (20, 100):
        items = http_get(base, params={
            "DPuG_ID": "EA-001", "page_size": page_size,
            "sort_by": "Date", "order_by": "DESC"}).json().get("items", [])
        hits = da_items(items)
        tags = sorted({m.group(0) for it in items
                       for m in [re.search(r"SEM-[A-Z0-9]+",
                                           str(it.get("ResourceName") or ""))]
                       if m})
        log(f"semopx: page_size {page_size} - {len(items)} items, "
            f"{len(hits)} DA, auction tags seen: {tags}")
        if hits:
            chosen = hits[0]
            break
    if not chosen:
        raise ValueError("semopx: no _SEM-DA_ resource in EA-001 listing")

    resource = chosen.get("ResourceName") or chosen.get("_id")
    log("semopx: resolved", resource)
    body = http_get(f"https://reports.semopx.com/documents/{resource}")
    parsed = parse_semopx_csv(body.text,
                              trade_day=semopx_trade_day(chosen))

    if not parsed["markets"]:
        log("semopx: CSV parse empty - first 800 chars:", body.text[:800])
        raise ValueError("semopx CSV parse failed - inspect log")

    def avg(currency):
        vals = [v for mk, cur in parsed["markets"].items()
                for c, series in cur.items() if c == currency
                for v in series]
        return round(statistics.mean(vals), 2) if vals else None

    # All-island hourly EUR price: NI-DA and ROI-DA are the same
    # energy market in two currencies and settle at the same euro
    # price, so the euro series of whichever market carries it is the
    # island price. Mean where both are present rather than picking.
    hourly = dict(prev_series("semopx", "dam_hourly_eur_mwh"))
    fresh = {}
    for mk, cur in (parsed.get("series") or {}).items():
        for k, v in (cur.get("EUR") or {}).items():
            fresh.setdefault(k, []).append(v)
    for k, vals in fresh.items():
        hourly[k] = round(statistics.mean(vals), 2)
    # Bounded walk backwards. Soft: the daily document above is the
    # feed's job, this is opportunistic depth for B.2.3.
    try:
        floor = (today_utc() - dt.timedelta(days=HOURLY_MONTHS * 31)).isoformat()
        hourly.update(semopx_backfill(hourly, base, floor))
    except Exception as e:
        log(f"semopx: backfill skipped ({e.__class__.__name__})")
    out = {
        "dam_hourly_eur_mwh": trim_series(hourly),
        "dam_hourly_added": len(fresh),
        "dam_avg_eur_mwh": avg("EUR"),
        "dam_avg_gbp_mwh": avg("GBP"),
        "markets": {mk: {c: round(statistics.mean(v), 2)
                         for c, v in cur.items() if v}
                    for mk, cur in parsed["markets"].items()},
        "sem_fx_eur_gbp": parsed["fx_eur_gbp"],
        "auction": parsed["auction"],
        "trade_day": parsed["day"], "latest_day": parsed["day"],
        "source": ("SEMOpx day-ahead market results - dual currency "
                   "(EUR/GBP), incl. SEM trading-day FX rate"),
    }
    log("semopx: markets parsed:", list(parsed["markets"]))
    return out, recency_status(parsed["day"], 4)


def feed_oil_bulletin():
    """
    EU Weekly Oil Bulletin - Ireland heating gas oil, EUR/1000 L, with AND
    without taxes. Snapshot files; history accumulates in data.json across
    runs. Both sides of the border burn the same C2 kerosene; the series is
    treated as the ROI heating-oil price level (see FEED_FLAGS).
    """
    page = http_get(
        "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
    ).text

    import openpyxl

    def fetch_ireland(with_tax):
        url = resolve_oil_bulletin_url(page, with_tax=with_tax)
        if not url:
            seen = [urllib.parse.unquote(u)[:110] for u in
                    re.findall(r'href="([^"]+)"', page)
                    if ".xlsx" in urllib.parse.unquote(u).lower()]
            log(f"oil_bulletin: no '{'with' if with_tax else 'without'} "
                "taxes' xlsx resolved - decoded links:", seen)
            return None, None
        log("oil_bulletin: resolved", urllib.parse.unquote(url)[:120])
        wb = openpyxl.load_workbook(
            io.BytesIO(http_get(url, timeout=180).content),
            read_only=True, data_only=True)
        for ws in wb.worksheets:
            d, v = parse_bulletin_rows(ws.iter_rows(values_only=True))
            if v is not None:
                return d, v
        for ws in wb.worksheets:
            head = [r for _, r in zip(range(6), ws.iter_rows(values_only=True))]
            log(f"oil_bulletin: sheet '{ws.title}' first rows:", head)
        return None, None

    d_wt, v_wt = fetch_ireland(True)
    d_nt, v_nt = fetch_ireland(False)
    if v_wt is None:
        raise ValueError("oil bulletin with-taxes parse failed - see log")
    for label, v in (("with", v_wt), ("without", v_nt)):
        if v is not None and not 300 <= v <= 3000:
            log(f"oil_bulletin: WARNING - {label}-taxes {v} EUR/1000L "
                "outside plausible range, check column selection")
    if d_wt is None:
        log("oil_bulletin: no bulletin date found - using today, verify")
        d_wt = today_utc().isoformat()
    log(f"oil_bulletin: Ireland heating {v_wt} EUR/1000L with taxes, "
        f"{v_nt} without, at {d_wt}")

    series = prev_series("oil_bulletin", "roi_heating_gasoil_eur_per_1000l")
    series[d_wt] = v_wt

    # weekly-history workbook backfills the chart to full depth; the
    # snapshot value above stays authoritative for its own date
    hist_url = None
    for m in re.finditer(r'href="([^"]+)"', page):
        u = m.group(1)
        d = urllib.parse.unquote(u).lower()
        if ".xlsx" in d and ("histor" in d or "time series" in d):
            hist_url = u if u.startswith("http") \
                else "https://energy.ec.europa.eu" + u
            break
    if not hist_url:
        # stable UUID cited in every weekly newsletter since 2024 - the
        # landing page does not carry this link, the newsletter does
        hist_url = ("https://energy.ec.europa.eu/document/download/"
                    "906e60ca-8b6a-44e7-8589-652854d2fd3f_en?filename="
                    "Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx")
        log("oil_bulletin: history link not on page - using stable "
            "newsletter URL")
    if hist_url:
        log("oil_bulletin: history resolved",
            urllib.parse.unquote(hist_url)[:110])
        try:
            hwb = openpyxl.load_workbook(
                io.BytesIO(http_get(hist_url, timeout=240).content),
                read_only=True, data_only=True)
            hist, hist_nt = {}, {}
            for ws in hwb.worksheets:
                title = ws.title.lower()
                if "price" not in title:
                    continue
                got = parse_bulletin_history_rows(ws.iter_rows(values_only=True))
                if not got:
                    continue
                if "wo" in title.split() or "without" in title:
                    if len(got) > len(hist_nt):
                        hist_nt = got
                elif len(got) > len(hist):
                    hist = got
            if hist:
                log(f"oil_bulletin: history {len(hist)} Ireland weeks "
                    f"with tax, {len(hist_nt)} ex tax, "
                    f"{min(hist)}..{max(hist)}")
                merged = dict(hist)
                merged.update(series)   # snapshot/current values win
                series = merged
                globals()["_hist_nt_pending"] = hist_nt
            else:
                for ws in hwb.worksheets[:3]:
                    head = [r for _, r in zip(range(5),
                                              ws.iter_rows(values_only=True))]
                    log(f"oil_bulletin: history sheet '{ws.title}' rows:", head)
                log("oil_bulletin: history parse empty - dumps above, "
                    "snapshot-only this run")
        except Exception as e:
            log(f"oil_bulletin: history fetch/parse failed "
                f"({e.__class__.__name__}: {e}) - snapshot-only this run")
    series = trim_series(clip_days(series))
    series_nt = prev_series("oil_bulletin",
                            "roi_heating_gasoil_eur_per_1000l_ex_tax")
    hist_nt = globals().pop("_hist_nt_pending", None)
    if hist_nt:
        merged_nt = dict(hist_nt)
        merged_nt.update(series_nt)
        series_nt = merged_nt
    if v_nt is not None:
        series_nt[d_nt or d_wt] = v_nt
    series_nt = trim_series(clip_days(series_nt))
    out = {
        "roi_heating_gasoil_eur_per_1000l": series,
        "roi_heating_gasoil_eur_per_1000l_ex_tax": series_nt,
        "latest_value": v_wt,
        "latest_value_ex_tax": v_nt,
        "latest_day": max(series),
        "source": ("European Commission Weekly Oil Bulletin - prices with "
                   "and without taxes"),
    }
    return out, recency_status(out["latest_day"], 10)


CCNI_ARCHIVE_URL = ("https://www.consumercouncil.org.uk/home-heating/"
                    "price-checker/archive")
# Bands for the volume ratios within one published row. CCNI's own
# record contains rows that fail these: 18 Nov 2021 carries a 900 L
# figure 9.7% above both its neighbours while its 300 and 500 L
# figures are identical to the previous day's to the penny. The gate
# does NOT reject - the series is published and we do not get to
# overrule it - it names the rows so a run that uses them is not
# silent about it. Bands are wide enough to leave the 2026 rows alone,
# where the volume discount genuinely moved.
CCNI_RATIO_BANDS = {"500_300": (1.40, 1.68), "900_500": (1.66, 1.82)}


def ccni_ratio_gate(series, label):
    """
    Name rows whose litre ratios are not internally plausible.

    A single day's 300/500/900 L figures come from one survey of the
    same suppliers, so their ratios are near-constant even as the
    level swings. A row that breaks that is a transcription fault in
    one cell, not a market movement - and it is invisible in the
    900 L series alone, which is the only one the site prices on.
    """
    days = sorted(set(series.get("300l", {})) & set(series.get("500l", {}))
                  & set(series.get("900l", {})))
    bad = []
    for d in days:
        a, b, c = (series["300l"][d], series["500l"][d], series["900l"][d])
        if not a or not b:
            continue
        r53, r95 = b / a, c / b
        lo53, hi53 = CCNI_RATIO_BANDS["500_300"]
        lo95, hi95 = CCNI_RATIO_BANDS["900_500"]
        if not (lo53 <= r53 <= hi53) or not (lo95 <= r95 <= hi95):
            bad.append((d, a, b, c, r53, r95))
    if bad:
        log(f"ccni_oil: {len(bad)} {label} row(s) fail the litre-ratio "
            f"gate - published as-is, listed so they are not silent")
        for d, a, b, c, r53, r95 in bad[:5]:
            log(f"ccni_oil:   {d} {a}/{b}/{c} "
                f"(500:300 {r53:.3f}, 900:500 {r95:.3f})")
    return bad


def feed_ccni_oil():
    """
    Consumer Council NI home heating oil, from TWO pages.

    The daily checker (Mon-Fri) is the recent detail; its embedded
    chart reaches back only a few months. The weekly ARCHIVE carries
    the whole published record - 277 rows to 2021 as of 15 Aug 2026 -
    in an embedded chart of exactly the same shape, so it parses
    through the same functions. That archive is what lets the NI oil
    series reach a 24-month window; the daily page alone cannot.

    (An earlier docstring here said "weekly page confirmed
    chart-free". That was wrong, and it is the reason the archive was
    written off as needing a scraper for months.)

    The two are kept apart in the payload and merged for consumers.
    Where a day appears in both, the DAILY reading wins: it is the
    same survey at finer resolution, and their overlap is logged every
    run rather than silently reconciled.
    """
    url = "https://www.consumercouncil.org.uk/home-heating/price-checker/daily"
    page = http_get(url).text
    # per-chart diagnostics: the page can embed more than one litre-labelled
    # chart, and 14/15 Jul runs stored different values for the same date -
    # log every candidate so the series identity question is answerable
    for i, arr in enumerate(extract_chart_data_arrays(page)):
        if arr and isinstance(arr[0], list) \
                and any("litre" in str(c).lower() for c in arr[0]):
            body = [r for r in arr[1:] if r]
            log(f"ccni_oil: chart {i} header={arr[0]} n={len(body)} "
                f"first={body[0] if body else None} "
                f"last={body[-1] if body else None}")
    parsed = parse_ccni_series(page)
    n = sum(len(v) for v in parsed.values())
    if not n:
        arrays = extract_chart_data_arrays(page)
        log("ccni_oil: no litre-labelled chart; "
            f"{len(arrays)} chart array(s) found, first headers:",
            [a[0] for a in arrays[:3] if a])
        raise ValueError("ccni_oil: no series parsed - inspect log")
    log(f"ccni_oil: {n} datapoints across "
        f"{[k for k, v in parsed.items() if v]}")

    # ---- the weekly archive, the deep half of the record ------------
    # SOFT: a failure here loses reach, not the run. The daily page
    # already priced every recent week before this feed existed, so a
    # bad archive fetch must not take the live figures down with it.
    archive = {"300l": {}, "500l": {}, "900l": {}}
    try:
        apage = http_get(CCNI_ARCHIVE_URL).text
        archive = parse_ccni_series(apage)
        an = sum(len(v) for v in archive.values())
        a9 = archive.get("900l") or {}
        log(f"ccni_oil: archive {an} datapoints, 900l {len(a9)} weeks "
            f"{min(a9) if a9 else '-'}..{max(a9) if a9 else '-'}")
        if not an:
            log("ccni_oil: WARNING archive page parsed no litre-labelled "
                "chart - reach falls back to the daily page")
    except Exception as exc:
        log(f"ccni_oil: archive fetch failed ({exc.__class__.__name__}) - "
            "keeping the daily series; reach is short until it returns")

    ccni_ratio_gate(parsed, "daily")
    ccni_ratio_gate(archive, "archive")

    # Overlap cross-check: the two pages describe the same survey, so
    # a disagreement is a fact about CCNI's publishing, not a rounding
    # question. Logged, never averaged.
    both = sorted(set(parsed.get("900l", {})) & set(archive.get("900l", {})))
    if both:
        gaps = [(d, archive["900l"][d], parsed["900l"][d]) for d in both
                if abs(parsed["900l"][d] - archive["900l"][d]) > 0.01]
        log(f"ccni_oil: daily/archive overlap {len(both)} day(s) "
            f"{both[0]}..{both[-1]}, {len(gaps)} disagree")
        for d, av, dv in gaps[:5]:
            log(f"ccni_oil:   {d} archive {av} vs daily {dv}")

    merged, conflicts = {}, 0
    for k, new in parsed.items():
        old = prev_series("ccni_oil", "series_gbp", "daily", k)
        for d, v in new.items():
            if d in old and old[d] and abs(v - old[d]) / old[d] > 0.05:
                conflicts += 1
                if conflicts <= 3:
                    log(f"ccni_oil: SERIES BREAK {k} {d}: stored "
                        f"{old[d]} -> page {v}")
        # archive first, daily last: the daily reading wins a tie
        old.update(archive.get(k) or {})
        old.update(new)
        merged[k] = trim_series(clip_days(old))
    if conflicts:
        log(f"ccni_oil: {conflicts} same-date value conflicts vs stored "
            "history - series identity unstable, new values kept")
    out = {"series_gbp": {"daily": merged,
                          "archive_weekly": {k: v for k, v in
                                             archive.items() if v}},
           "series_conflicts_this_run": conflicts,
           "daily_page_days": sorted(parsed.get("900l") or {})}

    all_days = [d for s in merged.values() for d in s]
    out["latest_day"] = max(all_days) if all_days else None
    d9 = merged.get("900l") or {}
    if d9:
        log(f"ccni_oil: merged 900l {len(d9)} day(s) "
            f"{min(d9)}..{max(d9)} (retention cap "
            f"{SERIES_KEEP_DAYS} days)")
    out["source"] = ("Consumer Council for Northern Ireland home heating oil "
                     "price checker - daily (Mon-Fri) and weekly archive, "
                     "NI average, 300/500/900 L")
    return out, recency_status(out["latest_day"], 7)



# ---------------------------------------------------------------- probes
# Structured probes (soft): fetch and log response shape only, so a run
# log establishes what each source actually serves before anything
# depends on it - the pattern that adopted the EirGrid CO2 series.
# Candidates from Paul Deane's source list, 27 Jul 2026.

# The ENTSOG probe is log-only and its finding is analysed
# fortnightly, but it polled six points every run - 4.5 of a 7-minute
# build on 13 Aug 2026, and 5.5 on a bad day. Its own retention is 25
# days, so a twice-weekly poll loses nothing, and the standing NI-exit
# finding it exists to witness moves at the pace of a monthly gas
# balance rather than a daily one.
ENTSOG_POLL_WEEKDAYS = (0, 3)      # Monday and Thursday


def feed_entsog_probe():
    """ENTSOG Transparency Platform - GB<->IE / GB<->GB(NI)
    interconnection points. Goal: observe Moffat / SNIP physical flows
    as (a) an independent NI gas measurement and (b) a cross-check of
    the gni_live jurisdiction finding (Total LDM 25% above ROI LDM)."""
    # Round 3 (30 Jul 2026): schema discovered - filter
    # operatorpointdirections by tSOCountry / adjacentCountry, list
    # Ireland-facing points, then sample one day of Physical Flow.
    out = {"source": "ENTSOG Transparency Platform (probe, round 3)"}
    # Twice weekly. On other days the retained series is returned
    # untouched, so the feed still reports and nothing downstream
    # notices - it simply does not spend four minutes re-measuring a
    # standing finding.
    if today_utc().weekday() not in ENTSOG_POLL_WEEKDAYS:
        keep = {k: prev_series("entsog_probe", k)
                for k in (PREVIOUS_FEEDS.get("entsog_probe") or {})
                if k.endswith("_gwh_daily")}
        prev_latest = ((PREVIOUS_FEEDS.get("entsog_probe") or {})
                       .get("latest_day"))
        if keep:
            out.update(keep)
            out["latest_day"] = prev_latest
            log(f"entsog_probe: skipped - polls "
                f"{'/'.join(dt.date(2026, 1, 5 + d).strftime('%a') for d in ENTSOG_POLL_WEEKDAYS)}"
                f", {len(keep)} series retained, latest {prev_latest}")
            return out, "lagging"
    pts = {}
    for q in ("tSOCountry=IE", "adjacentCountry=IE"):
        try:
            r = http_get("https://transparency.entsog.eu/api/v1/"
                         f"operatorpointdirections?{q}&limit=300",
                         timeout=90)
            rows = (r.json() or {}).get("operatorpointdirections", [])
            for row in rows:
                k = row.get("pointKey")
                if k:
                    pts[k] = (row.get("pointLabel"),
                              row.get("tSOCountry"),
                              row.get("adjacentCountry"),
                              row.get("directionKey"),
                              row.get("crossBorderPointType"))
            log(f"entsog_probe: {q} -> {len(rows)} rows")
        except Exception as e:
            log(f"entsog_probe: {q} {e.__class__.__name__}: {e}")
    for k, v in sorted(pts.items())[:12]:
        log(f"entsog_probe: point {k}: label={v[0]} tso={v[1]} "
            f"adj={v[2]} dir={v[3]} type={v[4]}")
    out["ie_points"] = sorted(pts)

    # Round 4a: the SNIP (Scotland->NI) is UK->UK, invisible to both
    # IE filters - label sweep of a large page for NI-relevant points.
    try:
        r = http_get("https://transparency.entsog.eu/api/v1/"
                     "operatorpointdirections?tSOCountry=UK"
                     "&limit=3000", timeout=120)
        rows = (r.json() or {}).get("operatorpointdirections", [])
        WANT = ("twynholm", "ballylumford", "brighouse", "belfast",
                "larne", "islandmagee", "scotland", "snip",
                "moffat", "south north", "gormanston", "corrib",
                "inch", "bellanaboy")
        hits = {}
        for row in rows:
            lbl = (row.get("pointLabel") or "").lower()
            if any(w in lbl for w in WANT):
                hits[row.get("pointKey")] = (row.get("pointLabel"),
                                             row.get("directionKey"),
                                             row.get("adjacentCountry"),
                                             row.get("adjacentZones"))
        log(f"entsog_probe: UK sweep {len(rows)} rows, "
            f"NI-relevant hits {len(hits)}")
        for k, v in sorted(hits.items())[:10]:
            log(f"entsog_probe: NI point {k}: label={v[0]} dir={v[1]} "
                f"adj={v[2]} zones={v[3]}")
        out["ni_candidate_points"] = sorted(hits)
    except Exception as e:
        log(f"entsog_probe: UK sweep {e.__class__.__name__}: {e}")

    # Round 6 (30 Jul 2026): Moffat combined, Twynholm and Greater
    # Belfast answer with sane values; the island-side mirror points
    # (Moffat IE/NI, South North) return non-JSON error documents -
    # their TSOs appear not to publish Physical Flow, so the UK-side
    # points are the observables. Retain daily series for the points
    # that answer: this is NI's first observed gas record, banking
    # from today. Jurisdiction confrontation runs on SNIP flow, the
    # measurable NI-bound artery (South North, ROI->NI, unmeasured).
    SHORT = {"ITP-00090": ("moffat_combined", "Moffat (combined)"),
             "ITP-00077": ("twynholm_snip", "Twynholm (SNIP)"),
             "DIS-00015": ("greater_belfast", "Greater Belfast"),
             "ITP-00495": ("moffat_ie", "Moffat (IE)"),
             "ITP-00496": ("moffat_ni", "Moffat (NI)"),
             "ITP-00222": ("south_north", "South North CSEP")}
    frm = (today_utc() - dt.timedelta(days=10)).isoformat()
    to = today_utc().isoformat()
    means = {}
    for pk, (slug, lbl) in SHORT.items():
        got = {}
        for dk in ("exit", "entry"):
            try:
                r2 = http_get(
                    "https://transparency.entsog.eu/api/v1/"
                    "operationaldata"
                    f"?pointKey={pk}&indicator=Physical%20Flow"
                    f"&periodType=day&directionKey={dk}"
                    f"&from={frm}&to={to}&limit=30", timeout=90)
                ctype = r2.headers.get("content-type", "?")
                if "json" not in ctype:
                    log(f"entsog_probe: {lbl} [{dk}] non-JSON "
                        f"({r2.status_code}, {ctype[:30]})")
                    continue
                j2 = r2.json() or {}
                key = next((k for k in j2
                            if isinstance(j2[k], list)), None)
                for x in (j2.get(key) or []):
                    try:
                        d = str(x.get("periodFrom", ""))[:10]
                        v = float(x["value"]) / 1e6
                        if d:
                            got[d] = got.get(d, 0.0) + v
                    except (TypeError, ValueError, KeyError):
                        continue
                if got:
                    break
            except Exception as e:
                log(f"entsog_probe: {lbl} [{dk}] "
                    f"{e.__class__.__name__}")
        if got:
            ser = prev_series("entsog_probe", f"{slug}_gwh_daily")
            ser.update(got)
            out[f"{slug}_gwh_daily"] = trim_series(ser)
            means[slug] = sum(got.values()) / len(got)
            log(f"entsog_probe: {lbl}: {len(got)} days, mean "
                f"{means[slug]:.1f} GWh/d, series "
                f"{len(out[f'{slug}_gwh_daily'])}d retained")
    snip = means.get("twynholm_snip")
    if snip is not None:
        sn = means.get("south_north")
        ni_bound = snip + (sn or 0.0)
        note = ("SNIP + South North" if sn is not None
                else "SNIP alone; South North (ROI->NI) unmeasured")
        log(f"entsog_probe: jurisdiction confrontation - NI-bound "
            f"{ni_bound:.1f} GWh/d ({note}) vs gni_live Total-ROI "
            f"LDM difference ~35.8 GWh/d "
            + ("-> magnitudes MATCH: independent witness for the "
               "MATERIAL verdict"
               if abs(ni_bound - 35.8) < 12 else
               "-> gap remains: scope needs a second look"))
    out["latest_day"] = today_utc().isoformat()
    return out, "ok"



def feed_eirgrid_probe():
    """Discovery dispatch for the v7 hourly engine (A'.2). One run,
    four chart types x three regions, answering the questions that
    stand between us and the 13-month hourly store:

      1. HISTORY DEPTH - does the endpoint serve a year, or only a
         rolling window? This is the single unknown that decides
         whether the store backfills or fills forward.
      2. WIND and SOLAR - both are on the endpoint we already use for
         demand and co2; we have simply never asked. Solar matters:
         AIRAA shows 3,260 MW in Ireland for 2026 rising to 6,880 by
         2031, and Energy-Charts cannot see it at all (which is the
         likeliest cause of the sem_mix indigenous-share shortfall).
      3. GRANULARITY - 15-minute rows aggregate to hourly MEANS, never
         samples; confirm the interval actually returned.
      4. REGIONALITY - is wind/solar published per jurisdiction as
         demand is, or all-island only?

    Soft, log-only. Nothing downstream depends on it.
    """
    # ROUND 2 (7 Aug 2026). Round 1's 400s were a parameter error, not
    # a missing series: `areas` must match the chart (demand ->
    # demandactual, co2 -> co2intensity), and I omitted it for wind,
    # solar and co2. Round 1 did establish: 15-minute interval,
    # 2,845/2,880 rows valued (98.8%), all three regions served for
    # demand, and a month returned regardless of the span requested -
    # so depth is the walk-back question tested below.
    out = {"source": "EirGrid Smart Grid Dashboard (probe, round 3)"}
    end = today_utc()

    def dmy(d):
        return d.strftime("%d-%b-%Y").replace(" 0", " ")

    def call(chart, region, areas, frm, to, rng="month"):
        r = http_get(EIRGRID_ENDPOINT, params={
            "region": region, "chartType": chart, "dateRange": rng,
            "dateFrom": dmy(frm), "dateTo": dmy(to), "areas": areas,
        }, timeout=120).json()
        rows = (r or {}).get("Rows", []) or []
        st = [x.get("EffectiveTime") for x in rows if x.get("EffectiveTime")]
        vals = [x.get("Value") for x in rows if x.get("Value") is not None]
        return rows, st, vals

    # --- A: correct areas names for wind and solar
    CAND = {"wind": ["windactual", "windforecast", "generationactual"],
            "solar": ["solaractual", "solarforecast"],
            "co2": ["co2intensity"], "demand": ["demandactual"]}
    good = {}
    for chart, names in CAND.items():
        for areas in names:
            try:
                rows, st, vals = call(chart, "ALL", areas,
                                      end - dt.timedelta(days=8), end)
                if rows:
                    good[chart] = areas
                    log(f"eirgrid_probe: {chart}/ALL areas={areas} OK - "
                        f"{len(rows)} rows, {len(vals)} valued, "
                        f"{st[0]} .. {st[-1]}")
                    break
                log(f"eirgrid_probe: {chart} areas={areas} 0 rows")
            except Exception as e:
                log(f"eirgrid_probe: {chart} areas={areas} "
                    f"{e.__class__.__name__}")

    # --- B: regionality of whichever series answered
    for chart, areas in good.items():
        if chart in ("demand", "co2"):
            continue
        for region in ("ROI", "NI"):
            try:
                rows, st, vals = call(chart, region, areas,
                                      end - dt.timedelta(days=8), end)
                log(f"eirgrid_probe: {chart}/{region} {len(rows)} rows, "
                    f"{len(vals)} valued")
            except Exception as e:
                log(f"eirgrid_probe: {chart}/{region} "
                    f"{e.__class__.__name__}")

    # --- C: THE DEPTH QUESTION. Can past months be walked? If a
    # request centred on a historic month returns that month's data,
    # the 13-month store backfills by chunked walking. If it returns
    # the current month (or nothing), the store fills forward instead
    # and the seasonal exhibit waits for the calendar.
    areas = good.get("demand", "demandactual")
    for back in (2, 6, 12):
        m_end = end - dt.timedelta(days=30 * back)
        m_start = m_end - dt.timedelta(days=27)
        try:
            rows, st, vals = call("demand", "ALL", areas, m_start, m_end)
            first = st[0] if st else "-"
            last = st[-1] if st else "-"
            hit = ("HISTORIC DATA RETURNED - walk-back works"
                   if st and str(m_end.year) in str(first)
                   and first[:6].lower() in
                   dmy(m_start).lower()[:6] + dmy(m_end).lower()[:6]
                   else "check span against request")
            log(f"eirgrid_probe: WALK-BACK {back}mo "
                f"(asked {dmy(m_start)}..{dmy(m_end)}) -> "
                f"{len(rows)} rows, got {first} .. {last} [{hit}]")
        except Exception as e:
            log(f"eirgrid_probe: WALK-BACK {back}mo "
                f"{e.__class__.__name__}: {e}")

    # --- C2: THE CARBON DEPTH QUESTION (round 3, 8 Aug 2026).
    #
    # This one decides the shape of the 24-month view. The hourly
    # store holds 13 months and the daily feed 391 days, so every week
    # before about mid-July 2025 has no own-week grid intensity from
    # anything retained. If co2intensity walks back two years, the
    # extension keeps the back-look's central claim - each week priced
    # at its own week's carbon. If it does not, those weeks need an
    # annual or monthly intensity and the window has to say so.
    #
    # RETRIED AND REPEATED, deliberately. On 8 Aug 2026 this endpoint
    # returned 0 rows for co2intensity and demandactual within seconds
    # of serving wind and solar, and the main feed pulled its 30 days
    # on the same run. A single zero is therefore not evidence of
    # depth - it is as likely to be the endpoint being flaky for that
    # chart at that moment. Nothing here is called EMPTY until two
    # independent attempts agree.
    co2_areas = good.get("co2", "co2intensity")
    depth = {}
    for back in (12, 18, 24):
        m_end = end - dt.timedelta(days=30 * back)
        m_start = m_end - dt.timedelta(days=27)
        got = []
        for attempt in (1, 2):
            try:
                rows, st, vals = call("co2", "ALL", co2_areas,
                                      m_start, m_end)
                got.append((len(rows), len(vals),
                            st[0] if st else "-", st[-1] if st else "-"))
                if rows:
                    break
                time.sleep(1.5)
            except Exception as e:
                got.append((0, 0, f"{e.__class__.__name__}", "-"))
                time.sleep(1.5)
        best = max(got, key=lambda g: g[0])
        # A returned span that lands in the requested month is the
        # real test; the endpoint is known to serve the current month
        # regardless of what was asked, which would look like success.
        want = {dmy(m_start)[3:], dmy(m_end)[3:]}
        landed = any(w.lower() in str(best[2]).lower() for w in want)
        if best[0] and landed:
            verdict = "HISTORIC CARBON RETURNED"
        elif best[0]:
            verdict = "rows returned but span is NOT the month asked for"
        else:
            verdict = f"EMPTY on {len(got)} attempts - treat as unavailable"
        depth[f"{back}mo"] = verdict
        log(f"eirgrid_probe: CARBON DEPTH {back}mo "
            f"(asked {dmy(m_start)}..{dmy(m_end)}) -> "
            f"{best[0]} rows, {best[1]} valued, got {best[2]} .. {best[3]} "
            f"[{verdict}]")
    out["carbon_depth"] = depth
    reach = [k for k, v in depth.items() if v == "HISTORIC CARBON RETURNED"]
    log("eirgrid_probe: CARBON DEPTH verdict - "
        + (f"own-week carbon reaches {max(reach, key=lambda k: int(k[:-2]))}"
           if reach else
           "no historic carbon at any depth tested; a 24-month window "
           "would need an annual or monthly intensity and must say so"))

    # --- D: does dateRange=year serve more than a month?
    try:
        rows, st, vals = call("demand", "ALL", areas,
                              end - dt.timedelta(days=365), end,
                              rng="year")
        log(f"eirgrid_probe: dateRange=year -> {len(rows)} rows, "
            f"{st[0] if st else '-'} .. {st[-1] if st else '-'}")
    except Exception as e:
        log(f"eirgrid_probe: dateRange=year {e.__class__.__name__}: {e}")

    out["latest_day"] = end.isoformat()
    return out, "ok"


def feed_sem_mix():
    """Energy-Charts (Fraunhofer ISE) public_power for Ireland (SEM,
    all-island) - adopted 30 Jul 2026 after the probe confirmed 13
    half-hourly production types. Computes a daily indigenous share of
    electricity: wind + hydro RoR + peat + Others x other-indigenous
    anchor, over generation plus imports; gas-fired generation is
    credited at the gas indigenous anchor for cross-fuel consistency.
    Pumped storage (both directions) is recycled load - excluded."""
    out = {"source": ("energy-charts.info public_power IE - "
                      "SEM all-island, half-hourly")}
    end = today_utc()
    start = end - dt.timedelta(days=35)
    r = http_get("https://api.energy-charts.info/public_power"
                 f"?country=ie&start={start.isoformat()}"
                 f"&end={end.isoformat()}", timeout=90)
    j = r.json() or {}
    stamps = j.get("unix_seconds") or []
    types = {p.get("name"): p.get("data") or []
             for p in (j.get("production_types") or [])}
    a_ind = ANCHORS["indigenous"]
    W_IND = {"Wind onshore": 1.0, "Hydro Run-of-River": 1.0,
             "Fossil peat": a_ind.get("peat", 1.0),
             "Others": a_ind.get("other", 0.9),
             "Fossil gas": a_ind.get("gas", 0.0),
             "Fossil oil": a_ind.get("oil", 0.0)}
    GEN = ["Wind onshore", "Hydro Run-of-River", "Fossil peat",
           "Others", "Fossil gas", "Fossil oil"]
    days = {}
    for i, ts in enumerate(stamps):
        d = dt.datetime.fromtimestamp(ts, dt.timezone.utc)\
            .date().isoformat()
        rec = days.setdefault(d, {"ind": 0.0, "gen": 0.0,
                                  "imp": 0.0, "n": 0})
        vals = {t: (types.get(t) or [None])[i]
                if i < len(types.get(t) or []) else None for t in GEN}
        if any(v is None for v in vals.values()):
            continue
        gen = sum(vals.values())
        ind = sum(v * W_IND[t] for t, v in vals.items())
        x = (types.get("Cross border electricity trading")
             or [None])[i] if i < len(
                 types.get("Cross border electricity trading") or []) \
            else None
        imp = max(x, 0.0) if x is not None else 0.0
        rec["ind"] += ind
        rec["gen"] += gen
        rec["imp"] += imp
        rec["n"] += 1
    ser = prev_series("sem_mix", "indigenous_share_daily")
    for d, rec in days.items():
        if rec["n"] >= 40:   # >=20h of half-hours = a complete-ish day
            denom = rec["gen"] + rec["imp"]
            if denom > 0:
                ser[d] = round(100.0 * rec["ind"] / denom, 2)
    out["indigenous_share_daily"] = trim_series(ser)
    ds = sorted(out["indigenous_share_daily"])
    log(f"sem_mix: {len(days)} days fetched, "
        f"{len(ds)} complete retained"
        + (f", latest {ds[-1]} at "
           f"{out['indigenous_share_daily'][ds[-1]]}%" if ds else ""))
    if not ds:
        raise RuntimeError("sem_mix: no complete days")
    out["latest_day"] = ds[-1]
    return out, recency_status(out["latest_day"], 4)


def feed_gb_oil():
    """
    GB heating-oil context line, two strategies (SOFT feed):
      A. BoilerJuice /kerosene-prices/ server-rendered sentence - present
         on legacy edge renders, absent on the modern template; tried with
         two user agents since the edge appears to vary by client.
      B. DESNZ/gov.uk monthly petroleum products table - official, stable,
         resolved from the statistics landing page at runtime. First
         contact logs candidate links, sheet names and header rows so the
         parser can be pinned in one iteration.
    History accumulates across runs whichever strategy lands.
    """
    # --- strategy A: BoilerJuice sentence. Diagnosis 18 Jul 2026: the CDN
    # serves non-browser clients an archived template (observed: Oct 2021,
    # (c) 2004-2021) while browsers and search crawlers receive the live
    # page with today's sentence. Countermeasures: browser headers, a
    # cache-busting query so no cached variant matches, and a freshness
    # gate - a parsed date older than 7 days is a fossil, rejected and
    # logged, never stored.
    A_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/126.0.0.0 Safari/537.36",
        "Referer": "https://www.google.com/",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-GB,en;q=0.9",
        "Cache-Control": "no-cache", "Pragma": "no-cache",
    }
    fresh_floor = (today_utc() - dt.timedelta(days=7)).isoformat()
    for url in ("https://www.boilerjuice.com/kerosene-prices/",
                "https://www.boilerjuice.com/heating-oil-prices/"):
        for attempt in range(2):
            try:
                page = http_get(url, headers=A_HEADERS, retries=1,
                                params={"cb": int(time.time()) + attempt}
                                ).text
            except Exception as e:
                log(f"gb_oil: A fetch failed on {url.rsplit('/', 2)[-2]} "
                    f"({e.__class__.__name__})")
                break
            d, ppl = parse_gb_oil_page(page)
            if ppl is None:
                log(f"gb_oil: A no sentence on {url.rsplit('/', 2)[-2]} "
                    f"attempt {attempt + 1} - page {len(page)} chars")
                if attempt == 0:
                    i = page.lower().find("verage")
                    if i >= 0:
                        log("gb_oil: A context around 'average':",
                            re.sub(r"\s+", " ",
                                   page[max(0, i - 80):i + 320]))
                    for pat in ("pence", "chart", "props",
                                "application/json"):
                        j = page.lower().find(pat)
                        if j >= 0:
                            log(f"gb_oil: A first '{pat}' at {j}:",
                                re.sub(r"\s+", " ", page[j:j + 240]))
                continue
            if d is None:
                d = today_utc().isoformat()
            if d < fresh_floor:
                log(f"gb_oil: A fossil template on "
                    f"{url.rsplit('/', 2)[-2]} - sentence dated {d}, "
                    f"{ppl} p/L - rejected, cache-busting again")
                continue
            log(f"gb_oil: A (BoilerJuice) {ppl} p/L at {d}")
            return _gb_oil_out(d, ppl,
                               "BoilerJuice UK daily average (lowest quote "
                               "per postcode district, incl 5% VAT)")
    log("gb_oil: A exhausted (fossil or sentence-less variants only) - "
        "falling through to DESNZ")

    # --- strategy B: DESNZ "Oil and petroleum products monthly
    # statistics" (QEP 4.1.1: monthly typical retail prices incl.
    # standard grade burning oil - GB home heating oil). Canonical page
    # pinned 18 Jul 2026; xlsx links resolved from it at runtime.
    # Diagnostics-first: sheet names and header rows are logged before
    # the parser commits, so a format change costs one log read.
    try:
        page = http_get(
            "https://www.gov.uk/government/statistical-data-sets/"
            "oil-and-petroleum-products-monthly-statistics").text
    except Exception as e:
        log(f"gb_oil: B page fetch failed ({e.__class__.__name__}: {e})")
        return _gb_oil_stale()
    links = re.findall(
        r'href="(https://assets\.publishing\.service\.gov\.uk/'
        r'[^"]+\.xlsx)"', page)
    pref = [u for u in links if "4.1.1" in u or "411" in u
            or "petroleum" in u.lower()] or links
    log(f"gb_oil: B found {len(links)} xlsx links; trying "
        f"{pref[0].rsplit('/', 1)[-1] if pref else 'none'}")
    if not pref:
        return _gb_oil_stale()
    try:
        import openpyxl
        blob = http_get(pref[0], timeout=120).content
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True,
                                    read_only=True)
    except Exception as e:
        log(f"gb_oil: B download/open failed "
            f"({e.__class__.__name__}: {e})")
        return _gb_oil_stale()
    series = {}
    for ws in wb.worksheets:
        got = parse_desnz_burning_oil_rows(
            ws.iter_rows(values_only=True))
        if got:
            log(f"gb_oil: B parsed {len(got)} months from sheet "
                f"'{ws.title}', latest {max(got)} = "
                f"{got[max(got)]} p/L")
            series = got
            break
        else:
            hdr = [list(r)[:8] for _i, r in
                   zip(range(3), ws.iter_rows(values_only=True))]
            log(f"gb_oil: B sheet '{ws.title}' no burning-oil parse; "
                f"first rows: {hdr}")
    if not series:
        return _gb_oil_stale()
    prev = prev_series("gb_oil", "gb_ppl_daily")
    prev.update(series)
    latest = max(series)
    return ({"gb_ppl_daily": trim_series(clip_days(prev)),
             "latest_day": latest,
             "latest_ppl": series[latest],
             "source": ("DESNZ QEP 4.1.1 - typical retail price, "
                        "standard grade burning oil, monthly, incl. "
                        "taxes"),
             },
            recency_status(latest, 100))


def parse_desnz_burning_oil_rows(rows) -> dict:
    """
    QEP 4.1.1-style sheet -> {iso_date (mid-month): pence_per_litre}.
    Finds a header row containing a burning-oil/kerosene column, then
    reads month rows (datetime cells or 'January 2026' text) with a
    plausible price (25-250 p/L). Pure, unit tested against a fixture;
    real-format confirmation comes from the first run's diagnostics.
    """
    MONTHS = {m.lower(): i + 1 for i, m in enumerate(
        ["January", "February", "March", "April", "May", "June", "July",
         "August", "September", "October", "November", "December"])}
    idx = None
    series = {}
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        if idx is None:
            for i, s in enumerate(strs):
                sl = s.lower()
                if "burning oil" in sl or "kerosene" in sl:
                    idx = i
                    break
            continue
        row_date = None
        for c in cells:
            if isinstance(c, dt.datetime):
                row_date = c.date().replace(day=15).isoformat()
                break
            if isinstance(c, dt.date):
                row_date = c.replace(day=15).isoformat()
                break
            if isinstance(c, str):
                parts = c.strip().split()
                if len(parts) == 2 and parts[0].lower() in MONTHS \
                        and parts[1].isdigit():
                    row_date = (f"{int(parts[1]):04d}-"
                                f"{MONTHS[parts[0].lower()]:02d}-15")
                    break
        if row_date is None:
            continue
        try:
            v = float(cells[idx])
        except (TypeError, ValueError, IndexError):
            continue
        if 25 <= v <= 250:
            series[row_date] = round(v, 2)
    return series


def _gb_oil_stale():
    series = prev_series("gb_oil", "gb_ppl_daily")
    return ({"gb_ppl_daily": series,
             "latest_day": max(series) if series else None,
             "source": "GB heating oil - awaiting source",
             }, "stale")


def _gb_oil_out(d, ppl, source):
    series = prev_series("gb_oil", "gb_ppl_daily")
    series[d] = ppl
    series = trim_series(clip_days(series))
    return ({"gb_ppl_daily": series,
             "latest_day": max(series) if series else None,
             "source": source},
            recency_status(max(series) if series else None, 40))


# ------------------------------------------------- analysis (pure functions)

def space_heat_split(gas_daily: dict, hdd_daily: dict):
    """
    Space-heat sensitivity of daily gas demand. Primary estimator is
    within-class (monthly) centring - within-month deviations of demand on
    deviations of HDD - which removes seasonal confounds (holidays, school
    terms, baseload drift) that bias the naive slope. The naive OLS is
    retained for reference. See tests/test_synthetic.py.
    """
    days = sorted(set(gas_daily) & set(hdd_daily))
    if len(days) < 30:
        return None
    x = [hdd_daily[d] for d in days]
    y = [gas_daily[d] for d in days]
    n = len(days)
    mx, my = statistics.mean(x), statistics.mean(y)

    def ols(xs, ys):
        mxx, myy = statistics.mean(xs), statistics.mean(ys)
        sxx = sum((xi - mxx) ** 2 for xi in xs)
        if sxx == 0:
            return None, None
        slope = sum((xi - mxx) * (yi - myy)
                    for xi, yi in zip(xs, ys)) / sxx
        return slope, myy - slope * mxx

    naive_slope, _ = ols(x, y)

    # within-class centring: subtract each month's own means
    from collections import defaultdict
    bym = defaultdict(list)
    for d, xi, yi in zip(days, x, y):
        bym[d[:7]].append((xi, yi))
    xd, yd = [], []
    for pts in bym.values():
        mxm = statistics.mean(p[0] for p in pts)
        mym = statistics.mean(p[1] for p in pts)
        for xi, yi in pts:
            xd.append(xi - mxm)
            yd.append(yi - mym)
    sxx = sum(v * v for v in xd)
    if sxx == 0:
        return None
    slope = sum(a * b for a, b in zip(xd, yd)) / sxx
    ss_res = sum((b - slope * a) ** 2 for a, b in zip(xd, yd))
    ss_tot = sum(b * b for b in yd) or 1e-9
    dof = max(1, len(xd) - len(bym) - 1)
    return {"slope_gwh_per_hdd": round(slope, 3),
            "baseload_gwh_per_day": round(my - slope * mx, 2),
            "r2_within_month": round(1 - ss_res / ss_tot, 3),
            "residual_se_gwh_per_day": round((ss_res / dof) ** 0.5, 2),
            "n_days": n,
            "method": "within-month centred OLS (within-class centring)",
            "naive_slope_gwh_per_hdd": round(naive_slope, 3)
            if naive_slope is not None else None}


def derive_gas_calibration(reg, hdd_roi, anchors=None):
    """
    Calibration disclosure: the regression-implied annual gas space heat
    (slope x trailing-year ROI degree days) against the anchor-implied
    figure (ROI buildings heat x gas share x space-heat fraction, input
    basis). Ratio published with a +/-10% gate; a miss is disclosed, not
    hidden - the two measures differ in scope (distribution-metered gas
    vs whole-anchor) and the ratio quantifies exactly that.
    """
    a = anchors or ANCHORS
    if not reg or not hdd_roi:
        return None
    days = sorted(hdd_roi)[-365:]
    if len(days) < 300:
        return None
    annual_hdd = sum(hdd_roi[d] for d in days)
    implied = reg["slope_gwh_per_hdd"] * annual_hdd
    j = a["roi"]
    anchor = ((j["residential_heat_twh"] + j["services_heat_twh"])
              * j["fuel_shares"]["gas"] * a["space_heat_fraction"] * 1000.0)
    ratio = implied / anchor if anchor else None
    return {"implied_annual_space_heat_gwh": round(implied, 0),
            "anchor_annual_space_heat_gwh": round(anchor, 0),
            "ratio": round(ratio, 2), "gate": "0.90-1.10",
            "within_gate": bool(0.90 <= ratio <= 1.10),
            "note": ("Scopes differ: the regression sees "
                     "distribution-metered gas; the anchor is the whole "
                     "buildings-gas estimate. The ratio measures the "
                     "difference and is published either way. A ratio "
                     "below one is expected - much of services-sector "
                     "gas is daily-metered and outside the NDM series "
                     "the regression sees.")}



# ------------------------------------------------------ weekly back-look
def ni_bridge_margin(feeds):
    """
    Calibrated NI-vs-bulletin margin, p/L: mean over the CCNI overlap
    of (observed NI 900L p/L) minus (bulletin ex-tax x week FX x 1.05
    VAT). Requires >=20 overlapping days. Bridged weeks reconstruct
    pre-CCNI NI prices as bridge + margin, dagger.
    """
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily") or {}).get("900l") or {}
    ext = ((feeds.get("oil_bulletin") or {})
           .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
    if not (ccni and ext and fxs):
        return None
    ext_days = sorted(ext)
    diffs = []
    for d in sorted(ccni):
        if d not in fxs:
            continue
        prior = [x for x in ext_days if x <= d]
        if not prior:
            continue
        est = ext[prior[-1]] / 1000.0 * fxs[d] * 100.0 * 1.05
        diffs.append(ccni[d] / 9.0 - est)
    if len(diffs) < 20:
        return None
    return round(sum(diffs) / len(diffs), 2)


def report_skips(skips, built):
    """
    Say which weeks were dropped and why, grouped by reason.

    A silent `continue` is the worst failure mode this pipeline has:
    the record simply comes out shorter, which reads as a smaller
    number rather than an error, and there is nothing in the log to
    contradict it. Sixteen weeks quietly missing from a 104-week
    extension would look exactly like a 88-week extension.
    """
    if not skips:
        log(f"history: {built} weeks built, none skipped")
        return
    by_reason = {}
    for w, r in skips:
        by_reason.setdefault(r, []).append(w)
    log(f"history: WARNING {len(skips)} week(s) skipped of "
        f"{built + len(skips)} attempted")
    for r, ws in sorted(by_reason.items()):
        ws = sorted(ws)
        span = ws[0] if len(ws) == 1 else f"{ws[0]}..{ws[-1]}"
        log(f"history:   {len(ws)} week(s) [{span}] - {r}")


def week_inputs(feeds, w_end, skips=None):
    """
    Per-week prices, fx, carbon and tariffs, or None if the week
    cannot be priced.

    `skips` is an optional list this appends (week, reason) to. It
    exists because returning None silently makes a short back-look
    look like a smaller number rather than an error - a week that
    cannot be built simply is not there, and nothing says so. The
    only silent decline is a week outside the window by design.
    """
    """Per-week pricing context for a calendar week ending w_end (Sun):
    NI oil (CCNI 900L weekly mean, p/L), ROI oil (bulletin week), fx
    (weekly mean of daily ECB), electricity EF (weekly CI mean or
    anchor), tariffs (period resolver). Returns None if a required
    input is absent - the week is not built."""
    tar = tariffs_for(w_end)
    if tar is None:
        if skips is not None:
            skips.append((w_end, "before the tariff table starts "
                                 f"({TARIFF_HISTORY[0][0]}) - extend "
                                 "TARIFF_HISTORY to price it"))
        return None

    def _skip(reason):
        if skips is not None:
            skips.append((w_end, reason))
        return None

    w_start = (dt.date.fromisoformat(w_end)
               - dt.timedelta(days=6)).isoformat()
    days = [(dt.date.fromisoformat(w_start) + dt.timedelta(days=i))
            .isoformat() for i in range(7)]
    if w_end < HISTORY_START:
        return None
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily") or {}).get("900l") or {}
    ni_vals = [ccni[d] / 9.0 for d in days if d in ccni]
    # Which CCNI page priced this week. The merged series is one dict
    # by design - every consumer reads it unchanged - but a week
    # carried entirely by the weekly archive is a weekly mean of ONE
    # reading, not of five, and that is worth recording rather than
    # inferring later from the date.
    dp = (feeds.get("ccni_oil") or {}).get("daily_page_days")
    if dp is None:
        # Field arrives with 5.10.0. A payload without it is not a
        # week priced by the archive - it is a week whose provenance
        # was never recorded, and saying "archive" would be a claim.
        ni_src = "ccni"
    else:
        from_daily = [d for d in days if d in ccni and d in set(dp)]
        ni_src = ("ccni" if len(from_daily) == len(ni_vals)
                  else "ccni archive (weekly)" if not from_daily
                  else "ccni (daily + archive)")
    if not ni_vals:
        # pre-CCNI weeks: bridge from the bulletin ex-tax series
        # (same cargoes, 5% VAT) plus the overlap-calibrated margin
        m = ni_bridge_margin(feeds)
        ext = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
        fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
        e_days = [d for d in sorted(ext) if d <= w_end]
        f_vals = [fxs[d] for d in days if d in fxs]
        if m is None or not e_days or not f_vals:
            return _skip(
                "NI oil: no CCNI reading and the bulletin bridge could "
                "not be built" + (" (no overlap margin)" if m is None else "")
                + ("" if e_days else " (no ex-tax bulletin week)")
                + ("" if f_vals else " (no FX for the week)"))
        fx_w = sum(f_vals) / len(f_vals)
        ni_vals = [ext[e_days[-1]] / 1000.0 * fx_w * 100.0 * 1.05 + m]
        ni_src = "bridged (bulletin ex-tax + calibrated margin, dagger)"
    bull = ((feeds.get("oil_bulletin") or {})
            .get("roi_heating_gasoil_eur_per_1000l") or {})
    bull_nt = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    b_days = [d for d in sorted(bull) if d <= w_end]
    if not b_days:
        return _skip("ROI oil: no EU bulletin week at or before this date")
    fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
    fx_vals = [fxs[d] for d in days if d in fxs]
    fx = (sum(fx_vals) / len(fx_vals)) if fx_vals \
        else (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855
    co2 = ((feeds.get("eirgrid") or {})
           .get("co2_intensity_g_per_kwh") or {})
    ci_vals = [co2[d] for d in days if d in co2]
    ef = round(sum(ci_vals) / len(ci_vals), 1) if len(ci_vals) >= 4 \
        else None
    if ef is None and w_end < LIVE_FROM:
        # A reconstructed week priced at the carbon ANCHOR would carry
        # today's grid intensity wearing last year's date - the clamp
        # pattern, landing on emissions instead of price. Refuse it.
        # The record then extends itself as the carbon backfill
        # reaches further, and never contains a week it cannot date.
        return _skip("no grid carbon for this week - the daily feed "
                     "retains 50 days and the hourly store 13 months, "
                     "so an older week needs the EirGrid backfill to "
                     "have reached it")
    nd_week, _nd_sem = nondom_for(w_end, ((feeds.get("ecb_fx") or {})
                                          .get("eur_gbp_semester")))
    if nd_week is None:
        return _skip("non-domestic: no published REMM semester at or "
                     f"before this week (first is "
                     f"{sorted(NONDOM_SEMESTERS)[0]}) - extend "
                     "NONDOM_SEMESTERS to price it")
    return {"ni_oil_ppl": round(sum(ni_vals) / len(ni_vals), 2),
            "roi_oil_eur_1000l": bull[b_days[-1]],
            "fx": round(fx, 5), "ef_electricity": ef,
            "ef_source": ("weekly grid CI" if ef else "anchor"),
            "ni_oil_source": ni_src,
            # Reconstructed weeks are priced from published tariffs and
            # the hourly store rather than observed as they happened.
            # They are perfectly good weeks; they are just not LIVE
            # ones, and the milestone counts live.
            "live": w_end >= LIVE_FROM,
            "nondom": nd_week,
            "tariffs": tar}


# Wire format for the history block. Content is HISTORY_SCHEMA; this
# is how it is written, and the two are deliberately orthogonal - a
# re-encoding is not a restatement and must not trigger one.
HISTORY_ENCODING = "columnar-1"


def compact_history(entries):
    """
    Array of week objects -> columnar. Every key is written once
    instead of once per week, which is most of the file: the entries
    are wide and shallow, so at 52 weeks the key strings outweigh the
    numbers they label. Recurses into the ni/roi/fuels sub-blocks.
    """
    if not entries:
        return {"encoding": HISTORY_ENCODING, "n": 0, "cols": {}}

    def cols_of(objs):
        keys = []
        for o in objs:
            for k in o:
                if k not in keys:
                    keys.append(k)
        out = {}
        for k in keys:
            vals = [o.get(k) for o in objs]
            if any(isinstance(v, dict) for v in vals):
                out[k] = cols_of([v if isinstance(v, dict) else {}
                                  for v in vals])
            else:
                out[k] = vals
        return out

    return {"encoding": HISTORY_ENCODING, "n": len(entries),
            "cols": cols_of(entries)}


def expand_history(doc):
    """
    Columnar OR the legacy list -> list of week objects.

    Both shapes are accepted, and that is not politeness. index.html
    publishes the moment Pages deploys while data.json only changes at
    the next build, so for up to a day the new front end reads the old
    payload - and the pipeline reads its own previous output the same
    way. Either side must tolerate either shape.
    """
    if isinstance(doc, list):
        return doc
    if not isinstance(doc, dict) or not doc.get("cols"):
        return []
    n = int(doc.get("n") or 0)

    def rows(cols):
        out = [{} for _ in range(n)]
        for k, v in cols.items():
            if isinstance(v, dict):
                sub = rows(v)
                for i in range(n):
                    out[i][k] = sub[i]
            else:
                # Nulls are written back, not skipped. ef_electricity
                # is legitimately None in a week with too few carbon
                # observations, and dropping the key instead of the
                # value would make the round trip inexact for no gain.
                for i in range(n):
                    out[i][k] = v[i] if i < len(v) else None
        return out

    return rows(doc["cols"])


def build_history(feeds, anchors=None):
    """UK-pattern weekly history: complete calendar weeks (Mon-Sun),
    hero combined four + what-if twins per entry, frozen after the two
    most recent, capped at HISTORY_MAX, append-or-update by
    week_ending. Records the GNI feed's empirical window each run
    (gas_window) per the porting handover.

    DEPTH IS NOT BOUNDED HERE. This said "capped at 60 ... bounded by
    the PRICE feeds (CCNI from 2026-02-26)"; both were wrong by
    5.16.0. The 60 was a literal in the loop below, and CCNI stopped
    being the constraint when the weekly archive landed at 5.10.0 and
    took the series back to 2023. What binds now, in order: the carbon
    backfill's reach, then the tariff table floor - and a week that
    cannot be priced is refused by name rather than filled in."""
    def fuel_sub(b):
        """Compact per-fuel in/useful for the windowed energy bars.
        Keys stay short - this rides in every history entry, three
        times over (island, NI, ROI)."""
        bf = b.get("by_fuel") or {}
        cl = b.get("cooling") or {}
        out = {f: {"i": round(v.get("in_gwh", 0.0), 1),
                   "u": round(v.get("useful_gwh", 0.0), 1)}
               for f, v in bf.items()}
        out["cool"] = {"i": round(cl.get("elec_gwh", 0.0), 1),
                       "u": round(cl.get("served_gwh",
                                         cl.get("elec_gwh", 0.0)), 1)}
        return out

    def jur_sub(b):
        # (JUR_SPLIT_KEYS is module-level so the tests can name it)
        return {
            "purchased_gwh": b["combined"]["purchased_gwh"],
            "served_gwh": b["combined"]["served_gwh"],
            "indigenous_pct": b["combined"]["indigenous_share_pct"],
            "bill_eur_m": b["combined"]["bill_eur_m"],
            "bill_gbp_m": b["combined"]["bill_gbp_m"],
            "emissions_kt": b["combined"]["emissions_kt_co2"],
            "fuels": fuel_sub(b),
            "wf_purchased_gwh": b["what_if_combined"]["purchased_gwh"],
            "wf_indigenous_pct":
                b["what_if_combined"]["indigenous_share_pct"],
            "wf_bill_eur_m": b["what_if_combined"]["bill_eur_m"],
            "wf_bill_gbp_m": b["what_if_combined"]["bill_gbp_m"],
            "wf_emissions_kt":
                b["what_if_combined"]["emissions_kt_co2"],
            # schema 5: the same heat/cold splits the island entry
            # carries, so a jurisdiction window can break its cards
            # down instead of falling back to a bare total. Sterling
            # AND euro, because the NI view prices in sterling and
            # must not convert components in the browser.
            **{k: b["combined"][k] for k in JUR_SPLIT_KEYS},
            **{"wf_" + k: b["what_if_combined"][k]
               for k in JUR_SPLIT_KEYS},
        }

    skips = []
    prev = list(expand_history((PREVIOUS_DERIVED or {})
                               .get("history") or []))
    prev_schema = int((PREVIOUS_DERIVED or {})
                      .get("history_schema") or 1)
    prev_epoch = int((PREVIOUS_DERIVED or {})
                     .get("anchor_epoch") or 1)
    re_anchor = prev_epoch < ANCHOR_EPOCH
    if re_anchor and prev:
        log(f"history: anchor epoch {prev_epoch} -> {ANCHOR_EPOCH}, "
            f"re-anchoring {len(prev)} stored weeks onto the new "
            f"basis (own-week prices retained)")
    frozen = {e["week_ending"]: e for e in prev[:-2]} if len(prev) > 2 \
        else {}
    hdd = (feeds.get("hdd") or {}).get("hdd_island") or {}
    if not hdd:
        return prev
    today = today_utc()
    last_sun = today - dt.timedelta(days=(today.isoweekday() % 7) or 7)
    out = []
    # HISTORY_MAX weeks offered, not a literal. This read
    # `range(59, -1, -1)` while HISTORY_MAX was 120 and the cap was
    # applied afterwards to a list that could never exceed 60 - so the
    # record sat at 60 weeks and the log line "60 weeks built, none
    # skipped" was reporting the loop bound rather than any data
    # limit. It also shortened panel 1's sparklines, which slice
    # whatever the record holds. Weeks that genuinely cannot be built
    # are still refused BY NAME further down, which is what stops this
    # from silently inventing depth.
    for k in range(HISTORY_MAX - 1, -1, -1):
        w_end = (last_sun - dt.timedelta(weeks=k)).isoformat()
        if w_end in frozen:
            e = frozen[w_end]
            needs = (prev_schema < HISTORY_SCHEMA or re_anchor
                     or "ni" not in e or "roi" not in e)
            if needs:
                # Restatement (schema policy, 1 Aug 2026 handover):
                # recompute through derive_hero with the entry's
                # STORED inputs - ef_electricity injected so weeks
                # whose grid CI has rolled out of retention keep the
                # factor they were built with; oil, fx and tariffs
                # come from the retained series and the resolver.
                # Existing fields must re-round identically (audited
                # below); only new fields appear. Sub-block splits:
                # DEFERRED - combined-level is the like-for-like
                # scope; extend when a panel displays them.
                ctx = week_inputs(feeds, w_end)
                if ctx is not None:
                    if e.get("ef_electricity"):
                        ctx = {**ctx,
                               "ef_electricity": e["ef_electricity"],
                               "ef_source": e.get("ef_source",
                                                  ctx["ef_source"])}
                    h2 = derive_hero(feeds, anchors,
                                     week_ctx={"week_ending": w_end,
                                               **ctx})
                    if h2 is not None:
                        drift = max(
                            abs(h2["combined"]["purchased_gwh"]
                                - e["purchased_gwh"]),
                            abs(h2["combined"]["bill_eur_m"]
                                - e["bill_eur_m"]),
                            abs(h2["combined"]["emissions_kt_co2"]
                                - e["emissions_kt"]))
                        if drift > 0.5 and not re_anchor:
                            log(f"history: restatement drift {w_end} "
                                f"{drift:.1f} - stored values kept "
                                f"frozen, review")
                        e = dict(e)
                        C2, W2 = h2["combined"], h2["what_if_combined"]
                        if re_anchor:
                            # basis change: rewrite the stored values
                            # too, so totals and splits stay on one
                            # footing across the whole series
                            e.update({
                                "purchased_gwh": C2["purchased_gwh"],
                                "served_gwh": C2["served_gwh"],
                                "indigenous_pct":
                                    C2["indigenous_share_pct"],
                                "bill_eur_m": C2["bill_eur_m"],
                                "bill_gbp_m": C2["bill_gbp_m"],
                                "emissions_kt": C2["emissions_kt_co2"],
                                "wf_purchased_gwh": W2["purchased_gwh"],
                                "wf_indigenous_pct":
                                    W2["indigenous_share_pct"],
                                "wf_bill_eur_m": W2["bill_eur_m"],
                                "wf_bill_gbp_m": W2["bill_gbp_m"],
                                "wf_emissions_kt":
                                    W2["emissions_kt_co2"],
                                "ni": jur_sub(h2["ni"]),
                                "roi": jur_sub(h2["roi"]),
                                "fuels": fuel_sub(h2),
                            })
                        for k in ("heat_gwh", "cold_gwh",
                                  "bill_heat_eur_m", "bill_cold_eur_m",
                                  "bill_heat_gbp_m", "bill_cold_gbp_m",
                                  "emissions_heat_kt",
                                  "emissions_cold_kt"):
                            e[k] = C2[k]
                            e["wf_" + k] = W2[k]
                        e.setdefault("fx_eur_gbp", ctx["fx"])
                        # Assign, never setdefault: a block that
                        # already exists from an older schema must be
                        # REFRESHED so it gains new sub-fields. This
                        # is safe and idempotent because everything
                        # is recomputed from the entry's own stored
                        # inputs.
                        e["fuels"] = fuel_sub(h2)
                        # NOT setdefault: these keys already exist
                        # from earlier schemas, so a defaulting write
                        # silently skipped the per-fuel addition.
                        # A schema migration refreshes the block.
                        e["ni"] = jur_sub(h2["ni"])
                        e["roi"] = jur_sub(h2["roi"])
                        e["live"] = w_end >= LIVE_FROM
                    else:
                        log(f"history: {w_end} unrecomputable - "
                            f"kept at prior schema (caption degrades)")
                else:
                    log(f"history: {w_end} outside recompute window - "
                        f"kept at prior schema (caption degrades)")
            out.append(e)
            continue
        ctx = week_inputs(feeds, w_end, skips)
        if ctx is None:
            continue
        h = derive_hero(feeds, anchors, week_ctx={"week_ending": w_end,
                                                  **ctx})
        if h is None:
            skips.append((w_end, "derive_hero declined - check the HDD "
                                 "series has 200+ days"))
            continue
        C, WFC = h["combined"], h["what_if_combined"]
        out.append({
            "week_ending": w_end,
            "purchased_gwh": h["combined"]["purchased_gwh"],
            "served_gwh": h["combined"]["served_gwh"],
            "indigenous_pct": h["combined"]["indigenous_share_pct"],
            "bill_eur_m": h["combined"]["bill_eur_m"],
            "bill_gbp_m": h["combined"]["bill_gbp_m"],
            "emissions_kt": h["combined"]["emissions_kt_co2"],
            "wf_purchased_gwh": h["what_if_combined"]["purchased_gwh"],
            "wf_indigenous_pct":
                h["what_if_combined"]["indigenous_share_pct"],
            "wf_bill_eur_m": h["what_if_combined"]["bill_eur_m"],
            "wf_bill_gbp_m": h["what_if_combined"]["bill_gbp_m"],
            "wf_emissions_kt":
                h["what_if_combined"]["emissions_kt_co2"],
            "hdd": h["hdd_week"],
            "ef_electricity": ctx["ef_electricity"] or None,
            "ef_source": ctx["ef_source"],
            "heat_gwh": C["heat_gwh"], "cold_gwh": C["cold_gwh"],
            "bill_heat_eur_m": C["bill_heat_eur_m"],
            "bill_cold_eur_m": C["bill_cold_eur_m"],
            "bill_heat_gbp_m": C["bill_heat_gbp_m"],
            "bill_cold_gbp_m": C["bill_cold_gbp_m"],
            "emissions_heat_kt": C["emissions_heat_kt"],
            "emissions_cold_kt": C["emissions_cold_kt"],
            "wf_heat_gwh": WFC["heat_gwh"],
            "wf_cold_gwh": WFC["cold_gwh"],
            "wf_bill_heat_eur_m": WFC["bill_heat_eur_m"],
            "wf_bill_cold_eur_m": WFC["bill_cold_eur_m"],
            "wf_bill_heat_gbp_m": WFC["bill_heat_gbp_m"],
            "wf_bill_cold_gbp_m": WFC["bill_cold_gbp_m"],
            "wf_emissions_heat_kt": WFC["emissions_heat_kt"],
            "wf_emissions_cold_kt": WFC["emissions_cold_kt"],
            "fx_eur_gbp": ctx["fx"],
            "fuels": fuel_sub(h),
            "ni_oil_source": ctx["ni_oil_source"],
            "live": ctx["live"],
            "ni": jur_sub(h["ni"]),
            "roi": jur_sub(h["roi"]),
        })
    report_skips(skips, len(out))
    return out[-HISTORY_MAX:]


def derive_hero(feeds, anchors=None, week_ctx=None):
    """
    Weekly hero four-stat + what-if, per jurisdiction and all-island.
    Each jurisdiction is shaped by its own HDD series (island fallback);
    the island block is the sum of the two, so the toggle views always
    reconcile. Scaffold estimator - pure, unit tested. Top-level keys
    carry the island values; per-jurisdiction blocks under "roi"/"ni".
    """
    a = anchors or ANCHORS
    hddf = feeds.get("hdd") or {}
    island_hdd = hddf.get("hdd_island") or {}
    if len(island_hdd) < 200:
        return None
    W = week_ctx or {}
    fx = W.get("fx") or (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855
    shf = a["space_heat_fraction"]

    if W:
        # historic week: prices, fx, EF and tariffs from the week itself
        oil_eur_kwh = W["roi_oil_eur_1000l"] / 1000.0 \
            / a["kerosene_kwh_per_litre"]
        oil_gbp_kwh = W["ni_oil_ppl"] / 100.0 \
            / a["kerosene_kwh_per_litre"]
    else:
        oil_eur_kwh = None
        ob = (feeds.get("oil_bulletin") or {}).get("latest_value")
        if ob:
            oil_eur_kwh = ob / 1000.0 / a["kerosene_kwh_per_litre"]
        oil_gbp_kwh = None
        ccni = ((feeds.get("ccni_oil") or {}).get("series_gbp") or {}).get(
            "daily", {}).get("900l") or {}
        if ccni:
            oil_gbp_kwh = ccni[max(ccni)] / 900.0 \
                / a["kerosene_kwh_per_litre"]

    def hdd_stats(series):
        days = sorted(series)
        if W:
            w_end = W["week_ending"]
            wk = [d for d in days if d <= w_end][-7:]
            yr = [d for d in days if d <= w_end][-365:]
            if len(wk) < 7 or len(yr) < 200:
                return (0.0, 0.0, w_end)
            return (sum(series[d] for d in wk),
                    sum(series[d] for d in yr), w_end)
        wk = days[-7:]
        return (sum(series[d] for d in wk),
                sum(series[d] for d in days[-365:]), wk[-1])

    def jur_block(jur, cur, hdd_series):
        hdd_week, hdd_year, week_end = hdd_stats(hdd_series)
        if hdd_year <= 0:
            return None
        j = a[jur]
        heat_twh = j["residential_heat_twh"] + j["services_heat_twh"]

        def week_input_gwh(annual_twh):
            annual_gwh = annual_twh * 1000.0
            return annual_gwh * ((1 - shf) / 52.0
                                 + shf * hdd_week / hdd_year)

        # Heat-pump share of the electricity line. The stock is an
        # annual quantity; its delivered heat is shaped by the same
        # weekly profile as everything else, so the split holds
        # week to week. hp_useful is delivered heat; hp_elec is the
        # purchased part; the remainder is ambient heat harvested
        # from the environment - free, and never purchased.
        hpa = a.get("heat_pumps") or {}
        hpj = hpa.get(jur) or {}
        hp_annual_twh = (
            (hpj.get("households", 0) * hpj.get("census_uplift", 1.0)
             + hpj.get("nondom_equivalent", 0))
            * hpa.get("delivered_mwh_per_dwelling", 9.0) / 1e6)
        spf = j.get("ashp_climate_spf") or a["ashp"].get("climate_spf") \
            or 2.85
        hp_useful = week_input_gwh(hp_annual_twh)
        hp_elec = hp_useful / spf
        hp_ambient = hp_useful - hp_elec

        inp_t = useful_t = indig_t = kt_t = bill_t = 0.0
        by_fuel = {}
        for fuel, share in j["fuel_shares"].items():
            inp = week_input_gwh(heat_twh) * share
            eff = a["efficiency"][fuel]
            useful = inp * eff
            if fuel == "electricity":
                # split the line: resistive keeps the remainder, the
                # heat-pump part is carried separately, and the
                # ambient harvest is an out-bar-only entry (no input)
                hp_in = min(hp_elec, inp)
                by_fuel["heatpump"] = {"in_gwh": round(hp_in, 1),
                                       "useful_gwh": round(hp_in, 1)}
                by_fuel["ambient"] = {
                    "in_gwh": 0.0,
                    "useful_gwh": round(hp_in * (spf - 1.0), 1)}
                inp = inp - hp_in
                useful = inp * eff
            by_fuel[fuel] = {"in_gwh": round(inp, 1),
                             "useful_gwh": round(useful, 1)}
            indig = useful * (
                j["gas_indigenous"] if fuel == "gas" else
                j["elec_indigenous"] if fuel == "electricity" else
                a["indigenous"].get(fuel, 0.0))
            kt = inp * EF[fuel] / 1000.0
            if fuel == "oil":
                price = oil_eur_kwh if cur == "eur" else oil_gbp_kwh
                if price is None:
                    price = 0.11 if cur == "eur" else 0.09  # dagger fallback
            else:
                table = ((W.get("tariffs") or {}).get(cur)
                         or (a["retail_eur_per_kwh"] if cur == "eur"
                             else a["retail_gbp_per_kwh"]))
                price = table.get(fuel, table["gas"])
                if fuel in ("gas", "electricity"):
                    nd = _nondom(a, W)["eur" if cur == "eur"
                                        else "gbp"][fuel]
                    ds = a["dom_share"][jur][fuel]
                    price = ds * price + (1 - ds) * nd
            inp_t += inp
            useful_t += useful
            if fuel == "electricity":
                # heat-pump electricity and its ambient harvest join
                # the totals: purchased counts only the electricity,
                # delivered counts electricity x SPF
                inp_t += by_fuel["heatpump"]["in_gwh"]
                useful_t += (by_fuel["heatpump"]["useful_gwh"]
                             + by_fuel["ambient"]["useful_gwh"])
            indig_t += indig
            kt_t += kt
            bill_t += inp * price   # GWh x cur/kWh = millions of cur

        bill_eur = bill_t if cur == "eur" else bill_t / fx
        bill_gbp = bill_t * fx if cur == "eur" else bill_t

        dom_elec = ((W.get("tariffs") or {}).get(cur, {})
                    .get("electricity")
                    or (a["retail_eur_per_kwh"]["electricity"]
                        if cur == "eur"
                        else a["retail_gbp_per_kwh"]["electricity"]))
        nondom_elec = _nondom(a, W)["eur" if cur == "eur"
                                    else "gbp"]["electricity"]
        # cooling is wholly non-domestic (UK convention)
        elec_price = nondom_elec
        # what-if heat-pump electricity blends at the domestic share of
        # delivered heat (UK convention)
        heat_dom = j["residential_heat_twh"] / max(
            j["residential_heat_twh"] + j["services_heat_twh"], 1e-9)
        wf_elec_price = heat_dom * dom_elec + (1 - heat_dom) * nondom_elec

        # what-if: 20% of useful heat moves to geothermal heat pumps
        spf = a["geothermal_spf"]
        moved = useful_t * 0.20
        elec_in = moved / spf
        ambient = moved - elec_in
        scale = 0.80
        wf_bill_native = bill_t * scale + elec_in * wf_elec_price
        wf = {
            "heat_purchased_gwh": round(inp_t * scale + elec_in, 1),
            "indigenous_share_pct": round(100 * (
                indig_t * scale + ambient
                + elec_in * j["elec_indigenous"]) / max(useful_t, 1e-9), 1),
            "bill_eur_m": round(wf_bill_native if cur == "eur"
                                else wf_bill_native / fx, 1),
            "bill_gbp_m": round(wf_bill_native * fx if cur == "eur"
                                else wf_bill_native, 1),
            "emissions_kt_co2": round(
                kt_t * scale
                + elec_in * EF["electricity"] / 1000.0, 1),
            "geothermal_spf": spf,
        }
        # cooling: cold-economy census, weekly. Flat loads spread
        # evenly; the ROI comfort load follows the live ODH26 record
        # (30% dagger ventilation floor) once a season of it exists.
        cc = a["cool"]
        if jur == "roi":
            dc = (cc["roi_elec_twh"] * cc["dc_share_of_roi_elec"]
                  * cc["dc_cooling_share"])
            l = cc["loads_twh"]
            comfort_a = l["comfort"]
            flat_a = dc + l["refrigeration"] + l["process"]
        else:
            comfort_a, flat_a = 0.0, cc["loads_twh"]["ni_all"]
        sf = cc["cooling_service_factor"]
        if jur == "roi":
            l2 = dict(l); l2["dc"] = dc
            num = (l2["dc"] * sf["dc"] + l2["refrigeration"]
                   * sf["refrigeration"] + l2["process"] * sf["process"]
                   + l2["comfort"] * sf["comfort"])
            blend = num / (l2["dc"] + l2["refrigeration"]
                           + l2["process"] + l2["comfort"])
        else:
            blend = sf["ni_all"]
        cool_week = flat_a * 1000.0 / 52.0
        if comfort_a > 0:
            frac = odh_frac if odh_frac is not None else 1.0 / 52.0
            cool_week += comfort_a * 1000.0 * (0.3 / 52.0 + 0.7 * frac)
        cool_bill_native = cool_week * elec_price
        cool_kt = cool_week * EF["electricity"] / 1000.0
        cool_indig = cool_week * j["elec_indigenous"]
        cool_served = cool_week * blend
        cooling = {
            "elec_gwh": round(cool_week, 1),
            "served_gwh": round(cool_served, 1),
            "service_factor": round(blend, 2),
            "bill_eur_m": round(cool_bill_native if cur == "eur"
                                else cool_bill_native / fx, 1),
            "bill_gbp_m": round(cool_bill_native * fx if cur == "eur"
                                else cool_bill_native, 1),
            "emissions_kt_co2": round(cool_kt, 1),
            "comfort_shaped_by_odh": bool(comfort_a > 0
                                          and odh_frac is not None),
        }
        combined = {
            "purchased_gwh": round(inp_t + cool_week, 1),
            "served_gwh": round(useful_t + cool_served, 1),
            "bill_eur_m": round((bill_t if cur == "eur"
                                 else bill_t / fx)
                                + cooling["bill_eur_m"], 1),
            "bill_gbp_m": round((bill_t * fx if cur == "eur"
                                 else bill_t)
                                + cooling["bill_gbp_m"], 1),
            "emissions_kt_co2": round(kt_t + cool_kt, 1),
            # heat/cold splits (schema 2): cold = the cold-economy
            # electricity; heat = the fuel side. Pairs reconcile per
            # currency within rounding.
            "heat_gwh": round(inp_t, 1),
            "cold_gwh": round(cool_week, 1),
            "bill_heat_eur_m": round(bill_t if cur == "eur"
                                     else bill_t / fx, 1),
            "bill_cold_eur_m": cooling["bill_eur_m"],
            "bill_heat_gbp_m": round(bill_t * fx if cur == "eur"
                                     else bill_t, 1),
            "bill_cold_gbp_m": cooling["bill_gbp_m"],
            "emissions_heat_kt": round(kt_t, 1),
            "emissions_cold_kt": round(cool_kt, 1),
            # delivered basis: purchased electricity carries the grid
            # share; the balance of the cooling service is ambient
            # rejection, indigenous by convention
            "indigenous_share_pct": round(100 * (
                indig_t + cool_indig + (cool_served - cool_week))
                / max(useful_t + cool_served, 1e-9), 1),
        }

        # peak winter week for the for-scale line
        days = sorted(hdd_series)[-365:]
        pk, pk_end = 0.0, None
        for i in range(6, len(days)):
            w = sum(hdd_series[d] for d in days[i - 6:i + 1])
            if w > pk:
                pk, pk_end = w, days[i]
        peak = None
        if pk_end and hdd_year > 0:
            annual_gwh = heat_twh * 1000.0
            peak = {"week_ending": pk_end, "hdd": round(pk, 1),
                    "heat_purchased_gwh": round(
                        annual_gwh * ((1 - shf) / 52.0
                                      + shf * pk / hdd_year), 1)}
        # combined what-if: heat side as computed in wf; cooling side
        # moves 20% of load to ground-coupled systems at the dagger
        # saving factor - service unchanged, electricity cut, the
        # avoided share ambient and indigenous
        save = cc["ground_cooling_saving"]
        cool_wf_elec = cool_week * (1 - 0.20 * save)
        wf_cool_bill = cool_wf_elec * elec_price
        wf_heat_indig_abs = wf["indigenous_share_pct"] / 100.0 * useful_t
        wf_combined = {
            "purchased_gwh": round(wf["heat_purchased_gwh"]
                                   + cool_wf_elec, 1),
            "bill_eur_m": round(wf["bill_eur_m"]
                                + (wf_cool_bill if cur == "eur"
                                   else wf_cool_bill / fx), 1),
            "bill_gbp_m": round(wf["bill_gbp_m"]
                                + (wf_cool_bill * fx if cur == "eur"
                                   else wf_cool_bill), 1),
            "emissions_kt_co2": round(
                wf["emissions_kt_co2"]
                + cool_wf_elec * EF["electricity"] / 1000.0, 1),
            "indigenous_share_pct": round(100 * (
                wf_heat_indig_abs
                + cool_wf_elec * j["elec_indigenous"]
                + (cool_served - cool_wf_elec))
                / max(useful_t + cool_served, 1e-9), 1),
            # what-if splits: cold at (1 - R x ground_cooling_saving)
            # - deliberately NOT the UK's COP-20 arithmetic, per the
            # what-if parameters addendum to the cross-calibration
            "heat_gwh": round(wf["heat_purchased_gwh"], 1),
            "cold_gwh": round(cool_wf_elec, 1),
            "bill_heat_eur_m": round(wf["bill_eur_m"], 1),
            "bill_cold_eur_m": round(wf_cool_bill if cur == "eur"
                                     else wf_cool_bill / fx, 1),
            "bill_heat_gbp_m": round(wf["bill_gbp_m"], 1),
            "bill_cold_gbp_m": round(wf_cool_bill * fx if cur == "eur"
                                     else wf_cool_bill, 1),
            "emissions_heat_kt": round(wf["emissions_kt_co2"], 1),
            "emissions_cold_kt": round(
                cool_wf_elec * EF["electricity"] / 1000.0, 1),
        }
        return {
            "heat_purchased_gwh": round(inp_t, 1),
            "heat_delivered_gwh": round(useful_t, 1),
            "cooling": cooling,
            "combined": combined,
            "what_if_combined": wf_combined,
            "by_fuel": by_fuel,
            "peak_week": peak,
            "indigenous_share_pct": round(
                100 * indig_t / max(useful_t, 1e-9), 1),
            "bill_eur_m": round(bill_eur, 1),
            "bill_gbp_m": round(bill_gbp, 1),
            "emissions_kt_co2": round(kt_t, 1),
            "hdd_week": round(hdd_week, 1),
            "hdd_year": round(hdd_year, 1),
            "week_ending": week_end,
            "what_if_20pct_geothermal": wf,
            "_raw": {"useful": useful_t, "indig": indig_t},
        }

    # Weekly comfort-cooling share of the year. The fraction is only
    # meaningful against a FULL year of overheating: with a partial
    # series the denominator omits the rest of the season and the
    # summer weeks inflate (observed 18 Jul 2026: a 61-day series made
    # one July week ~14% of "annual"). Hero shaping therefore requires
    # >=300 days; until then comfort is flat here. The cold-economy
    # panel keeps its 60-day rule - its supply and demand shapes share
    # one window, so partial coverage stays internally consistent.
    # electricity emission factor: live all-island grid intensity
    # (trailing 14-day mean) once at least 7 days exist; anchor
    # otherwise. The 280 g anchor predates the EirGrid restoration;
    # live summer intensity runs ~210-220 g.
    # Electricity indigenous share: HELD AT ANCHOR. The sem_mix live
    # share failed its cross-examination on first adoption (30 Jul
    # 2026): 14-day mean 31.2% against anchors of 41.3/46% during a
    # fortnight whose live grid intensity (~179 g/kWh) implies ~half
    # of generation zero-carbon. Two identified biases: the IE feed
    # has no Solar category, and the cross-border sign convention is
    # unverified (exports may be inflating the denominator). The feed
    # keeps collecting; the diagnostic below logs the divergence each
    # run; the anchor is not replaced until the live share reconciles
    # with the CI evidence.
    ind_src = "anchor (SEAI RES-E / DfE)"
    if not W:
        sm = ((feeds.get("sem_mix") or {})
              .get("indigenous_share_daily") or {})
        sdays = sorted(sm)[-14:]
        if len(sdays) >= 7:
            live_ind = sum(sm[d] for d in sdays) / len(sdays)
            co2d = ((feeds.get("eirgrid") or {})
                    .get("co2_intensity_g_per_kwh") or {})
            cdd = sorted(co2d)[-14:]
            ci_note = ""
            if len(cdd) >= 7:
                ci = sum(co2d[d] for d in cdd) / len(cdd)
                # CI-implied zero-carbon share against a ~420 g/kWh
                # dagger fossil-fleet average - coarse, diagnostic only
                zc = max(0.0, min(1.0, 1.0 - ci / 420.0)) * 100.0
                ci_note = f"; CI-implied zero-carbon ~{zc:.0f}%"
            log(f"sem_mix diagnostic: 14d live share {live_ind:.1f}% "
                f"vs anchors ROI 41.3 / NI 46{ci_note} - held at "
                f"anchor pending solar/sign validation")
        log(f"hero: elec indigenous share - {ind_src}")

    EF = dict(a["ef_g_per_kwh"])
    if W:
        ef_src = W["ef_source"]
        if W.get("ef_electricity"):
            EF["electricity"] = W["ef_electricity"]
    else:
        co2 = ((feeds.get("eirgrid") or {})
               .get("co2_intensity_g_per_kwh") or {})
        cdays = sorted(co2)[-14:]
        ef_src = "anchor"
        if len(cdays) >= 7:
            EF["electricity"] = round(
                sum(co2[d] for d in cdays) / len(cdays), 1)
            ef_src = f"live grid intensity, {len(cdays)}-day mean"
        log(f"hero: electricity EF {EF['electricity']} g/kWh ({ef_src})")

    odh = {} if W else (hddf.get("odh26_island") or {})
    odh_frac = None
    odays = sorted(odh)[-365:]
    if len(odays) >= 300:
        oy = sum(odh[d] for d in odays)
        if oy > 0:
            odh_frac = sum(odh[d] for d in odays[-7:]) / oy

    roi = jur_block("roi", "eur", hddf.get("hdd_roi") or island_hdd)
    ni = jur_block("ni", "gbp", hddf.get("hdd_ni") or island_hdd)
    if not (roi and ni):
        return None

    def sum_blocks(x, y):
        useful = x["_raw"]["useful"] + y["_raw"]["useful"]
        indig = x["_raw"]["indig"] + y["_raw"]["indig"]
        bf = {}
        for src_b in (x, y):
            for f, v in (src_b.get("by_fuel") or {}).items():
                slot = bf.setdefault(f, {"in_gwh": 0.0, "useful_gwh": 0.0})
                slot["in_gwh"] = round(slot["in_gwh"] + v["in_gwh"], 1)
                slot["useful_gwh"] = round(
                    slot["useful_gwh"] + v["useful_gwh"], 1)
        pk = None
        if x.get("peak_week") and y.get("peak_week"):
            pk = {"week_ending": max(x["peak_week"]["week_ending"],
                                     y["peak_week"]["week_ending"]),
                  "heat_purchased_gwh": round(
                      x["peak_week"]["heat_purchased_gwh"]
                      + y["peak_week"]["heat_purchased_gwh"], 1)}
        def addd(key, fields):
            return {f: round(x[key][f] + y[key][f], 1) for f in fields}
        cooling = addd("cooling", ("elec_gwh", "served_gwh",
                                   "bill_eur_m", "bill_gbp_m",
                                   "emissions_kt_co2"))
        cooling["comfort_shaped_by_odh"] = (
            x["cooling"]["comfort_shaped_by_odh"]
            or y["cooling"]["comfort_shaped_by_odh"])
        combined = addd("combined", ("purchased_gwh", "served_gwh",
                                     "bill_eur_m", "bill_gbp_m",
                                     "emissions_kt_co2",
                                     "heat_gwh", "cold_gwh", "bill_heat_eur_m", "bill_cold_eur_m", "bill_heat_gbp_m", "bill_cold_gbp_m", "emissions_heat_kt", "emissions_cold_kt"))
        served = x["combined"]["served_gwh"] + y["combined"]["served_gwh"]
        combined["indigenous_share_pct"] = round(
            (x["combined"]["indigenous_share_pct"]
             * x["combined"]["served_gwh"]
             + y["combined"]["indigenous_share_pct"]
             * y["combined"]["served_gwh"]) / max(served, 1e-9), 1)
        wfc = addd("what_if_combined", ("purchased_gwh", "bill_eur_m",
                                        "bill_gbp_m", "emissions_kt_co2",
                                        "heat_gwh", "cold_gwh", "bill_heat_eur_m", "bill_cold_eur_m", "bill_heat_gbp_m", "bill_cold_gbp_m", "emissions_heat_kt", "emissions_cold_kt"))
        wfc["indigenous_share_pct"] = round(
            (x["what_if_combined"]["indigenous_share_pct"]
             * x["combined"]["served_gwh"]
             + y["what_if_combined"]["indigenous_share_pct"]
             * y["combined"]["served_gwh"]) / max(served, 1e-9), 1)
        out = {
            "cooling": cooling,
            "combined": combined,
            "what_if_combined": wfc,
            "by_fuel": bf,
            "peak_week": pk,
            "heat_purchased_gwh": round(
                x["heat_purchased_gwh"] + y["heat_purchased_gwh"], 1),
            "heat_delivered_gwh": round(
                x["heat_delivered_gwh"] + y["heat_delivered_gwh"], 1),
            "indigenous_share_pct": round(100 * indig / max(useful, 1e-9), 1),
            "bill_eur_m": round(x["bill_eur_m"] + y["bill_eur_m"], 1),
            "bill_gbp_m": round(x["bill_gbp_m"] + y["bill_gbp_m"], 1),
            "emissions_kt_co2": round(
                x["emissions_kt_co2"] + y["emissions_kt_co2"], 1),
        }
        wf = {}
        for k in ("heat_purchased_gwh", "bill_eur_m", "bill_gbp_m",
                  "emissions_kt_co2"):
            wf[k] = round(x["what_if_20pct_geothermal"][k]
                          + y["what_if_20pct_geothermal"][k], 1)
        wf["geothermal_spf"] = a["geothermal_spf"]
        wf["indigenous_share_pct"] = round(
            (x["what_if_20pct_geothermal"]["indigenous_share_pct"]
             * x["_raw"]["useful"]
             + y["what_if_20pct_geothermal"]["indigenous_share_pct"]
             * y["_raw"]["useful"]) / max(useful, 1e-9), 1)
        out["what_if_20pct_geothermal"] = wf
        return out

    island = sum_blocks(roi, ni)
    hdd_week, hdd_year, week_end = hdd_stats(island_hdd)
    for b in (roi, ni):
        b.pop("_raw", None)

    out = dict(island)
    out.update({
        "week_ending": week_end,
        "hdd_week": round(hdd_week, 1), "hdd_year": round(hdd_year, 1),
        "roi": roi, "ni": ni,
        "ef_electricity_g_per_kwh": EF["electricity"],
        "ef_electricity_source": ef_src,
        "elec_indigenous_source": ind_src,
        "basis": ("Scaffold estimator (dagger throughout) - annual anchors "
                  "shaped by each jurisdiction's weekly HDD; SEAI 2024, "
                  "DfE/NISRA, Causeway estimates. Oil, the island's "
                  "majority heating fuel, is modelled from annual "
                  "anchors, not metered - its weekly estimates carry a "
                  "materially wider band than gas. Hot water is carried as a flat "
                  "22.4% of annual (SEAI National Heat Study, Aug 2026 "
                  "re-anchoring); space heat follows the week's "
                  "degree days. Bills are sector-blended: services gas and "
                  "electricity, and all cooling, price at non-domestic "
                  "rates (dagger, pending Eurostat band prices); oil "
                  "prices identically across sectors. VAT convention: "
                  "domestic rates include VAT (5% NI, 9% ROI), "
                  "non-domestic rates exclude it, because businesses "
                  "recover input VAT - Eurostat level 3 and level 2 "
                  "respectively. NI gas and electricity are all-in "
                  "rates at the Utility Regulator's own consumption "
                  "basis, gas weighted across SSE Airtricity (Greater "
                  "Belfast and West) and Firmus (Ten Towns) by "
                  "regulated customer count; ROI is a standard unit "
                  "rate without standing charges, so the two "
                  "jurisdictions are internally consistent but not "
                  "like-for-like at component level (open item). The "
                  "Domestic rates in BOTH jurisdictions are "
                  "all-in effective rates at a stated consumption, "
                  "including VAT and standing charges - NI from the "
                  "Utility Regulator's regulated bills, ROI from the "
                  "Eurostat band price (S2 2024) stepped by the "
                  "supplier announcements. The residual difference is "
                  "scope rather than basis: NI is incumbent-weighted "
                  "regulated, ROI a market-wide average including "
                  "discounts, and the two do not bias in one "
                  "direction (dagger). Cooling is the "
                  "cold-economy "
                  "census (dagger loads beside the CSO data-centre "
                  "anchor), flat across the year with the comfort share "
                  "following live overheating degree-hours once a season "
                  "exists - a cold-economy scope, wider than comfort-only national "
                  "cooling lines and not one-to-one comparable with them; the "
                  "data-centre line counts cooling electricity (~14% of "
                  "the fleet's draw, SEAI), not the whole draw, and "
                  "delivered cooling uses the SEAI National Heat "
                  "Study's useful-to-final ratios. "
                  "Electricity emissions use the live all-island grid "
                  "intensity when available. Challenge and input "
                  "welcome at contact@causewaygt.com"),
        "anchors_used": a,
    })
    return out


def derive_ashp_spf(hdd_daily: dict, anchors=None):
    """
    Air-source heat pump seasonal performance from the HDD series itself.
    For heating days T_out = base - HDD, so the demand-weighted source
    temperature is base - sum(h^2)/sum(h) over the trailing year. COP is a
    Carnot fraction at 45C flow with a defrost derate, blended (harmonic,
    energy-weighted) with a DHW share at 55C. All parameters dagger - see
    ANCHORS["ashp"]. Returns None without a season of data.
    """
    a = (anchors or ANCHORS)
    p = a["ashp"]
    days = sorted(hdd_daily)[-365:]
    hs = [hdd_daily[d] for d in days if hdd_daily[d] > 0]
    if len(hs) < 60 or sum(hs) < 200:
        return None
    w = sum(hs)
    t_src = HDD_BASE_C - sum(h * h for h in hs) / w

    def cop(source_c, flow_c, derate=1.0):
        lift = flow_c - source_c
        if lift <= 5:
            lift = 5.0
        return p["carnot_fraction"] * (flow_c + 273.15) / lift * derate

    # defrost_factor is the site's own Gaussian, centred on the 2 C
    # frost band - NOT the flat defrost_derate anchor, which applied a
    # 10% penalty as a STEP across 0-7 C and switched it off entirely
    # below zero, so modelled COP rose as it got colder. Both the cost
    # panel and this now use the same curve.
    space = cop(t_src, p["flow_c"], defrost_factor(t_src))
    dhw = cop(p["dhw_source_c"], p["dhw_flow_c"])
    sh = p["dhw_share"]
    spf = 1.0 / ((1 - sh) / space + sh / dhw)
    return {"spf": round(spf, 2),
            "demand_weighted_source_c": round(t_src, 1),
            "space_cop": round(space, 2), "dhw_cop": round(dhw, 2),
            "params": p}


# ------------------------------------------------------- Phase B.2.1
# The de-rated all-island dispatchable block, MW. AIRAA Appendix 3
# registered capacity x Table 5.18 availability factors, of which
# ~1,490 MW is run-hour-limited. Dagger. Observed all-island peak was
# 7,502 MW on 8 Jan 2025, which is the sanity rail: a computed demand
# far above that on a mild hour means the shaping is wrong, not that
# the island is short of plant.
# ROUTE TIERS - aligned with the UK sibling, 8 Aug 2026, so the two
# grid layers are read against the same ladder. Field-observed
# in-situ figures, not brochure SCOPs:
#   ASHP  2.80  Energy Systems Catapult, Electrification of Heat median
#   GSHP  3.24  Energy Systems Catapult, in-situ GSHP average
#   NET   5.00  networked geothermal on a shared ambient loop
# The Irish site keeps its own Carnot-fraction engine for the AIR
# route, because the whole point of that column is that air-source
# performance collapses in the hour that binds - a flat 2.80 would
# hide exactly the effect being measured. ASHP_SPF is therefore the
# seasonal comparator, not the hourly one, and the two are logged
# side by side. Ground and network are flat by construction: a
# borehole field does not care what the air is doing.
ASHP_SPF = 2.80
GSHP_SPF = 3.24
GEO_NETWORK_SCOP = 5.00

GRID_BLOCK_MW = 8595
GRID_RUN_HOUR_LIMITED_MW = 1490
OBSERVED_PEAK_MW = 7502
OBSERVED_PEAK_AT = "2025-01-08"


def hourly_heat_mw(store, anchors=None):
    """
    Island building-heat demand per hour, MW of USEFUL heat.

    Shaped the way the weekly hero shapes a week, one level finer:
    hot water flat, space heat following each hour's share of the
    store's own heating-degree total. Degree hours come from temp_ai,
    so this is the island's own weather rather than a profile.

    Returns {hour_key: MW} or {} if the store cannot support it.
    """
    a = anchors or ANCHORS
    temps = (expand_hourly(store) or {}).get("temp_ai") or {}
    if len(temps) < 24 * 300:
        return {}
    # useful heat, not input: the anchors are fuel input and each fuel
    # burns at its own efficiency, so convert before shaping.
    useful_twh = 0.0
    for jur in ("ni", "roi"):
        j = a[jur]
        heat = j["residential_heat_twh"] + j["services_heat_twh"]
        useful_twh += heat * sum(sh * a["efficiency"][f]
                                 for f, sh in j["fuel_shares"].items())
    n = len(temps)
    dh = {k: max(0.0, HDD_BASE_C - v) for k, v in temps.items()}
    total_dh = sum(dh.values())
    if total_dh <= 0:
        return {}
    # The store is ~13 months, so scale the annual anchor to its span
    # rather than assuming a calendar year.
    span_years = n / (365.25 * 24)
    dhw_mwh = useful_twh * (1 - a["space_heat_fraction"]) * 1e6 * span_years
    space_mwh = useful_twh * a["space_heat_fraction"] * 1e6 * span_years
    flat = dhw_mwh / n
    return {k: flat + space_mwh * dh[k] / total_dh for k in temps}


GRID_WHATIF_SHARE = 0.20


def derive_grid_views(store, anchors=None):
    """
    The three series Panel 3 plots, at the resolution each view needs.

    NOT the whole hourly store. The live week is 168 hourly points, the
    daily view 90, the falcon 24 monthly - about 280 rows against the
    store's 9,384 hours, so the payload carries what is drawn rather
    than what was computed. Each row is the island's delivered heat and
    the electricity the SITE'S OWN 20% what-if would draw by route,
    netted of the resistive heating already in observed demand.

    The binding-hour panel above solves for the share that fits; this
    is the fixed 20% the rest of the site uses, so the two answer
    different questions and must not be read as one.
    """
    a = anchors or ANCHORS
    heat = hourly_heat_mw(store, a)
    if not heat:
        return None
    ser = expand_hourly(store) or {}
    temps = ser.get("temp_ai") or {}
    demand = ser.get("demand_ai") or {}
    hours = sorted(k for k in heat if k in demand)
    if len(hours) < 24 * 30:
        return None

    res = 0.0
    for jur in ("ni", "roi"):
        j = a[jur]
        h = j["residential_heat_twh"] + j["services_heat_twh"]
        res += h * j["fuel_shares"].get("electricity", 0.0) \
            * a["efficiency"].get("electricity", 1.0)
    tot = sum((a[j]["residential_heat_twh"] + a[j]["services_heat_twh"])
              * sum(sh * a["efficiency"][f]
                    for f, sh in a[j]["fuel_shares"].items())
              for j in ("ni", "roi"))
    res_share = (res / tot) if tot else 0.0
    s = GRID_WHATIF_SHARE
    # The SAME air COP the binding-hour panel uses, from the same
    # anchors. Two hourly COP models on one panel would be a bug
    # waiting to be found by a reader adding the numbers up.
    p = a["ashp"]

    def cop_at(t):
        lift = max(5.0, p["flow_c"] - t)
        return p["carnot_fraction"] * (p["flow_c"] + 273.15) / lift \
            * defrost_factor(t)

    def row(keys):
        """Mean MW across the keys, plus each route's what-if draw."""
        q = sum(heat[k] for k in keys) / len(keys)
        t = sum(temps.get(k, 5.0) for k in keys) / len(keys)
        d = sum(demand[k] for k in keys) / len(keys)
        cop = max(1.0, cop_at(t))
        disp = s * q * res_share
        return {
            "heat_mw": round(q), "temp_c": round(t, 1),
            "demand_mw": round(d),
            "air_source": round(s * q / cop - disp),
            "ground_source": round(s * q / GSHP_SPF - disp),
            "geothermal_network": round(s * q / GEO_NETWORK_SCOP - disp),
        }

    def stamp(k, n):
        return k[:n]

    hourly = [dict(row([k]), t=k) for k in hours[-168:]]
    days, months = {}, {}
    for k in hours:
        days.setdefault(stamp(k, 10), []).append(k)
        months.setdefault(stamp(k, 7), []).append(k)
    daily = [dict(row(v), t=d) for d, v in sorted(days.items())[-90:]]
    monthly = [dict(row(v), t=m) for m, v in sorted(months.items())[-24:]]
    # THE FALCON. Not two years of history - a calendar year, each
    # month filled with the LATEST COMPLETE instance of it. That is
    # how the UK sibling builds it, and it is why twelve complete
    # months is enough: Jan-Jul from this year, Aug-Dec from last.
    # A month is complete only if the store covers a full one, so the
    # first and last months of the span are dropped as partial.
    keys = sorted(months)
    complete = [m for m in keys[1:-1]]
    latest = {}
    for m in complete:
        latest[m[5:7]] = m          # later months overwrite earlier
    falcon = [dict(row(months[latest[f"{i:02d}"]]),
                   t=latest[f"{i:02d}"], m=f"{i:02d}")
              for i in range(1, 13) if f"{i:02d}" in latest]
    out = {"share": s, "hourly": hourly, "daily": daily,
           "monthly": monthly, "falcon": falcon,
           "falcon_complete": len(falcon),
           "span": [hours[0], hours[-1]]}
    log(f"grid views: {len(hourly)} hourly, {len(daily)} daily, "
        f"{len(monthly)} monthly, {len(falcon)}/12 falcon months at a "
        f"{int(s * 100)}% what-if ({hours[0]}..{hours[-1]})")
    return out


def derive_tightest_hour(store, feeds=None, anchors=None):
    """
    B.2.1 - the tightest hour. Log-only; nothing draws from it yet,
    by design: the three B.2 computations are run and logged before
    any panel is written, so the headline is chosen after the numbers
    are seen rather than before.

    Electrified heat is added to OBSERVED demand net of the resistive
    heating it displaces - that share is already in the demand series,
    so counting the heat pump without removing the immersion would
    double-count it.
    """
    a = anchors or ANCHORS
    heat = hourly_heat_mw(store, a)
    if not heat:
        return None
    ser = expand_hourly(store) or {}
    demand, temps = ser.get("demand_ai") or {}, ser.get("temp_ai") or {}
    hours = sorted(set(heat) & set(demand))
    if len(hours) < 24 * 300:
        return None

    p = ANCHORS["ashp"]

    def cop_at(t):
        lift = max(5.0, p["flow_c"] - t)
        return p["carnot_fraction"] * (p["flow_c"] + 273.15) / lift \
            * defrost_factor(t)

    # resistive share of useful heat, from the fuel shares
    res = 0.0
    for jur in ("ni", "roi"):
        j = a[jur]
        h = j["residential_heat_twh"] + j["services_heat_twh"]
        res += h * j["fuel_shares"].get("electricity", 0.0) \
            * a["efficiency"].get("electricity", 1.0)
    tot = sum((a[j]["residential_heat_twh"] + a[j]["services_heat_twh"])
              * sum(sh * a["efficiency"][f]
                    for f, sh in a[j]["fuel_shares"].items())
              for j in ("ni", "roi"))
    res_share = (res / tot) if tot else 0.0

    wind = ser.get("wind_ai") or {}
    solar = ser.get("solar_ai") or {}

    def ceiling_at(key):
        """The fleet that is actually there in that hour.

        De-rated dispatchable capacity PLUS the wind and solar the
        island actually generated - the ceiling breathes with the
        weather, because the hour that binds is cold AND still AND
        dark, and a flat block would credit wind that was not blowing
        or ignore wind that was. This is also what makes the figure
        comparable with the UK sibling, whose ceiling is defined the
        same way.
        """
        return GRID_BLOCK_MW + (wind.get(key) or 0.0) + (solar.get(key) or 0.0)

    # --- how far can heat be electrified before the fleet fills?
    # Solving for the SHARE is the right question, and the one the UK
    # sibling asks. Fixing a share instead forces a choice between the
    # site's own 20% what-if and a 100% ceiling that appears nowhere
    # else, and the answer swings entirely on which is picked. The
    # netting is linear in the share - displaced resistive scales with
    # it - so share = headroom / added-at-100%, exactly.
    routes = ("air_source", "ground_source", "geothermal_network")
    best = {r: None for r in routes}
    rows = []
    for k in hours:
        q = heat[k]
        displaced = q * res_share            # already drawn as MW today
        full = {"air_source": q / cop_at(temps.get(k, 5.0)) - displaced,
                "ground_source": q / GSHP_SPF - displaced,
                "geothermal_network": q / GEO_NETWORK_SCOP - displaced}
        head = ceiling_at(k) - demand[k]
        for r in routes:
            if full[r] <= 0:
                continue
            share = max(0.0, head / full[r])
            if best[r] is None or share < best[r][0]:
                best[r] = (share, k)
        rows.append((demand[k] + full["air_source"], k, q,
                     full["air_source"] + displaced,
                     full["ground_source"] + displaced,
                     full["geothermal_network"] + displaced))
    worst = max(rows)
    total_mw, k, q, ashp, gshp, net = worst
    added = {"air_source": round(ashp - q * res_share),
             "ground_source": round(gshp - q * res_share),
             "geothermal_network": round(net - q * res_share)}
    # Headroom PER ROUTE. The first live run reported it for the air
    # route alone, which meant the actual result - that only the
    # network route fits under the block - had to be worked out by
    # hand from three other numbers. The route ordering IS the
    # finding; it should not need subtracting.
    totals = {r: round(demand[k]) + v for r, v in added.items()}
    ceil_k = round(ceiling_at(k))
    headroom = {r: ceil_k - t for r, t in totals.items()}
    fits_share = {r: (round(100 * best[r][0], 1) if best[r] else None)
                  for r in routes}
    fits_hour = {r: (best[r][1] if best[r] else None) for r in routes}
    fits = [r for r, h in headroom.items() if h >= 0]
    out = {
        "hour": k, "observed_mw": round(demand[k]),
        "total_mw": totals, "headroom_by_route_mw": headroom,
        "routes_that_fit": fits,
        # The heat system against the power system, same hour, same
        # units. This is the quotable one and it was computed all
        # along without ever being printed.
        "heat_vs_block_ratio": round(q / GRID_BLOCK_MW, 2),
        "air_c": temps.get(k), "useful_heat_mw": round(q),
        "added_mw": added,
        "hour_cop_air": round(cop_at(temps.get(k, 5.0)), 2),
        "spf": {"air_seasonal": ASHP_SPF, "ground": GSHP_SPF,
                "network": GEO_NETWORK_SCOP},
        "total_with_air_source_mw": round(total_mw),
        "block_mw": GRID_BLOCK_MW,
        "ceiling_mw": ceil_k,
        "wind_solar_mw": round(ceil_k - GRID_BLOCK_MW),
        "share_that_fits_pct": fits_share,
        "share_binding_hour": fits_hour,
        "headroom_mw": headroom["air_source"],
        "observed_peak_mw": OBSERVED_PEAK_MW,
        "hours_considered": len(hours),
        "basis": ("Hourly useful heat from temp_ai degree hours (hot "
                  "water flat, space heat degree-shaped), through a "
                  "Carnot-fraction COP at each hour's own air "
                  "temperature, NETTED of the resistive heating "
                  "already in observed demand, added to observed "
                  "all-island demand. Block is the de-rated "
                  "dispatchable capacity (dagger); no panel drawn."),
    }
    log(f"B.2.1 how far can heat be electrified inside today's fleet? "
        f"(ceiling = {GRID_BLOCK_MW} MW de-rated block + the wind and "
        f"solar actually generated)")
    for r in routes:
        pct, hr = fits_share[r], fits_hour[r]
        log(f"B.2.1   {r:<19} {pct:>6}% of island heat fits "
            f"[binds {hr}]"
            + ("  - ALL OF IT, with room over" if pct and pct >= 100 else ""))
    log(f"B.2.1 tightest hour for added load: {k} at {out['air_c']} C - "
        f"observed {out['observed_mw']} MW, ceiling {ceil_k} MW "
        f"({out['wind_solar_mw']} MW of it wind and solar)")
    for r in ("air_source", "ground_source", "geothermal_network"):
        h = headroom[r]
        log(f"B.2.1   at 100%: {r:<19} +{added[r]:>5} MW -> "
            f"{totals[r]:>6} MW, headroom {h:+6} MW  "
            f"[{'FITS' if h >= 0 else 'EXCEEDS THE CEILING'}]")
    log(f"B.2.1   routes that fit: "
        f"{', '.join(fits) if fits else 'NONE'}")
    log(f"B.2.1   useful heat in that hour {out['useful_heat_mw']} MWth "
        f"against an {GRID_BLOCK_MW} MW electrical block - the island's "
        f"heat system is {out['heat_vs_block_ratio']}x its power system")
    log(f"B.2.1   air COP in that hour {out['hour_cop_air']} against a "
        f"seasonal {ASHP_SPF} - the gap IS the argument; ground {GSHP_SPF}, "
        f"network {GEO_NETWORK_SCOP} are flat by construction")
    log(f"B.2.1   observed peak on record {OBSERVED_PEAK_MW} MW "
        f"({OBSERVED_PEAK_AT}) - this hour is "
        f"{round(100 * out['observed_mw'] / OBSERVED_PEAK_MW)}% of it, "
        f"which is the sanity rail: a binding hour far below that "
        f"would mean the shaping picked a mild hour")
    if min(headroom.values()) < 0:
        log("B.2.1   NOTE headroom is negative - a full electrification "
            "of heat exceeds the ceiling in this hour. That is a "
            "scenario result, not a forecast: nothing here phases the "
            "conversion or counts storage, diversity or demand response.")
    return out


# Comparator national buildings heat, input basis, for the calibrated
# chart. SHARED WITH THE UK SIBLING and identical to its NAT_HEAT_TWH:
# France ~350 TWh (IEA / Heat Roadmap order), Netherlands ~115 TWh,
# Sweden ~80 TWh (Swedish Energy Agency / ODYSSEE). Estimates, daggered.
# Per-capita watts, computed from the anchors above rather than typed,
# so re-anchoring a capacity cannot leave the headline stale.
GEO["per_capita_w"]["roi"] = round(
    GEO["roi"]["capacity_mwth"] * 1e6 / (GEO["population_m"]["roi"] * 1e6))
GEO["per_capita_w"]["ni"] = round(
    GEO["ni_capacity_mwth_est"] * 1e6 / (GEO["population_m"]["ni"] * 1e6))

GEO_NAT_HEAT_TWH = {"France": 350.0, "Netherlands": 115.0, "Sweden": 80.0}
# The register's inclusion threshold. Both exclusions sit below it -
# Randalstown 44 kW, Strabane 18 kW - which is what fixes it at 45.
GEO_NI_REGISTER_KW = 45
GEO_SOURCES = {
    "roi": "WGC2026 Country Update: Ireland \u2014 Ireland, Blake, "
           "Pasquali, Dunphy & Hunter Williams, June 2026",
    # The register is Causeway's own compilation, currently OUT FOR
    # COMMENT among Northern Ireland practitioners rather than
    # published. The site cites it as a source and reports totals
    # derived from it; it does not reproduce the entries, and should
    # not until the review closes.
    "ni_register": "Causeway Energies register of Northern Ireland "
                   "ground-source schemes above 45 kW, compiled site "
                   "by site and currently circulating for comment "
                   "among Northern Ireland practitioners",
    "ni_domestic": "MCS certification records, ~386\u2013450 units, plus "
                   "a pre-certification estimate \u2014 Causeway "
                   "triangulation",
    "comparators": "EGC 2025 country updates (Sanner et al., Tables "
                   "3\u20134, end-2024) \u2014 shared with the UK sibling",
}
# NI Energy Strategy "The Path to Net Zero Energy" (DfE, Dec 2021): a
# 25% reduction in energy use from buildings and industry by 2030,
# expressed as 8,000 GWh of savings. It is the ONLY quantified energy
# target in Northern Ireland that geothermal can be measured against -
# there is no NI renewable heat target of any kind, and RED III's
# heating sub-targets do not bind NI.
#
# It also suits the mechanism: the target counts energy SAVED, and a
# heat pump saves purchased energy by construction. Delivery stands at
# 90 GWh of 8,000 as at March 2025 - 1% - on about GBP 107m spent since
# 2020 (NI Audit Office, 21 Oct 2025).
GEO_NI_EE_TARGET_GWH = 8000.0
GEO_NI_EE_ACHIEVED_GWH = 90.0


# ---------------------------------------------------------------- VFM
# Panel 6: value for money. Two counterfactuals, two jurisdictions, two
# appraisal rulebooks, and every shared quantity READ from the panels
# that already publish it rather than restated here. Restating is how
# the data-centre cooling share and the geothermal EERs drifted between
# panels twice this week; asserts below fail the build if they diverge.
VFM_COUNTERFACTUALS = ("bau", "ashp")
# BAU is the incumbent the sector actually runs on - oil-dominated in
# Northern Ireland, mixed in the Republic. ASHP is the electrified
# alternative. THEY GIVE OPPOSITE SIGNS ON CAPACITY, which is the
# single thing most likely to be misread:
#   against BAU  geothermal ADDS electrical peak, because the load was
#                not electric at all
#   against ASHP geothermal SAVES peak, because it draws the same heat
#                at a higher SPF
# Panel 3 measures this directly at the binding hour. Any capacity
# benefit therefore belongs to the ASHP lever alone and must never be
# shown against BAU.
VFM_CAPACITY_APPLIES_TO = ("ashp",)

# All-island SEM capacity auction outturn, EUR per MW-year, de-rated.
# THE SINGLE LARGEST DIVERGENCE FROM THE UK SIBLING, which uses a
# GB-derived composite of about GBP 75/kW/yr. The SEM is one all-island
# market and clears several times higher.
VFM_SEM_CAPACITY = {
    "t4_2028_29": 149960.0,      # auction 20 Dec 2024, final 16 Jan 2025
    "t4_2029_30": 135499.99,     # auction 26 Mar 2026, provisional 14 Apr
    "t1_2025_26": 90000.0,       # auction 22 May 2025, final 1 Jul 2025
    "net_cone_2028_29": 113355.0,   # SEM Committee parameters
    "existing_cap_price_cap": 55678.0,   # 0.5x Net CONE - the low case
    "source": "SEMO capacity auction reports; SEM Committee parameters",
}
# GB, for the reconciliation note only - never used in an Irish figure.
VFM_GB_CAPACITY_GBP_KW = {"t4_2028_29": 60.00, "t4_2029_30": 27.10}

# Two rulebooks, and they differ in ways that change the answer.
VFM_APPRAISAL = {
    "roi": {
        "name": "Infrastructure Guidelines, DPENDR, Dec 2023",
        "discount": "4% real, flat",
        "discount_flat_pct": 4.0,
        "carbon_eur_t": {2024: 322, 2030: 408, 2040: 604, 2050: 890},
        "carbon_basis": "shadow price, constant 2018 euro, "
                        "target-consistent TIMES-Ireland",
        "sppf_pct": 130.0,       # applies to the EXCHEQUER-funded share
        "spl_pct": (80.0, 100.0),
        "optimism_bias": "no national table - UK Mott MacDonald as a "
                         "flagged proxy",
    },
    "ni": {
        "name": "Better Business Cases NI, on HM Treasury Green Book",
        "discount": "declining STPR: 3.5% to yr 30, 3.0% to 75, "
                    "2.5% thereafter",
        "discount_declining": ((30, 3.5), (75, 3.0), (None, 2.5)),
        "carbon_basis": "DESNZ carbon values",
        "optimism_bias": "Mott MacDonald, 66% upper bound for "
                         "non-standard civil engineering",
        "overlays": "Section 75 equality, Rural Needs Act (NI) 2016, "
                    "Climate Change Act (NI) 2022",
    },
}
# THE CARBON TAX IS A TRANSFER, the shadow price is the resource cost,
# and they are never added. The tax appears only in the private
# bill-saving view, flagged as a transfer, exactly as network charges,
# levies, supplier margin and VAT do.
VFM_TRANSFERS = ("carbon tax", "network charges", "policy levies",
                 "supplier margin", "VAT")

# The one number the island does not publish. Every available figure is
# a European or GB proxy, and it defines one whole end of the
# counterfactual lever - so it is carried as a RANGE and flagged, not
# picked.
VFM_ASHP_ENERGY_CENTRE_EUR_KWTH = {
    "low": 856.25,    # PyPSA technology-data 2030, central air heat pump
    "high": 906.10,   # HIR Hamburg Research Institute, EUR2020
    "gb_bracket_gbp_kw": (500.0, 800.0),   # CIBSE Journal
    "spf_range": (2.8, 3.6),
    "flow_c": (70, 90),
    "note": "NO Irish or Northern Irish outturn is published. This is "
            "the highest-value data gap in the whole appraisal.",
}


# The scenario Panel 6 prices. NOT the fifth the rest of the site
# prices, and the difference is an order of magnitude - so it is named
# differently everywhere it appears.
#
#   Panels 1-4  a fifth of ALL delivered building heat: 5.07 TWh in the
#               Republic, 2.18 in the North. A scale argument.
#   Panel 6     a fifth of the DISTRICT HEATING scenario: 0.54 and 0.23
#               TWh. An appraisal question, answerable against the two
#               rulebooks, inside a commitment already made.
#
# The Republic's 2.7 TWh by 2030 is a government commitment under the
# Climate Action Plan - "up to", and running at roughly 20-32% of
# trajectory, so the outturn case matters as much as the target.
#
# NORTHERN IRELAND HAS NO DISTRICT HEATING TARGET AT ALL. Its 1.16 TWh
# is what it WOULD have at the Republic's proportion of building heat
# (10.7%). That is lent ambition, not policy, and the panel says so as
# the point rather than as a caveat: the only way to give the North a
# number here is to borrow the South's.
# THE SCENARIO IS THE TEN-YEAR BUILD, NOT THE 2030 MILESTONE.
#
# 2.7 TWh by 2030 is the Climate Action Plan commitment and it is
# 10.7% of the Republic's building heat - barely half what Britain has
# committed to. The UK sibling prices 20% of UK heat by 2050, about
# 87 TWh, from roughly 3% today.
#
# The like-for-like scenario is therefore the TEN-YEAR BUILD: 5.0 TWh,
# 19.7% of Irish building heat, which is essentially Britain's 20%.
# Three routes converged on it - continuing the rate the 2030 milestone
# implies, a decadal doubling of the kind Denmark and Sweden achieved,
# and Britain's own ambition scaled to Ireland - all within half a
# terawatt-hour of each other.
#
# 2030 IS THE MILESTONE ON THE WAY, at roughly half the ten-year
# figure, and it is running at 20-32% of trajectory. So the panel
# prices the ten-year build and carries the milestone as the near-term
# marker whose slippage is the argument for urgency.
VFM_DH_SCENARIO = {
    "roi_twh": 5.0,
    "roi_year": 2036,
    "roi_milestone_twh": 2.7,
    "roi_milestone_year": 2030,
    "roi_target_twh": 2.7,
    "roi_target_year": 2030,
    "roi_source": "Climate Action Plan 2025, up to 2.7 TWh by 2030",
    "roi_delivery_pct": (20.0, 32.0),   # projected share of trajectory
    "ni_basis": "lent - the Republic's share of building heat applied "
                "to Northern Ireland, which has set no target",
    # NO FIFTH INSIDE THE SCENARIO. The comparison is the WHOLE network
    # scenario supplied by geothermal against the SAME network supplied
    # by air-source, as the UK sibling does. The scenario is itself a
    # small fraction of building heat - 10.7% in the Republic - so
    # applying a further fifth would price something negligible.
    #
    # An earlier version took 20% of the scenario, giving 540 GWh, and
    # confused the site-wide what-if with the appraisal question. They
    # are different: the fifth elsewhere on the site is a scale
    # argument about all Irish heat; this is an appraisal of how one
    # committed programme is supplied.
    "network_load_hours": 4000,
}


# TWO STAGES, NEVER SUMMED.
#
#   Stage 1  BAU -> air-source network. The ELECTRIFICATION decision.
#            Carbon and fuel benefits, network and connection capital,
#            and a capacity COST, because electrifying heat adds winter
#            peak.
#   Stage 2  air-source network -> geothermal. THE SUBSURFACE
#            INCREMENT. Distribution and connections are common to both
#            and cancel. What remains is the SPF gain, a capacity
#            BENEFIT at SEM prices, the cooling increment, and the
#            subsurface capital and its risk.
#
# Reported separately because summing them would let geothermal bank
# the carbon saving of moving off oil and gas - most of which ANY heat
# pump delivers. The split leaves the subsurface to be judged on what
# it actually adds.
#
# It also settles the capacity sign structurally rather than by rule: a
# cost in stage 1, a benefit in stage 2, which is where the arithmetic
# puts it without anyone having to remember.
VFM_STAGES = ("electrify", "subsurface")

# The two jurisdictions differ by BLEND, not by kind. Both run open
# loop and standing column wells in ATES and BTES modes, both assume
# some waste heat recovery and storage, and the North adds a deeper
# Permo-Triassic HSA component on top. Source temperatures follow:
# 16.0 C in the Republic, 19.6 C once the deep fraction is blended in.
# Both figures are already on the site in NETWORK_MODEL and are read
# from there.
#
# TWO DIFFERENT RISKS, WHICH AN EARLIER VERSION CONFLATED:
#
#   SUCCESS RATE is pre-FID. Money spent on prospects that never reach
#   close, spread over the ones that do. A shallow scheme that tests
#   poorly gets REDESIGNED - more wells, different spacing, a bigger
#   store - so the spend converts into capex rather than being written
#   off. A deep doublet can die at the drill bit: if the reservoir is
#   not there at 2 km, no redesign recovers it. So the shallow fraction
#   runs at full success and the deep fraction does not.
#
#   SHORTFALL is post-commissioning. A built scheme underdelivering
#   against design - an aquifer yielding less than tested, thermal
#   breakthrough between wells, a store losing more than modelled, or,
#   specific to these blends, WASTE HEAT THAT DOES NOT ARRIVE in the
#   volume assumed. That last is not geological at all. BOTH
#   jurisdictions carry it; the North's range is wider and its default
#   higher because the deep fraction is the harder part to predict.
#
# The blend weight is the UK sibling's 40% deep. No reason to vary it
# yet, and if it moves it moves in both jurisdictions together.
VFM_DEEP_WEIGHT = 0.40
VFM_SUBSURFACE_CLASSES = {
    "shallow": {"devex_pct": 0.04, "success": 1.00,
                "note": "open loop, standing column wells, ATES and "
                        "BTES with waste heat recovery and storage - "
                        "pre-FID spend reduces design uncertainty "
                        "rather than mitigating a project-killing risk"},
    "deep": {"devex_pct": 0.10, "success": 0.80,
             "note": "Permo-Triassic hot sedimentary aquifer doublet - "
                     "found, not installed, and can be written off"},
}
VFM_JUR_MODEL = {
    "roi": {
        "class": "open loop and standing column wells, ATES and BTES, "
                 "with waste heat recovery and storage",
        "deep_weight": 0.0,
        "shortfall_default": 0.05,
        "shortfall_range": (0.0, 0.30),
    },
    "ni": {
        "class": "the same, with a deeper Permo-Triassic HSA "
                 "component blended in",
        "deep_weight": VFM_DEEP_WEIGHT,
        "shortfall_default": 0.10,
        "shortfall_range": (0.0, 0.45),
    },
}


# THE SUBSURFACE INCREMENT, the lever the whole panel turns on.
#
# Energy centre and subsurface over the air-source alternative, per kW
# of installed capacity at the reference sizing. DISTRIBUTION AND
# CONNECTIONS ARE COMMON TO BOTH ROUTES AND CANCEL - same mains, same
# trenching, same heat interface units - which is why this reduces to
# one number rather than two full cost stacks.
#
# The UK sibling carries GBP 14/kW on a 6-26 range. Ireland differs in
# both directions and the two effects partly offset:
#
#   UPWARD   no legacy oil and gas subsurface data of the kind that
#            de-risks British prospects, and a far smaller supply chain
#            - 20,128 systems in the Republic and a handful above 45 kW
#            in the North, against 55,210 in Britain. Mobilisation is a
#            larger share of a small programme.
#   DOWNWARD the Republic's blend is entirely shallow, where the UK's
#            is 40% intermediate doublet. Shallow plant is cheaper per
#            kW than a 2 km doublet.
#
# So the Republic sits below the UK's central figure and the North at
# roughly it, since the North IS the UK's blend on the same play.
# Currency: euro throughout, converted at the ECB rate the site already
# fetches, so the two panels can be compared.
# UNIT CAPITAL, ALL ON ONE BASIS: per kW of HEATING capacity, plant
# boundary (energy centre inclusive, distribution exclusive), EUR.
# Getting these onto a common basis is the whole difficulty - a bare
# ground loop set against a complete air-source energy centre produces
# a spuriously negative increment, which an earlier version of this
# model did produce.
#
# SHALLOW. Herrmann et al. 2026, RSER 226:116202, 133 ATES systems.
# Headline is 300 EUR/kW above 2 MW, but that is PER KW OF COMBINED
# HEATING AND COOLING - the paper adds an equivalent cooling capacity
# where only heating was stated - so on a heating basis a balanced
# system is about DOUBLE, near 600 EUR/kW. Scope is a whole ground
# system including heat pumps and surface plant, running from the
# wellhead to the building or heating station, so distribution is
# outside it. EUR2022.
#
# BUT THE DATASET IS 77% DUTCH, and the Netherlands has unconsolidated
# sand aquifers, a mature drilling supply chain and a national resource
# risk fund. The paper's own Danish systems run near 1,000 EUR/kW,
# which it attributes to different geology and an emerging market.
# Ireland is emerging and is not the Netherlands, so the Dutch figure
# is a best-case FLOOR and the Danish is the better central analogue.
# EVERY FIGURE ON A HEATING BASIS. Herrmann quotes all of them per kW
# of COMBINED heating and cooling, so each doubles for a balanced
# system. An earlier version corrected the Dutch plateau and then took
# the Danish figure UNCORRECTED, which halved the shallow central,
# collapsed the increment from about 1,000 to 119 EUR/kW and produced
# a 1.4-year payback. The implausible answer is what caught it.
VFM_SHALLOW_EUR_KW = {
    "dutch_floor": 600.0,      # 300 combined -> 600 heating
    "central": 1878.0,         # Danish 939 combined -> 1,878 heating
    "high": 2400.0,
    "source": "Herrmann et al. 2026, RSER 226:116202, 133 ATES "
              "systems; 300 EUR/kW combined heating and cooling above "
              "2 MW, restated to a heating basis; Danish subset near "
              "1,000 EUR/kW taken as the Irish central",
}
# DEEP. Todd et al., Geoenergy, Table 5, seven Permo-Triassic doublet
# cases from 10 to 1,800 m. GBP 1,550-2,179 per kW installed thermal,
# itemised into borehole heat exchanger, heat pump, ancillaries and
# development - a FULL installed cost including surface plant, so it is
# already on the same basis as the shallow and air-source figures.
# Notably FLAT with depth: the deepest case costs about the same per kW
# as the shallowest, because COP rises from 2.76 to 14.5.
#
# FIRST OF A KIND. A 500-1,250 MW programme is not first-of-a-kind by
# its end. Published learning rates for geothermal district heating run
# 5% per capacity doubling (conventional) to 10% (drilling-dominated),
# which over roughly seven doublings gives a 30-50% reduction. Carried
# at 30% as the conservative end.
VFM_DEEP_GBP_KW = {"low": 1550.0, "high": 2179.0, "mid": 1865.0,
                   "source": "Todd et al., Geoenergy, Table 5, seven "
                             "Permo-Triassic doublet cases"}
# LEARNING APPLIES TO THE DEEP FIGURE ONLY, and the asymmetry is
# deliberate but worth stating: Todd et al. is explicitly
# first-of-a-kind, while Herrmann's Danish systems are installed
# outturn from a market that already exists. So the deep figure is
# discounted and the shallow one is not.
#
# The consequence is that deep at nth-of-a-kind (EUR 1,527/kW) comes in
# BELOW shallow on the Danish basis (EUR 1,878/kW), which makes the
# North's blend cheaper per kW than the Republic's pure-shallow one -
# and, with SPF 5.0 against 4.0, gives it the shorter payback. That is
# a real consequence of the inputs rather than an error, but it turns
# on a learning rate applied to one side and not the other, so it is
# the first thing to test if the result is challenged.
VFM_FOAK_LEARNING = {"reduction": 0.30, "range": (0.25, 0.40),
                     "applies_to": "shallow and deep",
                     "note": "5-10% per capacity doubling over ~7 "
                             "doublings to a 500-1,250 MW programme. "
                             "Applied to BOTH classes from 22 Aug 2026, "
                             "each from its own Irish FOAK anchor, "
                             "ramped along the build years rather than "
                             "pricing the whole build at nth"}
# IRISH SHALLOW FOAK ANCHOR, set 22 Aug 2026. EUR 2,000/kW ALL-IN -
# the Causeway Hospital feasibility's SCW case (GBP 1,698/kW at
# 2.2 MW, EUR 1,950 at the ECB 2025-S2 semester mean), rounded.
# Boundary matches the increment: subsurface + ancillaries + heat pump
# package, 15% contingency inside, devex separate, no distribution.
# Stated AT ITS REFERENCE SIZING - the same wells serve 1.3 MW at
# EUR 3,300/kW, so this is not a scale-free constant. Developer
# estimate, daggered.
VFM_SHALLOW_FOAK_EUR_KW = 2000.0
# LEARNING APPLIES TO THE SUBSURFACE SHARE ONLY (22 Aug 2026). The
# heat pump package inside every all-in figure is the same mature
# technology as the flat air-source counterfactual, so it does not
# learn - discounting it would be learning on the component the
# comparison holds constant. Shares derived from the same workbook's
# own arithmetic (each component carries its 15% contingency):
#   shallow: (1,956,250 + 391,250) x 1.15 / 3,734,625 = 0.723
#   deep:    (1,670,500 + 167,050) x 1.15 / 3,421,938 = 0.618
# The deep share is a 1 km two-borehole scheme's; a 2 km HSA doublet's
# drilling share is plausibly higher, which would mean MORE learning -
# so 0.618 is the conservative side. Todd et al.'s author can set it.
VFM_SUBSURFACE_SHARE = {"shallow": 0.723, "deep": 0.618,
                        "source": "Causeway Hospital feasibility, SCW "
                                  "and Deep GHP capital stacks"}
# Effective programme reduction is therefore 30% x share: about 21.7%
# shallow and 18.5% deep. CONSEQUENCE DISCLOSED: shallow at nth is
# 2,000 x (1 - 0.30 x 0.723) = EUR 1,566 - still BELOW Herrmann's
# installed Danish outturn of 1,878, and the ramp mean sits below the
# flat Danish figure the Republic was priced at before, so this change
# RAISES the ROI BCR. VFM_SHALLOW_NTH_FLOOR stops learning at Danish
# outturn instead.
VFM_SHALLOW_NTH_FLOOR = None   # DECIDED 22 Aug 2026: stays None. The
                               # Danish comparator is not accepted as a
                               # bound on an Irish programme; Herrmann
                               # remains in the payload as a
                               # cross-reference only
# WASTE-HEAT DISPLACEMENT (22 Aug 2026). A fraction of the programme
# couples to waste heat - data centres, energy-from-waste, power
# stations - and DRILLS LESS: subsurface capital is replaced by heat
# exchangers and a connection to the source. CAPEX AVOIDANCE ONLY.
# The operating upside of warm sources is already inside the class
# SPFs and blended source temperatures, and the risk of waste heat not
# arriving is already named in the shortfall lever's definition - so
# counting anything but capital here would double-dip.
# Applies to the SHALLOW class only (waste-heat coupling is the
# 5G/ambient-loop move; the North's deep fraction is untouched) and to
# the subsurface share only, BEFORE learning - the remaining
# subsurface still learns, the connection kit, like the heat pump,
# stays flat. Defaults are Causeway judgement pending a source
# inventory, daggered and ranged.
# CONSTRAINED WIND (NI ONLY, 23 Aug 2026). Wind wasted in Derry and
# Antrim because it cannot LEAVE - local network constraint, not
# system curtailment - overlaps the Permo-Trias play fairway. A
# storage-capable geothermal fleet sited there charges in constrained
# hours: for that coincident fraction, the fleet's electricity falls
# from LRVC toward the near-zero cost of wind otherwise spilled.
# TRANSFERS EXCLUDED, per the panel's convention: dispatch-down
# PAYMENTS avoided are a consumer-to-generator transfer and are NOT
# the benefit - the resource saving is, and the payments falling is a
# consequence for the asks, not a priced stream. NO DOUBLE-DIP: the
# SPF gain prices using LESS electricity; this prices the remaining
# electricity COSTING less. Delivery risk (controllable charging
# actually happening) sits inside the shortfall lever like waste heat.
# CURTAILMENT stays excluded - the morning's anti-correlation finding
# stands; only the local-constraint component rides here.
# Heat per Irish home per year, MWh. SEAI Energy in Ireland puts
# residential heat demand at 22.3 TWh (2024) across roughly 1.85m
# occupied dwellings = 12.1 MWh; SEAI's residential statistics give
# 17.15 MWh of energy per home (2022) of which 74% is non-electric,
# about 12.7 MWh. Two routes, same answer to the nearest MWh, so the
# figure is published rather than judged - the dagger on the panel
# marks the CONVERSION (rejected heat to homes served), not this.
HOME_HEAT_MWH = 12.0

VFM_CONSTRAINED_WIND = {
    # CONSTRAINT ENERGY IS READ FROM OUR OWN SERIES, not asserted -
    # docs/dispatch_down_monthly.json, EirGrid's half-hourly wind
    # files, TRANS_CONSTR_MWH only (the transmission-constraint reason
    # code), NI, for the stated basis year below. This replaced a
    # derived 0.43 TWh (2.9 TWh x 22.0% x ~0.68) that UNDERSTATED it:
    # the actual NI constraint share of dispatch-down is 83-89%, not
    # 68%, because loss-of-tie-line dispatch-down is classed as NI
    # constraint and is the dominant component.
    #
    # A STATED YEAR, not "latest" or a trailing window: dispatch-down
    # is trending hard (409 GWh in 2023, 817 in 2024, 563 in 2025) and
    # a moving basis on an advocacy-adjacent stream is cherry-pickable.
    # 2025 is the most recent COMPLETE year and matches the 22.0%
    # already cited on the grid panel. Moving this to 2026 is a
    # decision, and it edits the test that pins it.
    "constraint_basis_year": "2025",
    # siting share x hour overlap x storage dispatchability. ONE LOW
    # CONSERVATIVE NUMBER, accepted 25 Aug 2026 rather than decomposed:
    # the hour-overlap term is empirical from the same series (half the
    # spill falls outside the heating season, 44% between midnight and
    # six), the dispatchability term is a design-basis judgement. Still
    # daggered, still hard-wired - the panel holds at four sliders.
    "coincidence": 0.15, "coincidence_range": (0.0, 0.40),
    # constraint is not permanent: the North-West reinforcement
    # pipeline erodes it. Linear decay to zero over this horizon -
    # judgement, disclosed - so the stream does NOT run flat for
    # sixty years.
    "erosion_years": 25,
    "applies": "ni only - the Republic's constraint geography does "
               "not overlap its shallow class the same way",
}
VFM_WASTE_HEAT = {
    "share": {"roi": 0.15, "ni": 0.10},
    "share_range": (0.0, 0.40),
    "connection_relcost": 0.30,
    "connection_relcost_range": (0.15, 0.60),
    "note": "share of shallow-class schemes coupled to waste heat; "
            "connection_relcost is the interconnection's cost as a "
            "fraction of the subsurface it displaces. Causeway "
            "judgement, ACCEPTED 25 Aug 2026 as deliberately small "
            "and conservative rather than derived from a source "
            "inventory - the shares understate the coupling "
            "opportunity and are meant to",
}
# AIR-SOURCE COUNTERFACTUAL. Danish Energy Agency technology catalogue
# via PyPSA, and HIR Hamburg. The DEA boundary is the complete energy
# centre - heat pump, air evaporators, civil works, buildings, grid
# connection, commissioning - delivered to the nearest district heating
# network, so distribution is outside it and VAT is excluded. Same
# basis as both geothermal figures.
VFM_AIRSOURCE_EUR_KW = {"low": 856.25, "high": 906.10, "mid": 881.18,
                        "source": "PyPSA technology-data 2030 central "
                                  "air heat pump (Danish Energy Agency "
                                  "catalogue) and HIR Hamburg"}
GBP_EUR = 1.17

# CARBON: TWO PRICES THAT MUST NEVER BE ADDED.
#
# The site already holds CARBON_STEPS - the Finance Act 2020 carbon TAX
# trajectory, EUR 26 to 100 a tonne. That is a TRANSFER: money moving
# from a household to the exchequer. It belongs in the bill panels,
# where it does, and in Panel 6's private view flagged as a transfer.
# It is NOT a resource benefit.
#
# The RESOURCE cost of a tonne emitted is the shadow price, and the two
# jurisdictions publish different ones on different bases. Adding tax
# to shadow price would double-count: the tax is a policy instrument
# that partially internalises the very damage the shadow price
# measures.
#
# ROI: Infrastructure Guidelines (DPENDR, Dec 2023), constant 2018
# euro, target-consistent on a TIMES-Ireland basis. NI: DESNZ carbon
# values, which are on a different basis again - a traded/non-traded
# convergence rather than a target-consistent shadow price - so the two
# columns are NOT directly comparable and the panel says so.
VFM_SHADOW_CARBON = {
    "roi": {2024: 322.0, 2030: 408.0, 2040: 604.0, 2050: 890.0,
            "unit": "EUR/tCO2e, constant 2018 euro",
            "source": "Infrastructure Guidelines, DPENDR, Dec 2023, "
                      "target-consistent TIMES-Ireland basis"},
    "ni": {2024: 277.0, 2050: 398.0,
           "unit": "GBP/tCO2e",
           "source": "DESNZ carbon values, as the UK sibling uses"},
}


def vfm_shadow_carbon(jur, year):
    """Shadow price of carbon, interpolated. NOT the carbon tax."""
    t = VFM_SHADOW_CARBON[jur]
    yrs = sorted(k for k in t if isinstance(k, int))
    if year <= yrs[0]:
        return t[yrs[0]]
    if year >= yrs[-1]:
        return t[yrs[-1]]
    for i in range(1, len(yrs)):
        if year <= yrs[i]:
            a, b = yrs[i - 1], yrs[i]
            f = (year - a) / (b - a)
            return t[a] + f * (t[b] - t[a])
    return t[yrs[-1]]


# THE OPERATORS' OWN PLANNING BASIS, and the gap in it.
#
# EirGrid and SONI's Tomorrow's Energy Scenarios 2023 databook, IE
# Demand sheet, Table 6.4, gives TWO NUMBERS: air source 2.6, ground
# source 2.94. That is the entire heat-pump treatment. Searching all
# twenty-four sheets for "seasonal", "SCOP", "weather", "temperature",
# "degree", "peak heat", "coincidence" or "diversity" returns NOTHING.
#
# So a flat COP is applied as a divisor from heat demand to electricity
# demand, in four scenarios, out to 2050 - with no stated source
# temperature, flow temperature or season, and no treatment of what
# happens on the coldest evening. That same demand forecast feeds the
# capacity requirement the SEM auction procures against.
#
# THE OMISSION CUTS BOTH WAYS, which is why it is worth stating rather
# than merely disagreeing with. A flat 2.6 UNDERSTATES air-source peak
# draw, because real units do worse when it is cold and damp - our own
# model puts an Irish air-source unit at 2.32 at the tightest hour, 12%
# more peak demand than a flat 2.6 implies. And a flat 2.94
# UNDERSTATES the ground-source advantage, because ground source does
# not degrade at all. The planning basis therefore underestimates both
# the problem and the solution.
#
# OUR FIGURES ARE NOT THE SAME MACHINES. 2.94 is fair for a domestic
# ground-source unit. NETWORK_MODEL's 4.0 and 5.0 are network scale -
# low flow temperatures, a stable source, and in the North a warmer
# Permo-Triassic source. The panel says so explicitly rather than
# leaving a reviewer to assume we have simply inflated TES.
VFM_TES_COP = {
    "air_source": 2.6, "ground_source": 2.94,
    "table": "Tomorrow's Energy Scenarios 2023 databook, IE Demand, "
             "Table 6.4",
    "seasonal_treatment": None,
    "peak_treatment": None,
    "searched": ("seasonal", "SCOP", "weather", "temperature", "degree",
                 "peak heat", "coincidence", "diversity"),
    "note": "two numbers, no stated conditions, applied flat to 2050",
}
# TES's own carbon path, from the same databook - Tables 7.11 and 7.12,
# megatonnes, against annual demand from Figure 6.1. Ireland's average
# intensity in Self-Sustaining is +14 g/kWh in 2035 and NEGATIVE
# thereafter, because the scenarios reach a carbon-negative power
# system. That all but extinguishes the carbon stream in this panel
# within a decade, and it is the operators' own forecast.
VFM_TES_CARBON = {
    "ie_mt": {"self_sustaining": {2035: 0.98, 2040: -0.70,
                                  2045: -0.74, 2050: -1.24},
              "gas_evolution": {2035: 1.13, 2040: 0.50,
                                2045: 0.0, 2050: 0.0},
              "constrained_growth": {2035: 2.49, 2040: 1.63,
                                     2045: 0.97, 2050: -0.40}},
    "ni_mt": {"self_sustaining": {2035: 0.15, 2040: -0.33,
                                  2045: -0.39, 2050: -0.59}},
    "ie_avg_g_kwh_ss": {2035: 14.1, 2040: -8.6, 2045: -8.9, 2050: -14.7},
    "source": "TES 2023 databook, Tables 7.11 and 7.12 with Figure 6.1",
    "caveat": "AVERAGE intensity, not marginal. A saved kilowatt-hour "
              "displaces the marginal generator, which may still be gas "
              "when the annual total is negative - and the negative "
              "totals imply removals whose accounting treatment "
              "matters. Marginal is the right basis and TES does not "
              "publish it.",
}


# RUNNING COST: TWO PRICES, AND THE GAP BETWEEN THEM IS TRANSFERS.
#
# The RESOURCE saving is the long-run variable cost of the electricity
# not generated - fuel, carbon and variable operating cost on the
# marginal plant. That is what the Green Book and the Infrastructure
# Guidelines both want, and it is what society actually saves.
#
# The BILL saving is the retail price, which the site already derives
# per jurisdiction and sector. The difference between the two is
# network charges, policy levies, supplier margin and VAT - all
# TRANSFERS under both rulebooks. The UK sibling found roughly half its
# private return was exactly these, so the resource figure is
# materially smaller than the bill and the panel reports both,
# separately labelled, never summed.
#
# LRVC is a dagger. Neither jurisdiction publishes a long-run variable
# cost series for electricity, so it is anchored on the wholesale
# element of the retail price - and the SEM's own energy market is the
# closest observable, though it contains scarcity rent that a long-run
# cost does not.
# Long-run variable cost, p/kWh, on the UK sibling's own anchor years.
# NORTHERN IRELAND USES THESE DIRECTLY - it is the same market for
# appraisal purposes and the Green Book applies. THE REPUBLIC IS
# PINNED TO THE EURO EQUIVALENT for now, which is a placeholder with a
# reason rather than a source: the Infrastructure Guidelines name no
# national energy price series at all, so there is nothing Irish to
# use, and a neighbouring market's published figure beats a number
# invented here.
#
# IT REPLACES EUR 95, WHICH I ASSERTED. That figure had no derivation
# - the note claimed it was anchored on retail and SEM prices but it
# was not computed from either. It sat below the Irish wholesale
# average of EUR 113.83 (2025, kilowatt.ie, the series the
# dispatch-down work already uses) and a third below the UK's own
# figure, with nothing supporting the position. Running cost is the
# largest DURABLE stream in this panel - carbon collapses within a
# decade - so the whole answer scales linearly with it.
#
# STILL TO DO: build it from the marginal plant. Fuel plus carbon plus
# variable O&M on Irish gas generation is computable from the gas
# price this site already fetches, a CCGT efficiency and a carbon
# price. That would make it derived and testable rather than borrowed.
VFM_LRVC_YEARS = (2026, 2030, 2035, 2040, 2045, 2050, 2055, 2060)
VFM_LRVC_P_KWH = (14.306, 12.530, 12.763, 12.695, 12.565,
                  12.273, 12.273, 12.273)
VFM_LRVC_SOURCE = ("UK Heat Split sibling, long-run variable cost "
                   "series on its own anchor years; NI direct, ROI "
                   "pinned to the euro equivalent pending an Irish "
                   "derivation")


def vfm_lrvc(year, jur):
    """LRVC interpolated. GBP/MWh for NI, EUR/MWh for ROI."""
    ys, vs = VFM_LRVC_YEARS, VFM_LRVC_P_KWH
    if year <= ys[0]:
        v = vs[0]
    elif year >= ys[-1]:
        v = vs[-1]
    else:
        v = vs[-1]
        for i in range(1, len(ys)):
            if year <= ys[i]:
                f = (year - ys[i - 1]) / (ys[i] - ys[i - 1])
                v = vs[i - 1] + f * (vs[i] - vs[i - 1])
                break
    gbp_mwh = v * 10.0
    return gbp_mwh * (GBP_EUR if jur == "roi" else 1.0)


VFM_LRVC_EUR_MWH = {"central": round(vfm_lrvc(2030, "roi"), 1),
                    "low": round(vfm_lrvc(2050, "roi"), 1),
                    "high": round(vfm_lrvc(2026, "roi"), 1),
                    "note": VFM_LRVC_SOURCE}


# ENERGY PRICE FORECASTS. A sixty-year appraisal cannot run on today's
# prices, and each rulebook names its own source:
#
#   NI   DESNZ energy price projections, required by the Green Book.
#        Published as low/central/high paths to 2050 in real terms.
#   ROI  the Infrastructure Guidelines name NO national series. SEAI's
#        National Energy Projections are the fallback the evidence base
#        recommends, and they are a modelling output rather than an
#        appraisal price set - so the Republic's price path is a
#        weaker input than the North's, which is the opposite of the
#        usual asymmetry on this site.
#
# TODAY'S PRICES COME FROM THE COST PANEL, not from a second
# derivation. derive_heat_cost_series already computes the daily cost
# of a useful kWh by route, per jurisdiction, on live feeds - so
# Panel 6 reads that and the two panels cannot disagree about what
# heat costs. What Panel 6 adds is the forward path and the transfer
# split.
VFM_PRICE_FORECAST = {
    "ni": {"source": "DESNZ energy price projections",
           "status": "published appraisal price set, Green Book "
                     "requires it",
           "paths": ("low", "central", "high")},
    "roi": {"source": "SEAI National Energy Projections",
            "status": "FALLBACK - the Infrastructure Guidelines name "
                      "no national energy price series, so this is a "
                      "modelling output pressed into appraisal use",
            "paths": ("central",)},
    "today_from": "derive_heat_cost_series - the cost panel's own "
                  "daily cost of a useful kWh by route",
}
# THE TWO STAGES NEED DIFFERENT PRICES, which is why the stage split
# matters for more than presentation:
#
#   Stage 1  BAU -> air-source network. Needs a FUEL price (gas in the
#            Republic's networked areas, oil across much of the North)
#            AND an electricity price. The saving is fuel avoided less
#            electricity drawn, so it turns on the RATIO between two
#            forecasts - and that ratio is the single most uncertain
#            thing in the electrification case.
#   Stage 2  air-source -> geothermal. BOTH ROUTES ARE ELECTRIC, so
#            the price level largely cancels and only the SPF gap
#            matters. That makes the subsurface stage far more robust
#            to price forecasting than the electrification stage.
#
# Worth stating on the panel: the stage we are advocating is the one
# least exposed to the forecast nobody can make.
VFM_STAGE_PRICES = {
    "electrify": ("fuel", "electricity"),
    "subsurface": ("electricity",),
}


# CAPITAL PHASING AND OPTIMISM BIAS, both ported from the UK sibling
# because neither jurisdiction gives us a better basis.
#
# PHASING. Capital spread evenly over a ten-year build and discounted
# at mid-year; benefits ramp linearly with the fleet, so year five of a
# ten-year build delivers half. Nothing in the Irish rulebooks
# specifies a build period, so the sibling's ten years is the anchor -
# and it is arguably generous to Northern Ireland, which has 4.5 MWth
# installed and no supply chain to ramp.
#
# OPTIMISM BIAS. THE REPUBLIC PUBLISHES NO TABLE. The Infrastructure
# Guidelines treat it as a process requirement, not a number: "full
# risk assessment and consideration of REMAINING optimism bias" at
# Final Business Case, the word "remaining" implying a quantitative
# risk assessment has already absorbed most of it. The centrally
# specified parameters - shadow carbon, the shadow price of labour at
# 80-100% - live in a separate document and optimism bias is not among
# them. So the Green Book's Mott MacDonald bands are used in BOTH
# jurisdictions: they apply directly in the North, and in the Republic
# as a flagged proxy.
#
# AND NEITHER RULEBOOK ADJUSTS BENEFITS. The UK's own Department for
# Transport work covers cost, time AND benefits, but no Irish
# requirement for a downward benefit adjustment was found. That is
# recorded as unresolved rather than assumed absent - it may sit in
# sectoral guidance not reached.
VFM_BUILD_YEARS = 10
VFM_HORIZON_YEARS = 60
VFM_OPTIMISM = {
    "default_pct": 50.0, "min_pct": 0.0, "max_pct": 66.0,
    "applies_to": "capital only",
    "benefits_adjustment": None,
    "source": "HM Treasury Green Book, Mott MacDonald bands - 66% is "
              "the upper bound for non-standard civil engineering, "
              "mitigating toward 40-50% at outline business case",
    "roi_basis": "PROXY - the Infrastructure Guidelines publish no "
                 "optimism bias table, only a process requirement to "
                 "consider remaining bias after risk assessment",
    "benefits_note": "no Irish requirement for a downward benefit "
                     "adjustment was found; recorded as unresolved",
}
# Discount: the Republic flat, the North declining. Both from their own
# rulebooks, and this is one place the two genuinely differ.
VFM_DISCOUNT = {
    "roi": {"kind": "flat", "rate": 0.04},
    "ni": {"kind": "declining",
           "bands": ((30, 0.035), (75, 0.030), (None, 0.025))},
}


def vfm_discount_factor(t, jur):
    """Discount factor at year t. Flat 4% in the Republic; the Green
    Book's declining STPR in the North."""
    d = VFM_DISCOUNT[jur]
    if d["kind"] == "flat":
        return 1.0 / ((1.0 + d["rate"]) ** t)
    f, k = 1.0, 1
    while k <= int(t):
        r = 0.035 if k <= 30 else (0.030 if k <= 75 else 0.025)
        f /= (1.0 + r)
        k += 1
    fr = t - int(t)
    if fr:
        f /= (1.035 ** fr)
    return f


# TES Self-Sustaining, Ireland, average intensity in gCO2/kWh. The
# 2026-30 values interpolate from peer-reviewed modelling to TES's own
# 2035 figure. NEGATIVE FROM 2040 on 300 MW of unbuilt BECCS.
VFM_GRID_INTENSITY = ((2026, 320.0), (2030, 290.0), (2035, 14.1),
                      (2040, -8.6), (2045, -8.9), (2050, -14.7),
                      (2060, -14.7))
VFM_BECCS_MW = 300
VFM_BECCS_NOTE = ("the sign flip after 2040 rests on 300 MW of "
                  "bioenergy with carbon capture, against zero "
                  "biomass-only capacity in every scenario - no "
                  "operating plant on this island, no licensed CO2 "
                  "storage, no transport infrastructure")


def vfm_grid_intensity(year):
    """Average grid intensity, gCO2/kWh, on the TES path."""
    p = VFM_GRID_INTENSITY
    if year <= p[0][0]:
        return p[0][1]
    if year >= p[-1][0]:
        return p[-1][1]
    for i in range(1, len(p)):
        if year <= p[i][0]:
            a, b = p[i - 1], p[i]
            f = (year - a[0]) / (b[0] - a[0])
            return a[1] + f * (b[1] - a[1])
    return p[-1][1]


# COOLING IMPROVES BOTH SIDES AT ONCE - more benefit AND less
# capital - which is the UK sibling's framing and it is right.
#
# An earlier version here counted only the capital half, on the
# argument that counting both would double-count one physical fact.
# THAT WAS WRONG. They are two facts: not BUYING a chiller and not
# RUNNING it are separate savings and both are real. The operating
# half is the larger of the two, because capital is once and the
# electricity is every summer for sixty years.
#
# CAPITAL: avoided chillers, netted off the cost.
#
# A ground loop that already rejects heat can absorb it too, so the
# chillers those buildings would otherwise need are never bought. That
# is capital AVOIDED, so it enters as a negative increment on the cost
# side rather than as a fourth benefit - which keeps the benefit bar to
# three streams and stops the same physical fact being counted twice.
#
# THE IRISH CASE IS MUCH SMALLER THAN THE BRITISH ONE, for two
# structural reasons and neither is about ambition:
#
#   A HEAT NETWORK MOSTLY SERVES HOMES, AND IRISH HOMES DO NOT COOL.
#   72% of the Republic's building heat is residential. Only about
#   1.38 TWh of the 5.0 TWh scenario lands in services buildings -
#   the offices, hotels and hospitals that have cooling at all.
#
#   AND IRISH COMFORT COOLING IS SMALL TO BEGIN WITH. Panel 4 puts
#   Tier 1 at 1.25 TWh against 13.95 of process; with the mixed tier
#   it is 1.94 TWh. At the network's share that is about 0.38 TWh of
#   cooling service that could ride the same loops.
#
# SO THE DEFAULT IS 12%, NOT THE UK'S 30%. Three things cut it: the
# residential share above, the free-cooling finding from Panel 4 (an
# Irish office already meets much of its cooling with fresh air, so
# the chiller displaced is smaller than a British equivalent), and the
# fact that only a 5G ambient loop can carry cooling at all - which is
# the Republic's blend but only part of the North's.
# OPERATING: the same cooling delivered at a far better efficiency.
# An air-cooled chiller runs near EER 3; rejecting to ground or to a
# store is circulation work, at an effective ratio well into the
# teens. The gap is the saving, valued at the same long-run variable
# cost as the heat side.
VFM_COOLING = {
    "connections_pct": 12.0, "min_pct": 0.0, "max_pct": 60.0,
    "chiller_eur_kw": 620.0,
    "chiller_range": (400.0, 900.0),
    "chiller_eer": 3.0,
    "ground_cooling_eer": 15.0,
    "note": "avoided air-cooled chiller capital where the loop serves "
            "both. Set the lever to zero and the heat case stands on "
            "its own.",
    "roi_only_caveat": "only an ambient loop can carry cooling at "
                       "network flow temperatures; that is the "
                       "Republic's blend but only the shallow part of "
                       "the North's, so the North's share is capped at "
                       "its non-deep fraction",
}


def derive_vfm_cooling(anchors=None):
    """
    Avoided chiller capital, as a NEGATIVE increment.

    Sized from the cooling that could physically ride the same loops -
    Panel 4's comfort and mixed tiers at the network's share of
    building heat - times the connection lever, times a chiller capital
    rate.
    """
    a = anchors or ANCHORS
    sc = derive_vfm_scenario(a)
    ct = derive_cooling_tiers()
    tt = ct["tier_totals_twh"]
    comfort_twh = tt["1"] + tt["m"]          # Tier 1 plus mixed
    roi_del = sc["jur"]["roi"]["delivered_twh"]
    pct = VFM_COOLING["connections_pct"] / 100.0
    out = {"lever": dict(VFM_COOLING),
           "comfort_cooling_twh": round(comfort_twh, 2), "jur": {}}
    for k, v in sc["jur"].items():
        share = v["network_twh"] / v["delivered_twh"]
        # the Republic's comfort cooling scaled to each jurisdiction by
        # its heat, since neither publishes a cooling split of its own
        avail = comfort_twh * (v["delivered_twh"] / roi_del) * share
        # only the non-deep fraction of the North's blend can carry it
        deep_w = VFM_JUR_MODEL[k]["deep_weight"]
        carriable = avail * (1.0 - deep_w)
        served = carriable * pct
        mw = served * 1e6 / VFM_DH_SCENARIO["network_load_hours"]
        avoided = mw * VFM_COOLING["chiller_eur_kw"] / 1000.0
        # and the electricity those chillers never draw
        elec_chiller = served / VFM_COOLING["chiller_eer"]
        elec_ground = served / VFM_COOLING["ground_cooling_eer"]
        elec_saved = elec_chiller - elec_ground
        out["jur"][k] = {
            "elec_saved_twh": round(elec_saved, 4),
            "available_twh": round(avail, 3),
            "carriable_twh": round(carriable, 3),
            "served_twh": round(served, 3),
            "chiller_mw_avoided": round(mw, 1),
            "avoided_capital_eur_m": round(avoided, 1),
        }
    log(f"vfm cooling: {VFM_COOLING['connections_pct']:.0f}% of "
        f"connections take cooling too (UK sibling 30%, lower here "
        f"because a network mostly serves homes and Irish homes do not "
        f"cool); "
        + "; ".join(f"{k.upper()} avoids {v['chiller_mw_avoided']:.0f} MW "
                    f"of chillers = EUR {v['avoided_capital_eur_m']:.0f}m"
                    for k, v in out["jur"].items()))
    return out


def derive_vfm_phased(anchors=None):
    """
    The appraisal, phased and discounted - capital over a ten-year
    build, benefits ramping with the fleet, over sixty years.

    Optimism bias applies to CAPITAL ONLY, because neither rulebook
    specifies a benefit adjustment and inventing one would be our
    thumb on the scale in the other direction.
    """
    a = anchors or ANCHORS
    sc = derive_vfm_scenario(a)
    inc = derive_vfm_increment(a)
    st = derive_vfm_stages(a)
    run = derive_vfm_running(a)
    carb = derive_vfm_carbon(a)
    cool = derive_vfm_cooling(a)
    ob = VFM_OPTIMISM["default_pct"] / 100.0
    out = {"build_years": VFM_BUILD_YEARS,
           "horizon_years": VFM_HORIZON_YEARS,
           "optimism": dict(VFM_OPTIMISM),
           "discount": {k: dict(v) for k, v in VFM_DISCOUNT.items()},
           "jur": {}}
    for k in ("roi", "ni"):
        mw = sc["jur"][k]["plant_mw"]
        per_kw = inc["jur"][k]["central"]["increment_eur_kw"]
        ramp = inc["jur"][k]["ramp"]["increment_eur_kw_by_year"]
        devex = st["jur"][k]["devex_social_pct"]
        # RAMPED CAPITAL (22 Aug 2026): each build-year tranche is
        # priced at that year's increment - FOAK in year one, nth in
        # the final year - instead of the whole build at nth. Early
        # tranches are dearer AND less discounted, so this is stricter
        # than the flat treatment on both counts.
        capex = mw * per_kw / 1000.0        # EUR m, mean - for display
        cool_m = cool["jur"][k]["avoided_capital_eur_m"]
        cap_pv = 0.0
        for t in range(1, VFM_BUILD_YEARS + 1):
            tranche = (mw / VFM_BUILD_YEARS) * ramp[t - 1] / 1000.0
            tranche_net = tranche - cool_m / VFM_BUILD_YEARS
            cap_pv += (tranche_net * (1.0 + devex) * (1.0 + ob)
                       * vfm_discount_factor(t - 0.5, k))
        capex_net = capex - cool_m
        capex_all = capex_net * (1.0 + devex) * (1.0 + ob)
        # Avoided capacity, DERIVED - it was hard-coded, and when the
        # scenario moved from the 2030 milestone to the ten-year build
        # the cost doubled while this did not, dropping both BCRs. The
        # audit had flagged it as a shortcut; it bit within the hour.
        #
        # A geothermal network draws less at the system peak because
        # its source does not freeze. Air source at the tightest hour
        # is 2.32, our own model on Irish weather; ground source holds
        # its seasonal figure. Valued at Net CONE, the avoided COST of
        # capacity, not the auction clearing price.
        p_ash = ANCHORS["ashp"]
        t_peak = 0.21                       # Panel 3 tightest hour
        lift = max(5.0, p_ash["flow_c"] - (t_peak - AIR_APPROACH_C))
        cop_peak = (p_ash["carnot_fraction"] * (p_ash["flow_c"] + 273.15)
                    / lift * defrost_factor(t_peak))
        avoided_mw = mw / cop_peak - mw / NETWORK_MODEL[k]["spf"]
        cap_eur_m = (avoided_mw
                     * VFM_SEM_CAPACITY["net_cone_2028_29"] / 1e6)
        # the cooling electricity those chillers never draw, at the
        # same long-run variable cost as the heat side
        cool_elec = cool["jur"][k]["elec_saved_twh"]
        cool_run_m = cool_elec * 1e6 * vfm_lrvc(2030, k) / 1e6
        # SUBSURFACE SHORTFALL, applied 22 Aug 2026. Post-commissioning
        # underdelivery against design - defined since the lever was
        # built, never applied until now, and it flattered us.
        #
        # It scales BENEFITS and leaves capital at the RATED capacity,
        # following the UK sibling: a shortfall is capacity paid for and
        # not received. The counterfactual carries no equivalent,
        # because the atmosphere does not run down.
        #
        # FLAT (1-s) on every benefit stream. The UK reduces effective
        # sizing and reads energy off a duration curve, so its shortfall
        # bites less than (1-s) on energy and exactly (1-s) on capacity.
        # This scenario is energy-defined at 4,000 hours with no sizing
        # lever, so there is no curve to read - the flat treatment is
        # the conservative port, and deliberately harder on us.
        #
        # The gap is filled by the COUNTERFACTUAL, not by gas as in the
        # UK, because stage 2 compares against an air-source network. So
        # the benefit is forgone; no extra carbon penalty beyond scaling.
        #
        # AND THE AVOIDED CHILLER CAPITAL (cool_m) IS NOT SCALED. Under
        # the fixed-build framing capital sits at rated capacity on both
        # sides: those chillers are genuinely never bought, however the
        # ground performs. Scaling it too would charge the shortfall
        # against cooling twice - once on the operating benefit below,
        # once on capital above.
        shortfall = st["jur"][k]["shortfall_default"]
        keep = 1.0 - shortfall
        # gross_annual is kept because the stream SPLIT divides by it.
        # Dividing by the scaled figure leaves the fractions summing to
        # 1/keep, inflating every stream by the same factor.
        gross_annual = (run["jur"][k]["resource_eur_m_yr"]["central"]
                        + cap_eur_m + cool_run_m)
        annual = gross_annual * keep
        # CARBON, now summed rather than merely computed. It is the one
        # stream that does not run flat: it is worth most now and
        # extinguishes within a decade on the operators' own path, so
        # it is discounted year by year rather than annualised.
        #
        # TES Self-Sustaining, IE average intensity. The 2026-30 values
        # interpolate from peer-reviewed modelling (~290 g/kWh in 2030)
        # to TES's own 2035 figure; TES publishes megatonnes, not
        # intensities, so the early years - which carry most of the
        # value - are our construction.
        #
        # AND IT GOES NEGATIVE AFTER 2040 ON UNBUILT BECCS. The sign
        # flip depends on 300 MW of bioenergy with carbon capture,
        # against zero biomass-only capacity in every scenario. There
        # is no operating BECCS plant on this island, no licensed CO2
        # storage and no transport infrastructure. A reader who assumes
        # the grid simply gets clean would miss that the negative years
        # rest on one unbuilt technology.
        saved_twh = carb["jur"][k]["elec_saved_twh"] * keep
        carb_pv = 0.0
        for t in range(1, VFM_HORIZON_YEARS + 1):
            y = 2026 + t - 1
            ramp = min(t, VFM_BUILD_YEARS) / VFM_BUILD_YEARS
            tco2 = saved_twh * ramp * 1e9 * vfm_grid_intensity(y) / 1e6
            carb_pv += (tco2 * vfm_shadow_carbon(k, y) / 1e6
                        * vfm_discount_factor(t, k))
        flat_pv = sum(annual * min(t, VFM_BUILD_YEARS) / VFM_BUILD_YEARS
                      * vfm_discount_factor(t, k)
                      for t in range(1, VFM_HORIZON_YEARS + 1))
        # CONSTRAINED WIND (NI only). Fleet draw = network heat / SPF,
        # capped at the constraint energy; coincident fraction priced
        # at that year's LRVC; linear decay over the reinforcement
        # horizon; ramped with the build; scaled by keep like every
        # benefit. See VFM_CONSTRAINED_WIND for the exclusions.
        cw_pv = 0.0
        cw_constraint = vfm_constraint_twh()
        if k == "ni" and cw_constraint:
            cw = VFM_CONSTRAINED_WIND
            draw_twh = min(sc["jur"][k]["network_twh"]
                           / st["jur"][k]["spf"],
                           cw_constraint)
            for t in range(1, VFM_HORIZON_YEARS + 1):
                decay = max(0.0, 1.0 - (t - 1) / cw["erosion_years"])
                if decay <= 0.0:
                    break
                ramp = min(t, VFM_BUILD_YEARS) / VFM_BUILD_YEARS
                cw_pv += (keep * cw["coincidence"] * draw_twh * decay
                          * ramp * vfm_lrvc(2026 + t - 1, k)
                          * vfm_discount_factor(t, k))
        ben_pv = flat_pv + carb_pv + cw_pv
        f_run = (run["jur"][k]["resource_eur_m_yr"]["central"]
                 / max(gross_annual, 1e-9))
        f_cap = cap_eur_m / max(gross_annual, 1e-9)
        f_cool = cool_run_m / max(gross_annual, 1e-9)
        run_pv = flat_pv * f_run
        cap_pv_ben = flat_pv * f_cap
        cool_pv_ben = flat_pv * f_cool
        out["jur"][k] = {
            "plant_mw": mw, "increment_eur_kw": per_kw,
            "capex_undiscounted_eur_m": round(capex, 1),
            "cooling_avoided_eur_m": round(cool_m, 1),
            "capex_net_of_cooling_eur_m": round(capex_net, 1),
            "cooling_pct_of_capex": round(100 * cool_m
                                          / max(capex, 1e-9), 1),
            "capex_with_devex_and_ob_eur_m": round(capex_all, 1),
            "capex_pv_eur_m": round(cap_pv, 1),
            "avoided_peak_mw": round(avoided_mw, 1),
            # Reported NET of the shortfall, so running + capacity +
            # cooling reconcile to annual_benefit_eur_m. The gross
            # figures sit beside them - what the stream is worth if the
            # ground performs to design.
            "capacity_eur_m_yr": round(cap_eur_m * keep, 1),
            "running_eur_m_yr": round(
                run["jur"][k]["resource_eur_m_yr"]["central"] * keep, 1),
            "capacity_gross_eur_m_yr": round(cap_eur_m, 1),
            "running_gross_eur_m_yr": round(
                run["jur"][k]["resource_eur_m_yr"]["central"], 1),
            "annual_benefit_eur_m": round(annual, 1),
            "carbon_pv_eur_m": round(carb_pv, 1),
            "benefit_pv_eur_m": round(ben_pv, 1),
            # FOUR streams. Cooling appears on BOTH sides - avoided
            # chillers net off the capital above, and the electricity
            # those chillers never draw is a benefit here. Two facts,
            # not one counted twice.
            "streams_pv_eur_m": {
                "running": round(run_pv, 1),
                "capacity": round(cap_pv_ben, 1),
                "constrained_wind": round(cw_pv, 1),
                "carbon": round(carb_pv, 1),
                "cooling": round(cool_pv_ben, 1)},
            "cooling_running_eur_m_yr": round(cool_run_m * keep, 1),
            "cooling_running_gross_eur_m_yr": round(cool_run_m, 1),
            "bcr": round(ben_pv / max(cap_pv, 1e-9), 2),
            # The lever, and what it costs us. bcr_before_shortfall is
            # the figure this panel published while the lever sat
            # unused, kept so the change is visible rather than silent.
            "shortfall_applied": shortfall,
            "shortfall_range": st["jur"][k]["shortfall_range"],
            "bcr_before_shortfall": round(
                ben_pv / keep / max(cap_pv, 1e-9), 2),
            "bcr_at_worst_shortfall": round(
                ben_pv / keep * (1 - st["jur"][k]["shortfall_range"][1])
                / max(cap_pv, 1e-9), 2),
            # The shortfall at which this case stops passing. If it
            # sits INSIDE the declared range, the range contains
            # failure and the panel must say so on its face.
            "shortfall_breakeven": round(
                1.0 - cap_pv / max(ben_pv / keep, 1e-9), 4),
        }
        # CLOSED-FORM COEFFICIENTS for the live levers (22 Aug 2026).
        # Every lever enters the discounted arithmetic linearly or
        # multiplicatively, so the browser never needs the appraisal -
        # it evaluates an exact closed form from these constants, and a
        # test pins the two implementations to each other at the corners
        # of the lever space. That makes drift between the page and this
        # function a tested invariant, not a standing worry.
        #
        #   m_wh   = 1 - wh_share*(1 - relcost)
        #   S0     = (1-w)*Ash*((1-ss) + ss*m_wh) + w*Adp
        #   S1     = (1-w)*Ash*ss*m_wh + w*Adp*ds
        #   sum(inc_t*df_t) = D0*(S0*capm - asc) - lr*D1*S1*capm
        #   cap_pv = ((mw/(1000*N))*sum - (cool_m/N)*D0)
        #            * (1+devex)*(1+ob)
        #   ben_pv = (1-s) * (flat_gross_pv + carbon_gross_pv)
        n_by = VFM_BUILD_YEARS
        d0 = sum(vfm_discount_factor(t - 0.5, k)
                 for t in range(1, n_by + 1))
        d1 = sum(((t - 1) / max(n_by - 1, 1))
                 * vfm_discount_factor(t - 0.5, k)
                 for t in range(1, n_by + 1))
        out["jur"][k]["coeffs"] = {
            "mw": mw, "build_years": n_by,
            "d0": round(d0, 6), "d1": round(d1, 6),
            "asc_eur_kw": VFM_AIRSOURCE_EUR_KW["mid"],
            "a_shallow_eur_kw": VFM_SHALLOW_FOAK_EUR_KW,
            "a_deep_eur_kw": round(VFM_DEEP_GBP_KW["mid"] * GBP_EUR, 2),
            "deep_weight": VFM_JUR_MODEL[k]["deep_weight"],
            "ss_shallow": VFM_SUBSURFACE_SHARE["shallow"],
            "ss_deep": VFM_SUBSURFACE_SHARE["deep"],
            "cool_capital_eur_m": cool_m,
            "devex": devex,
            # includes constrained wind - it scales only with keep,
            # so folding it here keeps the closed form exact with
            # no evaluator change
            "flat_gross_pv_eur_m": round((flat_pv + cw_pv) / keep, 3),
            "carbon_gross_pv_eur_m": round(carb_pv / keep, 3),
            "gbp_per_eur": round(1.0 / GBP_EUR, 6) if k == "ni" else None,
        }
        # CURRENCY (22 Aug 2026). The standing convention is each
        # jurisdiction in its own currency; the chain to here ran the
        # North in euro, a recorded breach. Every NI input is either
        # natively sterling converted in at the single ECB 2025-S2
        # semester rate, or natively euro (SEM Net CONE); all real
        # 2025. So a euro computation presented in sterling at that
        # same rate is EXACTLY the sterling computation - the rate
        # cancels in the BCR and native-sterling figures round-trip
        # unchanged. The legacy _eur_m keys stay until the renderer
        # migrates; the sterling block is authoritative for the North.
        out["jur"][k]["currency"] = "GBP" if k == "ni" else "EUR"
        if k == "ni":
            g = 1.0 / GBP_EUR
            out["jur"][k]["gbp"] = {
                "capex_pv_gbp_m": round(cap_pv * g, 1),
                "benefit_pv_gbp_m": round(ben_pv * g, 1),
                "annual_benefit_gbp_m": round(annual * g, 1),
                "capex_undiscounted_gbp_m": round(capex * g, 1),
                "cooling_avoided_gbp_m": round(cool_m * g, 1),
                "increment_gbp_kw": round(per_kw * g, 0),
                "streams_pv_gbp_m": {
                    "running": round(run_pv * g, 1),
                    "constrained_wind": round(cw_pv * g, 1),
                    "capacity": round(cap_pv_ben * g, 1),
                    "carbon": round(carb_pv * g, 1),
                    "cooling": round(cool_pv_ben * g, 1)},
                "rate": "ECB 2025-S2 semester mean, "
                        f"{round(1.0/GBP_EUR, 5)} GBP per EUR",
            }
    # THE LIVE LEVERS. Four headline - the four appraisal questions in
    # plain language - and three in a capital fold for the specialist.
    # Two hard-wired at their defaults: connection relcost (0.30,
    # Causeway judgement pending the source inventory - a daggered
    # default, deliberately NOT promoted to a midrange constant) and
    # cooling connections (its whole 0-60% swing moves the BCR ~0.02).
    #
    # The SHORTFALL IS A PROGRAMME MEAN, not a project number. What
    # survives aggregation across a hundred schemes is the correlated
    # remainder: systematic assessment bias, shared play risk, shared
    # completion practice, and the waste-heat volumes named in the
    # lever's own definition. It carries BOTH attrition (schemes that
    # commission and then close - 26% of 256 HSA projects failed, most
    # in the operational phase; Bremaud et al. 2025) and performance
    # shading of the survivors. Modern, remediated programmes imply an
    # attrition-equivalent of order 5-10%, which is where the defaults
    # sit. The tops of the ranges are common-mode stresses with named
    # precedents (Denmark 67%, Eromanga 87.5%), not expectations.
    out["levers"] = {
        "headline": [
            {"id": "shortfall", "label": "Subsurface shortfall, "
             "programme mean", "per_jur": True,
             "default": {j: st["jur"][j]["shortfall_default"]
                         for j in ("roi", "ni")},
             "range": {j: list(st["jur"][j]["shortfall_range"])
                       for j in ("roi", "ni")}},
            {"id": "optimism", "label": "Optimism bias",
             "per_jur": False,
             "default": VFM_OPTIMISM["default_pct"] / 100.0,
             "range": [VFM_OPTIMISM["min_pct"] / 100.0,
                       VFM_OPTIMISM["max_pct"] / 100.0]},
            {"id": "capm", "label": "Subsurface capital",
             "per_jur": False, "default": 1.0, "range": [0.75, 1.25]},
            {"id": "wh", "label": "Waste-heat coupling share",
             "per_jur": True,
             "default": dict(VFM_WASTE_HEAT["share"]),
             "range": {j: list(VFM_WASTE_HEAT["share_range"])
                       for j in ("roi", "ni")}},
        ],
        "capital_fold": [
            {"id": "a_sh", "label": "Shallow FOAK anchor (EUR/kW)",
             "default": VFM_SHALLOW_FOAK_EUR_KW,
             "range": [1500.0, 3300.0]},
            {"id": "a_dp", "label": "Deep FOAK anchor (EUR/kW)",
             "default": round(VFM_DEEP_GBP_KW["mid"] * GBP_EUR, 0),
             "range": [round(VFM_DEEP_GBP_KW["low"] * GBP_EUR, 0),
                       round(VFM_DEEP_GBP_KW["high"] * GBP_EUR, 0)]},
            {"id": "lr", "label": "Programme learning rate",
             "default": VFM_FOAK_LEARNING["reduction"],
             "range": list(VFM_FOAK_LEARNING["range"])},
        ],
        "hardwired": {"connection_relcost":
                      VFM_WASTE_HEAT["connection_relcost"],
                      "cooling_connections_pct": 12.0},
    }
    # WHAT IS IN THE ARITHMETIC AND WHAT IS NOT. Eight terms of about
    # twenty-five. Published so the panel cannot imply completeness.
    #
    # The previous version said "four" above a list of five, and the
    # list omitted carbon and cooling although both were summed. A
    # count asserted in a comment drifts from the list beneath it; the
    # test now counts the list instead.
    out["integrated"] = ["running cost at LRVC",
                         "avoided generation capacity",
                         "carbon at the shadow price",
                         "cooling, operating and avoided chiller "
                         "capital",
                         "subsurface increment",
                         "development capital divided by success",
                         "optimism bias on capital",
                         "subsurface shortfall on the benefit side"]
    out["not_integrated"] = {
        "stage2_benefits": ["avoided network reinforcement",
                            "air quality",
                            "interseasonal storage enabling waste-heat "
                            "recovery", "dispatch-down absorption",
                            "security of supply and indigenous share",
                            "residual value at 60 years"],
        "stage2_costs": ["operating and maintenance - EXCLUDED WITH A "
                         "BALANCE NOTE (22 Aug 2026): routine servicing "
                         "runs LOWER for geothermal (indoor water-to-"
                         "water plant on a steady source, no defrost "
                         "duty, no outdoor coils; the Causeway Hospital "
                         "feasibility carries 0.6-0.75 GBP/MWh against "
                         "roughly 2 EUR/MWh for the DEA air-source "
                         "centre), while the episodic subsurface "
                         "lifecycle runs HIGHER (well-pump and ESP "
                         "replacement on the deep leg, open-loop "
                         "scaling and dosing, workovers, licence fees "
                         "and monitoring). The two pull against each "
                         "other and the net is plausibly near zero; "
                         "excluded until the subsurface lifecycle "
                         "rates are set from developer data rather "
                         "than judgement"],
        # EDITED 23 Aug 2026 at Simon's direction: the UNRESOLVED
        # BOUNDARY sentence (source-pumping parasitics vs class SPFs)
        # and the heat-pump-replacement omission item were removed from
        # the page. The parasitics question itself is NOT resolved -
        # it lives here so it is not lost: if the class SPFs are
        # heat-pump-only, well and circulation pumping is an uncounted
        # operating cost and the running stream is flattered.
        "stage1": ["NOTHING IS BUILT - fuel avoided, electricity "
                   "drawn, carbon at the fossil margin, the capacity "
                   "COST of electrification, network reinforcement, "
                   "distribution and connection capital, building-side "
                   "works, boiler replacement cycles avoided"],
    }
    out["known_shortcuts"] = [
        "the shortfall is applied flat, not through a duration curve as "
        "the UK sibling does - harder on us, but it assumes every hour "
        "underdelivers equally",
        "ROI's deep weight is ZERO because the Early "
        "Carboniferous Limestone play remains unproven despite "
        "extensive research and, by our count, five wells to about "
        "1 km - there is no demonstrated intermediate play in "
        "Ireland. The current ~2 km Grangegorman well in Dublin is "
        "exploration of exactly this question, not evidence it is "
        "answered; the panel prices what is proven, and will move "
        "when the well does",
        "the avoided chiller capital is NOT scaled by the shortfall, "
        "on the ground that those chillers are never bought however "
        "the ground performs",
        "avoided distribution reinforcement is bounded rather than "
        "built: under GBP 11m for NI on RP7 unit rates, and "
        "transmission reinforcement sits outside the price control",
        "the long-run variable cost is borrowed from the UK sibling",
        "the ten-year build has no Irish basis and is generous to the "
        "North, which has no supply chain to ramp",
        "optimism bias applies to capital only; no Irish requirement "
        "for a benefit adjustment was found",
    ]
    log(f"vfm phased: {VFM_BUILD_YEARS}-year build, "
        f"{VFM_HORIZON_YEARS}-year horizon, {ob*100:.0f}% optimism on "
        f"capital only; "
        + "; ".join(f"{k.upper()} PV cost "
                    f"{v['capex_pv_eur_m']:.0f}m against benefit "
                    f"{v['benefit_pv_eur_m']:.0f}m, BCR {v['bcr']}"
                    for k, v in out["jur"].items()))
    return out


def derive_vfm_running(anchors=None):
    """
    The running-cost saving of the subsurface stage: electricity not
    drawn because the SPF is higher, valued twice.

    THIS IS NOW THE DURABLE STREAM. Carbon all but vanishes within a
    decade on the operators' own scenarios; capacity does not decay but
    is modest; this one runs for the whole sixty years and grows with
    the heat load.
    """
    a = anchors or ANCHORS
    c = derive_vfm_carbon(a)
    out = {"lrvc_eur_mwh": dict(VFM_LRVC_EUR_MWH), "jur": {}}
    for k, v in c["jur"].items():
        saved_mwh = v["elec_saved_twh"] * 1e6
        res = {lab: round(saved_mwh * VFM_LRVC_EUR_MWH[lab] / 1e6, 1)
               for lab in ("low", "central", "high")}
        out["jur"][k] = {
            "elec_saved_twh": v["elec_saved_twh"],
            "resource_eur_m_yr": res,
            "transfers_excluded": list(VFM_TRANSFERS),
        }
    log("vfm running: subsurface stage saves "
        + "; ".join(f"{k.upper()} EUR "
                    f"{v['resource_eur_m_yr']['central']}m a year at "
                    f"LRVC (range "
                    f"{v['resource_eur_m_yr']['low']}-"
                    f"{v['resource_eur_m_yr']['high']})"
                    for k, v in out["jur"].items())
        + " - RESOURCE cost, transfers excluded; the bill saving is "
          "larger and is reported separately")
    return out


def derive_vfm_carbon(anchors=None):
    """
    Carbon saved by the SUBSURFACE STAGE ONLY - geothermal against an
    air-source network, both electric.

    Stage one, moving off oil and gas, saves far more carbon - but any
    heat pump delivers that, and crediting it here would let the
    subsurface bank the electrification benefit. So this stage counts
    only the electricity NOT drawn because the SPF is higher, at the
    grid intensity of the year in question.

    GRID INTENSITY IS THE WHOLE STORY AND IT IS FALLING. On a
    decarbonising grid the carbon value of an efficiency gain shrinks
    every year - by 2050 an all-renewable grid makes it nearly zero.
    That cuts against our own argument and the panel should say so.
    """
    a = anchors or ANCHORS
    s = derive_vfm_scenario(a)
    out = {"stage": "subsurface", "jur": {}}
    for k, v in s["jur"].items():
        net = v["network_twh"]
        spf_geo = NETWORK_MODEL[k]["spf"]
        spf_as = SPF_ANCHORS["ashp"]
        # electricity saved per year, TWh
        saved = net / spf_as - net / spf_geo
        out["jur"][k] = {
            "network_twh": net,
            "elec_ashp_twh": round(net / spf_as, 3),
            "elec_geo_twh": round(net / spf_geo, 3),
            "elec_saved_twh": round(saved, 3),
            "saved_pct": round(100 * saved / (net / spf_as), 1),
        }
    log(f"vfm carbon: subsurface stage saves "
        + "; ".join(f"{k.upper()} {v['elec_saved_twh']} TWh of "
                    f"electricity a year ({v['saved_pct']}%)"
                    for k, v in out["jur"].items())
        + " - valued at the shadow price, NOT the carbon tax, and "
          "shrinking as the grid decarbonises")
    return out


def derive_vfm_increment(anchors=None):
    """
    The subsurface increment over an air-source network, per kW of
    heating capacity, blended by each jurisdiction's deep fraction.

    DISTRIBUTION AND CONNECTIONS CANCEL - same mains, same trenching,
    same heat interface units whichever source feeds them - which is
    why this reduces to one number rather than two cost stacks.

    The UK sibling carries GBP 875/kW here and it does not reconcile:
    a geothermal plant near GBP 1,200/kW less an air-source energy
    centre near GBP 750/kW, with distribution cancelling, gives about
    GBP 450/kW. GBP 875 credits air-source at only about GBP 325/kW,
    so it OVERSTATES geothermal's incremental capital - conservative,
    and against their own interest, but it should be a range rather
    than a default. Ours is computed from both sides so the netting is
    visible.
    """
    asc = VFM_AIRSOURCE_EUR_KW["mid"]
    deep_eur = VFM_DEEP_GBP_KW["mid"] * GBP_EUR
    lr = VFM_FOAK_LEARNING["reduction"]
    # Learning bites the SUBSURFACE SHARE of each all-in figure only -
    # the heat pump inside it is the same mature kit as the flat
    # air-source counterfactual. Effective reductions ~21.7% shallow,
    # ~18.5% deep at the workbook-derived shares.
    eff_sh = lr * VFM_SUBSURFACE_SHARE["shallow"]
    eff_dp = lr * VFM_SUBSURFACE_SHARE["deep"]
    deep_nth = deep_eur * (1 - eff_dp)
    sh_foak = VFM_SHALLOW_FOAK_EUR_KW
    sh_nth = sh_foak * (1 - eff_sh)
    if VFM_SHALLOW_NTH_FLOOR is not None:
        sh_nth = max(sh_nth, VFM_SHALLOW_NTH_FLOOR)
    out = {"air_source_eur_kw": asc,
           "shallow_eur_kw": dict(VFM_SHALLOW_EUR_KW),
           "subsurface_share": dict(VFM_SUBSURFACE_SHARE),
           "shallow_learn_eur_kw": {"foak": round(sh_foak, 0),
                                    "nth": round(sh_nth, 0),
                                    "effective_reduction":
                                        round(eff_sh, 4),
                                    "floor": VFM_SHALLOW_NTH_FLOOR,
                                    "danish_outturn":
                                        VFM_SHALLOW_EUR_KW["central"],
                                    "learning": lr},
           "deep_eur_kw": {"foak": round(deep_eur, 0),
                           "nth": round(deep_nth, 0),
                           "effective_reduction": round(eff_dp, 4),
                           "learning": lr},
           "jur": {}}
    for k, m in VFM_JUR_MODEL.items():
        w = m["deep_weight"]
        row = {}
        for lab, sh in (("floor", VFM_SHALLOW_EUR_KW["dutch_floor"]),
                        ("central", VFM_SHALLOW_EUR_KW["central"]),
                        ("high", VFM_SHALLOW_EUR_KW["high"])):
            for dl, dp in (("foak", deep_eur), ("nth", deep_nth)):
                blend = (1 - w) * sh + w * dp
                row[f"{lab}_{dl}"] = {
                    "blend_eur_kw": round(blend, 0),
                    "increment_eur_kw": round(blend - asc, 0)}
        # THE RAMP is the appraisal's central case from 22 Aug 2026:
        # both classes interpolate FOAK -> nth linearly along the build
        # years. Linear-in-year is the CONSERVATIVE shape - a learning
        # curve is convex in cumulative capacity, so most of the
        # reduction arrives early and a convex ramp would price the
        # build cheaper than this does.
        ramp = []
        n = VFM_BUILD_YEARS
        ss = VFM_SUBSURFACE_SHARE["shallow"]
        ds = VFM_SUBSURFACE_SHARE["deep"]
        # waste-heat displacement: this share of shallow schemes swaps
        # its subsurface for a connection at connection_relcost of the
        # displaced capital. Multiplier on the shallow subsurface:
        wh = VFM_WASTE_HEAT
        m_wh = 1.0 - wh["share"][k] * (1.0 - wh["connection_relcost"])
        for t in range(1, n + 1):
            f = (t - 1) / max(n - 1, 1)
            sh_t = sh_foak * ((1 - ss) + ss * m_wh * (1 - lr * f))
            dp_t = deep_eur * ((1 - ds) + ds * (1 - lr * f))
            blend_t = (1 - w) * sh_t + w * dp_t
            ramp.append(round(blend_t - asc, 1))
        row["ramp"] = {
            "increment_eur_kw_by_year": ramp,
            "year1": ramp[0], "final": ramp[-1],
            "mean": round(sum(ramp) / len(ramp), 1),
            "waste_heat_share": wh["share"][k],
            "waste_heat_relcost": wh["connection_relcost"],
            "waste_heat_subsurface_multiplier": round(m_wh, 4),
            "basis": "linear FOAK->nth over the build years, "
                     "subsurface-share learning both classes, "
                     "waste-heat displacement on the shallow "
                     "subsurface, from the Irish anchors"}
        out["jur"][k] = {"deep_weight": w, "cases": row,
                         "central": {"blend_eur_kw": round(
                             sum(ramp) / len(ramp) + asc, 0),
                             "increment_eur_kw": round(
                                 sum(ramp) / len(ramp), 0)},
                         "ramp": row["ramp"]}
    log(f"vfm increment: air-source {asc:.0f} EUR/kW flat; shallow "
        f"FOAK {sh_foak:.0f} -> nth {sh_nth:.0f} (Irish anchor, "
        f"subsurface-share learning); deep FOAK {deep_eur:.0f} -> nth "
        f"{deep_nth:.0f}. Ramped central increments (mean): ROI "
        f"{out['jur']['roi']['central']['increment_eur_kw']:+.0f}, NI "
        f"{out['jur']['ni']['central']['increment_eur_kw']:+.0f} "
        f"EUR/kW")
    return out


def derive_vfm_stages(anchors=None, geo=None):
    """
    The two-stage structure and the levers each jurisdiction is
    entitled to.

    Nothing is priced here - this publishes the SHAPE, so the front end
    and the tests agree on what the stages are and which levers belong
    where before any number is attached.
    """
    a = anchors or ANCHORS
    C = VFM_SUBSURFACE_CLASSES
    out = {"stages": list(VFM_STAGES), "jur": {},
           "classes": C, "deep_weight": VFM_DEEP_WEIGHT}
    for k, m in VFM_JUR_MODEL.items():
        nm = NETWORK_MODEL[k]
        w = m["deep_weight"]
        # social devex = devex / success, blended by weight. Money
        # spent on prospects that did not proceed is still a resource
        # cost per delivered megawatt.
        soc = ((1 - w) * C["shallow"]["devex_pct"] / C["shallow"]["success"]
               + w * C["deep"]["devex_pct"] / C["deep"]["success"])
        out["jur"][k] = {
            "class": m["class"],
            "deep_weight": w,
            "source_c": nm["source_c"],
            "spf": nm["spf"],
            "spf_counterfactual": SPF_ANCHORS["ashp"],
            "spf_gain": round(nm["spf"] / SPF_ANCHORS["ashp"], 2),
            "devex_social_pct": round(soc, 4),
            # BOTH carry a shortfall lever: productivity risk survives
            # commissioning, and part of it is not geological at all -
            # waste heat that does not arrive in the volume assumed.
            "shortfall_default": m["shortfall_default"],
            "shortfall_range": list(m["shortfall_range"]),
        }
    out["capacity_sign"] = {"electrify": -1, "subsurface": +1}
    log("vfm stages: BAU -> air-source network (electrify), then "
        "air-source -> geothermal (subsurface increment); never summed. "
        + "; ".join(
            f"{k.upper()} SPF {v['spf']} vs {v['spf_counterfactual']} "
            f"= {v['spf_gain']}x, {100*v['deep_weight']:.0f}% deep, "
            f"social devex {100*v['devex_social_pct']:.1f}%, shortfall "
            f"{100*v['shortfall_default']:.0f}% "
            f"({100*v['shortfall_range'][0]:.0f}-"
            f"{100*v['shortfall_range'][1]:.0f}%)"
            for k, v in out["jur"].items()))
    return out


def vfm_constraint_twh(year=None):
    """
    NI transmission-constraint energy for the stated basis year, TWh,
    read from our own dispatch-down series rather than asserted.

    Returns None if the file or the year is missing - the caller then
    prices no constrained-wind stream at all, which is the right
    failure: a stream whose volume cannot be sourced should not be in
    the appraisal.
    """
    y = year or VFM_CONSTRAINED_WIND["constraint_basis_year"]
    if not DD_PATH.exists():
        return None
    d = json.loads(DD_PATH.read_text())
    idx = [i for i, m in enumerate(d["months"]) if m.startswith(y)]
    if len(idx) != 12:
        log(f"constrained wind: {y} is not a complete year "
            f"({len(idx)} months) - stream declines")
        return None
    ni = d["jurisdictions"]["NI"]
    return sum(ni["cons"][i] for i in idx) / 1000.0


def derive_frontispiece(feeds, hcs=None, th=None, he=None,
                        hr=None, anchors=None):
    """
    Six figures that state the case, per jurisdiction.

    EVERY NUMBER HERE IS COMPUTED ELSEWHERE AND RESTATED. Nothing is
    derived for the first time in this function and nothing is
    hardcoded: it reads the scenario, stages, cooling and phased
    blocks, and the dispatch-down series, so a frontispiece figure
    cannot drift from the panel it summarises. If a panel moves, this
    moves with it or the build fails a test.

    Two of the six differ by jurisdiction ON PURPOSE. Figure 5 is the
    delivery gap in ROI - which has a target and is behind it - and
    the wasted wind in NI, which has no target but the worse network
    constraint. The UK sibling's "Britain is the outlier" figure does
    NOT transfer: Ireland's international comparison is weaker than
    its domestic one.
    """
    a = anchors or ANCHORS
    sc = derive_vfm_scenario(a)
    st = derive_vfm_stages(a)
    ph = derive_vfm_phased(a)
    ct = derive_cooling_tiers()
    dd = (json.loads(DD_PATH.read_text())
          if DD_PATH.exists() else None)
    if not (sc and st and ph):
        return None

    # figure 5, NI half: the most recent COMPLETE calendar year in the
    # series, named rather than "current" - dispatch-down is trending
    # and a moving basis would be cherry-pickable.
    ddyear, ddpct, ddcons = None, None, None
    if dd and dd.get("months"):
        yrs = sorted({m[:4] for m in dd["months"]})
        full = [y for y in yrs
                if sum(1 for m in dd["months"] if m.startswith(y)) == 12]
        if full:
            ddyear = full[-1]
            idx = [i for i, m in enumerate(dd["months"])
                   if m.startswith(ddyear)]
            ni = dd["jurisdictions"]["NI"]
            tot = sum(ni["dd"][i] for i in idx)
            av = sum(ni["avail"][i] for i in idx)
            cons = sum(ni["cons"][i] for i in idx)
            ddpct = round(100.0 * tot / av, 1) if av else None
            ddcons = round(100.0 * cons / tot) if tot else None

    # PANEL 2'S OWN TRAILING YEAR. Volume, spend and the price gap all
    # come from the same rows the cost chart draws, over the same 365
    # days, so a reader who checks the chart finds these figures in it.
    # Spend is the heat-weighted mean price times the volume, which is
    # the same arithmetic the chart's end labels print.
    def _year(j):
        if not hcs:
            return None
        vk, pk = "vol_" + j, j
        rows = [r for r in hcs[-365:] if r.get(vk) and r.get(pk)]
        if len(rows) < 300:
            return None
        gwh = sum(r[vk]["space"] + r[vk]["dhw"] for r in rows)
        if gwh <= 0:
            return None
        def wmean(route):
            return (sum(r[pk][route] * 10.0
                        * (r[vk]["space"] + r[vk]["dhw"]) for r in rows)
                    / gwh)
        net, ash = wmean("network"), wmean("ashp")
        blend = wmean("gas_boiler")  # the counterfactual actually bought
        return {"days": len(rows), "twh": gwh / 1000.0,
                "spend_bn": (blend * gwh * 1000.0) / 1e9,
                "network": net, "ashp": ash,
                "cheaper_pct": (100.0 * (ash - net) / ash) if ash else None}

    out = {"jur": {}}
    for j in ("roi", "ni"):
        heat = (a[j]["residential_heat_twh"]
                + a[j]["services_heat_twh"])
        fs = a[j]["fuel_shares"]
        comb = round(100.0 * (fs["oil"] + fs["gas"] + fs["peat"]), 1)
        s, t, p = sc["jur"][j], st["jur"][j], ph["jur"][j]
        cur = "GBP" if p.get("currency") == "GBP" else "EUR"

        yr = _year(j)
        thf = (th or {}).get("share_that_fits_pct") or {}
        # emissions cut against oil, from the emissions panel's
        # own per-route intensities - not recomputed here
        cut = None
        if he and he.get("routes"):
            byk = {r["key"]: r["g_per_useful_kwh"]
                   for r in he["routes"]}
            oil = byk.get("oil_boiler")
            net = byk.get("network") or byk.get("geothermal_network")
            if oil and net is not None and oil > 0:
                cut = 100.0 * (oil - net) / oil
        cy = "\u00a3" if cur == "GBP" else "\u20ac"
        where = "the Republic of Ireland" if j == "roi" \
            else "Northern Ireland"
        figs = [
            # SPEND, not volume, as the opening figure: what the place
            # PAYS is the number that lands. Both halves are panel 2's,
            # over its trailing year, so they reconcile against the
            # chart rather than against a separate anchor.
            {"n": 1,
             "v": (f"{cy}{yr['spend_bn']:.1f}bn" if yr
                   else f"{heat:.1f} TWh"),
             "unit": (f"on {yr['twh']:.1f} TWh" if yr else ""),
             # NOT "the biggest thing X buys": the Why heat footnote
             # says heat RIVALS transport as the largest service and
             # carries the smallest bill per unit delivered. The
             # frontispiece must not overstate what the footnote says.
             "claim": f"Heat rivals transport as the largest energy "
                      f"service {where} buys.",
             "body": ("Delivered heat over the last twelve months and "
                      "what it cost at the gas-boiler price most of it "
                      f"is actually bought at, of which {comb:.0f}% is "
                      "still combustion - oil, gas and peat burned in "
                      "buildings."),
             "to": "cost"},
            {"n": 2,
             "v": (f"{yr['cheaper_pct']:.0f}%" if yr and yr["cheaper_pct"]
                   else f"{t['spf']:.1f}"),
             "unit": ("cheaper heat" if yr and yr["cheaper_pct"]
                      else f"vs {t['spf_counterfactual']}"),
             "claim": "Geothermal heat is cheaper than air-source "
                      "heat, not just cleaner.",
             # PENCE PER kWh, not currency per MWh: a bill is read in
             # pence and the panel is arguing about bills. Same
             # numbers, divided by ten.
             "body": ((f"Per kWh delivered over the same twelve months, "
                       f"geothermal heat {yr['network'] / 10:.1f}"
                       + ("p" if cur == "GBP" else "c")
                       + f" against {yr['ashp'] / 10:.1f}"
                       + ("p" if cur == "GBP" else "c")
                       + " for an air-source heat pump network. "
                       if yr else "")
                      + "Clean heat from a source that does not cool "
                      "when the weather does."),
             "to": "cost"},
            # REPLACED the programme-scale figure, 26 Aug 2026. Scale
            # is a claim about ambition; this is a claim about the
            # thing being displaced, which is oil, and it is the
            # emissions panel's own arithmetic.
            {"n": 3,
             "v": (f"{cut:.0f}%" if cut else "-"),
             "unit": "less carbon than oil",
             "claim": "Against the oil boilers most of "
                      + ("Northern Ireland" if j == "ni"
                         else "the Republic")
                      + "'s heat still comes from.",
             "body": (("Grams of CO2 per useful kWh, geothermal network "
                       f"against an oil boiler, on today's grid. Oil is "
                       f"{100 * a[j]['fuel_shares']['oil']:.0f}% of "
                       + ("Northern Ireland" if j == "ni"
                          else "the Republic of Ireland")
                       + "'s building heat today, and the gap closes "
                       "further as the grid decarbonises.")
                      if cut else "Arrives with the next build."),
             "to": "cost"},
            # DISPLACES the cooling figure, 26 Aug 2026. The binding
            # hour is the harder test and the one a system operator
            # asks first. Ground routes clear it; air source does not,
            # which is the comparison that matters.
            {"n": 4,
             "v": (f"{thf['geothermal_network']:.0f}%"
                   if thf and thf.get("geothermal_network") else "-"),
             "unit": "fits at the tightest hour",
             "claim": "At the tightest hour on record, the ground "
                      "routes fit inside the grid. Air source does "
                      "not.",
             "body": (("All of " + where + "'s building heat, "
                       "electrified through networks, against the "
                       "headroom in the all-island block at the "
                       "tightest hour observed - there is no separate "
                       "northern ceiling, because there is one "
                       "dispatch. Ground source alone reaches "
                       + (f"{thf['ground_source']:.0f}%"
                          if thf.get("ground_source") else "less")
                       + ", air source "
                       + (f"{thf['air_source']:.0f}%"
                          if thf.get("air_source") else "less")
                       + ".") if thf else "Arrives with the next "
                      "build."),
             "to": "grid"},
        ]
        if j == "roi":
            # THE WIDE VERSION, by decision 26 Aug 2026. The
            # frontispiece states the PRIZE - all the heat data
            # centres reject, and what it would serve if recovered.
            # Panel 6 prices the conservative version: waste heat
            # enters the appraisal as capex avoidance on a small
            # coupled share, and nothing here changes that. The two
            # are different questions and the body says which is
            # which.
            dc = None
            if hr:
                dc = next((r for r in hr.get("rows", [])
                           if r["key"] == "datacentres"), None)
            # TWh -> MWh is 1e6, not 1e3. The first version divided
            # by a thousand too few and printed 1,000 homes for 6.4
            # TWh, which is off by three orders of magnitude.
            homes = (dc["rejected_twh"] * 1e6 / HOME_HEAT_MWH
                     if dc else None)
            figs.append(
                {"n": 5,
                 "v": (f"{dc['rejected_twh']:.1f}" if dc else "-"),
                 # SHORT unit. It carried the whole phrase and the
                 # claim wrapped onto the same line as the number.
                 "unit": "TWh rejected",
                 "claim": "Data centres already reject more heat than "
                          "the Republic's networks would need.",
                 "body": ((f"Growing to {COOL_DC_GROWTH[1]:.1f} TWh by "
                           f"{COOL_DC_GROWTH[2]}. Recovered, stored "
                           "and upgraded through geothermal networks, "
                           "today's rejection alone is the heat of "
                           f"about {homes / 1000:.0f},000 homes at "
                           f"{HOME_HEAT_MWH:.0f} MWh a year"
                           "\u2020. This is the whole prize, not a "
                           "forecast of recovery.")
                          if dc and homes else "Arrives with the next "
                          "build."),
                 "to": "cooling"})
        else:
            figs.append(
                {"n": 5, "v": f"{ddpct:.1f}", "unit": "% of wind spilled",
                 "claim": "Northern Ireland throws away wind it "
                          "cannot move.",
                 "body": (f"In {ddyear}, {ddcons}% of it transmission "
                          "constraint - wind that cannot leave where "
                          "it is generated, against roughly half that "
                          "share in the Republic. A heat load sited "
                          "inside the constraint absorbs it."),
                 "to": "grid"})
        figs.append(
            {"n": 6, "v": f"{p['bcr']:.2f}", "unit": "benefit-cost ratio",
             "claim": "Against an air-source-led network, the "
                      "subsurface investment pays for itself.",
             "body": ("By "
                      + ("Public Spending Code" if j == "roi"
                         else "Green Book")
                      + " conventions, after optimism bias on capital "
                      "and a programme shortfall. Measured against "
                      "air-source counterfactuals."
                      + (" Over a ten-year build extrapolated from "
                         "the government's own 2.7 TWh commitment for "
                         "2030 - a build starting now reaches 2.5 TWh "
                         "by then, most of the way to it."
                         if j == "roi" else "")),
             "to": "vfm"})
        out["jur"][j] = {"currency": cur, "figures": figs}

    log("frontispiece: six figures per jurisdiction, all restated "
        f"from the panels (NI wind basis {ddyear})")
    return out


def derive_vfm_scenario(anchors=None):
    """
    The policy scenario, per jurisdiction, and a fifth of it geothermal.

    SEAI's National Heat Study puts the technical potential for
    district heating at up to 54% of building heat, so a scenario at
    10.7% sits well inside it - the ceiling is SEAI's, not ours.
    """
    a = anchors or ANCHORS
    roi_del = (a["roi"]["residential_heat_twh"]
               + a["roi"]["services_heat_twh"]) \
        * a.get("delivered_over_input_roi", 0.8225)
    ni_del = (a["ni"]["residential_heat_twh"]
              + a["ni"]["services_heat_twh"]) \
        * a.get("delivered_over_input_ni", 0.8375)
    tgt = VFM_DH_SCENARIO["roi_twh"]
    ms = VFM_DH_SCENARIO["roi_milestone_twh"]
    prop = tgt / roi_del
    flh = VFM_DH_SCENARIO["network_load_hours"]
    lo, hi = VFM_DH_SCENARIO["roi_delivery_pct"]
    jur = {
        "roi": {"delivered_twh": round(roi_del, 2), "network_twh": tgt,
                "basis": "ten-year build", "committed": True,
                "milestone_twh": ms,
                "milestone_year": VFM_DH_SCENARIO["roi_milestone_year"]},
        "ni": {"delivered_twh": round(ni_del, 2),
               "network_twh": round(prop * ni_del, 2),
               "basis": "lent", "committed": False,
               "milestone_twh": round(ms / roi_del * ni_del, 2),
               "milestone_year": VFM_DH_SCENARIO["roi_milestone_year"]},
    }
    for k, v in jur.items():
        # the WHOLE scenario, either way it is supplied
        v["network_gwh"] = round(v["network_twh"] * 1000, 0)
        v["plant_mw"] = round(v["network_twh"] * 1e6 / flh, 0)
        v["network_pct_of_heat"] = round(100 * v["network_twh"]
                                         / v["delivered_twh"], 1)
        # the site-wide fifth, for contrast - a DIFFERENT question
        v["site_whatif_twh"] = round(0.20 * v["delivered_twh"], 2)
        v["vs_site_whatif"] = round(v["network_twh"]
                                    / v["site_whatif_twh"], 2)
    out = {
        "scenario": dict(VFM_DH_SCENARIO),
        "jur": jur,
        "load_hours": flh,
        "island_network_gwh": round(sum(v["network_gwh"]
                                        for v in jur.values()), 0),
        "island_plant_mw": round(sum(v["plant_mw"] for v in jur.values()), 0),
        "roi_outturn_twh": [round(tgt * lo / 100, 2),
                            round(tgt * hi / 100, 2)],
        "seai_technical_ceiling_pct": 54.0,
    }
    log(f"vfm scenario: ROI district heating target {tgt} TWh "
        f"({jur['roi']['network_pct_of_heat']}% of its building heat, "
        f"{jur['roi']['plant_mw']:.0f} MW at {flh} h); NI at the same "
        f"proportion would be {jur['ni']['network_twh']} TWh "
        f"({jur['ni']['plant_mw']:.0f} MW) - LENT, it has no target. "
        f"The WHOLE scenario is priced either way it is supplied - no "
        f"fifth inside it")
    return out


def derive_vfm_constants(anchors=None):
    """
    Panel 6's shared constants, and the cross-panel reads that keep it
    consistent with the rest of the site.

    Nothing here is computed twice: the dispatch-down series comes from
    Panel 3, the cooling service factors from Panel 1 and 4, the SPFs
    and the geothermal load hours from the anchors those panels use.
    """
    a = anchors or ANCHORS
    roi_del = (a["roi"]["residential_heat_twh"]
               + a["roi"]["services_heat_twh"]) \
        * a.get("delivered_over_input_roi", 0.8225)
    ni_del = (a["ni"]["residential_heat_twh"]
              + a["ni"]["services_heat_twh"]) \
        * a.get("delivered_over_input_ni", 0.8375)
    out = {
        "counterfactuals": list(VFM_COUNTERFACTUALS),
        "capacity_applies_to": list(VFM_CAPACITY_APPLIES_TO),
        "sem_capacity_eur_mw_yr": VFM_SEM_CAPACITY,
        "gb_capacity_gbp_kw_yr": VFM_GB_CAPACITY_GBP_KW,
        "appraisal": VFM_APPRAISAL,
        "transfers": list(VFM_TRANSFERS),
        "ashp_energy_centre": VFM_ASHP_ENERGY_CENTRE_EUR_KWTH,
        # read, not restated
        "spf": {"ashp": SPF_ANCHORS["ashp"], "gshp": SPF_ANCHORS["gshp"],
                "network_ni": NETWORK_MODEL["ni"]["spf"],
                "network_roi": NETWORK_MODEL["roi"]["spf"]},
        "delivered_heat_twh": {"roi": round(roi_del, 1),
                               "ni": round(ni_del, 1)},
        "whatif_share": 0.20,
        "horizon_years": 60,
    }
    log(f"vfm constants: SEM capacity "
        f"{VFM_SEM_CAPACITY['t4_2029_30']:,.0f} EUR/MW/yr against a GB "
        f"T-4 of GBP {VFM_GB_CAPACITY_GBP_KW['t4_2029_30']}/kW/yr; "
        f"capacity benefit applies to the ASHP counterfactual ONLY; "
        f"network-scale ASHP energy centre carried as a "
        f"{VFM_ASHP_ENERGY_CENTRE_EUR_KWTH['low']:.0f}-"
        f"{VFM_ASHP_ENERGY_CENTRE_EUR_KWTH['high']:.0f} EUR/kWth proxy "
        f"- no Irish figure is published")
    return out


def derive_geo_targets(anchors=None, geo=None):
    """
    What published policy the 20% what-if can be measured against.

    THE FINDING IS THAT ALMOST NOTHING EXISTS. Neither jurisdiction has
    set a geothermal deployment target of any kind; both treat it as a
    licensing matter. So the what-if is placed against the nearest
    quantified targets that geothermal could contribute to, each
    labelled with what it actually covers.
    """
    a = anchors or ANCHORS
    g = geo or GEO
    roi_del = (a["roi"]["residential_heat_twh"]
               + a["roi"]["services_heat_twh"]) \
        * a.get("delivered_over_input_roi", 0.8225)
    ni_del = (a["ni"]["residential_heat_twh"]
              + a["ni"]["services_heat_twh"]) \
        * a.get("delivered_over_input_ni", 0.8375)
    ni_in = a["ni"]["residential_heat_twh"] + a["ni"]["services_heat_twh"]
    share = 0.20
    # NI: a fifth of delivered heat, as energy SAVED against combustion
    d = share * ni_del
    comb = d / (ni_del / ni_in)
    saved = {k: round((comb - d / v) * 1000, 0) for k, v in
             (("gshp", SPF_ANCHORS["gshp"]),
              ("network", NETWORK_MODEL["ni"]["spf"]))}
    out = {
        "geothermal_targets": [],          # the point: this is empty
        "roi_delivered_twh": round(roi_del, 1),
        "ni_delivered_twh": round(ni_del, 1),
        "island_delivered_twh": round(roi_del + ni_del, 1),
        "whatif_share": share,
        "nearest": [
            {"jur": "ROI", "label": "District heating",
             "value": 2.7, "unit": "TWh/yr", "year": 2030,
             "status": "government commitment",
             "covers": "all heat sources, no geothermal share",
             "source": "Climate Action Plan 2025"},
            {"jur": "ROI", "label": "Heat pumps installed",
             "value": 680000, "unit": "units", "year": 2030,
             "status": "government commitment",
             "covers": "air and ground source, no ground-source share",
             "source": "Climate Action Plan 2025"},
            {"jur": "NI", "label": "Energy saved, buildings and industry",
             "value": GEO_NI_EE_TARGET_GWH, "unit": "GWh", "year": 2030,
             "status": "strategy target",
             "covers": "all savings measures, buildings AND industry",
             "source": "NI Energy Strategy, DfE 2021",
             "achieved_gwh": GEO_NI_EE_ACHIEVED_GWH},
        ],
        # The Republic's own comparison. Target-setting is a
        # jurisdictional act, so the panel has no all-island view: each
        # side is measured against what its own government committed to.
        "roi_vs_dh": {
            "fifth_twh": round(share * roi_del, 2),
            "dh_target_twh": 2.7,
            "multiple": round(share * roi_del / 2.7, 1),
        },
        "ni_energy_saved": {
            "target_gwh": GEO_NI_EE_TARGET_GWH,
            "achieved_gwh": GEO_NI_EE_ACHIEVED_GWH,
            "achieved_pct": round(100 * GEO_NI_EE_ACHIEVED_GWH
                                  / GEO_NI_EE_TARGET_GWH, 1),
            "whatif_delivered_twh": round(d, 2),
            "counterfactual_input_twh": round(comb, 2),
            "saved_gwh": saved,
            "saved_pct": {k: round(100 * v / GEO_NI_EE_TARGET_GWH, 0)
                          for k, v in saved.items()},
        },
    }
    log(f"geo targets: NO geothermal deployment target exists in either "
        f"jurisdiction. Nearest quantified NI target is "
        f"{GEO_NI_EE_TARGET_GWH:.0f} GWh saved by 2030, at "
        f"{out['ni_energy_saved']['achieved_pct']}% delivered; a fifth "
        f"of NI heat on geothermal would save "
        f"{saved['gshp']:.0f}-{saved['network']:.0f} GWh, or "
        f"{out['ni_energy_saved']['saved_pct']['gshp']:.0f}-"
        f"{out['ni_energy_saved']['saved_pct']['network']:.0f}% of it")
    return out


def derive_geo_hardware(anchors=None, geo=None):
    """
    The empty bar: installed hardware against the 20% what-if, and the
    same fleets as a share of each country's OWN buildings heat.

    Comparator constants are the UK sibling's, unchanged - EGC 2025
    country updates (Sanner et al., Tables 3-4, end-2024), GSHP fleet
    plus deep direct use. The Netherlands' deep capacity is almost
    entirely greenhouse heat rather than buildings, which the note says.

    THE IRISH BAR IS NOT AS EMPTY AS BRITAIN'S, and the panel should
    not pretend otherwise: the Republic runs 42 W per person against
    the UK's 13 and France's 34, on 20,128 installed systems. Northern
    Ireland runs 3 W. The interesting gap on this island is internal.
    """
    a = anchors or ANCHORS
    g = geo or GEO
    eflh = g["eflh_h"]
    roi_del = (a["roi"]["residential_heat_twh"]
               + a["roi"]["services_heat_twh"]) \
        * a.get("delivered_over_input_roi", 0.8225)
    ni_del = (a["ni"]["residential_heat_twh"]
              + a["ni"]["services_heat_twh"]) \
        * a.get("delivered_over_input_ni", 0.8375)
    isl_del = roi_del + ni_del
    # input basis for the calibrated shares, matching the UK sibling's
    # convention; the delivered-vs-input mismatch is stated on the page
    isl_in = (a["roi"]["residential_heat_twh"] + a["roi"]["services_heat_twh"]
              + a["ni"]["residential_heat_twh"]
              + a["ni"]["services_heat_twh"])
    inst = g["roi"]["capacity_mwth"] + g["ni_capacity_mwth_est"]
    # THE WHAT-IF USES IRELAND'S OWN LOAD HOURS, not a European
    # convention. Irish systems run 1,301 full-load hours against a
    # European average of 2,420, so serving a fifth of heat needs MORE
    # capacity here, not less - the hardware requirement rises by about
    # half against the old 2,000-hour figure. That is the conservative
    # direction and against our own interest, which is why it is right.
    # A purpose-built network would run more hours than a domestic
    # retrofit fleet; the panel says so rather than assuming it.
    flh = g["reference_output"]["Ireland"]["flh"]
    whatif = 0.20 * isl_del * 1e6 / flh
    comps = [{"name": k, "gshp_MWth": v["shallow"], "deep_MWth": v["deep"]}
             for k, v in g["reference_mwth"].items()]
    # REPORTED OUTPUT, no load-hour convention. Each country's GSHP
    # fleet delivered a stated number of GWh in 2024, and that against
    # its own buildings heat is the comparison - like for like, with
    # the largest assumption in this panel removed.
    ro = g["reference_output"]
    shares = [{"name": "Ireland", "national_heat_TWh": round(isl_in, 0),
               "output_gwh": ro["Ireland"]["gwh"],
               "flh": ro["Ireland"]["flh"],
               "share_pct": round(100 * ro["Ireland"]["gwh"] / 1000
                                  / isl_in, 2)}]
    for c in comps:
        nat = GEO_NAT_HEAT_TWH[c["name"]]
        o = ro[c["name"]]
        shares.append({"name": c["name"], "national_heat_TWh": nat,
                       "output_gwh": o["gwh"], "flh": o["flh"],
                       "share_pct": round(100 * o["gwh"] / 1000 / nat, 2)})
    pc = g["per_capita_w"]
    # PER JURISDICTION, so the panel can toggle. The contrast between
    # the two is then something the reader finds by switching rather
    # than something the copy asserts at them - which is a stronger way
    # to make it and a less partisan one.
    jur = {}
    for k, cap, del_twh, in_twh in (
            ("roi", g["roi"]["capacity_mwth"], roi_del,
             a["roi"]["residential_heat_twh"]
             + a["roi"]["services_heat_twh"]),
            ("ni", g["ni_capacity_mwth_est"], ni_del,
             a["ni"]["residential_heat_twh"]
             + a["ni"]["services_heat_twh"])):
        wi = 0.20 * del_twh * 1e6 / flh
        jur[k] = {
            "installed_MWth": round(cap, 1),
            "whatif_MWth": round(wi, 0),
            "multiple": round(wi / max(cap, 1e-9), 1),
            "delivered_TWh": round(del_twh, 1),
            "national_heat_TWh": round(in_twh, 1),
            # ROI uses its REPORTED output; NI has none published, so
            # its capacity is converted at Ireland's own load hours and
            # flagged as derived rather than reported.
            "output_gwh": (g["reference_output"]["Ireland"]["gwh"]
                           if k == "roi" else round(cap * flh / 1000, 1)),
            "output_reported": k == "roi",
            "share_pct": round(
                100 * ((g["reference_output"]["Ireland"]["gwh"]
                        if k == "roi" else cap * flh / 1000) / 1000)
                / in_twh, 2),
            "per_person_W": pc[k],
            "units": g["roi"]["units"] if k == "roi" else None,
            "population_m": g["population_m"][k],
        }
    out = {
        "jur": jur,
        "island_gshp_MWth": g["roi"]["capacity_mwth"],
        "island_ni_MWth": g["ni_capacity_mwth_est"],
        "island_total_MWth": round(inst, 1),
        "whatif_MWth": round(whatif, 0),
        "multiple": round(whatif / max(inst, 1e-9), 1),
        "eflh": eflh,
        "flh_ireland": flh,
        "flh_europe_avg": 2420,
        "output_source": g["reference_output_source"],
        "delivered_heat_TWh": round(isl_del, 1),
        "roi_units": g["roi"]["units"],
        "per_person_W": {"roi": pc["roi"], "ni": pc["ni"],
                         "sweden": g["reference_w_pp"]["Sweden"],
                         "france": g["reference_w_pp"]["France"],
                         "netherlands": g["reference_w_pp"]["Netherlands"]},
        "internal_gap": round(pc["roi"] / max(pc["ni"], 1e-9), 0),
        "sales_2025": g["egec_2025"]["ghp_sales_2025"],
        "comparators": comps,
        "share_of_national_heat": {"countries": shares, "whatif_pct": 20.0},
        "register_threshold_kw": GEO_NI_REGISTER_KW,
        "sources": GEO_SOURCES,
        "ni_register_confirmed": len([r for r in g["ni_register"]
                                      if r["confirmed"]]),
        "ni_register_total": len(g["ni_register"]),
        "ni_register_totals": g["ni_register_totals"],
        "ni_domestic": g["ni_domestic"],
    }
    log(f"geo hardware: island {inst:.0f} MWth installed against a "
        f"{whatif:.0f} MWth what-if - {out['multiple']}x. Per person ROI "
        f"{pc['roi']} W, NI {pc['ni']} W, an internal gap of "
        f"{out['internal_gap']:.0f}x; Sweden "
        f"{g['reference_w_pp']['Sweden']} W. Island fleet serves "
        f"{shares[0]['share_pct']}% of its own buildings heat against "
        f"Sweden's "
        f"{[c for c in shares if c['name'] == 'Sweden'][0]['share_pct']}%")
    return out


def derive_geo_percap(anchors=None, geo=None):
    """
    Ground-source Wth per person - installed today vs the capacity the
    hero's 20% what-if implies. Requirement = 20% of annual delivered
    (useful) buildings heat / equivalent full-load hours, per person.
    Pure, unit tested; all sizing parameters dagger.
    """
    a = anchors or ANCHORS
    g = geo or GEO
    eflh = g["eflh_h"]
    pop = g["population_m"]

    def useful_twh(jur):
        j = a[jur]
        heat = j["residential_heat_twh"] + j["services_heat_twh"]
        return heat * sum(sh * a["efficiency"][f]
                          for f, sh in j["fuel_shares"].items())

    def block(jur, current_mwth):
        u = useful_twh(jur)
        need_w = 0.20 * u * 1e12 / eflh          # W of capacity
        p = pop[jur] * 1e6
        return {"current_w_pp": round(current_mwth * 1e6 / p, 1),
                "whatif_w_pp": round(need_w / p, 0),
                "current_mwth": round(current_mwth, 1),
                "whatif_mwth": round(need_w / 1e6, 0),
                "useful_twh": round(u, 1)}

    roi = block("roi", g["roi"]["capacity_mwth"])
    ni = block("ni", g["ni_capacity_mwth_est"])
    pop_i = (pop["roi"] + pop["ni"]) * 1e6
    cur_i = (g["roi"]["capacity_mwth"] + g["ni_capacity_mwth_est"]) * 1e6
    need_i = 0.20 * (roi["useful_twh"] + ni["useful_twh"]) * 1e12 / eflh
    island = {"current_w_pp": round(cur_i / pop_i, 1),
              "whatif_w_pp": round(need_i / pop_i, 0),
              "current_mwth": round(cur_i / 1e6, 1),
              "whatif_mwth": round(need_i / 1e6, 0),
              "useful_twh": round(roi["useful_twh"] + ni["useful_twh"], 1)}
    return {"roi": roi, "ni": ni, "island": island, "eflh_h": eflh,
            "basis": ("20% of delivered buildings heat at "
                      f"{eflh} equivalent full-load hours - all sizing "
                      "parameters dagger; current NI capacity dagger. "
                      "Comparator bars: EGC 2025 country updates, data "
                      "year 2024. Challenge and input welcome at "
                      "contact@causewaygt.com")}


def derive_cool(feeds, anchors=None):
    """
    The cold economy - island cooling loads against the shape of heat
    demand. Flat loads (data centres, refrigeration, process, NI) run
    all year; comfort cooling is shaped by live overheating degree-hours
    (ODH26) when at least a season of the series exists, else treated
    flat with a note. Heat rejection applies per-load factors: vapour-
    compression loads reject compressor work plus the heat they pump.
    With annual totals normalised, the stranded share is the part of
    supply produced while heat demand runs below it. Pure, unit tested.
    """
    a = (anchors or ANCHORS)
    c = a["cool"]
    hddf = feeds.get("hdd") or {}
    hdd = hddf.get("hdd_island") or {}
    days = sorted(hdd)[-365:]
    if len(days) < 200:
        return None
    hs = [hdd[d] for d in days]
    H = sum(hs)
    if H <= 0:
        return None
    n = len(hs)

    loads = dict(c["loads_twh"])
    # DC line = COOLING electricity only (total x cooling share), not
    # the whole data-centre draw - see the correction note in ANCHORS
    dc_total = c["roi_elec_twh"] * c["dc_share_of_roi_elec"]
    loads["dc"] = round(dc_total * c["dc_cooling_share"], 2)
    rf = c["rejection_factor"]
    elec_total = round(sum(loads.values()), 1)
    reject_total = round(sum(loads[k] * rf[k] for k in loads), 1)

    # supply shape: flat loads spread 1/n; comfort follows ODH26
    comfort = loads["comfort"]
    flat = elec_total - comfort
    odh = hddf.get("odh26_island") or {}
    odh_days = [d for d in days if d in odh]
    odh_used = False
    supply = [flat / elec_total / n] * n
    if len(odh_days) >= 60 and sum(odh.get(d, 0.0) for d in days) > 0:
        O = sum(odh.get(d, 0.0) for d in days)
        base = 0.3   # dagger: ventilation floor of the comfort load
        for i, d in enumerate(days):
            shaped = (base / n + (1 - base) * odh.get(d, 0.0) / O)
            supply[i] += comfort / elec_total * shaped
        odh_used = True
    else:
        supply = [s + comfort / elec_total / n for s in supply]

    demand = [h / H for h in hs]
    stranded = sum(max(0.0, s - d) for s, d in zip(supply, demand))

    res_twh = (a["roi"]["residential_heat_twh"]
               + a["ni"]["residential_heat_twh"])
    r1 = lambda x: round(x, 1)
    return {
        "loads_twh": loads,
        "cooling_elec_twh": elec_total,
        "heat_rejected_twh": reject_total,
        "dc_twh": loads["dc"],
        "dc_share_pct": r1(100 * c["dc_share_of_roi_elec"]),
        "dc_share_2028_pct": r1(100 * c["dc_share_2028"]),
        "reject_vs_island_residential_pct": r1(
            100 * reject_total / res_twh),
        "stranded_summer_pct": r1(100 * stranded),
        "comfort_shaped_by_odh": odh_used,
        "dh_share_pct": r1(100 * c["dh_share_of_national_heat"]),
        "basis": ("DC electricity share CSO; other loads and rejection "
                  "factors dagger (refrigeration and comfort reject "
                  "compressor work plus pumped heat). Stranded share "
                  "computed from this site's own degree-day record"
                  + (", comfort load shaped by live overheating "
                     "degree-hours" if odh_used else
                     "; comfort treated flat pending a season of "
                     "overheating degree-hours") +
                  ". Challenge and input welcome at "
                  "contact@causewaygt.com"),
    }


# ------------------------------------------------- DHW vs space mode
# Every route performs WORSE on hot water than on space heating, and
# the site's own anchors say a July week is almost all hot water: hot
# water is a flat 22.4% of annual input while space heat is 77.6%
# HDD-shaped, so the space term collapses in summer and DHW is nearly
# all that is left. Pricing the summer half at an annual efficiency
# would flatter the heat pumps and flatter the boiler at exactly the
# point where the lines converge - which is where the argument is won
# or lost. So each route carries a PAIR.
#
#   oil boiler   space 0.82 / DHW 0.71. SAP 2012 Table 4b, p.207,
#                "Seasonal efficiency for gas and oil boilers", gives a
#                WINTER and a SUMMER figure for every archetype. Across
#                the modern oil stock - standard 1998+ 80/68,
#                condensing 84/72, condensing combi 82/73 - summer runs
#                85 to 89% of winter, mean 0.866. On this site's 0.82
#                winter anchor that is 0.71.
#   gas boiler   space 0.85 / DHW 0.75. Same table: regular condensing
#                84/74, condensing combi 84/75, regular non-condensing
#                74/64; mean ratio 0.879 on 0.85 gives 0.75.
#
# CORRECTED 14 Aug 2026, and the correction is large. These were 0.55
# and 0.68, resting on BRE STP09/B01's remark about "very low hot water
# efficiency (under 40% gross)" - a line describing TWO SPECIFIC oil
# boilers in a methodology paper, which I treated as a fleet floor and
# cited to Simon as one. Table 4b IS the fleet answer, and its worst
# archetype - a single-burner range cooker boiler at 47/37 - is a ratio
# of 0.79, still far above the 0.67 I was using. The summer oil penalty
# is real and roughly half what the panel was showing.
#
# Two things deliberately NOT applied. Table 4b is SAP's fallback for
# boilers absent from the Product Characteristics Database, so it is
# conservative and skewed to older plant; a BER-weighted PCDB figure
# would beat it, and BER records boiler make and model. And Table 4c
# deducts 5 points from BOTH figures where a regular boiler has no
# interlock, which is common in older Irish installations (dagger).
#   ashp         space from the Carnot engine at that day's air
#                temperature; DHW 1.70 - the MCS 031 Issue 4.0 default
#                (18 Mar 2025), cut from 1.75 in line with SAP 10.2.
#   gshp         space 3.24 / DHW 2.24 - Energy Systems Catapult
#                in-situ for space, MCS HPSPE default for DHW.
#   network      space 5.00 / DHW derived at the SAME ratio as GSHP
#                (2.24/3.24 = 0.691), because both are ground-coupled
#                with a constant source: the DHW penalty is the flow
#                temperature lift alone, not a source-side collapse.
#
# NOT SPLIT: EoH did not meter DHW separately, so there is no field
# DHW SPF - the heat-pump DHW figures are MCS design defaults. Said
# plainly rather than dressed as measurement.
# ------------------------------------------------- statutory wedges
# What has to be stripped to get from a retail price to an EX-TAX one
# - product plus network plus margin. Deliberately NOT called
# "wholesale": the EU bulletin's ex-tax oil line is product,
# distribution and margin, not a wholesale quote, and the same is true
# of a gas or electricity unit rate with its tax removed. Calling it
# wholesale would overclaim.
#
# ORDER OF OPERATIONS. VAT is the OUTERMOST layer in both
# jurisdictions - it is charged on the carbon-tax-inclusive price. So
# ex-VAT = retail / (1 + vat), THEN subtract the carbon component.
# Reversing that would understate the wedge.
#
# ROI, from Revenue: kerosene carries no non-carbon excise, a carbon
# component of EUR 160.81 per 1,000 L (16.081 c/L at EUR 63.50/t CO2),
# the NORA levy at 2 c/L, and VAT at 13.5%. Natural gas carries NGCT
# at EUR 11.48/MWh GCV = 1.148 c/kWh and VAT at 9%. Electricity
# carries VAT at 9% and a per-customer PSO levy that is NOT per kWh,
# so it is out of scope for a unit-rate strip and noted rather than
# subtracted.
#
# NI: kerosene is fully duty-rebated to nil with no carbon price; gas
# and electricity carry no carbon price and no Climate Change Levy
# (domestic use is excluded). VAT is 5% on all three. So ex-tax is
# simply retail / 1.05 - exact, not estimated.
#
# THE STEP THAT IS COMING. Ireland's carbon increase on non-propellant
# fuels normally lands on 1 May; for 2026 it was postponed to 14
# October, when the charge moves from EUR 63.50 to EUR 71.00 a tonne.
# That is a discontinuity in the middle of the record, so it is a
# dated table rather than a constant.
# The carbon component scales linearly with the charge per tonne, so
# the whole table derives from one anchored pair: at EUR 63.50/t
# Revenue publishes 16.081 c/L on kerosene and EUR 11.48/MWh GCV =
# 1.148 c/kWh on gas. Checked against the published prior rate - at
# EUR 56.00/t the gas figure comes out 1.012 against Revenue's 1.013,
# which is rounding, not a different method.
#
# WHY IT REACHES BACK TO 2020. Removing the back-look floor took the
# price panel to 122 weeks and 2024-04-15 on its first run, and the
# HDD record keeps growing. A table starting in 2025 would have
# clamped 55 of those weeks to a rate that did not exist yet - the
# 2024-25 charge was EUR 56.00, not EUR 63.50 - and done it silently.
# So the table covers the whole Finance Act 2020 trajectory, and
# anything earlier than its first row is REFUSED rather than clamped.
_CARBON_PER_TONNE = [
    ("2020-05-01", 26.00), ("2021-05-01", 33.50), ("2022-05-01", 41.00),
    ("2023-05-01", 48.50), ("2024-05-01", 56.00), ("2025-05-01", 63.50),
    # Normally 1 May; for 2026 the non-propellant step was postponed
    # to 14 October, so this row is the exception that proves the rule.
    ("2026-10-14", 71.00), ("2027-05-01", 78.50), ("2028-05-01", 86.00),
    ("2029-05-01", 93.50), ("2030-05-01", 100.00),
]
CARBON_STEPS = [(d, r, round(16.081 * r / 63.50, 3),
                 round(1.148 * r / 63.50, 4)) for d, r in _CARBON_PER_TONNE]

# ROI gas and electricity were cut from 13.5% to 9% on 1 May 2022 as a
# cost-of-living measure, extended to 31 Dec 2030. Kerosene never got
# the cut and stays at 13.5% - which is why the tax wedge on oil is
# heavier than on gas inside the same jurisdiction. NI is a flat 5%
# throughout.
VAT_STEPS = {
    "roi": {"oil": [("2000-01-01", 0.135)],
            "gas": [("2000-01-01", 0.135), ("2022-05-01", 0.09)],
            "electricity": [("2000-01-01", 0.135), ("2022-05-01", 0.09)]},
    "ni": {f: [("2000-01-01", 0.05)]
           for f in ("oil", "gas", "electricity")},
}


def vat_for(jur, fuel, date_iso):
    """The VAT rate in force on a date. A constant would have been
    wrong the moment the panel reached back past 1 May 2022."""
    rate = VAT_STEPS[jur][fuel][0][1]
    for frm, r in VAT_STEPS[jur][fuel]:
        if date_iso >= frm:
            rate = r
    return rate


VAT = {"roi": {"oil": 0.135, "gas": 0.09, "electricity": 0.09},
       "ni": {"oil": 0.05, "gas": 0.05, "electricity": 0.05}}
NORA_LEVY_C_PER_L = 2.0


def carbon_for(date_iso):
    """
    The carbon component in force on a date, or None before the table
    starts.

    None rather than a clamp, deliberately. Clamping is how 55 weeks
    got priced at a rate that had not been legislated yet, and it left
    no trace - the row simply looked like the others. A refusal
    propagates to a missing ex-tax figure, which the panel can grey
    out and a reader can see.
    """
    if date_iso < CARBON_STEPS[0][0]:
        return None
    row = CARBON_STEPS[0]
    for r in CARBON_STEPS:
        if date_iso >= r[0]:
            row = r
    return {"eur_per_tonne": row[1], "kerosene_c_per_l": row[2],
            "gas_c_per_kwh": row[3]}


def ex_tax(retail, jur, fuel, date_iso, kwh_per_litre=None):
    """
    Retail -> ex-tax, in the same units in. VAT first, then carbon.

    For oil, `retail` is per litre and the carbon and NORA components
    are per litre. For gas and electricity, per kWh. Returns None if
    the strip would go negative, which would mean a wrong rate rather
    than a cheap fuel.
    """
    net = retail / (1 + vat_for(jur, fuel, date_iso))
    if jur == "roi" and fuel in ("oil", "gas"):
        c = carbon_for(date_iso)
        if c is None:
            return None          # before the table: refuse, don't guess
        net -= (c["kerosene_c_per_l"] + NORA_LEVY_C_PER_L
                if fuel == "oil" else c["gas_c_per_kwh"])
    return round(net, 4) if net > 0 else None


DHW_MODE = {
    "oil_boiler": 0.71,        # SAP Table 4b ratio on the 0.82 anchor
    "gas_boiler": 0.75,        # SAP Table 4b ratio on the 0.85 anchor
    "ashp": 1.70,
    "gshp": 2.24,
    "network": round(5.00 * 2.24 / 3.24, 2),
}
# Worst archetype in Table 4b: single-burner range cooker boiler with
# a permanent pilot, 47 winter / 37 summer. Nothing in the Irish stock
# should price below it.
DHW_MODE_FLOOR_OIL = 0.37

# ZEROED, 14 Aug 2026. The fraction of oil-home hot water made by an
# electric immersion rather than the boiler. The mechanism is real -
# Irish oil households do switch in summer rather than fire a boiler
# at sub-40% for a cylinder, and at COP 1 that is DEARER than the
# inefficient boiler, so it pushed the oil line up rather than down.
#
# But at 0.30 it was adding 5.8 c to the summer oil figure, roughly
# 46% of which was electricity rather than oil, on no metered evidence
# at all - the research found consumer guidance and owner forums and
# nothing measured. It was the largest unevidenced term on the panel
# and it sat under its most striking feature. The UK sibling has no
# equivalent, so carrying it here also broke parity between the two.
#
# Left as a named constant rather than deleted: the code path stays,
# and a metered Irish or NI study of summer immersion use is what
# would turn it back on.
OIL_IMMERSION_DHW_SHARE = 0.0    # dagger - zeroed pending evidence


def sector_blend(jur, fuel, date_iso, anchors=None, nondom=None):
    """
    The price a route actually pays, blended across sectors.

    Every route on this panel was priced at pure DOMESTIC tariffs,
    which is wrong twice over. About a quarter of the island's
    building heat is services rather than residential, and the hero
    bill has always blended that - so the cost panel disagreed with
    the bill on the same page. And a heat NETWORK operator is not a
    household at all: it buys electricity on a commercial contract and
    would never pay a domestic tariff, which is the UK sibling's own
    correction.

    So the blend is not uniform:
      oil            one price - no non-domestic oil rate exists, and
                     kerosene is sold to both sectors on the same terms
      gas, ashp,     blended at the services share of island heat
      gshp           input (~26%), matching the hero bill
      network        100% NON-DOMESTIC, whoever the end customer is

    Non-domestic rates already EXCLUDE VAT by convention (businesses
    recover it), so the blend mixes a VAT-inclusive domestic rate with
    a VAT-exclusive commercial one - which is the point, not an error:
    it is what the heat actually costs to buy.
    """
    a = anchors or ANCHORS
    dom = tariffs_for(date_iso)
    if dom is None:
        return None
    cur = "gbp" if jur == "ni" else "eur"
    d = dom[cur][fuel] * 100
    nd = (nondom or {}).get(cur, {}).get(fuel)
    nd = nd * 100 if nd else None
    if nd is None:
        return {"domestic": d, "nondom": None, "blend": d, "network": d}
    r = sum(a[j]["residential_heat_twh"] for j in ("ni", "roi"))
    sv = sum(a[j]["services_heat_twh"] for j in ("ni", "roi"))
    w = sv / (r + sv) if (r + sv) else 0.0
    return {"domestic": d, "nondom": nd,
            "blend": round((1 - w) * d + w * nd, 4),
            "network": nd, "services_share": round(w, 4)}


# ---------------------------------------------------- the COP engine
# Ported from the UK sibling so the two dashboards compute the electric
# routes the same way. The Irish panel had ONE air-source SPF for the
# whole record, which made every electric line a flat multiple of the
# electricity price - they could never spread apart in cold weather,
# which is the entire point of the chart.
#
# Method, in the UK's order: assert the SPF anchor, then CALIBRATE the
# Carnot fraction so the trailing-year heat-weighted SPF reproduces it.
# That is the right way round. Assuming a fraction and computing an SPF
# (which is what I did by hand for the ROI figure) lets the two drift
# apart; calibrating means the headline number is the anchor by
# construction and the weather only redistributes it.
FLOW_MILD_C, FLOW_COLD_C = 30.0, 50.0      # weather compensation ends
FLOW_MILD_AT, FLOW_COLD_AT = 15.0, -5.0
DHW_FLOW_C = 52.0        # dagger - cylinder flow, year-round
MIN_LIFT_K = 8.0         # dagger - floor on Carnot lift
AIR_APPROACH_C = 3.0     # evaporator approach below air temperature
GROUND_SOURCE_C = 8.0    # 11 C ground less 3 C brine approach
DEFROST_MAX = 0.12       # peak fractional COP loss, centred ~2 C

# THE JURISDICTIONAL SPLIT. Permo-Triassic HSA is an NI play and has no
# onshore ROI equivalent at scale, so the two network routes are
# different machines and cannot share a number.
#   NI  - the UK sibling's blended model: UTES and intermediate-doublet
#         together, source 19.6 C, SPF 5.0.
#   ROI - a 5G ambient loop on seasonal storage alone, charged from
#         comfort cooling and process rejection. Source 16 C is the
#         SEASONAL MEAN of a store charged to 25-30 C in September and
#         depleting through the heating season; measured Dutch ATES
#         recovery of 68-87% is what makes the mean that rather than
#         the charge temperature. SPF 4.0, not the 4.2 I first derived
#         - a distributed ambient loop carries more circulation
#         pumping than the borehole array the fraction came from.
NETWORK_MODEL = {
    "ni": {"source_c": 19.6, "spf": 5.0,
           "note": "UTES and intermediate-doublet blend (Permo-Triassic "
                   "HSA), as the UK sibling"},
    "roi": {"source_c": 16.0, "spf": 4.0,
            "note": "5G ambient loop on seasonal storage, charged from "
                    "comfort cooling and process heat rejection"},
}
# Filled by derive_heat_cost_series and published in the payload, so
# the calibration board on the page shows what the log has always
# said. Module-level because the derivation is deep inside the cost
# series and returning it would change that function's contract.
CALIBRATION = {}

SPF_ANCHORS = {"ashp": 2.80, "gshp": 3.24}

# Hot-water share BY FUEL, from SEAI's residential end-use model (2022):
# oil space 1014.05 / water 299.27 / cooking 1.41; gas space 417.55 /
# water 157.79 / cooking 14.42. So oil is 22.8% water and gas 26.8%.
#
# WHERE THESE APPLY, and where they do NOT. They are the right shares
# for the ENERGY and BILL panels, where the question is how much of the
# oil actually burned went on hot water. They are the WRONG shares for
# the cost panel, where every route answers the same counterfactual -
# if you heated THIS building with X, what would it cost - so the share
# is a property of the building's demand, not of the fuel. Using 22.8%
# on the oil line and 26.8% on the gas line there would compare two
# different demand profiles and quietly break like-for-like. The cost
# panel uses the island share; these sit alongside it for the panels
# that need them.
DHW_SHARE_BY_FUEL = {"oil": 0.228, "gas": 0.268}

# Trailing-year HDD must land in this band or the build fails. Four
# lines, and the UK sibling's own history is the argument: it raised a
# regression window from 365 to 730 days, every quantity that treated
# the window as a year silently doubled, and all four of its test
# suites passed throughout - because they check structure and
# consistency, not whether an annual quantity spans a year. Ireland is
# milder than Britain, so the band sits lower: 1,900-2,900 brackets
# the observed all-island figure with room for a hard or a soft year.
# PROVISIONAL until the first live run reports the real figure - the
# synthetic fixture reads 2,920, which is a made-up year rather than a
# real one. Set this from observed all-island HDD once it is known.
HDD_YEAR_MIN, HDD_YEAR_MAX = 1700, 3000


# How long a single NI oil reading may be held forward. The CCNI
# daily checker publishes Mon-Fri and the weekly archive once a week,
# so 10 days covers a normal weekly cadence plus a missed publication
# without smearing one price across one of the archive's real gaps.
NI_OIL_HOLD_DAYS = 10

# The widest window the cost panel offers, in days. The retained
# series must cover this PLUS a trailing year, because every day needs
# a year of degree days behind it to know its own hot-water share.
WINDOW_MAX_DAYS = 730


def retention_span_gate(hdd_daily, label="island"):
    """
    Assert the retained record can support the widest window offered.

    SERIES_KEEP_DAYS is 1150 against 1095 needed - 55 days of margin,
    which is thin enough that a change to the window, to the trailing
    year, or to the retention constant could short the shaping without
    anything failing. A short record does not throw: it silently
    shapes space heat on a denominator that is a season rather than a
    year, and every figure downstream stays plausible. So this gate
    exists to make that change loud at the moment it is made rather
    than at the moment someone notices the chart.

    Returns the retained span in days, or None if it cannot be
    measured.
    """
    days = sorted(hdd_daily)
    if len(days) < 2:
        return None
    span = ((dt.date.fromisoformat(days[-1])
             - dt.date.fromisoformat(days[0])).days + 1)
    need = WINDOW_MAX_DAYS + 365
    if span < need:
        log(f"retention: WARNING {label} record spans {span} days, "
            f"{need} needed for a {WINDOW_MAX_DAYS}-day window plus its "
            f"trailing year - the widest window will under-fill and the "
            f"earliest days in it are shaped on a partial year. Raise "
            f"SERIES_KEEP_DAYS (currently {SERIES_KEEP_DAYS}) or narrow "
            f"WINDOW_MAX_DAYS")
    else:
        log(f"retention: {label} record spans {span} days, {need} needed "
            f"({span - need} days of margin)")
    return span


def hdd_year_gate(hdd_daily, label="island"):
    """Assert the trailing year of degree days is a plausible year."""
    days = sorted(hdd_daily)[-365:]
    if len(days) < 330:
        return None
    total = sum(hdd_daily[d] for d in days)
    ok = HDD_YEAR_MIN <= total <= HDD_YEAR_MAX
    log(f"hdd: trailing-year {label} HDD {total:.0f} "
        f"(gate {HDD_YEAR_MIN}-{HDD_YEAR_MAX}) - "
        + ("OK" if ok else "OUTSIDE THE GATE - an annual quantity is "
                          "probably not spanning a year"))
    return total


def scan_hdd_base(temp_daily, gas_daily, bases=(14.5, 15.5, 16.5)):
    """
    Which degree-day base fits Irish gas demand best?

    The base is fixed at 15.5 here and scanned in the UK sibling, which
    lands on 16.5 - so it is not safe to assume the Irish stock gives
    the same answer. Log-only: this reports the fit at each base and
    changes nothing, because moving the base moves every HDD-shaped
    figure on the site and that is a decision, not a tuning.
    """
    days = sorted(set(temp_daily) & set(gas_daily))
    if len(days) < 300:
        return None
    out = {}
    for b in bases:
        xs = [max(0.0, b - temp_daily[d]) for d in days]
        ys = [gas_daily[d] for d in days]
        n = len(xs)
        mx, my = sum(xs) / n, sum(ys) / n
        sxx = sum((x - mx) ** 2 for x in xs)
        sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
        if sxx <= 0:
            continue
        beta = sxy / sxx
        alpha = my - beta * mx
        ss_res = sum((y - (alpha + beta * x)) ** 2 for x, y in zip(xs, ys))
        ss_tot = sum((y - my) ** 2 for y in ys)
        out[b] = round(1 - ss_res / ss_tot, 4) if ss_tot > 0 else 0.0
    if out:
        best = max(out, key=out.get)
        log(f"hdd: base scan on {len(days)} days - "
            + ", ".join(f"{b} R2 {r}" for b, r in sorted(out.items()))
            + f"; best {best}, in use {HDD_BASE_C}"
            + ("" if best == HDD_BASE_C else
               " - MOVING IT WOULD MOVE EVERY HDD-SHAPED FIGURE, so it "
               "is a decision rather than a tuning"))
    return out


def flow_temp(t_out):
    """Weather-compensated SPACE flow. Hot water does not follow it."""
    f = min(1.0, max(0.0, (FLOW_MILD_AT - t_out)
                     / (FLOW_MILD_AT - FLOW_COLD_AT)))
    return FLOW_MILD_C + (FLOW_COLD_C - FLOW_MILD_C) * f


def defrost_factor(t_out):
    """Fraction of COP retained. Bell-shaped loss centred on 2 C, the
    humid frost band. Shape is a dagger; the level is re-anchored by
    the eta calibration afterwards, so only the shape matters here."""
    return 1.0 - DEFROST_MAX * math.exp(-((t_out - 2.0) / 3.0) ** 2)


def carnot_cop(t_flow, t_source):
    return (t_flow + 273.15) / max(MIN_LIFT_K, t_flow - t_source)


def route_cop(route, t_out, eta, jur="roi", dhw=False):
    """Point COP for one route on one day. `dhw` prices the hot-water
    leg at the cylinder flow rather than the compensated space flow -
    without that split, a mild summer day gives absurd COPs on a load
    that is entirely hot water."""
    tf = DHW_FLOW_C if dhw else flow_temp(t_out)
    if route == "ashp":
        return max(1.0, eta * carnot_cop(tf, t_out - AIR_APPROACH_C)
                   * defrost_factor(t_out))
    if route == "gshp":
        return max(1.0, eta * carnot_cop(tf, GROUND_SOURCE_C))
    if route == "network":
        return max(1.0, eta * carnot_cop(tf, NETWORK_MODEL[jur]["source_c"]))
    raise ValueError(route)


def calibrate_eta(route, days, jur="roi", anchors=None):
    """
    Solve the Carnot fraction so the HEAT-WEIGHTED SPF over the days
    given reproduces the route's anchor.

    Heat-weighted, not day-averaged: almost all the heat is drawn in
    the cold, so a mean over days would flatter every route. `days` is
    [(t_out, space_kwh, dhw_kwh), ...].
    """
    a = anchors or ANCHORS
    target = (NETWORK_MODEL[jur]["spf"] if route == "network"
              else SPF_ANCHORS[route])

    def spf(eta):
        heat = elec = 0.0
        for t, sp, dhw in days:
            if sp > 0:
                elec += sp / route_cop(route, t, eta, jur, dhw=False)
            if dhw > 0:
                elec += dhw / route_cop(route, t, eta, jur, dhw=True)
            heat += sp + dhw
        return (heat / elec) if elec > 0 else 0.0

    lo, hi = 0.05, 1.20
    for _ in range(60):
        mid = (lo + hi) / 2
        if spf(mid) < target:
            lo = mid
        else:
            hi = mid
    return round((lo + hi) / 2, 4)


def route_cost_useful(prices, ashp_spf, dhw_share, anchors=None):
    """
    Cost per useful kWh by route, in native minor units, with hot
    water and space heat priced in their own modes and blended at the
    week's own DHW share.

    prices: {"oil_per_kwh", "gas_per_kwh", "elec_per_kwh"} in minor
    units per kWh of INPUT. dhw_share: 0-1, the fraction of delivered
    heat that is hot water this week. Pure, unit tested.
    """
    a = anchors or ANCHORS
    w = min(1.0, max(0.0, dhw_share))
    oil, gas, elec = (prices["oil_per_kwh"], prices["gas_per_kwh"],
                      prices["elec_per_kwh"])
    # The network route buys commercially; the heat-pump routes are a
    # domestic/services blend. Falls back to the blended price when no
    # non-domestic rate is supplied, so callers need not change.
    elec_net = prices.get("elec_network_per_kwh", elec)

    cops = prices.get("cops")           # {route: (space_cop, dhw_cop)}

    def blend(space_perf, dhw_perf, price):
        return (1 - w) * price / space_perf + w * price / dhw_perf

    def blend_route(route, price, sp_fb, dhw_fb):
        """Per-day COPs when supplied, the flat SPF otherwise. The
        fallback is passed as its two performances rather than as a
        computed number, so it is not evaluated when the COPs are
        present - ashp_spf is legitimately None on the daily path."""
        c = (cops or {}).get(route)
        return blend(*(c if c else (sp_fb, dhw_fb)), price)

    # Oil hot water is part boiler, part immersion - and the immersion
    # runs on electricity at COP 1, so the leakage RAISES the cost
    # rather than lowering it.
    oil_dhw = ((1 - OIL_IMMERSION_DHW_SHARE) * oil / DHW_MODE["oil_boiler"]
               + OIL_IMMERSION_DHW_SHARE * elec / 1.0)
    return {
        "oil_boiler": round((1 - w) * oil / a["efficiency"]["oil"]
                            + w * oil_dhw, 2),
        "gas_boiler": round(blend(a["efficiency"]["gas"],
                                  DHW_MODE["gas_boiler"], gas), 2),
        "ashp": round(blend_route("ashp", elec, ashp_spf,
                                  DHW_MODE["ashp"]), 2),
        "gshp": round(blend_route("gshp", elec, GSHP_SPF,
                                  DHW_MODE["gshp"]), 2),
        "network": round(blend_route("network", elec_net, GEO_NETWORK_SCOP,
                                     DHW_MODE["network"]), 2),
        "dhw_share": round(w, 3),
    }


def week_dhw_share(hdd_daily, w_end, anchors=None):
    """
    What fraction of a week's delivered heat is hot water.

    The site's own shaping, one step further: hot water is flat across
    the year, space heat follows the week's share of the trailing
    year's heating degree days. So the DHW share is high in July and
    low in January, and it is computed rather than assumed.
    """
    a = anchors or ANCHORS
    days = sorted(d for d in hdd_daily if d <= w_end)
    if len(days) < 300:
        return None
    wk = [d for d in days if d > (dt.date.fromisoformat(w_end)
                                  - dt.timedelta(days=7)).isoformat()]
    year = days[-365:]
    hdd_wk = sum(hdd_daily[d] for d in wk)
    hdd_yr = sum(hdd_daily[d] for d in year)
    if hdd_yr <= 0 or len(wk) < 5:
        return None
    dhw = (1 - a["space_heat_fraction"]) / 52.0
    space = a["space_heat_fraction"] * hdd_wk / hdd_yr
    tot = dhw + space
    return (dhw / tot) if tot > 0 else None


def day_dhw_share(hdd_daily, day, anchors=None):
    """Fraction of a DAY's delivered heat that is hot water. Same
    shaping as the weekly version, one step finer."""
    a = anchors or ANCHORS
    days = sorted(d for d in hdd_daily if d <= day)
    # Needs a trailing year to know what share of the year's degree
    # days this one carries. Below ~200 the denominator is a season
    # rather than a year and the share is meaningless.
    if len(days) < 200:
        return None
    year = days[-365:]
    hdd_yr = sum(hdd_daily[d] for d in year)
    if hdd_yr <= 0:
        return None
    dhw = (1 - a["space_heat_fraction"]) / 365.0
    space = a["space_heat_fraction"] * hdd_daily.get(day, 0.0) / hdd_yr
    tot = dhw + space
    return (dhw / tot) if tot > 0 else None


def useful_heat_gwh_year(jur, anchors=None):
    """
    A jurisdiction's annual DELIVERED heat, GWh.

    The sector anchors are fuel INPUT and each fuel burns at its own
    efficiency, so they are converted before they can stand beside a
    price per MWh of delivered heat. Same conversion as
    hourly_heat_mw() uses - one definition of delivered heat on the
    site, not two.
    """
    a = anchors or ANCHORS
    j = a[jur]
    heat = j["residential_heat_twh"] + j["services_heat_twh"]
    return heat * sum(sh * a["efficiency"][f]
                      for f, sh in j["fuel_shares"].items()) * 1000.0


def day_delivered_heat(hdd_daily, day, jur, anchors=None):
    """
    GWh of delivered heat on one day, split space heat / hot water.

    The QUANTITY the cost panel's axis is charged on. Shaping is
    identical to day_dhw_share - hot water flat across the year, space
    heat by the day's share of the trailing year's degree days - so
    dhw / (space + dhw) reproduces that share exactly rather than
    approximately, and a test pins it.

    Shape is island-wide (one HDD series), scale is jurisdictional.
    Returns None on the same short-record guard as the share.
    """
    a = anchors or ANCHORS
    days = sorted(d for d in hdd_daily if d <= day)
    if len(days) < 200:
        return None
    year = days[-365:]
    hdd_yr = sum(hdd_daily[d] for d in year)
    if hdd_yr <= 0:
        return None
    u = useful_heat_gwh_year(jur, a)
    dhw = u * (1 - a["space_heat_fraction"]) / 365.0
    space = u * a["space_heat_fraction"] * hdd_daily.get(day, 0.0) / hdd_yr
    return {"space": round(space, 3), "dhw": round(dhw, 3)}


def derive_heat_cost_series(feeds, anchors=None):
    """
    DAILY cost of a useful kWh by route - the Irish equivalent of the
    UK sibling's panel, with oil as the main series.

    Daily, not weekly, because the electric routes are priced at each
    day's own COP and the whole argument is what happens on the cold
    days rather than the average of the days. The oil price is the
    only weekly input (the EU bulletin publishes weekly), so it is
    STEP-HELD across the week rather than interpolated - a price that
    did not move should not appear to.
    """
    a = anchors or ANCHORS
    hddf = feeds.get("hdd") or {}
    hdd_i = hddf.get("hdd_island") or {}
    temp = {"roi": hddf.get("temp_roi") or {}, "ni": hddf.get("temp_ni") or {}}
    if len(hdd_i) < 300 or not temp["roi"]:
        return None
    bull = ((feeds.get("oil_bulletin") or {})
            .get("roi_heating_gasoil_eur_per_1000l") or {})
    bull_nt = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily", {}).get("900l") or {})
    if not bull:
        return None
    kwh_l = a["kerosene_kwh_per_litre"]

    # Calibrate once, over the trailing year, per jurisdiction.
    etas = {}
    for jur in ("roi", "ni"):
        t = temp[jur]
        cal = []
        for d in sorted(t)[-365:]:
            v = day_delivered_heat(hdd_i, d, jur, a)
            if v is None:
                continue
            cal.append((t[d], v["space"], v["dhw"]))
        # A full year is what you want - the fraction is calibrated
        # against the whole seasonal swing - but a short record should
        # degrade rather than return nothing, and say so.
        if len(cal) < 150:
            log(f"heat cost: cannot calibrate {jur} - only {len(cal)} "
                "days carry a hot-water share; needs 150")
            return None
        if len(cal) < 330:
            log(f"heat cost: WARNING calibrating {jur} on {len(cal)} days, "
                "not a full year - the fraction is biased toward whichever "
                "season the record covers")
        etas[jur] = {r: calibrate_eta(r, cal, jur, a)
                     for r in ("ashp", "gshp", "network")}
    hdd_year_gate(hdd_i)
    retention_span_gate(hdd_i)
    try:
        gni = ((feeds.get("gni_live") or {}).get("ndm_gwh")
               or (feeds.get("gni_live") or {}).get("total_gwh") or {})
        if gni and hddf.get("temp_roi"):
            scan_hdd_base(hddf["temp_roi"], gni)
    except Exception as exc:
        log(f"hdd: base scan skipped ({exc.__class__.__name__})")
    log("heat cost: calibrated Carnot fractions "
        + "; ".join(f"{j} " + ", ".join(f"{r} {v}" for r, v in e.items())
                    for j, e in etas.items())
        + f" (anchors ashp {SPF_ANCHORS['ashp']}, gshp "
          f"{SPF_ANCHORS['gshp']}, network ni "
          f"{NETWORK_MODEL['ni']['spf']} / roi "
          f"{NETWORK_MODEL['roi']['spf']})")
    spread = max(v for e in etas.values() for v in e.values()) / \
        min(v for e in etas.values() for v in e.values())
    # Published, not just logged. The calibration is the answer to
    # "how do you know the COP model is right", so it belongs on the
    # page rather than in a run log nobody sees. Anchors travel with
    # the fractions because the fraction alone means nothing without
    # the SPF it was solved to reproduce.
    CALIBRATION.clear()
    CALIBRATION.update({
        "gate": 1.15,
        "spread": round(spread, 4),
        "jurisdictions": {
            j: {r: {"eta": e[r],
                    "spf_anchor": (NETWORK_MODEL[j]["spf"] if r == "network"
                                   else SPF_ANCHORS[r]),
                    "source_c": (NETWORK_MODEL[j]["source_c"]
                                 if r == "network" else
                                 (GROUND_SOURCE_C if r == "gshp" else None))}
                for r in ("ashp", "gshp", "network")}
            for j, e in etas.items()},
    })
    if spread > 1.15:
        log(f"heat cost: WARNING calibrated fractions spread {spread:.2f}x "
            "- more than 15% apart means a source temperature and an SPF "
            "anchor that do not describe the same machine")

    def cops_for(jur, t_out):
        e = etas[jur]
        return {r: (route_cop(r, t_out, e[r], jur, dhw=False),
                    route_cop(r, t_out, e[r], jur, dhw=True))
                for r in ("ashp", "gshp", "network")}

    # THE MODE REGIME IS SEASONAL, NOT DAILY. A boiler in January does
    # not drop to summer cycling efficiency because one day was mild -
    # it is still running its space-heating circuit. The sub-40% hot
    # water efficiency arises from a REGIME: the boiler off for space
    # heat, cycling for a small cylinder load. So the mode blend runs
    # on a trailing 28-day share while the heat-pump COPs stay on the
    # day's own temperature, which is instantaneous physics.
    #
    # Before this the oil line sawtoothed 15 to 28 c between adjacent
    # days, which is not a price signal, it is a shaping artefact.
    MODE_SMOOTH_DAYS = 28
    shares = {}
    for d in sorted(temp["roi"]):
        sh = day_dhw_share(hdd_i, d, a)
        if sh is not None:
            shares[d] = sh
    sm_days = sorted(shares)
    smooth = {}
    for i, d in enumerate(sm_days):
        w = [shares[x] for x in sm_days[max(0, i - MODE_SMOOTH_DAYS + 1):i + 1]]
        smooth[d] = sum(w) / len(w)

    weeks = sorted(bull)
    # Sorted once: the NI step-hold below scans it per day.
    ccni_days = sorted(ccni)
    out, unpriced = [], 0
    for day in sorted(temp["roi"]):
        if day < weeks[0]:
            continue
        share = day_dhw_share(hdd_i, day, a)
        if share is None:
            continue
        t = tariffs_for(day)
        if t is None:
            unpriced += 1
            continue
        # step-held: the most recent bulletin week at or before today
        wk = max((w for w in weeks if w <= day), default=None)
        if wk is None:
            continue
        nd, _ = nondom_for(day, ((feeds.get("ecb_fx") or {})
                                 .get("eur_gbp_semester")))
        if nd is None:
            # Before the first published REMM semester. The services
            # share of the gas and heat-pump routes, and the whole of
            # the network route, have no anchor - so the day is not
            # priced at all rather than priced on a borrowed one.
            unpriced += 1
            continue
        # daily share for the caption, smoothed share for the money
        mode = smooth.get(day, share)
        row = {"day": day, "dhw_share": round(share, 3),
               "dhw_mode": round(mode, 3),
               "t_roi": temp["roi"].get(day)}
        # The quantity behind the price. Emitted for BOTH jurisdictions
        # unconditionally - it depends on degree days and the sector
        # anchors, not on whether that day carries an NI oil price, so
        # it does not belong inside the NI branch below.
        for jur in ("roi", "ni"):
            v = day_delivered_heat(hdd_i, day, jur, a)
            if v:
                row["vol_" + jur] = v
        oil_c_l = bull[wk] / 10.0
        oil_ex = (bull_nt[wk] / 10.0 if wk in bull_nt
                  else ex_tax(oil_c_l, "roi", "oil", day))
        gb = sector_blend("roi", "gas", day, a, nd)
        eb = sector_blend("roi", "electricity", day, a, nd)
        cr = cops_for("roi", temp["roi"][day])
        pr = {"oil_per_kwh": oil_c_l / kwh_l, "gas_per_kwh": gb["blend"],
              "elec_per_kwh": eb["blend"],
              "elec_network_per_kwh": eb["network"], "cops": cr}
        # THREE SERVICES, not one. Space heating and hot water are
        # different questions and the answers diverge: on space heat a
        # heat pump rides the weather-compensated flow down to 30 C,
        # on hot water the cylinder pins it at 52 C whatever the
        # weather. The geothermal advantage is STEADY on hot water and
        # SWINGING on space heat - and it is air source whose case
        # collapses on the cylinder, not the ground-coupled routes.
        # "As delivered" is what a household actually pays: the two
        # blended at the season's own share.
        row["roi"] = route_cost_useful(pr, None, mode, a)
        row["roi_space"] = route_cost_useful(pr, None, 0.0, a)
        row["roi_dhw"] = route_cost_useful(pr, None, 1.0, a)
        gd_x = ex_tax(gb["domestic"], "roi", "gas", day)
        en_x = ex_tax(eb["domestic"], "roi", "electricity", day)
        c = carbon_for(day)
        gn_x = (gb["nondom"] - c["gas_c_per_kwh"]) if c and gb["nondom"] else None
        w_sv = gb.get("services_share", 0.0)
        if oil_ex and gd_x and en_x and gn_x:
            row["roi_ex_tax"] = route_cost_useful(
                {"oil_per_kwh": oil_ex / kwh_l,
                 "gas_per_kwh": (1 - w_sv) * gd_x + w_sv * gn_x,
                 "elec_per_kwh": (1 - w_sv) * en_x + w_sv * eb["nondom"],
                 "elec_network_per_kwh": eb["nondom"], "cops": cr},
                None, mode, a)
        # STEP-HELD, exactly as the ROI bulletin week is above. The
        # NI series is daily only back to the CCNI daily checker's
        # start; behind that it is the weekly archive, one reading a
        # week. Reading `day in ccni` priced one day in seven and left
        # the NI line dotted - 144 of 375 days on the 15 Aug run
        # against ROI's 375 - which looks like missing data rather
        # than a weekly survey. Holding the reading forward is the
        # same treatment, and the same claim, as ROI oil.
        #
        # CAPPED at NI_OIL_HOLD_DAYS. The archive has real gaps - a
        # 26-day one after its first row, a 21-day one in 2023 - and
        # smearing one reading across a month would invent a flat
        # price rather than admit a hole.
        ni_day = max((d for d in ccni_days if d <= day), default=None)
        if ni_day is not None \
                and (dt.date.fromisoformat(day)
                     - dt.date.fromisoformat(ni_day)).days \
                <= NI_OIL_HOLD_DAYS \
                and temp["ni"].get(day) is not None:
            ppl = ccni[ni_day] * 100 / 900
            gbn = sector_blend("ni", "gas", day, a, nd)
            ebn = sector_blend("ni", "electricity", day, a, nd)
            cn = cops_for("ni", temp["ni"][day])
            row["t_ni"] = temp["ni"][day]
            pn = {"oil_per_kwh": ppl / kwh_l, "gas_per_kwh": gbn["blend"],
                  "elec_per_kwh": ebn["blend"],
                  "elec_network_per_kwh": ebn["network"], "cops": cn}
            row["ni"] = route_cost_useful(pn, None, mode, a)
            row["ni_space"] = route_cost_useful(pn, None, 0.0, a)
            row["ni_dhw"] = route_cost_useful(pn, None, 1.0, a)
            row["ni_ex_tax"] = route_cost_useful(
                {"oil_per_kwh": ex_tax(ppl, "ni", "oil", day) / kwh_l,
                 "gas_per_kwh": ((1 - w_sv) * ex_tax(gbn["domestic"], "ni",
                                                     "gas", day)
                                 + w_sv * gbn["nondom"]),
                 "elec_per_kwh": ((1 - w_sv) * ex_tax(ebn["domestic"], "ni",
                                                      "electricity", day)
                                  + w_sv * ebn["nondom"]),
                 "elec_network_per_kwh": ebn["nondom"], "cops": cn},
                None, mode, a)
        out.append(row)
    if not out:
        return None
    last = out[-1]
    for j in ("roi", "ni"):
        if j in last:
            r = last[j]
            log(f"heat cost ({j} {last['day']}, {last.get('t_' + j)} C, "
                f"DHW {last['dhw_share']}): oil {r['oil_boiler']} / gas "
                f"{r['gas_boiler']} / ashp {r['ashp']} / gshp {r['gshp']} "
                f"/ network {r['network']} per useful kWh")
    if unpriced:
        log(f"heat cost: {unpriced} day(s) NOT priced - before the tariff "
            f"table starts ({TARIFF_HISTORY[0][0]})")
    for j in ("roi", "ni"):
        v = last.get("vol_" + j)
        if v:
            log(f"heat cost: delivered heat {j} {last['day']} "
                f"{v['space'] + v['dhw']:.1f} GWh "
                f"(space {v['space']:.1f} / hot water {v['dhw']:.1f}); "
                f"annual anchor {useful_heat_gwh_year(j, a):.0f} GWh")
    log(f"heat cost: {len(out)} days priced "
        f"({sum(1 for r in out if 'ni' in r)} with NI), "
        f"{out[0]['day']}..{out[-1]['day']}; oil price step-held weekly")
    return out


DD_PATH = ROOT / "docs" / "dispatch_down_monthly.json"
# Wind dispatch-down, monthly, by jurisdiction and reason code, from
# EirGrid's own half-hourly DD files (DD-HH-<year>.xlsx on
# eirgrid.ie/grid/system-and-renewable-data-reports). Shipped as a
# STATIC file rather than fetched: the half-hourly downloads sit behind
# a JavaScript accordion and carry version suffixes that change without
# notice (V7, v10), so a guessed URL would rot silently. Closed years
# never change. The current year is refreshed by re-running
# tools/dd_convert.py against a freshly downloaded file.
#
# Wind only. Solar coverage does not start until 2023 and solar is a
# tenth of the volume, so including it would mean a series whose
# denominator changes shape midway.
DD_REASONS = [
    ("trans", "Transmission constraint", "constraint"),
    ("test", "TSO testing", "constraint"),
    ("hifrq", "High frequency / minimum generation", "curtailment"),
    ("snsp", "SNSP limit", "curtailment"),
    ("rocof", "RoCoF / inertia", "curtailment"),
    ("other", "Other reductions", "other"),
]


def derive_dispatch_down(anchors=None):
    """
    Wind dispatch-down by month, jurisdiction and reason, plus the heat
    each spilled GWh could have made by route.

    AN ENERGY-SCALE STATEMENT, NOT A DISPATCH CLAIM. It is the heat that
    volume of electricity could have produced, not heat the system would
    have delivered: it takes no account of whether the spill coincided
    with heat demand, and adding a large flexible load would itself
    change the dispatch. The seasonal SPFs are used rather than an
    hourly COP because the hourly store holds 13 months and this series
    runs to 2021 - the hourly refinement is available only for the last
    year and would make the series inconsistent with itself.
    """
    a = anchors or ANCHORS
    if not DD_PATH.exists():
        log("dispatch down: no monthly file - panel will decline")
        return None
    d = json.loads(DD_PATH.read_text())
    months = d["months"]
    spf = {"ashp": SPF_ANCHORS["ashp"], "gshp": SPF_ANCHORS["gshp"],
           "network": round((a["ni"]["residential_heat_twh"]
                             + a["ni"]["services_heat_twh"]
                             + a["roi"]["residential_heat_twh"]
                             + a["roi"]["services_heat_twh"]) /
                            ((a["ni"]["residential_heat_twh"]
                              + a["ni"]["services_heat_twh"])
                             / NETWORK_MODEL["ni"]["spf"]
                             + (a["roi"]["residential_heat_twh"]
                                + a["roi"]["services_heat_twh"])
                             / NETWORK_MODEL["roi"]["spf"]), 3)}
    out = {"months": months, "unit": "GWh", "technology": "Wind",
           "reasons": [{"key": k, "label": lab, "group": g}
                       for k, lab, g in DD_REASONS],
           "spf": spf, "jurisdictions": {}}
    if d.get("price_month_mean"):
        out["price_month_mean"] = d["price_month_mean"]
        out["price_unit"] = "EUR/MWh"
    for j, block in d["jurisdictions"].items():
        heat = {r: [round(v * s, 1) for v in block["dd"]]
                for r, s in spf.items()}
        rate = [round(100 * dd / av, 2) if av else None
                for dd, av in zip(block["dd"], block["avail"])]
        # WHAT THE SPILLED ENERGY WAS WORTH: volume in GWh times the
        # price in the half-hours it was actually spilling. Not a
        # payment, and not the constraint payment - who captures it
        # depends on the arrangement, and differs by reason. Constrained
        # wind with firm access is already compensated, so absorbing it
        # saves the system operator and consumers; curtailed wind is
        # not, so absorbing it is revenue the generator keeps.
        val = {}
        for k in ("dd", "cons", "curt"):
            pk = block.get("price_" + k)
            if not pk:
                continue
            val[k] = [round(v * p / 1000.0, 2) if (p and v) else 0.0
                      for v, p in zip(block[k], pk)]
        # the same volume valued at the month's plain average, which is
        # the naive figure this panel exists to correct
        naive = [round(v * p / 1000.0, 2) if p else 0.0
                 for v, p in zip(block["dd"], d.get("price_month_mean")
                                 or [None] * len(months))]
        out["jurisdictions"][j] = dict(block, rate_pct=rate, heat=heat,
                                       value_eur_m=val,
                                       value_naive_eur_m=naive)
    for j, b in out["jurisdictions"].items():
        tot = sum(b["dd"])
        line = (f"dispatch down: {j} wind {tot:.0f} GWh over {len(months)} "
                f"months, {100 * sum(b['cons']) / max(tot, 1):.0f}% "
                f"constraint; at network SPF {spf['network']} that is "
                f"{sum(b['heat']['network']) / 1000:.1f} TWh of heat")
        if b.get("value_eur_m"):
            v = sum(b["value_eur_m"]["dd"])
            n = sum(b["value_naive_eur_m"])
            line += (f"; worth EUR {v:.0f}m at the prices of its own "
                     f"hours against EUR {n:.0f}m at monthly averages "
                     f"({100 * v / max(n, 1):.0f}%)")
        log(line)
    return out


# Worked examples of absorbing dispatched-down wind. WORKED EXAMPLES,
# not measurements - a different kind of claim from every other panel
# on the site, and labelled as such. They size the SINK ("what scale of
# load would it take") rather than claim a benefit, because sizing is
# robust to the coincidence objection and a benefit claim is not: half
# the spill lands outside the heating season and only 44% of it in the
# small hours, so nothing here should be read as heat delivered.
#
# The domestic figures are Agbonaye, Keatley, Huang, Odiase & Hewitt
# (2022), Renewable Energy 190:487-500, doi 10.1016/j.renene.2022.03.131
# - same jurisdiction, four SONI constraint groups, 2019 dispatch-down,
# spatial and hourly. Quoted as the paper's own results, not re-derived.
ODD_HOSPITAL = {
    # ERIC 2024/25, 1,104 English acute sites, mean 211 kWh/m2 - TOTAL
    # energy, so the heat share is ours and daggered. NHS Scotland and
    # NI do not publish an equivalent series.
    "eui_kwh_m2": 211,
    "heat_share": 0.60,
    "floor_m2": 55000,
    "source": "ERIC 2024/25 acute mean, heat share \u2020",
}
AGBONAYE = {
    "subscribers": 250000,
    "constraint_cut_pct": 67,
    "curtailment_cut_pct": 74,
    "household_saving_gbp": 220,
    "farm_10mw_gbp": 19400,
    "operator_saving_pct": 78,
    "cite": "Agbonaye, Keatley, Huang, Odiase & Hewitt (2022), "
            "Renewable Energy 190:487\u2013500",
}


# Irish cooling. SEAI Comprehensive Assessment Technical Annex (ERM for
# SEAI, May 2026), Figure 7, 2023 base.
#
# THE BAR IS COOLING SERVICE THROUGHOUT. That took work to establish.
# SEAI's own figures are a MIXTURE of service and electricity, and the
# 2019 National Heat Study is mixed differently from the 2025 Annex:
#
#   sector        2019 study      2025 Annex
#   commercial    service (2.07)  service (2.08)
#   public        service (2.58)  service (2.50)
#   industry      ELECTRICITY     ELECTRICITY  (ratio exactly 1.00)
#   agriculture   ELECTRICITY     not listed
#   data centres  ELECTRICITY     service (2.00)
#
# Verified by back-calculating each sector's electricity from its
# emissions bar: commercial, public, industry and agriculture all imply
# 317-324 gCO2/kWh against the final-energy column, which is the 2019
# grid intensity - four sectors agreeing to 2% confirms the emissions
# and energy tables are one model. Test the same emissions against
# FIGURE 8 instead and the implied intensity runs 123 to 333, which is
# how we know Figure 8 is not one quantity.
#
# SEAI says why industry is unconverted: the boundary "between passive
# heat loss and active cooling requiring energy input is unclear", so
# cooling is "represented as electricity consumption for cooling rather
# than total thermal energy removed".
#
# So ONE judgement is needed, not three: an EER for industry.
# Both panels read ANCHORS["cool"]["cooling_service_factor"], where
# each figure is documented. Mirrored here as named constants so the
# cooling derivation reads plainly, with an assert below that fails the
# build if the two ever drift apart.
COOL_INDUSTRY_EER = ANCHORS["cool"]["cooling_service_factor"]["process"]
# DATA CENTRES: an EFFECTIVE EER of ~6, not SEAI's borrowed 2.0.
#
# SEAI's own report says why 2.0 is wrong: for data centres it reports
# ELECTRICITY for cooling rather than cooling delivered, "because there
# is little publicly available information regarding the efficiency of
# cooling techniques used in data centres". It never models free
# cooling as a mechanism at all - its 0.4 TWh is PUE arithmetic, total
# electricity times overhead, of which 70% assumed to be cooling. The
# 2025 Annex then divides that by the COMMERCIAL sector's average,
# which it admits is borrowed.
#
# But the heat removed from the white space is essentially everything
# except the cooling plant's own draw. Two routes agree:
#   6.4 TWh electricity, 14% on cooling -> 5.5 TWh removed / 0.9 = 6.1
#   IT load at PUE 1.15-1.25            -> 5.1-5.6 / 0.9 = 5.7-6.2
# In Ireland's climate that is what free cooling looks like in a
# number: nearly all the heat leaves through economisers and dry
# coolers with only fan and pump work behind it. Uptime Institute puts
# a closed-loop adiabatic site in a cool climate on free cooling 90-95%
# of the year; Microsoft reports mechanical cooling under 2% of the
# year in Ireland; Digital Realty's Profile Park runs with no
# compressor cooling at all.
COOL_DC_EFFECTIVE_EER = ANCHORS["cool"]["cooling_service_factor"]["dc"]
# WHICH IS WHY DATA CENTRES ARE EXCLUDED FROM THE GEOTHERMAL WHAT-IF.
# Ground cooling cannot beat free air in this climate: the competitor
# is not a running compressor, it is a fan moving 10 degC air. Nothing
# is displaced, so nothing is claimed. Retail refrigeration, industrial
# process cooling and comfort cooling in offices are different - free
# cooling is a data centre design, not a supermarket one - and they
# stay in.
COOL_WHATIF_EXCLUDE = ("datacentres",)
# 3.0, bracketed by SEAI's own internal anchors rather than by an
# outside source: commercial 2.08 and public 2.50 are SEAI's, and
# process chillers run at steadier load and higher utilisation than a
# commercial fleet full of part-load and older plant, so above both.
# The upper marker is Barth et al. (2025), which quantifies Manhattan
# at 10.0 TWh of cooling on 2.82 TWh of electricity - an EER of 3.5 on
# explicit assumptions.
#
# Data centres are NOT re-judged. The 2025 Annex already carries them
# at 2.00, and overriding a published service figure with our own EER
# would be a larger claim than converting a sector SEAI declined to
# convert at all. We think 2.00 is low for an Irish fleet where free
# cooling carries most hours - that argument belongs in the geothermal
# section, not in a silent constant.
COOL_TIERS_2023 = [
    ("datacentres", "Data centres", None, 0.9, "process"),
    ("industry", "Industry", None, 0.8, "process"),
    ("commercial", "Commercial", 7.5, 3.6, "mixed"),
    ("public", "Public", 0.5, 0.2, "mixed"),
]
_sf = ANCHORS["cool"]["cooling_service_factor"]
assert abs(7.5 / 3.6 - _sf["comfort"]) < 0.02, (
    "Panel 1's comfort/refrigeration factor no longer matches SEAI's "
    "commercial ratio")
assert abs(0.5 / 0.2 - _sf["public"]) < 0.02, (
    "Panel 1's public factor no longer matches SEAI's public ratio")
# Retail is 73.1% of commercial cooling - SOURCED, from Figure 13 of
# the 2019 study, which disaggregates commercial and public by building
# activity and reconciles to Figure 8 within 0.1%. Retail 4,345 GWh of
# 5,946. That replaces the judgement band we previously carried, whose
# upper end (72%) turned out to be almost exactly right.
#
# The band that remains sits INSIDE retail, because a supermarket runs
# refrigeration and shop comfort cooling off the same site. Offices and
# education - 1,092 GWh - are the cleanest Tier 1 available: no
# refrigeration argument is possible for them.
# The activity split, from Figures 52/53 of the 2019 study. These are
# HARD SEGMENTS, not a gradient: SEAI disaggregates commercial and
# public by building activity and the totals reconcile with its sector
# figures to within 0.1%, so the boundary can be drawn rather than
# guessed.
#
# SEAI ATTRIBUTES RETAIL'S DOMINANCE ITSELF: "The cooling demand from
# retail archetypes exceeds that of all other archetypes combined. This
# points to the large amount of energy used for cooling in the retail
# sector, which is likely attributable to refrigeration." That is a
# sourced attribution, not our judgement, and it is why retail sits in
# Tier 0.
#
# FIGURE 53 ADDS CONCENTRATION, which the totals alone conceal. Per
# archetype: warehouse and storage 570 MWh - the highest of any
# activity on ~150 buildings, which is what a cold store looks like -
# against retail's 168 MWh spread across some 26,000. Education 305 and
# hotel 377 are also intense per building and small in total.
#
# AND ONLY 62 OF 181 ARCHETYPES HAVE ANY COOLING DEMAND AT ALL. Most of
# the Irish non-domestic stock has none, which is the Tier 2 condition
# showing up inside SEAI's own model.
#
# tier: 0 process, 1 comfort, m genuinely mixed under one roof
COOL_ACTIVITY_2019 = [
    ("Retail", 4345, 168, "comm", 0),
    ("Restaurant/public house", 371, 50, "comm", 0),
    ("Warehouse and storage", 88, 570, "comm", 0),
    ("Hotel", 377, 377, "comm", "m"),
    ("Healthcare", 241, 121, "pub", "m"),
    ("Office (commercial)", 765, 101, "comm", 1),
    ("Office (public)", 295, 205, "pub", 1),
    ("Education", 32, 305, "pub", 1),
]
# SEAI holds commercial and public cooling constant to 2050 at
# archetype level - "the cooling demand for each archetype in the
# commercial and public sectors between now and 2050 is therefore
# assumed to be constant" - which is exactly the "held, not forecast"
# treatment these bars already gave them, now sourced rather than ours.
COOL_COMMERCIAL_RETAIL_BAND = (0.60, 0.85)   # inside retail, dagger
COOL_PUBLIC_PROCESS_BAND = (0.25, 0.55)
# THE TIERS CUT ACROSS SEAI'S SECTORS. A hospital's imaging suites,
# laboratory and blood refrigeration run regardless of the weather and
# are Tier 0; only ward and office comfort is Tier 1. Same for an
# airport. That is the panel's point rather than a caveat: those are
# the sites where one borefield serves both, and where the seasonal
# balance it needs comes closest to striking itself.
COOL_TIER_DEFS = [
    ("tier0", "Tier 0 \u00b7 process",
     "Runs regardless of the weather. Data centres, industrial plant, "
     "retail and food refrigeration, cold stores \u2014 and inside "
     "hospitals and airports, the imaging, laboratory, theatre and "
     "equipment loads that do not stop in January."),
    ("tier1", "Tier 1 \u00b7 comfort, equipped",
     "Cooling delivered by installed plant in buildings that have it, "
     "driven by weather and occupancy. Offices, shops and public "
     "buildings with high internal gains."),
    ("tier2", "Tier 2 \u00b7 comfort, unequipped",
     "Buildings that overheat and have no cooling to draw. Outside "
     "every figure on these bars, because a consumption survey cannot "
     "count non-consumption. SEAI records residential cooling as zero "
     "and expects it to stay zero to 2050; Irish overheating research "
     "finds a large fraction of the stock already past comfort "
     "thresholds. Both are true, and the gap between them is this "
     "tier."),
]
# EirGrid contracted-demand trajectory: data centre electricity 9.4 TWh
# in 2025 to 14.6 TWh in 2034, via CRU. 2023 was ~6.4 TWh, so the block
# scales by 14.6/6.4. CONTRACTED demand, not a growth assumption - but
# it predates the CRU's 2025 connection policy (80% additional
# renewables, six-year glide path), so it may prove high.
COOL_DC_GROWTH = (6.4, 14.6, 2034)
COOL_DC_COOLING_SHARE = 0.14
# 0.14 is stated in the National Heat Study in plain text - "cooling is
# responsible for only a small proportion of total electricity use by
# data centres, at approximately 14%" - and confirmed twice over: the
# archetype table averages there, and back-calculating the emissions
# bar at the grid intensity gives 280 GWh, which is 14.0% of the ~2.0
# TWh consumed in 2019. Reported as ~10% twice in drafting, from a 2023
# numerator over a 2025 denominator. It is 14%.
assert ANCHORS["cool"]["dc_cooling_share"] == COOL_DC_COOLING_SHARE, (
    "Panel 1 and Panel 4 disagree on the data-centre cooling share")
COOL_NI_ALL_TWH = 1.2
COOL_SCOPE = "Republic of Ireland"
# Direct ground cooling: circulation only, no compressor. Dagger, and
# the same undefined-boundary problem the peer review flagged on the
# network SPF - what pumping is inside it is not settled, so this is
# deliberately conservative against the "order of 20" once carried by
# the UK sibling.
COOL_GEO_EER = 15.0
COOL_WHATIF_SHARE = 0.20


# Heat REJECTED, and what a store can recover of it.
#
# Everything that goes into cooling comes out as heat: the service
# removed from the space PLUS the electricity that drove the removal.
# For a data centre there is no cooling service in that sense - the
# whole facility draw leaves as heat, IT load included.
#
# THE TWO SOURCES BANK DIFFERENTLY, and that is the panel's point.
# Comfort cooling rejects ONLY in summer, exactly when nothing wants
# heat, so without a store every unit is lost. Process refrigeration
# and data centres reject CONTINUOUSLY - the winter half can go
# straight into a network, as Tallaght already does, and it is only the
# summer half that strands. A store therefore does not merely bank
# summer heat: it lifts annual utilisation of a continuous source from
# roughly half to nearly all. That is the multiplicative effect.
COOL_SUMMER_FRACTION = 0.50   # dagger - Oct-Mar is the heating season
COOL_UTES_ROUNDTRIP = 0.70    # dagger - literature range 0.50-0.80
COOL_UTES_RANGE = (0.50, 0.80)


def derive_heat_rejected(ct, anchors=None):
    """
    A fifth of Irish cooling, as heat rejected and heat recovered.

    Data centres ARE included here, unlike the cooling what-if in the
    bars above. There the question was whether ground cooling displaces
    a compressor, and in this climate it does not. Here the question is
    what happens to the heat, and a data centre rejects it whether it
    used a compressor or a fan.
    """
    a = anchors or ANCHORS
    share = COOL_WHATIF_SHARE
    rows = []
    for t in ct["tiers"]:
        if t["key"] == "datacentres":
            # the whole facility draw leaves as heat, not just the
            # cooling block - IT load included
            dc_elec = COOL_DC_GROWTH[0]
            rej = dc_elec
            cont = True
            lab = "Data centres"
        else:
            rej = t["service_twh"] + t["elec_twh"]
            cont = t["key"] != "public"
            lab = t["label"]
        rows.append({"key": t["key"], "label": lab,
                     "rejected_twh": round(rej, 2), "continuous": cont,
                     "summer_twh": round(
                         rej * (COOL_SUMMER_FRACTION if cont else 1.0), 2)})
    banked = sum(r["summer_twh"] for r in rows) * share
    rec = banked * COOL_UTES_ROUNDTRIP
    lo, hi = COOL_UTES_RANGE
    ni_heat = (a["roi"]["residential_heat_twh"]
               + a["roi"]["services_heat_twh"]) * 1000 \
        * a.get("delivered_over_input_roi", 0.8225)
    out = {"share": share, "rows": rows,
           "rejected_twh": round(sum(r["rejected_twh"] for r in rows)
                                 * share, 2),
           "banked_twh": round(banked, 2),
           "recovered_twh": round(rec, 2),
           "recovered_range_twh": [round(banked * lo, 2),
                                   round(banked * hi, 2)],
           "roundtrip": COOL_UTES_ROUNDTRIP,
           "roundtrip_range": list(COOL_UTES_RANGE),
           "summer_fraction": COOL_SUMMER_FRACTION,
           "share_of_roi_heat_pct": round(100 * rec * 1000
                                          / max(ni_heat, 1), 1)}
    log(f"heat rejected: a fifth rejects {out['rejected_twh']} TWh, of "
        f"which {out['banked_twh']} TWh strands in summer; at a "
        f"{int(COOL_UTES_ROUNDTRIP*100)}% round trip "
        f"{out['recovered_twh']} TWh returns for winter heating - "
        f"{out['share_of_roi_heat_pct']}% of the Republic's building "
        f"heat")
    return out


def derive_cooling_tiers():
    """
    Four bars, and the units change halfway - which is the point.

      1  2023 cooling SERVICE, every sector on the same basis
      2  2034 service, data centre block projected, rest held
      3  the ELECTRICITY that 2034 service takes
      4  the same, with a fifth of the service on ground cooling

    Bars 1-2 are what buildings and plant receive; 3-4 are what is
    bought to deliver it. The drop from 3 to 4 is the dividend.
    """
    lo, hi, yr = COOL_DC_GROWTH
    factor = hi / lo
    tiers = []
    for k, lab, svc, elec, grp in COOL_TIERS_2023:
        judged = svc is None
        if judged:
            eer_used = (COOL_DC_EFFECTIVE_EER if k == "datacentres"
                        else COOL_INDUSTRY_EER)
            service = round(elec * eer_used, 2)
        else:
            service = svc
        eer = round(service / elec, 2) if elec else None
        proj = round(service * factor, 2) if k == "datacentres" else service
        pel = round(elec * factor, 2) if k == "datacentres" else elec
        tiers.append({"key": k, "label": lab, "group": grp,
                      "service_twh": service, "elec_twh": elec,
                      "eer": eer, "eer_is_ours": judged,
                      "service_proj_twh": proj, "elec_proj_twh": pel,
                      "held": k != "datacentres"})
    svc23 = round(sum(t["service_twh"] for t in tiers), 1)
    svc34 = round(sum(t["service_proj_twh"] for t in tiers), 1)
    el34 = round(sum(t["elec_proj_twh"] for t in tiers), 2)
    # the what-if: a fifth of each sector's projected service moved to
    # ground cooling, which still pumps but does not compress
    saved = 0.0
    for t in tiers:
        if not t["eer"] or t["key"] in COOL_WHATIF_EXCLUDE:
            continue
        moved = COOL_WHATIF_SHARE * t["service_proj_twh"]
        saved += moved / t["eer"] - moved / COOL_GEO_EER
    el34_geo = round(el34 - saved, 2)
    off, run = {}, 0.0
    for t in tiers:
        off[t["key"]] = round(run, 2)
        run += t["service_twh"]
    # Hard segments. Commercial and public are scaled SEPARATELY onto
    # their own 2023 sector totals, because the two sectors did not
    # grow at the same rate and one blended factor would misplace the
    # boundary we are drawing.
    sect = {"comm": 7.5, "pub": 0.5}
    base = {g: sum(v for _, v, _, gg, _ in COOL_ACTIVITY_2019 if gg == g)
            for g in sect}
    eer_of = {"comm": 2.08, "pub": 2.50}
    segs = []
    for name, gwh, per, g, tier in COOL_ACTIVITY_2019:
        svc = round(gwh / base[g] * sect[g], 2)
        segs.append({"label": name, "service_twh": svc,
                     "elec_twh": round(svc / eer_of[g], 2),
                     "per_archetype_mwh": per, "sector": g, "tier": tier})
    act_tot = sum(v for _, v, _, _, _ in COOL_ACTIVITY_2019)
    out = {
        "base_year": 2023, "proj_year": yr, "scope": COOL_SCOPE,
        "unit": "TWh", "tiers": tiers,
        "service_twh": svc23, "service_proj_twh": svc34,
        "elec_proj_twh": el34, "elec_proj_geo_twh": el34_geo,
        "geo_saving_twh": round(saved, 2),
        "geo_saving_pct": round(100 * saved / max(el34, 1e-9), 1),
        "geo_eer": COOL_GEO_EER, "whatif_share": COOL_WHATIF_SHARE,
        "industry_eer": COOL_INDUSTRY_EER,
        "dc_eer": COOL_DC_EFFECTIVE_EER,
        "whatif_excluded": list(COOL_WHATIF_EXCLUDE),
        "offsets_twh": off,
        "commercial_retail_band_twh": [
            round(7.5 * 0.731 * f, 2) for f in COOL_COMMERCIAL_RETAIL_BAND],
        "public_process_band_twh": [
            round(0.5 * f, 2) for f in COOL_PUBLIC_PROCESS_BAND],
        "retail_share_of_commercial": 0.731,
        "segments": segs,
        "activity_total_gwh": act_tot,
        # THE WHOLE BAR, not just the activity segments. This
        # previously summed segs only, so it reported Tier 0 as 6.06
        # and silently left out data centres and industry - both Tier
        # 0, and between them more than half of it. The note beneath
        # the chart was worded "across commercial and public", so it
        # was not false, but the headline tier split was nowhere on
        # the page and the published figure invited the wrong reading.
        "tier_totals_twh": {
            str(t): round(
                sum(x["service_twh"] for x in segs if x["tier"] == t)
                + sum(x["service_twh"] for x in tiers
                      if t == 0 and x["key"] in ("datacentres", "industry")),
                2)
            for t in (0, 1, "m")},
        "tier_totals_segments_twh": {
            str(t): round(sum(x["service_twh"] for x in segs
                              if x["tier"] == t), 2)
            for t in (0, 1, "m")},
        "archetypes_with_cooling": [62, 181],
        "ni_all_twh": COOL_NI_ALL_TWH,
        "dc_cooling_share": COOL_DC_COOLING_SHARE,
        "tier_defs": [{"key": k, "label": lab, "text": txt}
                      for k, lab, txt in COOL_TIER_DEFS],
        "source": "SEAI Comprehensive Assessment Technical Annex 2025, "
                  "Figure 7 (2023); activity split from National Heat "
                  "Study Report 1, Figure 13 (2019); EirGrid contracted "
                  "demand via CRU",
    }
    log(f"cooling tiers: service {svc23} TWh in 2023 -> {svc34} by {yr}; "
        f"electricity {el34} TWh, with a {int(COOL_WHATIF_SHARE*100)}% "
        f"geothermal what-if {el34_geo} TWh "
        f"(-{out['geo_saving_pct']}%); industry EER "
        f"{COOL_INDUSTRY_EER} is ours, the rest are SEAI's; retail is "
        f"{out['retail_share_of_commercial']*100:.1f}% of commercial "
        f"and the activity split is HARD. WHOLE BAR: Tier 0 "
        f"{out['tier_totals_twh']['0']} TWh, mixed "
        f"{out['tier_totals_twh']['m']}, Tier 1 "
        f"{out['tier_totals_twh']['1']} TWh; of which commercial and "
        f"public alone are {out['tier_totals_segments_twh']['0']} / "
        f"{out['tier_totals_segments_twh']['m']} / "
        f"{out['tier_totals_segments_twh']['1']}")
    return out


def derive_odd_examples(dd, anchors=None):
    """
    What scale of load it would take to absorb Northern Ireland's
    constrained wind, worked two ways.

    NI because that is where the spill is both largest as a share and
    overwhelmingly LOCAL constraint - the only kind a local load can
    address. The arithmetic is deliberately simple and the point is the
    ORDER OF MAGNITUDE: institutional anchor loads cannot absorb this
    volume, which is why the published answer is an aggregation of a
    quarter of a million households rather than a list of large sites.
    """
    if not dd or "NI" not in dd.get("jurisdictions", {}):
        return None
    a = anchors or ANCHORS
    b = dd["jurisdictions"]["NI"]
    months = dd["months"]
    # ROLLING TWELVE MONTHS, not a calendar year. It keeps the panel
    # current without waiting for a year to close, and it lands on the
    # same window as the hourly store, so a spill-weighted COP computed
    # later covers exactly these months rather than a different set.
    yr = list(range(len(months)))[-12:]
    cons = sum(b["cons"][i] for i in yr)          # GWh electricity
    spf = dd["spf"]["network"]
    heat = cons * spf                             # GWh of heat
    ni_heat = (a["ni"]["residential_heat_twh"]
               + a["ni"]["services_heat_twh"]) * 1000 \
        * a.get("delivered_over_input_ni", 0.8375)
    # PER ROUTE. The same spilled electricity makes different amounts
    # of heat, so the share of NI's heat it could cover differs by
    # route - and that is a discriminator this panel can carry today.
    # The SECOND discriminator, the COP each route sees IN THE HOURS
    # THE WIND IS ACTUALLY SPILLED, needs the hourly store joined to
    # the spill half-hours and is not built yet; the hourly record
    # covers 13 months against this series' five years.
    routes = {}
    for r, s_ in (("ashp", SPF_ANCHORS["ashp"]),
                  ("gshp", SPF_ANCHORS["gshp"]),
                  ("network", spf)):
        h = cons * s_
        routes[r] = {"spf": s_, "heat_gwh": round(h, 1),
                     "share_pct": round(100 * h / max(ni_heat, 1), 1)}
    hosp = ODD_HOSPITAL
    hosp_gwh = hosp["eui_kwh_m2"] * hosp["heat_share"] \
        * hosp["floor_m2"] / 1e6
    out = {
        "basis_from": months[yr[0]],
        "basis_to": months[yr[-1]],
        "basis_months": len(yr),
        "constrained_gwh": round(cons, 1),
        "heat_gwh": round(heat, 1),
        "ni_delivered_heat_gwh": round(ni_heat, 0),
        "share_of_ni_heat_pct": round(100 * heat / max(ni_heat, 1), 1),
        "routes": routes,
        "hospital": dict(hosp, heat_gwh=round(hosp_gwh, 1),
                         equivalent=round(heat / max(hosp_gwh, 0.1))),
        "domestic": dict(AGBONAYE),
    }
    log(f"odd examples: NI constrained wind {cons:.0f} GWh in the "
        f"{len(yr)} months {out['basis_from']}..{out['basis_to']} -> "
        f"{heat:.0f} GWh of heat at SPF {spf}, "
        f"{out['share_of_ni_heat_pct']:.0f}% of NI delivered building "
        f"heat; equivalent to {out['hospital']['equivalent']} hospitals "
        f"of {hosp_gwh:.1f} GWh \u2020 - which is why the published "
        f"answer is {AGBONAYE['subscribers']:,} households, not a list "
        f"of large sites")
    return out


def derive_heat_emissions(feeds, anchors=None):
    """
    gCO2e per USEFUL kWh by route, all-island.

    ALL-ISLAND ON PURPOSE, with no jurisdiction toggle. Emissions per
    useful kWh are the combustion factor over the boiler efficiency,
    or the grid intensity over the route's COP. The factors do not
    change at the border, the efficiencies are shared, and the grid is
    a single all-island market - so an NI/ROI split would draw the
    same bars three times. That is the finding rather than a gap: the
    PRICE answer differs sharply across the border and the EMISSIONS
    answer does not at all.

    The network route's SPF is jurisdictional by design (5.0 NI, 4.0
    ROI). For one island figure they combine as a HEAT-WEIGHTED
    HARMONIC mean, which is the right way to average efficiencies over
    a population - not the arithmetic mean, which would flatter it.
    """
    a = anchors or ANCHORS
    co2 = ((feeds.get("eirgrid") or {})
           .get("co2_intensity_g_per_kwh") or {})
    cdays = sorted(co2)[-14:]
    if len(cdays) < 7:
        return None
    grid = round(sum(co2[d] for d in cdays) / len(cdays), 1)

    def heat(j):
        return a[j]["residential_heat_twh"] + a[j]["services_heat_twh"]
    w_ni, w_roi = heat("ni"), heat("roi")
    net_spf = round((w_ni + w_roi) /
                    (w_ni / NETWORK_MODEL["ni"]["spf"]
                     + w_roi / NETWORK_MODEL["roi"]["spf"]), 3)

    rows = [
        ("gas_boiler", "Gas boiler",
         a["ef_g_per_kwh"]["gas"] / a["efficiency"]["gas"], None),
        ("oil_boiler", "Oil boiler",
         a["ef_g_per_kwh"]["oil"] / a["efficiency"]["oil"], None),
        ("resistive", "Resistive electric", grid, 1.0),
        ("ashp", "Air-source heat pump", grid / SPF_ANCHORS["ashp"],
         SPF_ANCHORS["ashp"]),
        ("gshp", "Ground source", grid / SPF_ANCHORS["gshp"],
         SPF_ANCHORS["gshp"]),
        ("network", "Geothermal heat network", grid / net_spf, net_spf),
    ]
    out = {"grid_g_per_kwh": grid, "grid_days": len(cdays),
           "network_spf_island": net_spf,
           "routes": [{"key": k, "label": lab, "g_per_useful_kwh": round(v, 1),
                       "spf": spf} for k, lab, v, spf in rows]}
    log("heat emissions: grid {} g/kWh ({}-day mean), island network SPF "
        "{} - ".format(grid, len(cdays), net_spf)
        + ", ".join(f"{r['label']} {r['g_per_useful_kwh']:.0f}"
                    for r in out["routes"]))
    return out


def heat_gap_from_cost_series(rows, window=730):
    """
    The masthead ticker, taken from the SAME engine as panel 2.

    It used to come from derive_heat_gap(), which is the original
    calculation and never got panel 2's changes: one geothermal SPF of
    4.0 for both jurisdictions where the panel now uses 5.0 in the
    North, a single oil-boiler efficiency of 0.82 where the panel
    prices hot water at 0.71, and no hot-water blending at all. The
    two disagreed by 13-20% on identical routes on the same day, with
    the ticker sitting above the panel that contradicted it.

    Also returns the gap's MEDIAN over the window. A spot gap alone
    reads as a standing fact; on this record oil climbs steeply from
    November 2025, so today's gap runs well above its own two-year
    norm and the ticker was stating a war-driven condition as
    structural.
    """
    out = {}
    for jur in ("ni", "roi"):
        have = [r for r in rows if r.get(jur)]
        if not have:
            continue
        last = have[-1][jur]
        gaps = sorted(r[jur]["oil_boiler"] - r[jur]["network"]
                      for r in have[-window:])
        n = len(gaps)
        med = gaps[n // 2] if n % 2 else (gaps[n // 2 - 1] + gaps[n // 2]) / 2
        out[jur] = {
            "oil_boiler": last["oil_boiler"],
            "gas_boiler": last["gas_boiler"],
            "ashp": last["ashp"],
            "gshp": last["gshp"],
            # key kept so the front end keeps reading; the value is
            # now the panel's jurisdictional network model rather than
            # a single 4.0 anchor
            "geothermal_spf40": last["network"],
            "gap_now": round(last["oil_boiler"] - last["network"], 2),
            "gap_median": round(med, 2),
            "gap_days": n,
            "day": have[-1]["day"],
        }
    return out


def derive_heat_gap(feeds, anchors=None):
    """
    Cost of useful heat by route, per jurisdiction, standard tariffs -
    plus the break-even SPF against the incumbent oil boiler. Pure,
    unit tested. Native currency minor units (p or c) per useful kWh.
    """
    a = anchors or ANCHORS
    fx = (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855

    ccni = ((feeds.get("ccni_oil") or {}).get("series_gbp") or {}).get(
        "daily", {}).get("900l") or {}
    oil_ni_ppl = ccni[max(ccni)] * 100 / 900 if ccni else None
    ob = (feeds.get("oil_bulletin") or {}).get("latest_value")
    oil_roi_cpl = ob * 100 / 1000 if ob else None
    kwh_l = a["kerosene_kwh_per_litre"]
    eff_oil, eff_gas = a["efficiency"]["oil"], a["efficiency"]["gas"]
    spf_geo = a["geothermal_spf"]
    hddf = feeds.get("hdd") or {}
    ashp_ni = derive_ashp_spf(hddf.get("hdd_ni") or {}, a)
    ashp_roi = derive_ashp_spf(hddf.get("hdd_roi") or {}, a)
    fallback = {"spf": 2.8}   # field-trial median, dagger
    ashp_ni = ashp_ni or fallback
    ashp_roi = ashp_roi or fallback

    def jur(oil_pl, elec, gas, ashp):
        if oil_pl is None:
            return None
        oil_useful = oil_pl / kwh_l / eff_oil
        r2 = lambda x: round(x, 2)
        return {
            "oil_boiler": r2(oil_useful),
            "gas_boiler": r2(gas * 100 / eff_gas),
            "ashp": r2(elec * 100 / ashp["spf"]),
            "ashp_spf": ashp["spf"],
            "ashp_model": {k: v for k, v in ashp.items() if k != "params"},
            "geothermal_spf40": r2(elec * 100 / spf_geo),
            "breakeven_spf_vs_oil": r2(elec * 100 / oil_useful),
            "breakeven_spf_vs_gas": r2(elec * 100 / (gas * 100 / eff_gas)),
            "inputs": {"oil_per_litre": round(oil_pl, 2),
                       "electricity_per_kwh": elec, "gas_per_kwh": gas},
        }

    ni = jur(oil_ni_ppl, a["retail_gbp_per_kwh"]["electricity"],
             a["retail_gbp_per_kwh"]["gas"], ashp_ni)
    roi = jur(oil_roi_cpl, a["retail_eur_per_kwh"]["electricity"],
              a["retail_eur_per_kwh"]["gas"], ashp_roi)
    if not (ni and roi):
        return None
    return {
        "ni": ni, "roi": roi, "fx_eur_gbp": fx,
        "geo_spf": spf_geo,
        "basis": ("Standard tariffs, July 2026 pass (Power NI/UR review, "
                  "ROI standard 24h rates) - dagger; time-of-use and night "
                  "tariffs materially lower for heat-pump households. Oil "
                  "prices live. ASHP SPF is modelled from each "
                  "jurisdiction's HDD-weighted climate (Carnot-fraction, "
                  "defrost derate, DHW share - all dagger), calibrated to "
                  "GB field-trial medians. Kerosene 10.35 kWh/L; boiler "
                  "efficiencies 82%/85% dagger. Challenge and input "
                  "welcome at contact@causewaygt.com"),
    }


# ---------------------------------------------------------------- assembly

FEEDS = {
    "eirgrid": feed_eirgrid,
    "hdd": feed_hdd,
    "ecb_fx": feed_ecb_fx,
    "gni_ckan": feed_gni_ckan,
    "semopx": feed_semopx,
    "oil_bulletin": feed_oil_bulletin,
    "gni_live": feed_gni_live,
    "ccni_oil": feed_ccni_oil,
    "gb_oil": feed_gb_oil,
    "entsog_probe": feed_entsog_probe,
    "sem_mix": feed_sem_mix,
    "eirgrid_probe": feed_eirgrid_probe,
}



# ------------------------------------------------------- hourly store
# EirGrid serves a ~30-day window of 15-minute rows ending at dateTo,
# for chartType x region, with `areas` matching the chart. Probe round
# 2 (7 Aug 2026) confirmed historic windows return correctly-dated
# data a year back, so the 13-month store backfills by chunked
# walking. dateRange=year returns nothing - month chunks are the
# mechanism. All-island scope, per the v7 grid-layer decision.
HOURLY_SERIES = {
    # WHAT THESE SERIES CONTAIN. Written here, at the point of load,
    # because getting it wrong cost the UK sibling three published
    # claims and cost this one an entire withdrawn analysis.
    #
    # demandactual - EirGrid's own words: "the electricity production
    #   required to meet national electricity consumption, including
    #   system losses, but net of generators' requirements." It is
    #   met BY grid-connected generation, so grid solar and wind are
    #   already inside it. It is NOT reduced by them.
    #
    # solaractual - "the total electricity production of large scale
    #   solar farms on the system. Small scale embedded solar is NOT
    #   included."
    #
    # THEREFORE: NEVER ADD solar_ai TO demand_ai. That is a double
    # count, not a reconstruction. The UK's NESO series is the
    # opposite case - its demand figure EXCLUDES embedded generation,
    # so there it must be added back once. Ported reasoning does not
    # survive the crossing.
    #
    # The Irish problem is the reverse and has no fix in this data:
    # small-scale embedded solar (~310 MW, tens of thousands of roofs)
    # IS invisible to demandactual and there is NO published series to
    # add back. So Irish underlying demand is understated on bright
    # days by an unknown and growing amount, which biases any DAYLIGHT
    # cooling estimate downward. Night-time estimators are unaffected,
    # because solar is zero - which is why the weekday/weekend night
    # placebo is the route that survives.
    "demand_ai": ("demand", "ALL", "demandactual"),
    "wind_ai": ("wind", "ALL", "windactual"),
    "solar_ai": ("solar", "ALL", "solaractual"),
    "co2_ai": ("co2", "ALL", "co2intensity"),
}


def _hour_key(ts):
    """15-minute stamp -> UTC hour key. EirGrid stamps are local
    clock; the weekly layer already treats these as day-local, and the
    grid layer needs consistency with it rather than with UTC."""
    for f in ("%d-%b-%Y %H:%M:%S", "%d-%B-%Y %H:%M:%S"):
        try:
            return dt.datetime.strptime(ts, f).strftime("%Y-%m-%dT%H")
        except ValueError:
            continue
    return None


def hourly_from_rows(rows, min_quarters=3):
    """15-minute rows -> hourly MEANS (never samples). An hour needs
    at least min_quarters of its four values or it is dropped, so a
    partial hour cannot masquerade as a low one."""
    buckets = {}
    for r in rows or []:
        v = r.get("Value")
        if v is None:
            continue
        k = _hour_key(str(r.get("EffectiveTime") or ""))
        if k:
            buckets.setdefault(k, []).append(float(v))
    return {k: round(sum(v) / len(v), 2)
            for k, v in buckets.items() if len(v) >= min_quarters}


def fetch_hourly_chunk(chart, region, areas, end_day):
    """One ~30-day window ending end_day."""
    def dmy(d):
        return d.strftime("%d-%b-%Y").replace(" 0", " ")
    payload = http_get(EIRGRID_ENDPOINT, params={
        "region": region, "chartType": chart, "dateRange": "month",
        "dateFrom": dmy(end_day - dt.timedelta(days=27)),
        "dateTo": dmy(end_day), "areas": areas,
    }, timeout=120).json()
    return hourly_from_rows((payload or {}).get("Rows", []))


def weighted_hourly_temp(payload, names, weights):
    """
    Open-Meteo hourly payloads -> population-weighted island air
    temperature per hour, keyed 'YYYY-MM-DDTHH'.

    The divisor is the weight ACTUALLY PRESENT in each hour, not the
    full 1.0. A station missing an hour then leaves the island mean
    unbiased instead of dragging it toward zero, which a plain
    weighted sum would do silently and only in the hours where a feed
    gap already exists. Pure, unit tested.
    """
    locs = payload if isinstance(payload, list) else [payload]
    acc, wt = {}, {}
    for name, loc in zip(names, locs):
        w = weights.get(name, 0.0)
        if w <= 0:
            continue
        hh = (loc or {}).get("hourly", {}) or {}
        for ts, t in zip(hh.get("time", []) or [],
                         hh.get("temperature_2m", []) or []):
            if t is None:
                continue
            k = str(ts)[:13]
            acc[k] = acc.get(k, 0.0) + w * float(t)
            wt[k] = wt.get(k, 0.0) + w
    return {k: round(acc[k] / wt[k], 2) for k in acc if wt[k] > 0}


def fetch_hourly_temp(previous, floor_key, end_day):
    """
    Island hourly temperature across the store's window, merged onto
    whatever previous runs retained.

    TIMEZONE. Requested on Europe/Dublin clock, NOT UTC. _hour_key
    deliberately keys the EirGrid series on its own local stamps, and
    the daily HDD feed asks Open-Meteo for UTC. Joining those two
    would put temperature an hour out of step with demand from late
    March to late October - silently, and in exactly the direction
    that misaligns an evening peak with the temperature that caused
    it. Both sides of this store are therefore local clock.

    The autumn fold repeats one local hour; the later value wins, in
    both series, for one hour a year.
    """
    names = list(STATIONS)
    lats = ",".join(str(STATIONS[n][0]) for n in names)
    lons = ",".join(str(STATIONS[n][1]) for n in names)
    weights = {n: STATIONS[n][2] for n in names}
    have = dict(previous or {})

    ends, cursor = [], end_day
    while len(ends) < HOURLY_TEMP_CHUNKS_PER_RUN:
        ends.append(cursor)
        cursor = cursor - dt.timedelta(days=HOURLY_TEMP_CHUNK_DAYS)
        if cursor.strftime("%Y-%m-%dT00") < floor_key:
            break

    for i, e in enumerate(ends):
        start = e - dt.timedelta(days=HOURLY_TEMP_CHUNK_DAYS - 1)
        span_start = start.strftime("%Y-%m-%dT00")
        span_end = e.strftime("%Y-%m-%dT23")
        covered = sum(1 for k in have if span_start <= k <= span_end)
        # Expect only the part of the chunk AT OR ABOVE the retention
        # floor. The oldest chunk always straddles it - roughly a
        # quarter of its hours are discarded on write - so measured
        # against the whole chunk its coverage could never reach the
        # skip threshold, and it was re-fetched on every run forever
        # for a month of data already held.
        expect = sum(1 for h in range(HOURLY_TEMP_CHUNK_DAYS * 24)
                     if (start + dt.timedelta(hours=h)
                         ).strftime("%Y-%m-%dT%H") >= floor_key)
        expect = max(expect, 1)
        # Newest chunk always re-fetched (ERA5 revises); older chunks
        # skipped only once nearly whole, so gaps converge over runs
        # rather than becoming permanent.
        if i and covered >= expect * 0.98:
            continue
        try:
            payload = http_get(
                "https://archive-api.open-meteo.com/v1/archive", params={
                    "latitude": lats, "longitude": lons,
                    "start_date": start.isoformat(),
                    "end_date": e.isoformat(),
                    "hourly": "temperature_2m",
                    "timezone": "Europe/Dublin",
                    # 90, not 240. The first live run (7 Aug 2026)
                    # spent 256 s on a ReadTimeout and the identical
                    # retry then succeeded in three seconds, so a
                    # long timeout buys nothing here but wasted
                    # runner minutes - the retry budget is what
                    # actually absorbs the transient.
                }, timeout=90).json()
            got = weighted_hourly_temp(payload, names, weights)
            have.update(got)
            log(f"hourly: temp_ai archive {start.isoformat()}.."
                f"{e.isoformat()} {len(got)}h")
        except Exception as exc:
            log(f"hourly: temp_ai archive chunk to {e.isoformat()} "
                f"{exc.__class__.__name__} - gap left for next run")
        time.sleep(0.4)

    # ERA5 lags ~5 days; the forecast endpoint covers the tail. Archive
    # values are authoritative where both exist.
    try:
        tail = weighted_hourly_temp(http_get(
            "https://api.open-meteo.com/v1/forecast", params={
                "latitude": lats, "longitude": lons,
                "past_days": 10, "forecast_days": 1,
                "hourly": "temperature_2m",
                "timezone": "Europe/Dublin",
            }, timeout=120).json(), names, weights)
        added = 0
        for k, v in tail.items():
            if k not in have:
                have[k] = v
                added += 1
        log(f"hourly: temp_ai forecast tail +{added}h")
    except Exception as exc:
        log(f"hourly: temp_ai forecast tail unavailable "
            f"({exc.__class__.__name__}) - archive only")
    return have


def _naive_hour(key):
    """Hour key -> datetime, parsed as written. The keys are IRISH
    LOCAL CLOCK, so this is not a UTC instant and must never be
    treated as one; it exists only to give the keys a total order and
    a spacing, which is all the array encoding needs."""
    return dt.datetime.strptime(key, "%Y-%m-%dT%H")


def compact_hourly(series):
    """
    {name: {hour_key: value}} -> (t0, n, {name: [value|None] * n}).

    Position i is the key t0 + i hours, formatted the same way. On the
    spring transition the local clock skips an hour, which shows up
    here as one null a year in every series - the alternative, a
    stored key list, costs about 140 kB to avoid a gap that is already
    indistinguishable from a feed gap. The autumn fold repeats a local
    hour, but the source dict has already collapsed it to one entry,
    so offsets stay unique in both directions.
    """
    keys = set()
    for v in series.values():
        keys |= set(v or {})
    if not keys:
        return None, 0, {k: [] for k in series}
    t0 = min(keys)
    h0 = _naive_hour(t0)
    n = int((_naive_hour(max(keys)) - h0).total_seconds() // 3600) + 1
    out = {}
    for name, v in series.items():
        arr = [None] * n
        for k, val in (v or {}).items():
            i = int((_naive_hour(k) - h0).total_seconds() // 3600)
            if 0 <= i < n:
                arr[i] = val
        out[name] = arr
    return t0, n, out


def expand_hourly(doc):
    """
    Read a store document of ANY schema back into
    {name: {hour_key: value}}.

    Schema 1 and 2 wrote a dict per series; schema 3 writes flat
    arrays against `t0`. Both are accepted, so the run that first
    writes schema 3 still inherits the schema-2 file already in the
    repo instead of refilling 13 months from empty.
    """
    ser = (doc or {}).get("series") or {}
    if not ser:
        return {}
    sample = next(iter(ser.values()))
    if isinstance(sample, dict):
        return {k: dict(v or {}) for k, v in ser.items()}
    t0 = doc.get("t0")
    if not t0:
        log("hourly: array-form store without t0 - previous state "
            "discarded, refilling")
        return {}
    h0 = _naive_hour(t0)
    out = {}
    for name, arr in ser.items():
        d = {}
        for i, val in enumerate(arr or []):
            if val is None:
                continue
            d[(h0 + dt.timedelta(hours=i)).strftime("%Y-%m-%dT%H")] = val
        out[name] = d
    return out


def daily_ci_from_hourly(store, min_hours=20):
    """
    Daily mean grid carbon intensity from the hourly store.

    The daily EirGrid feed retains 50 days. The backfilled weeks are
    ten months old, so their carbon can only come from the hourly
    store's co2_ai series - and that store keeps 13 months, so its
    floor advances a day every day. Once a week is built and frozen
    its ef_electricity is stored and reused on restatement, so this is
    a one-time capture: after these weeks are in the record, the store
    rolling past them costs nothing.

    Returns {date: g/kWh}. Days with fewer than min_hours observations
    are dropped rather than averaged thin.
    """
    ser = (expand_hourly(store) or {}).get("co2_ai") or {}
    acc = {}
    for k, v in ser.items():
        acc.setdefault(k[:10], []).append(v)
    out = {d: round(sum(vs) / len(vs), 1)
           for d, vs in acc.items() if len(vs) >= min_hours}
    return out


def build_hourly_store(previous, feeds_now=None):
    """Walk back in ~28-day chunks until the window is covered, then
    keep only the most recent chunk (plus a 2-day revision re-fetch)
    on subsequent runs. Returns the store document or None."""
    prev = (previous or {})
    # NB: `prior`, not `prev_series` - there is a module-level
    # prev_series() helper for the weekly feeds and shadowing it here
    # would be a trap for the next edit.
    prior = expand_hourly(prev)
    series = {k: dict(prior.get(k) or {}) for k in HOURLY_SERIES}
    end = today_utc()
    floor = (end - dt.timedelta(days=30 * HOURLY_MONTHS)).strftime(
        "%Y-%m-%dT00")
    added = 0
    for name, (chart, region, areas) in HOURLY_SERIES.items():
        have = series[name]
        # chunk ends: newest first, walking back to the floor
        ends, cursor = [], end
        while len(ends) < HOURLY_CHUNKS_PER_RUN:
            ends.append(cursor)
            cursor = cursor - dt.timedelta(days=28)
            if cursor.strftime("%Y-%m-%dT00") < floor:
                break
        for e in ends:
            span_start = (e - dt.timedelta(days=27)).strftime("%Y-%m-%dT00")
            span_end = e.strftime("%Y-%m-%dT23")
            covered = sum(1 for k in have if span_start <= k <= span_end)
            # Newest chunk always re-fetched (2-day revision window).
            # Older chunks are skipped only once NEARLY whole: at the
            # original 600/672 threshold a chunk sitting at, say, 620
            # was never retried and its gap became permanent - which
            # is why carbon stalled at 93.6% on 7 Aug 2026. At 98% the
            # store converges over successive runs instead.
            if e != ends[0] and covered >= 660:
                continue
            got = {}
            for attempt in (1, 2):
                try:
                    got = fetch_hourly_chunk(chart, region, areas, e)
                    if got:
                        break
                except Exception as exc:
                    log(f"hourly: {name} chunk to {e.isoformat()} "
                        f"attempt {attempt} {exc.__class__.__name__}")
                if attempt == 1:
                    time.sleep(3)      # back off, then retry in-run
            if not got:
                log(f"hourly: {name} chunk to {e.isoformat()} "
                    f"EMPTY after 2 attempts - gap left for next run")
            # count what is NEW to the store, not what came back -
            # the newest chunk is deliberately re-fetched every run,
            # so len(got) double-counts an overlap the store already
            # holds and makes the run line meaningless as a measure
            # of progress.
            added += len(set(got) - set(have))
            have.update(got)
            time.sleep(0.4)            # throttle: ~56 chunk requests
                                       # per cold build trips limits
        before = len(prior.get(name) or {})
        series[name] = {k: v for k, v in sorted(have.items())
                        if k >= floor}
        after = len(series[name])
        # A series must never shrink except by the rolling floor.
        # 7 Aug 2026: demand_ai lost its oldest 600 hours in a single
        # run while the other series kept theirs - loud, not silent.
        if before and after < before - 24:
            log(f"hourly: WARNING {name} shrank {before}h -> {after}h "
                f"in one run - investigate before trusting the panel")
    # The grid series gate the store: with EirGrid down there is
    # nothing for a temperature series to be joined TO, and fetching
    # one would spend a minute of runner time on a document that is
    # not going to be written.
    if not any(len(series[k]) for k in HOURLY_SERIES):
        log("hourly: no data - store not written")
        return None

    # --- fifth series: island hourly temperature (Open-Meteo).
    prev_t = dict(prior.get("temp_ai") or {})
    try:
        t_all = fetch_hourly_temp(prev_t, floor, end)
    except Exception as exc:
        log(f"hourly: temp_ai unavailable ({exc.__class__.__name__}) - "
            "retaining previous")
        t_all = prev_t
    series["temp_ai"] = {k: v for k, v in sorted(t_all.items())
                         if k >= floor}
    # temp is fetched outside the EirGrid walk, so it was missing from
    # `added` and the run line read short (5,751 on 7 Aug 2026 while
    # temp alone had brought in 9,384 hours).
    added += len(set(series["temp_ai"]) - set(prev_t))
    if prev_t and len(series["temp_ai"]) < len(prev_t) - 24:
        log(f"hourly: WARNING temp_ai shrank {len(prev_t)}h -> "
            f"{len(series['temp_ai'])}h in one run - investigate "
            "before trusting the panel")

    # --- sixth series: SEMOpx day-ahead price, EUR/MWh.
    #
    # FILLS FORWARD ONLY. The SEMOpx report listing serves the recent
    # window; resolving an arbitrary historic trade day needs a filter
    # this pipeline has no evidence for, and guessing a parameter
    # against a live API is how the round-1 probe wasted a day. So the
    # series accumulates from the daily document already fetched, and
    # semopx_history_probe() asks the backfill question with logging
    # instead of assumptions.
    #
    # CONSEQUENCE, stated because it changes the plan: B.2.3 wants the
    # price in the tightest hour, and the tightest hour so far is
    # 5 Jan 2026. Until either the probe finds a way back or thirteen
    # months pass, B.2.3 can be computed on hours the store has priced,
    # not on the binding hour already found.
    price = dict(prev_series("semopx", "dam_hourly_eur_mwh"))
    price.update(((feeds_now or {}).get("semopx") or {})
                 .get("dam_hourly_eur_mwh") or {})
    series["price_ai"] = {k: v for k, v in sorted(price.items())
                          if k >= floor}

    counts = {k: len(v) for k, v in series.items()}
    spans = {k: (min(v), max(v)) for k, v in series.items() if v}
    ref = spans.get("demand_ai")
    # NEW hour-values, not "fetched". The EirGrid walk counted every
    # hour it received including re-fetched overlaps, while temp
    # counted only hours it did not already hold - two meanings in one
    # number, and the line read short on the day temp brought in 9,384
    # hours. Both sides now report new-to-the-store.
    log(f"hourly: {added} new hour-values this run; "
        + ", ".join(f"{k} {n}h" for k, n in counts.items())
        + (f"; span {ref[0]} .. {ref[1]}" if ref else ""))
    # completeness against the covered span (>=95% expected)
    # Completeness is judged PER SERIES against the reference span,
    # not on demand alone: the first store (7 Aug 2026) had demand at
    # 100% while carbon intensity reached only 86%, and a carbon
    # overlay drawn on that would have passed a demand-only gate.
    complete, heat_ready, price_ready, per_series = False, False, False, {}
    if ref:
        # Denominator is the INTENDED window (floor -> latest hour in
        # the store), not one series' own span. Using demand's span
        # let other series score 106.8% on 7 Aug 2026 when demand
        # shrank - a completeness figure above 100% is a bug signal,
        # so the denominator must not depend on the numerator.
        # ...and the window is the GRID window. temp_ai carries a
        # forecast tail that can run a day past the last EirGrid hour;
        # letting it set the denominator would depress every other
        # series' completeness for a reason that has nothing to do
        # with them.
        latest = max(mx for k, (_, mx) in spans.items()
                     if k in HOURLY_SERIES)
        h0 = dt.datetime.strptime(floor, "%Y-%m-%dT%H")
        h1 = dt.datetime.strptime(latest, "%Y-%m-%dT%H")
        expect = int((h1 - h0).total_seconds() // 3600) + 1
        for k, n in counts.items():
            sp = spans.get(k)
            log(f"hourly: {k} span "
                + (f"{sp[0]} .. {sp[1]}" if sp else "empty"))
            # Clamped at 100. temp_ai legitimately holds hours past
            # the grid window (whole local days, and a forecast tail),
            # so it scored 100.1% for a week - which reads as a bug
            # signal in a field whose whole job is to flag bugs. The
            # surplus is reported separately rather than inflating a
            # percentage that cannot exceed whole.
            pct = min(100.0, 100.0 * n / max(expect, 1))
            per_series[k] = round(pct, 1)
            extra = n - expect
            log(f"hourly: {k} completeness {pct:.1f}% of {expect}h "
                + ("- OK" if pct >= 95 else "- BELOW GATE")
                + (f" (+{extra}h beyond the grid window)"
                   if extra > 0 else ""))
        # the store is usable when the load/generation trio is whole;
        # carbon is an overlay and gates itself in the panel
        core = ("demand_ai", "wind_ai", "solar_ai")
        complete = all(per_series.get(k, 0) >= 95 for k in core)
        log(f"hourly: core trio {'complete' if complete else 'INCOMPLETE'}"
            f"; carbon overlay "
            f"{'available' if per_series.get('co2_ai', 0) >= 95 else 'withheld'}")
        # heat_ready is a SEPARATE gate, deliberately. The temperature
        # series is what the electrification computations need and the
        # existing panels do not; keeping it out of `complete` means
        # temp_ai filling over its first few runs can never withdraw
        # anything already shipping.
        heat_ready = complete and per_series.get("temp_ai", 0) >= 95
        # A third gate. price_ai starts empty and fills a day at a
        # time, so it must never be able to withdraw the heat layer or
        # the grid trio while it climbs.
        price_ready = complete and per_series.get("price_ai", 0) >= 95
        log(f"hourly: heat layer "
            f"{'ready' if heat_ready else 'not ready'} "
            f"(temp_ai {per_series.get('temp_ai', 0)}%)")
        log(f"hourly: price layer "
            f"{'ready' if price_ready else 'not ready'} "
            f"(price_ai {per_series.get('price_ai', 0)}% - fills forward, "
            f"see semopx_history_probe)")
    t0, n_hours, packed = compact_hourly(series)
    log(f"hourly: encoded {sum(len(v) for v in series.values())} values "
        f"as {len(packed)} arrays of {n_hours} from {t0}")
    return {"schema": HOURLY_SCHEMA,
            "t0": t0, "hours": n_hours,
            "heat_ready": heat_ready, "price_ready": price_ready,
            "generated": dt.datetime.now(dt.timezone.utc)
            .strftime("%Y-%m-%dT%H:%M:%SZ"),
            "months": HOURLY_MONTHS, "complete": complete,
            "completeness_pct": per_series,
            "encoding": ("series are flat arrays; position i is the "
                         "hour t0 + i, null where absent"),
            "basis": ("All-island 15-minute EirGrid series aggregated "
                      "to hourly means (>=3 of 4 quarters required). "
                      "Demand, wind and solar in MW; carbon intensity "
                      "in g CO2 per kWh. Source: EirGrid Smart Grid "
                      "Dashboard. temp_ai is population-weighted "
                      "island air temperature in degrees C, ERA5 via "
                      "Open-Meteo, weights as the daily HDD feed. "
                      "EVERY series is keyed on Irish local clock, "
                      "not UTC, so temperature and demand describe "
                      "the same hour year-round."),
            "series": packed}


def regenerate_panel6(path="docs/panel6.html"):
    """THE WORKING COPY'S GENERATOR - the function this file's header
    always claimed existed and did not, during which the copy was
    hand-regenerated four times in one day. It owns the three mutable
    parts of docs/panel6.html and treats the renderer as the static
    template it is:

      1. the VFM_PAYLOAD block, rebuilt from the derive functions
         (keys the derive chain does not produce are preserved, never
         dropped - beccs_mw and beccs_note live only in the page);
      2. the lever widget, injected verbatim from tools/vfm_levers.js
         + tools/panel6_widget.js between marker comments;
      3. the under-construction banner at the top of the body.

    Idempotent by construction: running it twice yields byte-identical
    output, and a test asserts exactly that.
    """
    import re as _re
    import pathlib as _pl
    root = _pl.Path(__file__).resolve().parent.parent
    page = root / path
    src = page.read_text()

    payload_m = _re.search(
        r'(<script>window\.VFM_PAYLOAD = )(\{.*?\})(;</script>)',
        src, _re.S)
    if not payload_m:
        raise RuntimeError("panel6: VFM_PAYLOAD block not found")
    old = json.loads(payload_m.group(2))
    vfm = old.setdefault("derived", {}).setdefault("vfm", {})
    preserved = set(vfm)
    vfm.update({
        "scenario": derive_vfm_scenario(),
        "stages": derive_vfm_stages(),
        "increment": derive_vfm_increment(),
        "carbon": derive_vfm_carbon(),
        "running": derive_vfm_running(),
        "cooling": derive_vfm_cooling(),
        "phased": derive_vfm_phased(),
        "constants": derive_vfm_constants(),
        "tes_cop": VFM_TES_COP,
        "tes_carbon": VFM_TES_CARBON,
        "lrvc": {"years": list(VFM_LRVC_YEARS),
                 "p_kwh": list(VFM_LRVC_P_KWH),
                 "source": VFM_LRVC_SOURCE},
    })
    if not preserved <= set(vfm):
        raise RuntimeError("panel6: a preserved payload key was lost")
    src = (src[:payload_m.start(2)]
           + json.dumps(old, separators=(", ", ": "))
           + src[payload_m.end(2):])

    # BANNER REMOVED 26 Aug 2026 by decision - the appraisal stands
    # without it and the method fold still records what is in the
    # arithmetic and what is not. The markers stay so the generator
    # keeps ownership of the slot: putting a banner back is a one-line
    # change here, not an edit to the shipped page.
    banner = "<!-- panel6-banner --><!-- /panel6-banner -->"
    if "<!-- panel6-banner -->" in src:
        src = _re.sub(r'<!-- panel6-banner -->.*?<!-- /panel6-banner -->',
                      lambda _m: banner, src, flags=_re.S)
    else:
        src = _re.sub(r'(<body[^>]*>)',
                      lambda _m: _m.group(1) + banner, src, count=1)

    widget = ("<!-- panel6-widget -->\n<script>\n"
              + (root / "tools" / "vfm_levers.js").read_text()
              + "\n"
              + (root / "tools" / "panel6_widget.js").read_text()
              + "</script>\n<!-- /panel6-widget -->")
    if "<!-- panel6-widget -->" in src:
        src = _re.sub(r'<!-- panel6-widget -->.*?<!-- /panel6-widget -->',
                      lambda _m: widget, src, flags=_re.S)
    else:
        src = src.replace("</body>", widget + "\n</body>")

    page.write_text(src)

    # THE PUBLIC PAGE carries the same widget (its payload arrives with
    # the daily data, so no payload injection there) - refreshed from
    # the same tools/ files so the two pages cannot diverge.
    idx = root / "docs" / "index.html"
    isrc = idx.read_text()
    if "<!-- panel6-widget -->" in isrc:
        isrc = _re.sub(r'<!-- panel6-widget -->.*?<!-- /panel6-widget -->',
                       lambda _m: widget, isrc, flags=_re.S)
    else:
        isrc = isrc.replace("</body>", widget + "\n</body>")
    idx.write_text(isrc)

    log(f"panel6: regenerated {path} + index widget - payload, "
        f"widget, banner")
    return path


def main():
    global PREVIOUS_FEEDS, PREVIOUS_DERIVED
    if DATA_PATH.exists():
        try:
            prev_doc = json.loads(DATA_PATH.read_text())
            PREVIOUS_FEEDS = prev_doc.get("feeds", {})
            PREVIOUS_DERIVED = prev_doc.get("derived", {})
        except Exception:
            log("warning - previous data.json unreadable, starting clean")

    feeds, failures = {}, []
    for name, fn in FEEDS.items():
        log(f"--- {name}")
        try:
            payload, status = fn()
            payload["status"] = status
            if name in FEED_FLAGS:
                payload["flags"] = FEED_FLAGS[name]
            payload["fetched_utc"] = dt.datetime.now(
                dt.timezone.utc).isoformat(timespec="seconds")
            feeds[name] = payload
            log(f"{name}: {status}, latest_day={payload.get('latest_day')}")
        except Exception as e:
            expected = (isinstance(e, NotImplementedError)
                        or name in EXPECTED_DOWN or name in SOFT_FEEDS)
            log(f"{name}: {'EXPECTED DOWN' if expected else 'FAILED'} - "
                f"{e.__class__.__name__}: {e}")
            if not expected:
                traceback.print_exc()
            prev = PREVIOUS_FEEDS.get(name, {})
            has_prev = bool(prev)
            prev["status"] = "stale"
            if name in EXPECTED_DOWN:
                prev["pending_note"] = EXPECTED_DOWN[name]
            if name in FEED_FLAGS:
                prev["flags"] = FEED_FLAGS[name]
            prev.setdefault("source", "previous run retained")
            feeds[name] = prev
            if not expected:
                # A failed feed with previous data in hand degrades to
                # stale-and-continue (transient 5xx like the CCNI 520 of
                # 27 Jul 2026 must not block the deploy); only a failure
                # with nothing to carry forward stays build-fatal.
                if has_prev:
                    log(f"{name}: previous data carried - degraded to "
                        f"stale, not build-fatal")
                else:
                    failures.append(name)

    gas = feeds.get("gni_ckan", {}).get("ndm_gwh") or {}
    hdd = feeds.get("hdd", {}).get("hdd_roi") or {}
    reg = space_heat_split(gas, hdd)
    derived = {"roi_space_heat_regression": reg} if reg else {}
    if reg:
        log("regression:", reg)
    apply_ie_fx(feeds)
    hero = derive_hero(feeds)
    if hero:
        derived["hero"] = hero
        log("hero:", {k: hero[k] for k in
                      ("week_ending", "heat_purchased_gwh",
                       "indigenous_share_pct", "bill_eur_m", "bill_gbp_m",
                       "emissions_kt_co2")})
    # derive_heat_gap is retained for its breakeven-SPF diagnostics and
    # its unit tests, but it NO LONGER feeds the masthead ticker - see
    # heat_gap_from_cost_series, wired after the cost series below.
    hg = derive_heat_gap(feeds)
    if hg:
        derived["heat_gap_diagnostic"] = hg
    # Carbon for the backfilled weeks. The daily feed keeps 50 days;
    # anything older comes from the hourly store, which is read here
    # BEFORE history so week_context can see it. Daily values win
    # where both exist - the store is the fallback, not the source.
    try:
        prev_store = (json.loads(HOURLY_PATH.read_text())
                      if HOURLY_PATH.exists() else {})
        hci = daily_ci_from_hourly(prev_store)
        if hci:
            eg = feeds.setdefault("eirgrid", {})
            ci = eg.setdefault("co2_intensity_g_per_kwh", {})
            added = sum(1 for d, v in hci.items() if ci.setdefault(d, v) is v)
            log(f"carbon: {len(hci)} daily means available from the "
                f"hourly store, {added} filled gaps in the daily feed "
                f"(span {min(hci)}..{max(hci)})")
    except Exception as exc:
        log(f"carbon: hourly-store fallback unavailable "
            f"({exc.__class__.__name__}) - backfilled weeks will use "
            "the anchor EF")
    derived["history"] = build_history(feeds)
    derived["history_schema"] = HISTORY_SCHEMA
    derived["anchor_epoch"] = ANCHOR_EPOCH
    # Two counters, deliberately. Weeks on record includes the weeks
    # reconstructed behind LIVE_FROM; weeks live counts only those
    # observed as they happened. The 52-live-weeks milestone is the
    # second number and is not reached by backfilling.
    _h = derived["history"] or []
    derived["weeks_on_record"] = len(_h)
    derived["weeks_live"] = sum(1 for e in _h if e.get("live"))
    derived["live_from"] = LIVE_FROM
    # Priced before compaction, while the history is still a list of
    # week objects - the columnar form is a wire format, not a
    # working one.
    try:
        hcs = derive_heat_cost_series(feeds)
        if hcs:
            # Columnar, same wire format and same encoder as the
            # history block. The cost rows are wide and shallow and
            # there are now up to 730 of them - measured at 1,135
            # bytes a row flat, which is 809 kB on top of a data.json
            # already past 600. Encoding is not a restatement: the
            # content schema is untouched and both readers accept
            # either shape, because index.html publishes on the Pages
            # deploy while data.json only changes at 04:17.
            _cflat = len(json.dumps(hcs, separators=(",", ":")))
            derived["heat_cost_series"] = compact_history(hcs)
            _ccols = len(json.dumps(derived["heat_cost_series"],
                                    separators=(",", ":")))
            log(f"heat cost: encoded {len(hcs)} days columnar, "
                f"{_cflat // 1024} kB -> {_ccols // 1024} kB "
                f"({100 * _ccols // max(_cflat, 1)}%)")
            # The masthead ticker, from the same rows the panel draws
            derived["heat_gap"] = heat_gap_from_cost_series(hcs)
            if CALIBRATION:
                derived["calibration"] = dict(CALIBRATION)
            for j, g in derived["heat_gap"].items():
                log(f"heat gap: {j} oil {g['oil_boiler']:.2f} vs network "
                    f"{g['geothermal_spf40']:.2f} per useful kWh - ground "
                    f"wins by {g['gap_now']:.2f} today against a median "
                    f"{g['gap_median']:.2f} over {g['gap_days']} days")
    except Exception as exc:
        log(f"heat cost: failed ({exc.__class__.__name__}) - "
            "the weekly tracker is unaffected")
    try:
        he = derive_heat_emissions(feeds)
        if he:
            derived["heat_emissions"] = he
    except Exception as exc:
        log(f"heat emissions: failed ({exc.__class__.__name__}) - "
            "the rest of the panel is unaffected")
    try:
        derived["cooling_tiers"] = derive_cooling_tiers()
        # Panel 6. Wholly offline - every input is an anchor, a
        # published constant or a cross-panel read, so it renders
        # without a single feed. That is deliberate at this stage: the
        # appraisal is being checked, not tracked.
        derived["vfm"] = {
            "scenario": derive_vfm_scenario(),
            "stages": derive_vfm_stages(),
            "increment": derive_vfm_increment(),
            "carbon": derive_vfm_carbon(),
            "running": derive_vfm_running(),
            "cooling": derive_vfm_cooling(),
            "phased": derive_vfm_phased(),
            "constants": derive_vfm_constants(),
            "tes_cop": VFM_TES_COP,
            "tes_carbon": VFM_TES_CARBON,
            "lrvc": {"years": list(VFM_LRVC_YEARS),
                     "p_kwh": list(VFM_LRVC_P_KWH),
                     "source": VFM_LRVC_SOURCE},
        }
        derived["heat_rejected"] = derive_heat_rejected(
            derived["cooling_tiers"])
    except Exception as exc:
        log(f"cooling tiers: failed ({exc.__class__.__name__})")
    try:
        dd = derive_dispatch_down()
        if dd:
            derived["dispatch_down"] = dd
            ex = derive_odd_examples(dd)
            if ex:
                derived["odd_examples"] = ex
    except Exception as exc:
        log(f"dispatch down: failed ({exc.__class__.__name__}) - "
            "the rest of the panel is unaffected")
    _flat = len(json.dumps(_h, separators=(",", ":")))
    derived["history"] = compact_history(_h)
    derived["history_encoding"] = HISTORY_ENCODING
    _cols = len(json.dumps(derived["history"], separators=(",", ":")))
    log(f"history: encoded {len(_h)} weeks columnar, "
        f"{_flat // 1024} kB -> {_cols // 1024} kB "
        f"({100 * _cols // max(_flat, 1)}%)")
    log(f"history: {derived['weeks_on_record']} weeks on record, "
        f"{derived['weeks_live']} live (from {LIVE_FROM}); "
        f"schema {HISTORY_SCHEMA}, anchor epoch {ANCHOR_EPOCH}")
    gw = sorted(((feeds.get("gni_live") or {}).get("ndm_gwh")
                 or {}))
    derived["gas_window"] = {"from": gw[0], "to": gw[-1],
                             "days": len(gw)} if gw else None
    # len() of the columnar block is its key count, not its week
    # count - it printed "3 complete weeks" for a 52-week record on
    # the first 4.24.0 run. Count the weeks, not the container.
    log(f"history: {derived['weeks_on_record']} complete weeks"
        + (f", gas_window {derived['gas_window']['days']}d"
           if derived["gas_window"] else ""))
    reg = derived.get("roi_space_heat_regression")
    cal = derive_gas_calibration(
        reg, (feeds.get("hdd") or {}).get("hdd_roi") or {})
    if cal:
        derived["gas_calibration"] = cal
        log("gas calibration: implied", cal["implied_annual_space_heat_gwh"],
            "vs anchor", cal["anchor_annual_space_heat_gwh"],
            "ratio", cal["ratio"],
            "(gate 0.90-1.10)" if cal["within_gate"] else
            "OUTSIDE gate 0.90-1.10 - disclosed")
    cool = derive_cool(feeds)
    if cool:
        derived["cool"] = cool
        log("cool: stranded summer share",
            cool["stranded_summer_pct"], "%")
        log("heat_gap: breakeven SPF vs oil - NI",
            hg["ni"]["breakeven_spf_vs_oil"], "ROI",
            hg["roi"]["breakeven_spf_vs_oil"])

    # the frontispiece renderer reads D.derived.frontispiece - it
    # belongs in DERIVED, not at document level beside why_heat,
    # which is a static constant and reads D.why_heat.
    fp = derive_frontispiece(feeds, locals().get("hcs"),
                             derived.get("tightest_hour"),
                             derived.get("heat_emissions"),
                             derived.get("heat_rejected"))
    if fp:
        derived["frontispiece"] = fp

    doc = {
        "pipeline_version": PIPELINE_VERSION,
        "built_utc": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "feeds": feeds,
        "derived": derived,
        "events": EVENTS,
        "geo": {**GEO, "percap": derive_geo_percap(),
                "hardware": derive_geo_hardware(),
                "targets": derive_geo_targets()},
        "why_heat": WHY_HEAT,
        "notes": ("Feed statuses - ok: fetched and current; lagging: fetched, "
                  "source publishes on a lag; stale: fetch failed, previous "
                  "values retained. Judgement figures are current Causeway "
                  "Energies estimates - challenge and input welcome at "
                  "contact@causewaygt.com"),
    }
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(doc, indent=1, sort_keys=True))

    # hourly store - separate file, separate schema, cannot corrupt
    # the weekly tracker if it fails
    try:
        prev_hourly = {}
        if HOURLY_PATH.exists():
            try:
                prev_hourly = json.loads(HOURLY_PATH.read_text())
            except Exception:
                log("hourly: previous store unreadable, rebuilding")
        try:
            semopx_history_probe()
        except Exception as exc:
            log(f"semopx_probe: failed ({exc.__class__.__name__})")
        try:
            semo_dispatch_probe()
        except Exception as exc:
            log(f"semo_probe: failed ({exc.__class__.__name__})")
        store = build_hourly_store(prev_hourly, feeds)
        # B.2.1 runs on the store we just wrote, log-only. Soft: a
        # failure here must never touch the weekly tracker.
        try:
            if store and store.get("heat_ready"):
                derived["tightest_hour"] = derive_tightest_hour(store)
                gv = derive_grid_views(store)
                if gv:
                    derived["grid_views"] = gv
            elif store:
                log("B.2.1 skipped - heat layer not ready "
                    f"(temp_ai {store.get('completeness_pct', {}).get('temp_ai')}%)")
        except Exception as exc:
            log(f"B.2.1 failed ({exc.__class__.__name__}) - log-only, "
                "weekly tracker unaffected")
        if store:
            HOURLY_PATH.write_text(json.dumps(store, separators=(",", ":")))
            log(f"wrote {HOURLY_PATH} "
                f"({HOURLY_PATH.stat().st_size // 1024} kB)")
    except Exception as exc:
        log(f"hourly: store step failed ({exc.__class__.__name__}: "
            f"{exc}) - weekly output unaffected")
    # RE-WRITE. data.json is serialised ABOVE, before the hourly block,
    # so anything the grid layer adds to `derived` - tightest_hour,
    # grid_views - lands after the file is already on disk and never
    # ships. That is why Panel 3 drew its decline messages while the
    # log showed B.2.1 and the falcon computing perfectly: the
    # renderers were right and the payload was empty.
    #
    # The first write stays where it is on purpose: it guarantees a
    # payload lands even if the hourly step throws. This adds the grid
    # layer when there is one, and says so, rather than moving the
    # write and making the weekly tracker depend on the hourly store.
    grid_keys = [k for k in ("tightest_hour", "grid_views")
                 if derived.get(k)]
    if derived.get("tightest_hour"):
        # THE FRONTISPIECE IS BUILT TWICE ON PURPOSE. Its binding-hour
        # figure needs derive_tightest_hour(), which runs in the hourly
        # block BELOW the first build - so the first pass publishes the
        # figure as pending and this one fills it in. Building it once,
        # early, is what shipped "Arrives with the next build" on a
        # panel whose data was sitting in the same payload.
        fp2 = derive_frontispiece(feeds, locals().get("hcs"),
                                  derived["tightest_hour"],
                                  derived.get("heat_emissions"),
                                  derived.get("heat_rejected"))
        if fp2:
            derived["frontispiece"] = fp2
            log("frontispiece: rebuilt with the binding-hour figure")
    if grid_keys:
        DATA_PATH.write_text(json.dumps(doc, indent=1, sort_keys=True))
        log(f"wrote {DATA_PATH} again with the grid layer "
            f"({', '.join(grid_keys)})")
    else:
        log("grid layer absent from the payload - Panel 3 will draw its "
            "decline message")
    log(f"wrote {DATA_PATH} ({DATA_PATH.stat().st_size/1024:.0f} kB)")

    if failures:
        log("hard failures:", failures)
        if NTFY_TOPIC:
            try:
                requests.post(f"https://ntfy.sh/{NTFY_TOPIC}",
                              data=f"ioi-heatsplit build: failed {failures}",
                              timeout=15)
            except Exception:
                pass
        sys.exit(1)


if __name__ == "__main__":
    main()
